from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from hyacinth.excel.contracts import ConversionProgress, ConversionResult, EngineName
from hyacinth.library.import_task import _copy_file, run_import_task
from hyacinth.tasks import TaskRequest
from hyacinth.tasks.worker import TaskCancelled
from hyacinth.versioning import MetadataStore

FIXTURES = Path(__file__).parents[1] / "fixtures"


class RecordingContext:
    def __init__(self) -> None:
        self.progress: list[tuple[float | None, str]] = []
        self.critical_messages: list[str] = []
        self.committed = False
        self.engine: EngineName | None = None

    def report_progress(self, progress: float | None, message: str = "") -> None:
        self.progress.append((progress, message))

    def check_cancelled(self) -> None:
        pass

    def commit(self) -> None:
        self.committed = True

    def set_engine(self, engine: EngineName) -> None:
        self.engine = engine

    @contextmanager
    def critical_section(self, message: str = "") -> Iterator[None]:
        self.critical_messages.append(message)
        yield


def _create_xlsx(path: Path, value: str = "风信子") -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = value
    workbook.save(path)
    workbook.close()


def _read_a1(path: Path) -> object:
    workbook = load_workbook(path, read_only=True)
    try:
        worksheet = workbook.active
        assert worksheet is not None
        return worksheet["A1"].value
    finally:
        workbook.close()


class WritingEngine:
    name = EngineName.PYTHON

    def __init__(self) -> None:
        self.source: Path | None = None

    def convert_xls_to_xlsx(
        self,
        source: Path,
        destination: Path,
        progress: ConversionProgress | None = None,
    ) -> ConversionResult:
        self.source = source
        _create_xlsx(destination, "已转换")
        return ConversionResult(engine=self.name, output_path=destination)


def test_xlsx_import_publishes_original_and_working_copy(tmp_path: Path) -> None:
    source = tmp_path / "销售报表.xlsx"
    library_root = tmp_path / "library"
    _create_xlsx(source)
    request = TaskRequest(
        task_id="import-1",
        name="导入工作簿",
        file_id="file-1",
        engine=None,
        operation="import-workbook",
        payload={"source_path": str(source), "library_root": str(library_root)},
    )
    context = RecordingContext()

    result = run_import_task(request, context)

    original = library_root / "files" / "file-1" / "original" / source.name
    working = library_root / "files" / "file-1" / "working" / "current.xlsx"
    assert result.file_id == "file-1"
    assert result.display_name == source.name
    assert result.original_path == original
    assert result.working_path == working
    assert original.exists()
    assert working.exists()
    assert result.root_version is not None
    assert result.root_version.parent_version_id is None
    assert result.root_version.name == "导入原始文件"
    assert result.root_version.snapshot_path.exists()
    assert result.root_version.snapshot_path != working
    assert (result.root_version.snapshot_path.parent / "manifest.json").exists()
    assert MetadataStore(library_root).list_workbooks() == (result,)
    assert _read_a1(working) == "风信子"
    assert _read_a1(result.root_version.snapshot_path) == "风信子"
    _create_xlsx(working, "工作副本已修改")
    assert _read_a1(result.root_version.snapshot_path) == "风信子"
    assert list((library_root / ".staging").iterdir()) == []
    assert context.committed is True
    assert context.progress[-1] == (1.0, "导入完成")


def test_import_rejects_unsupported_format_before_creating_library(tmp_path: Path) -> None:
    source = tmp_path / "销售报表.csv"
    source.write_text("name,value\n风信子,1", encoding="utf-8")
    library_root = tmp_path / "library"
    request = TaskRequest(
        task_id="import-csv",
        name="导入工作簿",
        file_id="file-csv",
        engine=None,
        operation="import-workbook",
        payload={"source_path": str(source), "library_root": str(library_root)},
    )

    with pytest.raises(ValueError, match="只支持 .xls 和 .xlsx"):
        run_import_task(request, RecordingContext())

    assert not library_root.exists()


def test_cancelled_import_removes_staging_directory(tmp_path: Path) -> None:
    source = tmp_path / "销售报表.xlsx"
    library_root = tmp_path / "library"
    _create_xlsx(source)
    request = TaskRequest(
        task_id="import-cancelled",
        name="导入工作簿",
        file_id="file-cancelled",
        engine=None,
        operation="import-workbook",
        payload={"source_path": str(source), "library_root": str(library_root)},
    )

    class CancelledContext(RecordingContext):
        def check_cancelled(self) -> None:
            raise TaskCancelled

    with pytest.raises(TaskCancelled):
        run_import_task(request, CancelledContext())

    assert not (library_root / "files" / "file-cancelled").exists()
    assert list((library_root / ".staging").iterdir()) == []


def test_xls_import_converts_copied_original_to_working_xlsx(tmp_path: Path) -> None:
    source = FIXTURES / "legacy-fidelity.xls"
    library_root = tmp_path / "library"
    request = TaskRequest(
        task_id="import-xls",
        name="导入工作簿",
        file_id="file-xls",
        engine=None,
        operation="import-workbook",
        payload={"source_path": str(source), "library_root": str(library_root)},
    )
    context = RecordingContext()
    engine = WritingEngine()

    result = run_import_task(request, context, select_engine=lambda: engine)

    original = library_root / "files" / "file-xls" / "original" / source.name
    assert engine.source is not None
    assert engine.source.name == source.name
    assert ".staging" in engine.source.parts
    assert original.read_bytes() == source.read_bytes()
    assert _read_a1(result.working_path) == "已转换"
    assert context.engine is EngineName.PYTHON


def test_file_copy_checks_cancellation_between_chunks(tmp_path: Path) -> None:
    source = tmp_path / "large.xlsx"
    destination = tmp_path / "copied.xlsx"
    source.write_bytes(b"x" * (3 * 1024 * 1024))

    class CancelDuringCopy(RecordingContext):
        def __init__(self) -> None:
            super().__init__()
            self.checks = 0

        def check_cancelled(self) -> None:
            self.checks += 1
            if self.checks == 2:
                raise TaskCancelled

    context = CancelDuringCopy()

    with pytest.raises(TaskCancelled):
        _copy_file(source, destination, context)

    assert 0 < destination.stat().st_size < source.stat().st_size


def test_metadata_failure_preserves_published_directory_for_recovery(tmp_path: Path) -> None:
    source = tmp_path / "销售报表.xlsx"
    library_root = tmp_path / "library"
    _create_xlsx(source)
    request = TaskRequest(
        task_id="import-metadata-failure",
        name="导入工作簿",
        file_id="file-metadata-failure",
        engine=None,
        operation="import-workbook",
        payload={"source_path": str(source), "library_root": str(library_root)},
    )

    class FailingStore:
        def record_import(self, _record: object) -> None:
            raise OSError("数据库写入失败")

    with pytest.raises(OSError, match="数据库写入失败"):
        run_import_task(
            request,
            RecordingContext(),
            metadata_store_factory=lambda _root: FailingStore(),
        )

    published_directory = library_root / "files" / request.file_id
    assert published_directory.is_dir()
    assert MetadataStore(library_root).reconcile_manifests() == 1
    assert MetadataStore(library_root).list_workbooks()[0].file_id == request.file_id
