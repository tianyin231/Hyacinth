import hashlib
import pickle
import time
from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from hyacinth.excel.contracts import EngineName
from hyacinth.processing import (
    SORT_PREVIEW_OPERATION,
    SortDirection,
    SortKey,
    SortPreviewResult,
    run_sort_preview_task,
    sort_preview_handlers,
    sort_preview_task,
)
from hyacinth.tasks import TaskEvent, TaskQueue, TaskRequest, TaskState
from hyacinth.tasks.worker import TaskCancelled


class SortPreviewTaskContext:
    def __init__(self) -> None:
        self.messages: list[str] = []
        self.committed = False
        self.engine: EngineName | None = None

    def report_progress(self, progress: float | None, message: str = "") -> None:
        self.messages.append(message)

    def check_cancelled(self) -> None:
        return

    def set_engine(self, engine: EngineName) -> None:
        self.engine = engine

    def commit(self) -> None:
        self.committed = True

    @contextmanager
    def critical_section(self, message: str = "") -> Iterator[None]:
        self.messages.append(message)
        yield


def _sortable_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "销售"
    sheet.append(["名称", "数量", "类别"])
    sheet.append(["apple", 30, "水果"])
    sheet.append(["banana", None, "水果"])
    sheet.append(["cherry", 10, "水果"])
    sheet.append(["durian", 20, "水果"])
    sheet.append(["fig", None, "水果"])
    sheet.append(["grape", 10, "水果"])
    workbook.save(path)
    workbook.close()


def _sort_request(
    source: Path,
    preview: Path,
    sort_keys: list[dict[str, object]],
    *,
    sheet_name: str = "销售",
    task_id: str = "sort-preview-1",
) -> TaskRequest:
    return TaskRequest(
        task_id=task_id,
        name="排序预览",
        file_id="file-1",
        engine=None,
        operation=SORT_PREVIEW_OPERATION,
        payload={
            "source_path": str(source),
            "preview_path": str(preview),
            "parent_version_id": "version-root-1",
            "sheet_name": sheet_name,
            "sort_keys": sort_keys,
        },
    )


def _rows(path: Path, sheet_name: str = "销售") -> list[tuple[object, ...]]:
    workbook = load_workbook(path)
    try:
        sheet = workbook[sheet_name]
        return list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def _temporary_path(preview: Path) -> Path:
    return preview.with_name(f".{preview.stem}.tmp.xlsx")


def _assert_no_temporary_files(preview: Path) -> None:
    assert not _temporary_path(preview).exists()
    assert not tuple(preview.parent.glob("*.tmp.xlsx"))


def test_sort_preview_handler_is_registered() -> None:
    assert sort_preview_handlers() == {SORT_PREVIEW_OPERATION: sort_preview_task}


def test_two_key_sort_directions_and_empty_at_bottom(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _sortable_workbook(source)
    preview = tmp_path / "preview.xlsx"
    request = _sort_request(
        source,
        preview,
        sort_keys=[
            {"column_index": 1, "direction": "desc"},
            {"column_index": 0, "direction": "asc"},
        ],
    )
    context = SortPreviewTaskContext()

    result = run_sort_preview_task(request, context)

    assert context.committed is True
    assert context.engine is EngineName.PYTHON
    assert preview.is_file()
    assert result.engine is EngineName.PYTHON
    assert result.sheet_name == "销售"
    assert result.data_rows == 6
    assert _rows(preview) == [
        ("名称", "数量", "类别"),
        ("apple", 30, "水果"),
        ("durian", 20, "水果"),
        ("cherry", 10, "水果"),
        ("grape", 10, "水果"),
        ("banana", None, "水果"),
        ("fig", None, "水果"),
    ]
    _assert_no_temporary_files(preview)


def test_ascending_sort_keeps_empty_values_at_bottom(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _sortable_workbook(source)
    preview = tmp_path / "preview.xlsx"
    request = _sort_request(
        source,
        preview,
        sort_keys=[{"column_index": 1, "direction": "asc"}],
    )

    run_sort_preview_task(request, SortPreviewTaskContext())

    assert _rows(preview) == [
        ("名称", "数量", "类别"),
        ("cherry", 10, "水果"),
        ("grape", 10, "水果"),
        ("durian", 20, "水果"),
        ("apple", 30, "水果"),
        ("banana", None, "水果"),
        ("fig", None, "水果"),
    ]


def test_sort_is_stable_for_equal_keys(tmp_path: Path) -> None:
    source = tmp_path / "stable.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "销售"
    sheet.append(["序号", "数量"])
    sheet.append(["A", 10])
    sheet.append(["B", 10])
    sheet.append(["C", 10])
    workbook.save(source)
    workbook.close()
    preview = tmp_path / "preview.xlsx"

    run_sort_preview_task(
        _sort_request(
            source,
            preview,
            sort_keys=[{"column_index": 1, "direction": "desc"}],
        ),
        SortPreviewTaskContext(),
    )

    assert _rows(preview) == [
        ("序号", "数量"),
        ("A", 10),
        ("B", 10),
        ("C", 10),
    ]


def test_sort_preview_moves_style_comment_hyperlink_with_record(
    tmp_path: Path,
) -> None:
    source = tmp_path / "styled.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "销售"
    sheet.append(["名称", "数量"])
    sheet.append(["alpha", 2])
    sheet.append(["beta", 1])
    fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    sheet["A2"].fill = fill
    sheet["B2"].number_format = "0.00%"
    sheet["A2"].comment = Comment("备注", "作者")
    sheet["B2"].hyperlink = "https://example.com/alpha"
    workbook.save(source)
    workbook.close()
    preview = tmp_path / "preview.xlsx"

    result = run_sort_preview_task(
        _sort_request(
            source,
            preview,
            sort_keys=[{"column_index": 1, "direction": "asc"}],
        ),
        SortPreviewTaskContext(),
    )

    assert result.data_rows == 2
    workbook = load_workbook(preview)
    try:
        sheet = workbook["销售"]
        assert sheet["A1"].value == "名称"
        assert sheet["A2"].value == "beta"
        assert sheet["B2"].value == 1
        assert sheet["B2"].number_format == "General"
        assert sheet["A2"].comment is None
        assert sheet["A3"].value == "alpha"
        assert sheet["B3"].value == 2
        assert sheet["A3"].fill.start_color.rgb == "FFFF0000"
        assert sheet["A3"].fill.patternType == "solid"
        assert sheet["B3"].number_format == "0.00%"
        assert sheet["A3"].comment is not None
        assert sheet["A3"].comment.text == "备注"
        assert sheet["B3"].hyperlink is not None
        assert sheet["B3"].hyperlink.target == "https://example.com/alpha"
        assert sheet["B3"].hyperlink.ref == "B3"
    finally:
        workbook.close()


def test_formula_cell_rejects_sort_preview(tmp_path: Path) -> None:
    source = tmp_path / "formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "销售"
    sheet.append(["名称", "数量"])
    sheet.append(["apple", "=1+1"])
    workbook.save(source)
    workbook.close()
    preview = tmp_path / "preview.xlsx"

    with pytest.raises(ValueError, match="公式"):
        run_sort_preview_task(
            _sort_request(
                source,
                preview,
                sort_keys=[{"column_index": 1, "direction": "asc"}],
            ),
            SortPreviewTaskContext(),
        )

    assert not preview.exists()
    _assert_no_temporary_files(preview)


def test_merged_cell_rejects_sort_preview(tmp_path: Path) -> None:
    source = tmp_path / "merged.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "销售"
    sheet.append(["名称", "数量"])
    sheet.append(["apple", 30])
    sheet.append(["banana", 20])
    sheet.merge_cells("A2:A3")
    workbook.save(source)
    workbook.close()
    preview = tmp_path / "preview.xlsx"

    with pytest.raises(ValueError, match="合并单元格"):
        run_sort_preview_task(
            _sort_request(
                source,
                preview,
                sort_keys=[{"column_index": 1, "direction": "asc"}],
            ),
            SortPreviewTaskContext(),
        )

    assert not preview.exists()


def test_mixed_key_type_rejects_sort_preview(tmp_path: Path) -> None:
    source = tmp_path / "mixed.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "销售"
    sheet.append(["名称", "数量"])
    sheet.append(["apple", 30])
    sheet.append(["banana", "三十"])
    workbook.save(source)
    workbook.close()
    preview = tmp_path / "preview.xlsx"

    with pytest.raises(ValueError, match="混合"):
        run_sort_preview_task(
            _sort_request(
                source,
                preview,
                sort_keys=[{"column_index": 1, "direction": "asc"}],
            ),
            SortPreviewTaskContext(),
        )

    assert not preview.exists()


def test_cancelled_sort_preview_leaves_no_result(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _sortable_workbook(source)
    preview = tmp_path / "preview.xlsx"
    request = _sort_request(
        source,
        preview,
        sort_keys=[{"column_index": 1, "direction": "asc"}],
    )

    class CancelledContext(SortPreviewTaskContext):
        def __init__(self) -> None:
            super().__init__()
            self.checks = 0

        def check_cancelled(self) -> None:
            self.checks += 1
            if self.checks >= 2:
                raise TaskCancelled

    with pytest.raises(TaskCancelled):
        run_sort_preview_task(request, CancelledContext())

    assert not preview.exists()
    _assert_no_temporary_files(preview)


def test_failed_sort_preview_leaves_no_result(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _sortable_workbook(source)
    preview = tmp_path / "preview.xlsx"
    request = _sort_request(
        source,
        preview,
        sort_keys=[{"column_index": 1, "direction": "asc"}],
        sheet_name="不存在",
    )

    with pytest.raises(ValueError, match="找不到工作表"):
        run_sort_preview_task(request, SortPreviewTaskContext())

    assert not preview.exists()
    _assert_no_temporary_files(preview)


def test_result_hash_matches_published_preview(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _sortable_workbook(source)
    preview = tmp_path / "preview.xlsx"
    request = _sort_request(
        source,
        preview,
        sort_keys=[
            {"column_index": 1, "direction": "desc"},
            {"column_index": 0, "direction": "asc"},
        ],
    )

    result = run_sort_preview_task(request, SortPreviewTaskContext())

    assert result.content_hash == hashlib.sha256(preview.read_bytes()).hexdigest()
    workbook = load_workbook(preview, read_only=True)
    try:
        assert workbook.sheetnames == ["销售"]
    finally:
        workbook.close()


def test_sort_preview_result_is_picklable(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _sortable_workbook(source)
    preview = tmp_path / "preview.xlsx"
    request = _sort_request(
        source,
        preview,
        sort_keys=[{"column_index": 1, "direction": "desc"}],
    )

    result = run_sort_preview_task(request, SortPreviewTaskContext())

    restored = pickle.loads(pickle.dumps(result))
    assert restored == result
    assert restored.preview_path == result.preview_path
    assert restored.source_path == result.source_path
    assert restored.parent_version_id == "version-root-1"
    assert restored.sort_keys == (SortKey(1, SortDirection.DESCENDING),)
    assert restored.engine is EngineName.PYTHON
    assert restored.content_hash == result.content_hash


def test_rejects_missing_or_invalid_sort_keys(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _sortable_workbook(source)
    preview = tmp_path / "preview.xlsx"

    cases: list[tuple[list[dict[str, object]] | None, str]] = [
        (None, "sort_keys"),
        (
            [
                {"column_index": 1, "direction": "asc"},
                {"column_index": 0, "direction": "asc"},
                {"column_index": 2, "direction": "asc"},
            ],
            "最多",
        ),
        (
            [
                {"column_index": 1, "direction": "asc"},
                {"column_index": 1, "direction": "desc"},
            ],
            "重复",
        ),
        ([{"column_index": 1, "direction": "sideways"}], "asc 或 desc"),
        ([{"column_index": -1, "direction": "asc"}], "大于等于 0"),
        ([{"column_index": "1", "direction": "asc"}], "整数"),
    ]
    for sort_keys, message in cases:
        payload: dict[str, object] = {
            "source_path": str(source),
            "preview_path": str(preview),
            "parent_version_id": "version-root-1",
            "sheet_name": "销售",
        }
        if sort_keys is not None:
            payload["sort_keys"] = sort_keys
        request = TaskRequest(
            task_id="sort-preview-invalid",
            name="排序预览",
            file_id="file-1",
            engine=None,
            operation=SORT_PREVIEW_OPERATION,
            payload=payload,
        )
        with pytest.raises(ValueError, match=message):
            run_sort_preview_task(request, SortPreviewTaskContext())
        assert not preview.exists()


def test_rejects_sort_key_column_out_of_range(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _sortable_workbook(source)
    preview = tmp_path / "preview.xlsx"

    with pytest.raises(ValueError, match="超出工作表范围"):
        run_sort_preview_task(
            _sort_request(
                source,
                preview,
                sort_keys=[{"column_index": 5, "direction": "asc"}],
            ),
            SortPreviewTaskContext(),
        )

    assert not preview.exists()


def test_task_queue_runs_sort_preview(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _sortable_workbook(source)
    preview = tmp_path / "preview.xlsx"
    queue = TaskQueue(sort_preview_handlers())
    try:
        queue.submit(
            _sort_request(
                source,
                preview,
                sort_keys=[
                    {"column_index": 1, "direction": "desc"},
                    {"column_index": 0, "direction": "asc"},
                ],
                task_id="queued-sort-preview",
            )
        )
        events: list[TaskEvent] = []
        deadline = time.monotonic() + 10
        while time.monotonic() < deadline:
            events.extend(queue.poll_events())
            if any(event.state in {TaskState.SUCCEEDED, TaskState.FAILED} for event in events):
                break
            time.sleep(0.01)

        succeeded = [event for event in events if event.state is TaskState.SUCCEEDED]
        terminal_evidence = [
            (event.state, event.message)
            for event in events
            if event.state in {TaskState.SUCCEEDED, TaskState.FAILED}
        ]
        assert len(succeeded) == 1, terminal_evidence
        assert isinstance(succeeded[0].result, SortPreviewResult)
        assert succeeded[0].result.engine is EngineName.PYTHON
        assert preview.is_file()
        assert _rows(preview)[1][0] == "apple"
    finally:
        assert queue.shutdown(timeout=5.0) is True
