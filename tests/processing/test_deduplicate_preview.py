import pickle
import time
from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from hyacinth.excel.contracts import EngineName
from hyacinth.processing import (
    DEDUPLICATE_PREVIEW_OPERATION,
    DeduplicatePreviewResult,
    DuplicateGroup,
    KeepDuplicate,
    deduplicate_preview_handlers,
    deduplicate_preview_task,
    run_deduplicate_preview_task,
)
from hyacinth.tasks import TaskEvent, TaskQueue, TaskRequest, TaskState
from hyacinth.tasks.worker import TaskCancelled


class RecordingContext:
    def __init__(self) -> None:
        self.messages: list[str] = []
        self.committed = False
        self.engine: EngineName | None = None

    def report_progress(self, progress: float | None, message: str = "") -> None:
        self.messages.append(message)

    def check_cancelled(self) -> None:
        return

    def set_engine(self, engine: EngineName) -> None:
        self.engine = engine

    def commit(self) -> None:
        self.committed = True

    @contextmanager
    def critical_section(self, message: str = "") -> Iterator[None]:
        self.messages.append(message)
        yield


def _create_workbook(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "销售"
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def _request(
    source: Path,
    preview: Path,
    *,
    key_columns: list[int] | None = None,
    keep: str = "first",
    ignore_case: bool = False,
    trim_whitespace: bool = False,
    task_id: str = "deduplicate-1",
) -> TaskRequest:
    return TaskRequest(
        task_id=task_id,
        name="删除重复行预览",
        file_id="file-1",
        engine=None,
        operation=DEDUPLICATE_PREVIEW_OPERATION,
        payload={
            "source_path": str(source),
            "preview_path": str(preview),
            "parent_version_id": "version-root-1",
            "sheet_name": "销售",
            "key_columns": key_columns or [],
            "keep": keep,
            "ignore_case": ignore_case,
            "trim_whitespace": trim_whitespace,
        },
    )


def _rows(path: Path) -> list[tuple[object, ...]]:
    workbook = load_workbook(path)
    try:
        return list(workbook["销售"].iter_rows(values_only=True))
    finally:
        workbook.close()


def test_deduplicate_preview_handler_is_registered() -> None:
    assert deduplicate_preview_handlers() == {
        DEDUPLICATE_PREVIEW_OPERATION: deduplicate_preview_task
    }


def test_all_columns_keep_first_and_empty_values_participate(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(
        source,
        [
            ["名称", "类别"],
            ["apple", "水果"],
            ["apple", "水果"],
            [None, "待定"],
            ["", "待定"],
            ["apple", "零食"],
        ],
    )
    context = RecordingContext()

    result = run_deduplicate_preview_task(_request(source, preview), context)

    assert context.committed is True
    assert context.engine is EngineName.PYTHON
    assert result.key_columns == ()
    assert result.keep is KeepDuplicate.FIRST
    assert result.deleted_rows == 2
    assert result.data_rows == 5
    assert result.duplicate_groups == (
        DuplicateGroup(kept_row=2, deleted_rows=(3,)),
        DuplicateGroup(kept_row=4, deleted_rows=(5,)),
    )
    assert _rows(preview) == [
        ("名称", "类别"),
        ("apple", "水果"),
        (None, "待定"),
        ("apple", "零食"),
    ]


def test_selected_key_keeps_last_and_normalizes_text_without_changing_it(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(
        source,
        [
            ["名称", "序号"],
            [" Alice ", 1],
            ["alice", 2],
            ["ALICE", 3],
            ["Bob", 4],
        ],
    )

    result = run_deduplicate_preview_task(
        _request(
            source,
            preview,
            key_columns=[0],
            keep="last",
            ignore_case=True,
            trim_whitespace=True,
        ),
        RecordingContext(),
    )

    assert result.duplicate_groups == (DuplicateGroup(kept_row=4, deleted_rows=(2, 3)),)
    assert _rows(preview) == [("名称", "序号"), ("ALICE", 3), ("Bob", 4)]


def test_key_comparison_is_type_sensitive(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(
        source,
        [["键", "序号"], [1, "数字"], ["1", "文本"], [1.0, "重复数字"]],
    )

    result = run_deduplicate_preview_task(
        _request(source, preview, key_columns=[0]),
        RecordingContext(),
    )

    assert result.duplicate_groups == (DuplicateGroup(kept_row=2, deleted_rows=(4,)),)
    assert _rows(preview) == [("键", "序号"), (1, "数字"), ("1", "文本")]


def test_retained_row_keeps_style_comment_and_hyperlink(tmp_path: Path) -> None:
    source = tmp_path / "styled.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(source, [["键", "值"], ["A", 1], ["A", 2], ["B", 3]])
    workbook = load_workbook(source)
    sheet = workbook["销售"]
    sheet["A3"].fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
    sheet["A3"].comment = Comment("保留", "测试")
    sheet["B3"].hyperlink = "https://example.com/kept"
    workbook.save(source)
    workbook.close()

    run_deduplicate_preview_task(
        _request(source, preview, key_columns=[0], keep="last"),
        RecordingContext(),
    )

    workbook = load_workbook(preview)
    try:
        kept = workbook["销售"]
        assert kept["A2"].value == "A"
        assert kept["B2"].value == 2
        assert kept["A2"].fill.fgColor.rgb == "FFFF0000"
        assert kept["A2"].comment is not None and kept["A2"].comment.text == "保留"
        assert kept["B2"].hyperlink is not None
        assert kept["B2"].hyperlink.target == "https://example.com/kept"
        assert kept["B2"].hyperlink.ref == "B2"
    finally:
        workbook.close()


def test_no_duplicates_does_not_publish_preview(tmp_path: Path) -> None:
    source = tmp_path / "unique.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(source, [["名称"], ["apple"], ["banana"]])

    with pytest.raises(ValueError, match="未发现重复行"):
        run_deduplicate_preview_task(_request(source, preview), RecordingContext())

    assert not preview.exists()
    assert not tuple(tmp_path.glob("*.tmp.xlsx"))


@pytest.mark.parametrize("unsafe", ["formula", "merged"])
def test_unsafe_structural_content_is_rejected(tmp_path: Path, unsafe: str) -> None:
    source = tmp_path / f"{unsafe}.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(source, [["名称", "值"], ["A", 1], ["A", 1]])
    workbook = load_workbook(source)
    sheet = workbook["销售"]
    if unsafe == "formula":
        sheet["B2"] = "=1+1"
    else:
        sheet.merge_cells("A2:A3")
    workbook.save(source)
    workbook.close()

    with pytest.raises(ValueError, match="公式|合并单元格"):
        run_deduplicate_preview_task(
            _request(source, preview, key_columns=[0]),
            RecordingContext(),
        )

    assert not preview.exists()


def test_invalid_parameters_are_rejected(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(source, [["名称"], ["A"], ["A"]])

    with pytest.raises(ValueError, match="超出工作表范围"):
        run_deduplicate_preview_task(
            _request(source, preview, key_columns=[2]),
            RecordingContext(),
        )
    with pytest.raises(ValueError, match="first 或 last"):
        run_deduplicate_preview_task(
            _request(source, preview, keep="middle"),
            RecordingContext(),
        )


def test_cancelled_task_removes_temporary_output(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(source, [["名称"], ["A"], ["A"]])

    class CancellingContext(RecordingContext):
        def __init__(self) -> None:
            super().__init__()
            self.checks = 0

        def check_cancelled(self) -> None:
            self.checks += 1
            if self.checks >= 2:
                raise TaskCancelled

    with pytest.raises(TaskCancelled):
        run_deduplicate_preview_task(_request(source, preview), CancellingContext())

    assert not preview.exists()
    assert not tuple(tmp_path.glob("*.tmp.xlsx"))


def test_worker_returns_picklable_deduplicate_result(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(source, [["名称"], ["A"], ["A"]])
    queue = TaskQueue(deduplicate_preview_handlers())
    try:
        queue.submit(_request(source, preview))
        events: list[TaskEvent] = []
        deadline = time.monotonic() + 10
        while time.monotonic() < deadline:
            events.extend(queue.poll_events())
            if any(event.state in {TaskState.SUCCEEDED, TaskState.FAILED} for event in events):
                break
            time.sleep(0.01)

        succeeded = [event for event in events if event.state is TaskState.SUCCEEDED]
        assert len(succeeded) == 1, [(event.state, event.message) for event in events]
        result = succeeded[0].result
        assert isinstance(result, DeduplicatePreviewResult)
        assert pickle.loads(pickle.dumps(result)) == result
    finally:
        assert queue.shutdown(timeout=5.0) is True
