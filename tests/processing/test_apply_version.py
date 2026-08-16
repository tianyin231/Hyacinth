import hashlib
import json
import time
from collections.abc import Iterator
from contextlib import contextmanager
from datetime import UTC, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from hyacinth.excel.contracts import EngineName
from hyacinth.processing import (
    APPLY_DEDUPLICATE_PREVIEW_OPERATION,
    APPLY_DELETE_BLANK_ROWS_PREVIEW_OPERATION,
    APPLY_FILTER_PREVIEW_OPERATION,
    APPLY_FIND_REPLACE_PREVIEW_OPERATION,
    APPLY_SORT_PREVIEW_OPERATION,
    APPLY_TRIM_PREVIEW_OPERATION,
    SAVE_MANUAL_EDITS_OPERATION,
    apply_deduplicate_preview_task,
    apply_delete_blank_rows_preview_task,
    apply_filter_preview_task,
    apply_find_replace_preview_task,
    apply_sort_preview_task,
    apply_trim_preview_task,
    apply_version_handlers,
    run_apply_deduplicate_preview_task,
    run_apply_delete_blank_rows_preview_task,
    run_apply_filter_preview_task,
    run_apply_sort_preview_task,
    run_save_manual_edits_task,
    save_manual_edits_task,
)
from hyacinth.tasks import TaskEvent, TaskQueue, TaskRequest, TaskState
from hyacinth.versioning import ImportedWorkbook, MetadataStore, VersionRecord


class RecordingContext:
    def __init__(self) -> None:
        self.committed = False
        self.engine: EngineName | None = None
        self.messages: list[str] = []

    def report_progress(self, progress: float | None, message: str = "") -> None:
        self.messages.append(message)

    def check_cancelled(self) -> None:
        return

    def set_engine(self, engine: EngineName) -> None:
        self.engine = engine

    def commit(self) -> None:
        self.committed = True

    @contextmanager
    def critical_section(self, message: str = "") -> Iterator[None]:
        self.messages.append(message)
        yield


def _create_xlsx(path: Path, rows: list[list[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "销售"
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def _value(path: Path, cell: str) -> object:
    workbook = load_workbook(path, read_only=True)
    try:
        sheet = workbook["销售"]
        return sheet[cell].value
    finally:
        workbook.close()


def _seed_library(root: Path) -> ImportedWorkbook:
    directory = root / "files/file-1"
    original = directory / "original/销售.xlsx"
    working = directory / "working/current.xlsx"
    snapshot = directory / "versions/version-1/snapshot.xlsx"
    rows: list[list[object]] = [["名称", "数量"], ["apple", 2], ["banana", 1]]
    _create_xlsx(original, rows)
    _create_xlsx(working, rows)
    _create_xlsx(snapshot, rows)
    version = VersionRecord(
        "version-1",
        "file-1",
        None,
        "导入原始文件",
        datetime(2026, 8, 15, 8, 0, tzinfo=UTC),
        "import",
        None,
        snapshot,
        hashlib.sha256(snapshot.read_bytes()).hexdigest(),
    )
    record = ImportedWorkbook("file-1", "销售.xlsx", original, working, version, version.created_at)
    MetadataStore(root).record_import(record)
    return record


def _request(root: Path, preview: Path) -> TaskRequest:
    return TaskRequest(
        task_id="apply-1",
        name="应用排序结果",
        file_id="file-1",
        engine=None,
        operation=APPLY_SORT_PREVIEW_OPERATION,
        payload={
            "library_root": str(root),
            "preview_path": str(preview),
            "preview_hash": hashlib.sha256(preview.read_bytes()).hexdigest(),
            "parent_version_id": "version-1",
            "version_id": "version-2",
            "sheet_name": "销售",
            "sort_keys": [{"column_index": 1, "direction": "asc"}],
        },
    )


def _deduplicate_request(root: Path, preview: Path) -> TaskRequest:
    return TaskRequest(
        task_id="apply-deduplicate-1",
        name="应用删除重复行结果",
        file_id="file-1",
        engine=None,
        operation=APPLY_DEDUPLICATE_PREVIEW_OPERATION,
        payload={
            "library_root": str(root),
            "preview_path": str(preview),
            "preview_hash": hashlib.sha256(preview.read_bytes()).hexdigest(),
            "parent_version_id": "version-1",
            "version_id": "version-2",
            "sheet_name": "销售",
            "key_columns": [0],
            "keep": "first",
            "ignore_case": True,
            "trim_whitespace": True,
            "duplicate_groups": 1,
            "deleted_rows": 1,
        },
    )


def _delete_blank_rows_request(root: Path, preview: Path) -> TaskRequest:
    return TaskRequest(
        task_id="apply-delete-blank-rows-1",
        name="应用删除空白行结果",
        file_id="file-1",
        engine=None,
        operation=APPLY_DELETE_BLANK_ROWS_PREVIEW_OPERATION,
        payload={
            "library_root": str(root),
            "preview_path": str(preview),
            "preview_hash": hashlib.sha256(preview.read_bytes()).hexdigest(),
            "parent_version_id": "version-1",
            "version_id": "version-2",
            "sheet_name": "销售",
            "key_columns": [0],
            "allow_unsafe": False,
            "compatibility_warning": False,
            "deleted_row_numbers": [3],
        },
    )


def _filter_request(root: Path, preview: Path) -> TaskRequest:
    return TaskRequest(
        task_id="apply-filter-1",
        name="应用条件筛选结果",
        file_id="file-1",
        engine=None,
        operation=APPLY_FILTER_PREVIEW_OPERATION,
        payload={
            "library_root": str(root),
            "preview_path": str(preview),
            "preview_hash": hashlib.sha256(preview.read_bytes()).hexdigest(),
            "parent_version_id": "version-1",
            "version_id": "version-2",
            "sheet_name": "销售",
            "conditions": [
                {
                    "column_index": 1,
                    "operator": "greater_than",
                    "value_type": "number",
                    "value": "1",
                    "second_value": None,
                }
            ],
            "connector": "and",
            "matched_rows": 1,
            "total_rows": 2,
        },
    )


def test_apply_handler_is_registered() -> None:
    assert apply_version_handlers() == {
        APPLY_SORT_PREVIEW_OPERATION: apply_sort_preview_task,
        APPLY_DEDUPLICATE_PREVIEW_OPERATION: apply_deduplicate_preview_task,
        APPLY_DELETE_BLANK_ROWS_PREVIEW_OPERATION: apply_delete_blank_rows_preview_task,
        APPLY_FILTER_PREVIEW_OPERATION: apply_filter_preview_task,
        APPLY_TRIM_PREVIEW_OPERATION: apply_trim_preview_task,
        APPLY_FIND_REPLACE_PREVIEW_OPERATION: apply_find_replace_preview_task,
        SAVE_MANUAL_EDITS_OPERATION: save_manual_edits_task,
    }


def test_manual_edits_create_child_snapshot_and_preserve_parent(tmp_path: Path) -> None:
    root = tmp_path / "library"
    parent = _seed_library(root)
    parent_version = parent.head_version
    assert parent_version is not None
    request = TaskRequest(
        task_id="manual-1",
        name="保存手动编辑",
        file_id=parent.file_id,
        engine=None,
        operation=SAVE_MANUAL_EDITS_OPERATION,
        payload={
            "library_root": str(root),
            "parent_version_id": parent_version.version_id,
            "version_id": "version-2",
            "edits": [
                {"sheet_name": "销售", "row": 1, "column": 0, "value": "pear"},
                {"sheet_name": "销售", "row": 1, "column": 1, "value": "42"},
                {"sheet_name": "销售", "row": 2, "column": 1, "value": "=40+2"},
            ],
        },
    )

    result = run_save_manual_edits_task(request, RecordingContext())

    child = result.head_version
    assert child is not None
    assert child.parent_version_id == parent_version.version_id
    assert child.operation == "manual-edit"
    assert child.name == "手动编辑"
    assert _value(child.snapshot_path, "A2") == "pear"
    assert _value(child.snapshot_path, "B2") == 42
    assert _value(child.snapshot_path, "B3") == "=40+2"
    assert _value(parent_version.snapshot_path, "A2") == "apple"
    assert result.working_path.read_bytes() == child.snapshot_path.read_bytes()
    assert MetadataStore(root).get_workbook(parent.file_id).head_version == child


def test_manual_edits_reject_empty_edit_list(tmp_path: Path) -> None:
    root = tmp_path / "library"
    parent = _seed_library(root)
    parent_version = parent.head_version
    assert parent_version is not None
    request = TaskRequest(
        task_id="manual-1",
        name="保存手动编辑",
        file_id=parent.file_id,
        engine=None,
        operation=SAVE_MANUAL_EDITS_OPERATION,
        payload={
            "library_root": str(root),
            "parent_version_id": parent_version.version_id,
            "version_id": "version-2",
            "edits": [],
        },
    )

    with pytest.raises(ValueError, match="edits"):
        run_save_manual_edits_task(request, RecordingContext())


def test_apply_promotes_exact_preview_to_child_snapshot_and_head(tmp_path: Path) -> None:
    root = tmp_path / "library"
    parent = _seed_library(root)
    preview = root / "files/file-1/.previews/preview-1/result.xlsx"
    _create_xlsx(preview, [["名称", "数量"], ["banana", 1], ["apple", 2]])
    preview_bytes = preview.read_bytes()
    context = RecordingContext()

    result = run_apply_sort_preview_task(_request(root, preview), context)

    child = result.head_version
    assert child is not None
    parent_version = parent.head_version
    assert parent_version is not None
    assert child.parent_version_id == parent_version.version_id
    assert child.snapshot_path.read_bytes() == preview_bytes
    assert result.working_path.read_bytes() == preview_bytes
    assert _value(result.working_path, "A2") == "banana"
    assert context.committed is True
    assert context.engine is EngineName.PYTHON
    store = MetadataStore(root)
    assert store.get_workbook("file-1") == result
    assert store.list_versions("file-1") == (parent_version, child)
    assert json.loads(child.parameters_json) == {
        "sheet_name": "销售",
        "sort_keys": [{"column_index": 1, "direction": "asc"}],
    }


def test_apply_deduplicate_preview_records_operation_and_statistics(tmp_path: Path) -> None:
    root = tmp_path / "library"
    _seed_library(root)
    preview = root / "files/file-1/.previews/preview-1/result.xlsx"
    _create_xlsx(preview, [["名称", "数量"], ["apple", 2]])
    preview_bytes = preview.read_bytes()

    result = run_apply_deduplicate_preview_task(
        _deduplicate_request(root, preview),
        RecordingContext(),
    )

    child = result.head_version
    assert child is not None
    assert child.name == "删除重复行"
    assert child.operation == "delete-duplicates"
    assert child.snapshot_path.read_bytes() == preview_bytes
    assert result.working_path.read_bytes() == preview_bytes
    assert json.loads(child.parameters_json) == {
        "sheet_name": "销售",
        "key_columns": [0],
        "keep": "first",
        "ignore_case": True,
        "trim_whitespace": True,
        "duplicate_groups": 1,
        "deleted_rows": 1,
    }


def test_apply_delete_blank_rows_records_rows_and_compatibility_mode(tmp_path: Path) -> None:
    root = tmp_path / "library"
    _seed_library(root)
    preview = root / "files/file-1/.previews/preview-1/result.xlsx"
    _create_xlsx(preview, [["名称", "数量"], ["apple", 2]])
    preview_bytes = preview.read_bytes()

    result = run_apply_delete_blank_rows_preview_task(
        _delete_blank_rows_request(root, preview),
        RecordingContext(),
    )

    child = result.head_version
    assert child is not None
    assert child.name == "删除空白行"
    assert child.operation == "delete-blank-rows"
    assert child.snapshot_path.read_bytes() == preview_bytes
    assert result.working_path.read_bytes() == preview_bytes
    assert json.loads(child.parameters_json) == {
        "sheet_name": "销售",
        "key_columns": [0],
        "allow_unsafe": False,
        "compatibility_warning": False,
        "deleted_row_numbers": [3],
        "deleted_rows": 1,
    }


def test_apply_filter_preview_records_conditions_and_statistics(tmp_path: Path) -> None:
    root = tmp_path / "library"
    _seed_library(root)
    preview = root / "files/file-1/.previews/preview-1/result.xlsx"
    _create_xlsx(preview, [["名称", "数量"], ["apple", 2], ["banana", 1]])
    workbook = load_workbook(preview)
    workbook["销售"].row_dimensions[3].hidden = True
    workbook.save(preview)
    workbook.close()
    preview_bytes = preview.read_bytes()

    result = run_apply_filter_preview_task(_filter_request(root, preview), RecordingContext())

    child = result.head_version
    assert child is not None
    assert child.name == "条件筛选"
    assert child.operation == "filter"
    assert child.snapshot_path.read_bytes() == preview_bytes
    assert result.working_path.read_bytes() == preview_bytes
    assert json.loads(child.parameters_json) == {
        "sheet_name": "销售",
        "conditions": [
            {
                "column_index": 1,
                "operator": "greater_than",
                "value_type": "number",
                "value": "1",
                "second_value": None,
            }
        ],
        "connector": "and",
        "matched_rows": 1,
        "total_rows": 2,
    }


def test_apply_rejects_unchanged_preview_without_child(tmp_path: Path) -> None:
    root = tmp_path / "library"
    parent = _seed_library(root)
    assert parent.head_version is not None
    preview = root / "files/file-1/.previews/preview-1/result.xlsx"
    preview.parent.mkdir(parents=True)
    preview.write_bytes(parent.head_version.snapshot_path.read_bytes())

    with pytest.raises(ValueError, match="没有变化"):
        run_apply_sort_preview_task(_request(root, preview), RecordingContext())

    assert MetadataStore(root).list_versions("file-1") == (parent.head_version,)
    assert not (root / "files/file-1/versions/version-2").exists()


def test_metadata_failure_preserves_child_manifest_for_recovery(tmp_path: Path) -> None:
    root = tmp_path / "library"
    parent = _seed_library(root)
    preview = root / "files/file-1/.previews/preview-1/result.xlsx"
    _create_xlsx(preview, [["名称", "数量"], ["banana", 1], ["apple", 2]])
    parent_working = parent.working_path.read_bytes()
    store = MetadataStore(root)

    class FailingStore:
        def get_workbook(self, file_id: str) -> ImportedWorkbook:
            return store.get_workbook(file_id)

        def record_child_version(
            self,
            version: VersionRecord,
            expected_parent_version_id: str,
        ) -> None:
            raise OSError("数据库写入失败")

    with pytest.raises(OSError, match="数据库写入失败"):
        run_apply_sort_preview_task(
            _request(root, preview),
            RecordingContext(),
            metadata_store_factory=lambda _root: FailingStore(),
        )

    child_directory = root / "files/file-1/versions/version-2"
    assert (child_directory / "snapshot.xlsx").is_file()
    assert (child_directory / "manifest.json").is_file()
    assert parent.working_path.read_bytes() == parent_working
    parent_version = parent.head_version
    assert parent_version is not None
    assert store.get_workbook("file-1").head_version == parent_version

    assert store.reconcile_manifests() == 1
    recovered_head = store.get_workbook("file-1").head_version
    assert recovered_head is not None and recovered_head.version_id == "version-2"
    assert parent.working_path.read_bytes() == preview.read_bytes()


def test_task_queue_applies_preview_in_worker_process(tmp_path: Path) -> None:
    root = tmp_path / "library"
    _seed_library(root)
    preview = root / "files/file-1/.previews/preview-1/result.xlsx"
    _create_xlsx(preview, [["名称", "数量"], ["banana", 1], ["apple", 2]])
    queue = TaskQueue(apply_version_handlers())
    try:
        queue.submit(_request(root, preview))
        events: list[TaskEvent] = []
        deadline = time.monotonic() + 10
        while time.monotonic() < deadline:
            events.extend(queue.poll_events())
            if any(event.state in {TaskState.SUCCEEDED, TaskState.FAILED} for event in events):
                break
            time.sleep(0.01)

        succeeded = [event for event in events if event.state is TaskState.SUCCEEDED]
        assert len(succeeded) == 1, [(event.state, event.message) for event in events]
        assert isinstance(succeeded[0].result, ImportedWorkbook)
        assert succeeded[0].result.head_version is not None
        assert succeeded[0].result.head_version.version_id == "version-2"
    finally:
        assert queue.shutdown(timeout=5.0) is True


def test_task_queue_saves_manual_edits_in_worker_process(tmp_path: Path) -> None:
    root = tmp_path / "library"
    parent = _seed_library(root)
    parent_version = parent.head_version
    assert parent_version is not None
    request = TaskRequest(
        task_id="manual-worker-1",
        name="保存手动编辑",
        file_id=parent.file_id,
        engine=None,
        operation=SAVE_MANUAL_EDITS_OPERATION,
        payload={
            "library_root": str(root),
            "parent_version_id": parent_version.version_id,
            "version_id": "version-2",
            "edits": [{"sheet_name": "销售", "row": 1, "column": 0, "value": "pear"}],
        },
    )
    queue = TaskQueue(apply_version_handlers())
    try:
        queue.submit(request)
        events: list[TaskEvent] = []
        deadline = time.monotonic() + 10
        while time.monotonic() < deadline:
            events.extend(queue.poll_events())
            if any(event.state in {TaskState.SUCCEEDED, TaskState.FAILED} for event in events):
                break
            time.sleep(0.01)

        succeeded = [event for event in events if event.state is TaskState.SUCCEEDED]
        assert len(succeeded) == 1, [(event.state, event.message) for event in events]
        assert isinstance(succeeded[0].result, ImportedWorkbook)
        assert _value(succeeded[0].result.working_path, "A2") == "pear"
    finally:
        assert queue.shutdown(timeout=5.0) is True
