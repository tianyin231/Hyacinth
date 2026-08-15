import pickle
import time
from collections.abc import Iterator
from contextlib import contextmanager
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table

from hyacinth.excel.contracts import EngineName
from hyacinth.processing import (
    FILTER_PREVIEW_OPERATION,
    FilterConnector,
    FilterPreviewResult,
    filter_preview_handlers,
    filter_preview_task,
    run_filter_preview_task,
)
from hyacinth.tasks import TaskEvent, TaskQueue, TaskRequest, TaskState
from hyacinth.tasks.worker import TaskCancelled


class RecordingContext:
    def __init__(self) -> None:
        self.messages: list[str] = []
        self.committed = False
        self.engine: EngineName | None = None

    def report_progress(self, progress: float | None, message: str = "") -> None:
        self.messages.append(message)

    def check_cancelled(self) -> None:
        return

    def set_engine(self, engine: EngineName) -> None:
        self.engine = engine

    def commit(self) -> None:
        self.committed = True

    @contextmanager
    def critical_section(self, message: str = "") -> Iterator[None]:
        self.messages.append(message)
        yield


def _create_workbook(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "销售"
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def _condition(
    column_index: int,
    operator: str,
    value_type: str = "text",
    value: str | None = None,
    second_value: str | None = None,
) -> dict[str, object]:
    return {
        "column_index": column_index,
        "operator": operator,
        "value_type": value_type,
        "value": value,
        "second_value": second_value,
    }


def _request(
    source: Path,
    preview: Path,
    conditions: list[dict[str, object]],
    *,
    connector: str = "and",
    task_id: str = "filter-1",
) -> TaskRequest:
    return TaskRequest(
        task_id=task_id,
        name="条件筛选预览",
        file_id="file-1",
        engine=None,
        operation=FILTER_PREVIEW_OPERATION,
        payload={
            "source_path": str(source),
            "preview_path": str(preview),
            "parent_version_id": "version-root-1",
            "sheet_name": "销售",
            "conditions": conditions,
            "connector": connector,
        },
    )


def _hidden_rows(path: Path) -> tuple[int, ...]:
    workbook = load_workbook(path)
    try:
        sheet = workbook["销售"]
        return tuple(
            row for row, dimension in sheet.row_dimensions.items() if dimension.hidden is True
        )
    finally:
        workbook.close()


def test_filter_preview_handler_is_registered() -> None:
    assert filter_preview_handlers() == {FILTER_PREVIEW_OPERATION: filter_preview_task}


def test_text_contains_hides_nonmatches_and_writes_native_filter(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(
        source,
        [["名称"], ["Red Apple"], ["banana"], ["APPLE pie"]],
    )
    context = RecordingContext()

    result = run_filter_preview_task(
        _request(source, preview, [_condition(0, "contains", value="apple")]),
        context,
    )

    assert result.matched_rows == 2
    assert result.total_rows == 3
    assert result.hidden_row_numbers == (3,)
    assert result.match_ratio == pytest.approx(2 / 3)
    assert result.connector is FilterConnector.AND
    assert context.committed is True
    assert context.engine is EngineName.PYTHON
    assert _hidden_rows(preview) == (3,)
    workbook = load_workbook(preview)
    try:
        sheet = workbook["销售"]
        assert sheet.auto_filter.ref == "A1:A4"
        custom_filter = sheet.auto_filter.filterColumn[0].customFilters
        assert custom_filter is not None
        assert custom_filter.customFilter[0].val == "*apple*"
    finally:
        workbook.close()


def test_cross_column_conditions_use_and_and_typed_number_comparison(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(
        source,
        [
            ["类别", "数量"],
            ["水果", 2],
            ["水果", 5],
            ["蔬菜", 8],
        ],
    )

    result = run_filter_preview_task(
        _request(
            source,
            preview,
            [
                _condition(0, "equal", value="水果"),
                _condition(1, "greater_than", "number", "3"),
            ],
        ),
        RecordingContext(),
    )

    assert result.matched_rows == 1
    assert result.hidden_row_numbers == (2, 4)
    workbook = load_workbook(preview)
    try:
        assert len(workbook["销售"].auto_filter.filterColumn) == 2
    finally:
        workbook.close()


def test_same_column_two_conditions_support_or(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(source, [["名称"], ["apple"], ["banana"], ["pear"]])

    result = run_filter_preview_task(
        _request(
            source,
            preview,
            [
                _condition(0, "equal", value="apple"),
                _condition(0, "equal", value="banana"),
            ],
            connector="or",
        ),
        RecordingContext(),
    )

    assert result.matched_rows == 2
    assert result.hidden_row_numbers == (4,)
    workbook = load_workbook(preview)
    try:
        custom_filters = workbook["销售"].auto_filter.filterColumn[0].customFilters
        assert custom_filters is not None
        assert custom_filters._and is False
        assert len(custom_filters.customFilter) == 2
    finally:
        workbook.close()


def test_date_between_is_inclusive_and_serialized_as_native_range(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(
        source,
        [["日期"], [date(2026, 8, 1)], [date(2026, 8, 15)], [date(2026, 9, 1)]],
    )

    result = run_filter_preview_task(
        _request(
            source,
            preview,
            [_condition(0, "between", "date", "2026-08-01", "2026-08-31")],
        ),
        RecordingContext(),
    )

    assert result.matched_rows == 2
    assert result.hidden_row_numbers == (4,)
    workbook = load_workbook(preview)
    try:
        custom_filters = workbook["销售"].auto_filter.filterColumn[0].customFilters
        assert custom_filters is not None and custom_filters._and is True
        assert [item.operator for item in custom_filters.customFilter] == [
            "greaterThanOrEqual",
            "lessThanOrEqual",
        ]
    finally:
        workbook.close()


def test_blank_filter_distinguishes_empty_from_whitespace(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(source, [["值"], [None], [""], [" "], ["数据"]])

    result = run_filter_preview_task(
        _request(source, preview, [_condition(0, "blank")]),
        RecordingContext(),
    )

    assert result.matched_rows == 2
    assert result.hidden_row_numbers == (4, 5)


def test_cross_column_or_is_rejected(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _create_workbook(source, [["A", "B"], ["x", "y"]])

    with pytest.raises(ValueError, match="同一列.*或者"):
        run_filter_preview_task(
            _request(
                source,
                tmp_path / "preview.xlsx",
                [
                    _condition(0, "equal", value="x"),
                    _condition(1, "equal", value="y"),
                ],
                connector="or",
            ),
            RecordingContext(),
        )


@pytest.mark.parametrize("unsafe", ["formula", "hidden", "filter", "table", "merged"])
def test_unsafe_existing_structures_are_rejected(tmp_path: Path, unsafe: str) -> None:
    source = tmp_path / f"{unsafe}.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(source, [["名称"], ["A"], ["B"]])
    workbook = load_workbook(source)
    sheet = workbook["销售"]
    if unsafe == "formula":
        sheet["A2"] = '=IF(TRUE,"A","B")'
    elif unsafe == "hidden":
        sheet.row_dimensions[2].hidden = True
    elif unsafe == "filter":
        sheet.auto_filter.ref = "A1:A3"
    elif unsafe == "table":
        sheet.add_table(Table(displayName="Table1", ref="A1:A3"))
    else:
        sheet.merge_cells("A2:A3")
    workbook.save(source)
    workbook.close()

    with pytest.raises(ValueError):
        run_filter_preview_task(
            _request(source, preview, [_condition(0, "equal", value="A")]),
            RecordingContext(),
        )

    assert not preview.exists()


def test_invalid_typed_values_and_same_column_between_combination_are_rejected(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    _create_workbook(source, [["值"], [1]])

    with pytest.raises(ValueError, match="无效的数字"):
        run_filter_preview_task(
            _request(source, tmp_path / "number.xlsx", [_condition(0, "equal", "number", "x")]),
            RecordingContext(),
        )
    with pytest.raises(ValueError, match="介于.*第二个"):
        run_filter_preview_task(
            _request(
                source,
                tmp_path / "between.xlsx",
                [
                    _condition(0, "between", "number", "1", "2"),
                    _condition(0, "equal", "number", "3"),
                ],
            ),
            RecordingContext(),
        )


def test_cancelled_task_removes_temporary_output(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(source, [["值"], ["A"], ["B"]])

    class CancellingContext(RecordingContext):
        def __init__(self) -> None:
            super().__init__()
            self.checks = 0

        def check_cancelled(self) -> None:
            self.checks += 1
            if self.checks >= 2:
                raise TaskCancelled

    with pytest.raises(TaskCancelled):
        run_filter_preview_task(
            _request(source, preview, [_condition(0, "equal", value="A")]),
            CancellingContext(),
        )

    assert not preview.exists()
    assert not tuple(tmp_path.glob("*.tmp.xlsx"))


def test_worker_returns_picklable_filter_result(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(source, [["值"], ["A"], ["B"]])
    queue = TaskQueue(filter_preview_handlers())
    try:
        queue.submit(_request(source, preview, [_condition(0, "equal", value="A")]))
        events: list[TaskEvent] = []
        deadline = time.monotonic() + 10
        while time.monotonic() < deadline:
            events.extend(queue.poll_events())
            if any(event.state in {TaskState.SUCCEEDED, TaskState.FAILED} for event in events):
                break
            time.sleep(0.01)

        succeeded = [event for event in events if event.state is TaskState.SUCCEEDED]
        assert len(succeeded) == 1, [(event.state, event.message) for event in events]
        result = succeeded[0].result
        assert isinstance(result, FilterPreviewResult)
        assert pickle.loads(pickle.dumps(result)) == result
    finally:
        assert queue.shutdown(timeout=5.0) is True
