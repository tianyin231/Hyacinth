from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path

import pytest
from openpyxl import Workbook

from hyacinth.excel.contracts import EngineName
from hyacinth.processing import (
    FIND_REPLACE_PREVIEW_OPERATION,
    FindReplaceMode,
    run_find_replace_preview_task,
)
from hyacinth.tasks import TaskRequest


class FindContext:
    def report_progress(self, progress: float | None, message: str = "") -> None:
        return

    def check_cancelled(self) -> None:
        return

    def set_engine(self, engine: EngineName) -> None:
        return

    def commit(self) -> None:
        return

    @contextmanager
    def critical_section(self, message: str = "") -> Iterator[None]:
        yield


def _seed(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "数据"
    sheet.append(["名称", "数量"])
    sheet.append(["Apple 销售", 3])
    sheet.append(["apple pie", 5])
    sheet.append(["Banana", 2])
    other = workbook.create_sheet("汇总")
    other["A1"] = "apple turnover"
    other["A2"] = '=A1&" apple "'
    workbook.save(path)
    workbook.close()


def _request(source: Path, preview: Path, **overrides: object) -> TaskRequest:
    payload: dict[str, object] = {
        "source_path": str(source),
        "preview_path": str(preview),
        "parent_version_id": "parent-1",
        "sheet_name": "数据",
        "all_sheets": False,
        "mode": "values",
        "find_text": "apple",
        "replace_text": "橙子",
        "match_case": False,
        "whole_cell": False,
        "trim_whitespace": False,
        "replace_all": True,
    }
    payload.update(overrides)
    return TaskRequest(
        task_id="task-1",
        name="查找与替换",
        file_id="file-1",
        engine=None,
        operation=FIND_REPLACE_PREVIEW_OPERATION,
        payload=payload,
    )


def test_replace_all_values_in_single_sheet(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _seed(source)

    result = run_find_replace_preview_task(_request(source, preview), FindContext())

    assert result.replaced == 2
    assert result.preview_path == preview
    from openpyxl import load_workbook

    workbook = load_workbook(preview)
    try:
        assert workbook["数据"]["A2"].value == "橙子 销售"
        assert workbook["数据"]["A3"].value == "橙子 pie"
        assert workbook["数据"]["A4"].value == "Banana"
        assert workbook["汇总"]["A1"].value == "apple turnover"
    finally:
        workbook.close()


def test_replace_all_sheets_and_case_sensitive(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _seed(source)
    request = _request(source, preview, all_sheets=True, match_case=True)

    result = run_find_replace_preview_task(request, FindContext())

    assert result.sheets == ("数据", "汇总")
    assert result.replaced == 2
    assert any(change.after == "橙子 turnover" for change in result.changes)
    assert all(change.before != "Apple 销售" for change in result.changes)


def test_whole_cell_match_replaces_entire_value(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _seed(source)
    request = _request(source, preview, whole_cell=True, find_text="apple pie")

    result = run_find_replace_preview_task(request, FindContext())

    assert result.replaced == 1
    assert result.changes[0].after == "橙子"


def test_find_only_returns_matches_without_preview_file(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _seed(source)
    request = _request(source, preview, replace_all=False)

    result = run_find_replace_preview_task(request, FindContext())

    assert result.preview_path is None
    assert result.replaced == 0
    assert len(result.changes) == 2
    assert not preview.exists()


def test_formula_mode_replaces_formula_text(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _seed(source)
    request = _request(
        source,
        preview,
        all_sheets=True,
        mode=FindReplaceMode.FORMULAS.value,
        find_text="apple",
    )

    result = run_find_replace_preview_task(request, FindContext())

    assert result.mode is FindReplaceMode.FORMULAS
    assert result.replaced == 1
    assert result.changes[0].after == '=A1&" 橙子 "'


def test_no_match_raises_and_empty_find_rejected(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _seed(source)
    with pytest.raises(ValueError, match="未找到匹配内容"):
        run_find_replace_preview_task(
            _request(source, tmp_path / "p.xlsx", find_text="不存在的词"), FindContext()
        )
    with pytest.raises(ValueError, match="查找内容不能为空"):
        run_find_replace_preview_task(
            _request(source, tmp_path / "p.xlsx", find_text=""), FindContext()
        )
