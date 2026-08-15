import pickle
import time
from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from hyacinth.excel.contracts import EngineName
from hyacinth.processing import (
    DELETE_BLANK_ROWS_PREVIEW_OPERATION,
    DeleteBlankRowsPreviewResult,
    delete_blank_rows_preview_handlers,
    delete_blank_rows_preview_task,
    run_delete_blank_rows_preview_task,
)
from hyacinth.tasks import TaskEvent, TaskQueue, TaskRequest, TaskState
from hyacinth.tasks.worker import TaskCancelled


class RecordingContext:
    def __init__(self) -> None:
        self.messages: list[str] = []
        self.committed = False
        self.engine: EngineName | None = None

    def report_progress(self, progress: float | None, message: str = "") -> None:
        self.messages.append(message)

    def check_cancelled(self) -> None:
        return

    def set_engine(self, engine: EngineName) -> None:
        self.engine = engine

    def commit(self) -> None:
        self.committed = True

    @contextmanager
    def critical_section(self, message: str = "") -> Iterator[None]:
        self.messages.append(message)
        yield


def _create_workbook(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "销售"
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def _request(
    source: Path,
    preview: Path,
    *,
    key_columns: list[int] | None = None,
    allow_unsafe: bool = False,
    task_id: str = "delete-blank-rows-1",
) -> TaskRequest:
    return TaskRequest(
        task_id=task_id,
        name="删除空白行预览",
        file_id="file-1",
        engine=None,
        operation=DELETE_BLANK_ROWS_PREVIEW_OPERATION,
        payload={
            "source_path": str(source),
            "preview_path": str(preview),
            "parent_version_id": "version-root-1",
            "sheet_name": "销售",
            "key_columns": key_columns or [],
            "allow_unsafe": allow_unsafe,
        },
    )


def _rows(path: Path) -> list[tuple[object, ...]]:
    workbook = load_workbook(path)
    try:
        return list(workbook["销售"].iter_rows(values_only=True))
    finally:
        workbook.close()


def test_delete_blank_rows_handler_is_registered() -> None:
    assert delete_blank_rows_preview_handlers() == {
        DELETE_BLANK_ROWS_PREVIEW_OPERATION: delete_blank_rows_preview_task
    }


def test_default_removes_only_fully_blank_rows_and_whitespace_text(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(
        source,
        [
            ["名称", "数量"],
            ["apple", 1],
            [None, None],
            [" \t\n", "　"],
            ["banana", 0],
            [None, False],
        ],
    )
    context = RecordingContext()

    result = run_delete_blank_rows_preview_task(_request(source, preview), context)

    assert context.committed is True
    assert context.engine is EngineName.PYTHON
    assert result.deleted_row_numbers == (3, 4)
    assert result.deleted_rows == 2
    assert result.data_rows == 5
    assert result.compatibility_warning is False
    assert _rows(preview) == [
        ("名称", "数量"),
        ("apple", 1),
        ("banana", 0),
        (None, False),
    ]


def test_selected_key_columns_can_delete_row_with_other_values(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(
        source,
        [
            ["客户", "日期", "备注"],
            [None, " ", "只看关键列时删除"],
            [None, "2026-08-15", "保留"],
            ["客户甲", None, "保留"],
        ],
    )

    result = run_delete_blank_rows_preview_task(
        _request(source, preview, key_columns=[0, 1]),
        RecordingContext(),
    )

    assert result.key_columns == (0, 1)
    assert result.deleted_row_numbers == (2,)
    assert _rows(preview) == [
        ("客户", "日期", "备注"),
        (None, "2026-08-15", "保留"),
        ("客户甲", None, "保留"),
    ]


def test_formula_is_not_blank_and_requires_explicit_compatibility_mode(
    tmp_path: Path,
) -> None:
    source = tmp_path / "formula.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(source, [["值"], [None], ["placeholder"], ["结束"]])
    workbook = load_workbook(source)
    workbook["销售"]["A3"] = '=""'
    workbook.save(source)
    workbook.close()

    with pytest.raises(ValueError, match="公式.*兼容模式"):
        run_delete_blank_rows_preview_task(_request(source, preview), RecordingContext())

    result = run_delete_blank_rows_preview_task(
        _request(source, preview, allow_unsafe=True),
        RecordingContext(),
    )

    assert result.deleted_row_numbers == (2,)
    assert result.compatibility_warning is True
    assert _rows(preview) == [("值",), ('=""',), ("结束",)]


def test_merged_cell_crossing_target_row_is_always_rejected(tmp_path: Path) -> None:
    source = tmp_path / "merged.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(source, [["名称"], [None], [None], ["结束"]])
    workbook = load_workbook(source)
    workbook["销售"].merge_cells("A2:A3")
    workbook.save(source)
    workbook.close()

    with pytest.raises(ValueError, match="合并单元格.*跨越待删除行"):
        run_delete_blank_rows_preview_task(
            _request(source, preview, allow_unsafe=True),
            RecordingContext(),
        )

    assert not preview.exists()


def test_style_only_rows_outside_data_region_are_not_selected(tmp_path: Path) -> None:
    source = tmp_path / "styled-tail.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(source, [["名称"], ["A"], [None], ["B"]])
    workbook = load_workbook(source)
    workbook["销售"]["A10"].fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
    workbook.save(source)
    workbook.close()

    result = run_delete_blank_rows_preview_task(_request(source, preview), RecordingContext())

    assert result.deleted_row_numbers == (3,)
    assert result.data_rows == 3
    workbook = load_workbook(preview)
    try:
        sheet = workbook["销售"]
        assert sheet["A2"].value == "A"
        assert sheet["A3"].value == "B"
        assert any(
            sheet.cell(row, 1).fill.fgColor.rgb == "FFFF0000" for row in range(4, sheet.max_row + 1)
        )
    finally:
        workbook.close()


def test_no_blank_rows_does_not_publish_preview(tmp_path: Path) -> None:
    source = tmp_path / "full.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(source, [["名称"], ["A"], ["B"]])

    with pytest.raises(ValueError, match="未发现空白行"):
        run_delete_blank_rows_preview_task(_request(source, preview), RecordingContext())

    assert not preview.exists()
    assert not tuple(tmp_path.glob("*.tmp.xlsx"))


def test_invalid_key_column_is_rejected(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(source, [["名称"], [None], ["A"]])

    with pytest.raises(ValueError, match="超出工作表范围"):
        run_delete_blank_rows_preview_task(
            _request(source, preview, key_columns=[2]),
            RecordingContext(),
        )


def test_cancelled_task_removes_temporary_output(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(source, [["名称"], [None], ["A"]])

    class CancellingContext(RecordingContext):
        def __init__(self) -> None:
            super().__init__()
            self.checks = 0

        def check_cancelled(self) -> None:
            self.checks += 1
            if self.checks >= 2:
                raise TaskCancelled

    with pytest.raises(TaskCancelled):
        run_delete_blank_rows_preview_task(_request(source, preview), CancellingContext())

    assert not preview.exists()
    assert not tuple(tmp_path.glob("*.tmp.xlsx"))


def test_worker_returns_picklable_blank_rows_result(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    preview = tmp_path / "preview.xlsx"
    _create_workbook(source, [["名称"], [None], ["A"]])
    queue = TaskQueue(delete_blank_rows_preview_handlers())
    try:
        queue.submit(_request(source, preview))
        events: list[TaskEvent] = []
        deadline = time.monotonic() + 10
        while time.monotonic() < deadline:
            events.extend(queue.poll_events())
            if any(event.state in {TaskState.SUCCEEDED, TaskState.FAILED} for event in events):
                break
            time.sleep(0.01)

        succeeded = [event for event in events if event.state is TaskState.SUCCEEDED]
        assert len(succeeded) == 1, [(event.state, event.message) for event in events]
        result = succeeded[0].result
        assert isinstance(result, DeleteBlankRowsPreviewResult)
        assert pickle.loads(pickle.dumps(result)) == result
    finally:
        assert queue.shutdown(timeout=5.0) is True
