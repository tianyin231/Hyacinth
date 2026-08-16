"""表格入口（右键菜单/功能区条）触发处理功能的端到端测试。"""

from collections.abc import Iterator
from contextlib import contextmanager
from datetime import UTC, datetime
from hashlib import sha256
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from PySide6.QtCore import Qt
from PySide6.QtWidgets import QCheckBox, QComboBox, QFrame, QLabel, QPushButton, QTableView
from pytestqt.qtbot import QtBot

from hyacinth.app import HyacinthMainWindow
from hyacinth.excel.contracts import EngineName
from hyacinth.library import ImportedWorkbook
from hyacinth.preview import BUILD_PREVIEW_INDEX_OPERATION, run_preview_index_task
from hyacinth.processing import (
    APPLY_FIND_REPLACE_PREVIEW_OPERATION,
    APPLY_TRIM_PREVIEW_OPERATION,
    DEDUPLICATE_PREVIEW_OPERATION,
    FIND_REPLACE_PREVIEW_OPERATION,
    SORT_PREVIEW_OPERATION,
    TRIM_PREVIEW_OPERATION,
    run_apply_find_replace_preview_task,
    run_apply_trim_preview_task,
    run_deduplicate_preview_task,
    run_find_replace_preview_task,
    run_trim_preview_task,
)
from hyacinth.tasks import TaskEvent, TaskRequest, TaskState
from hyacinth.versioning import MetadataStore, VersionRecord


class FakeApplicationTaskQueue:
    def __init__(self, events: list[Any]) -> None:
        self._events = events
        self.submitted: list[TaskRequest] = []

    def submit(self, request: TaskRequest) -> None:
        self.submitted.append(request)

    def push_event(self, event: Any) -> None:
        self._events.append(event)

    def poll_events(self) -> tuple[Any, ...]:
        events = tuple(self._events)
        self._events.clear()
        return events

    def cancel(self, task_id: str) -> bool:
        return True

    def shutdown(self, timeout: float = 1.0) -> bool:
        return True


class PreviewTaskContext:
    def report_progress(self, progress: float | None, message: str = "") -> None:
        return

    def check_cancelled(self) -> None:
        return

    def set_engine(self, engine: EngineName) -> None:
        return

    def commit(self) -> None:
        return

    @contextmanager
    def critical_section(self, message: str = "") -> Iterator[None]:
        yield


def _child(parent: object, child_type: type, name: str) -> Any:
    from PySide6.QtCore import QObject

    child = QObject.findChild(parent, child_type, name)  # type: ignore[arg-type]
    assert child is not None
    return child


def _preview_request_of(queue: FakeApplicationTaskQueue) -> TaskRequest:
    return next(
        request for request in queue.submitted if request.operation == BUILD_PREVIEW_INDEX_OPERATION
    )


def _last_request(queue: FakeApplicationTaskQueue, operation: str) -> TaskRequest:
    return [request for request in queue.submitted if request.operation == operation][-1]


def _push_success(
    queue: FakeApplicationTaskQueue,
    request: TaskRequest,
    result: object,
    engine: EngineName | None = None,
) -> None:
    queue.push_event(
        TaskEvent(
            request.task_id,
            TaskState.SUCCEEDED,
            request.name,
            request.file_id,
            engine,
            result=result,
        )
    )


def _seed_file(library_root: Path, rows: list[list[object]]) -> None:
    directory = library_root / "files/file-1"
    original = directory / "original/数据.xlsx"
    working = directory / "working/current.xlsx"
    snapshot = directory / "versions/version-1/snapshot.xlsx"
    for path in (original, working, snapshot):
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "数据"
        for row in rows:
            sheet.append(row)
        workbook.save(path)
        workbook.close()
    version = VersionRecord(
        "version-1",
        "file-1",
        None,
        "导入原始文件",
        datetime(2026, 8, 16, 8, 0, tzinfo=UTC),
        "import",
        None,
        snapshot,
        sha256(snapshot.read_bytes()).hexdigest(),
    )
    record = ImportedWorkbook(
        "file-1",
        "数据.xlsx",
        original,
        working,
        version,
        datetime(2026, 8, 16, 8, 0, tzinfo=UTC),
    )
    MetadataStore(library_root).record_import(record)


def _ready_window(
    qtbot: QtBot, tmp_path: Path, rows: list[list[object]]
) -> tuple[Path, FakeApplicationTaskQueue, HyacinthMainWindow]:
    from hyacinth.app import create_main_window

    library_root = tmp_path / "library"
    _seed_file(library_root, rows)
    task_queue = FakeApplicationTaskQueue([])
    window = create_main_window(task_queue=task_queue, library_root=library_root)
    qtbot.addWidget(window)
    window.show()
    preview_request = _preview_request_of(task_queue)
    preview = run_preview_index_task(preview_request, PreviewTaskContext())
    _push_success(task_queue, preview_request, preview)
    table = _child(window, QTableView, "preview-table")
    qtbot.waitUntil(lambda: table.model() is not None, timeout=2000)
    return library_root, task_queue, window


def _head_record(library_root: Path) -> VersionRecord | None:
    return MetadataStore(library_root).get_workbook("file-1").head_version


def _wait_second_preview_request(qtbot: QtBot, queue: FakeApplicationTaskQueue) -> None:
    _wait_preview_request_count(qtbot, queue, 2)


def _wait_preview_request_count(qtbot: QtBot, queue: FakeApplicationTaskQueue, count: int) -> None:
    qtbot.waitUntil(
        lambda: (
            len(
                [
                    request
                    for request in queue.submitted
                    if request.operation == BUILD_PREVIEW_INDEX_OPERATION
                ]
            )
            >= count
        ),
        timeout=2000,
    )


def test_trim_from_table_context_menu_creates_version(qtbot: QtBot, tmp_path: Path) -> None:
    library_root, task_queue, window = _ready_window(
        qtbot,
        tmp_path,
        [["名称", "数量"], ["  苹果  ", 3], [" 香蕉 ", 2]],
    )

    # 表格右键菜单“清除首尾空格…”先在功能区条下方展开参数行并预填选中列
    window._one_step_processing("trim", [0])
    params_bar = _child(window, QFrame, "processing-params-bar")
    qtbot.waitUntil(lambda: params_bar.isVisibleTo(window), timeout=500)
    assert "关键列 A" in _child(window, QLabel, "params-columns-label").text()
    qtbot.mouseClick(
        _child(window, QPushButton, "params-confirm-button"), Qt.MouseButton.LeftButton
    )  # type: ignore[no-untyped-call]

    trim_request = _last_request(task_queue, TRIM_PREVIEW_OPERATION)
    assert trim_request.payload["key_columns"] == [0]
    assert trim_request.payload["collapse_spaces"] is False
    result = run_trim_preview_task(trim_request, PreviewTaskContext())
    assert len(result.trimmed_cells) == 2
    _push_success(task_queue, trim_request, result)
    _wait_second_preview_request(qtbot, task_queue)

    temporary_request = _last_request(task_queue, BUILD_PREVIEW_INDEX_OPERATION)
    temporary_preview = run_preview_index_task(temporary_request, PreviewTaskContext())
    _push_success(task_queue, temporary_request, temporary_preview)
    qtbot.waitUntil(
        lambda: "将清理 2 个单元格" in _child(window, QLabel, "banner-message").text(),
        timeout=2000,
    )

    # 预览就绪后“查看明细”可用，并展示修改前后内容
    details_button = _child(window, QPushButton, "banner-details-button")
    assert details_button.isEnabled()
    qtbot.mouseClick(details_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]
    from PySide6.QtWidgets import QTableView as DetailsTable

    details_table = _child(window, DetailsTable, "processing-details-table")
    assert details_table.model() is not None
    assert details_table.model().rowCount() == 2
    assert "苹果" in str(details_table.model().data(details_table.model().index(0, 2)))

    qtbot.mouseClick(_child(window, QPushButton, "banner-apply-button"), Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]
    apply_request = _last_request(task_queue, APPLY_TRIM_PREVIEW_OPERATION)
    _push_success(
        task_queue,
        apply_request,
        run_apply_trim_preview_task(apply_request, PreviewTaskContext()),
        EngineName.PYTHON,
    )

    qtbot.waitUntil(lambda: _head_record(library_root) is not None, timeout=2000)
    head = _head_record(library_root)
    assert head is not None
    assert head.version_id != "version-1"
    assert head.operation == "trim-whitespace"


def test_temporary_preview_supports_manual_edits_on_apply(qtbot: QtBot, tmp_path: Path) -> None:
    """需求第 17 节：临时结果上可继续单元格编辑，统一随应用生成版本。"""
    from openpyxl import load_workbook
    from PySide6.QtCore import Qt

    library_root, task_queue, window = _ready_window(
        qtbot,
        tmp_path,
        [["名称", "数量"], ["  苹果  ", 3], [" 香蕉 ", 2]],
    )
    window._one_step_processing("trim", [0])
    qtbot.waitUntil(
        lambda: _child(window, QFrame, "processing-params-bar").isVisibleTo(window), timeout=500
    )
    qtbot.mouseClick(
        _child(window, QPushButton, "params-confirm-button"), Qt.MouseButton.LeftButton
    )  # type: ignore[no-untyped-call]
    trim_request = _last_request(task_queue, TRIM_PREVIEW_OPERATION)
    _push_success(
        task_queue, trim_request, run_trim_preview_task(trim_request, PreviewTaskContext())
    )
    _wait_second_preview_request(qtbot, task_queue)
    temporary_request = _last_request(task_queue, BUILD_PREVIEW_INDEX_OPERATION)
    temporary_preview = run_preview_index_task(temporary_request, PreviewTaskContext())
    _push_success(task_queue, temporary_request, temporary_preview)

    table = _child(window, QTableView, "preview-table")
    qtbot.waitUntil(lambda: table.model() is not None, timeout=2000)
    # 临时结果进入可编辑状态，编辑值先落在编辑会话中
    model = table.model()
    assert model is not None
    assert model.flags(model.index(1, 0)) & Qt.ItemFlag.ItemIsEditable
    assert model.setData(model.index(1, 0), "苹果·已编辑", Qt.ItemDataRole.EditRole)
    edits = window._workbook_preview.pending_edits()
    assert len(edits) == 1
    assert (edits[0].sheet_name, edits[0].row, edits[0].column) == ("数据", 1, 0)

    qtbot.mouseClick(_child(window, QPushButton, "banner-apply-button"), Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]
    apply_request = _last_request(task_queue, APPLY_TRIM_PREVIEW_OPERATION)
    assert apply_request.payload["edits"] == [
        {"sheet_name": "数据", "row": 1, "column": 0, "value": "苹果·已编辑"}
    ]
    applied = run_apply_trim_preview_task(apply_request, PreviewTaskContext())
    _push_success(task_queue, apply_request, applied, EngineName.PYTHON)

    qtbot.waitUntil(lambda: _head_record(library_root) is not None, timeout=2000)
    head = _head_record(library_root)
    assert head is not None and head.version_id != "version-1"
    # 等应用成功事件处理完毕（丢弃临时结果并清空编辑会话）后再检查状态
    qtbot.waitUntil(lambda: window._processing_result is None, timeout=2000)
    workbook = load_workbook(head.snapshot_path)
    try:
        assert workbook["数据"]["A2"].value == "苹果·已编辑"
    finally:
        workbook.close()
    # 应用成功后编辑会话清空，不再污染后续预览
    assert window._workbook_preview.pending_edits() == ()


def test_find_replace_only_and_apply_flow(qtbot: QtBot, tmp_path: Path) -> None:
    library_root, task_queue, window = _ready_window(
        qtbot,
        tmp_path,
        [["名称", "数量"], ["Apple 销售", 3], ["apple pie", 5]],
    )
    window._open_find_replace_dialog()
    parameters: dict[str, object] = {
        "all_sheets": False,
        "mode": "values",
        "find_text": "apple",
        "replace_text": "橙子",
        "match_case": False,
        "whole_cell": False,
        "trim_whitespace": False,
    }

    window._submit_find_replace_preview("数据", {**parameters, "replace_all": False})
    find_request = _last_request(task_queue, FIND_REPLACE_PREVIEW_OPERATION)
    find_result = run_find_replace_preview_task(find_request, PreviewTaskContext())
    assert find_result.preview_path is None
    _push_success(task_queue, find_request, find_result)
    qtbot.waitUntil(
        lambda: "找到 2 处匹配" in _child(window, QLabel, "find-dialog-status").text(),
        timeout=2000,
    )

    window._submit_find_replace_preview("数据", {**parameters, "replace_all": True})
    replace_request = next(
        request
        for request in task_queue.submitted
        if request.operation == FIND_REPLACE_PREVIEW_OPERATION
        and request.payload.get("replace_all")
    )
    replace_result = run_find_replace_preview_task(replace_request, PreviewTaskContext())
    assert replace_result.preview_path is not None
    _push_success(task_queue, replace_request, replace_result)
    _wait_second_preview_request(qtbot, task_queue)

    temporary_request = _last_request(task_queue, BUILD_PREVIEW_INDEX_OPERATION)
    temporary_preview = run_preview_index_task(temporary_request, PreviewTaskContext())
    _push_success(task_queue, temporary_request, temporary_preview)

    apply_button = _child(window, QPushButton, "banner-apply-button")
    qtbot.waitUntil(apply_button.isEnabled, timeout=2000)
    qtbot.mouseClick(apply_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]
    apply_request = _last_request(task_queue, APPLY_FIND_REPLACE_PREVIEW_OPERATION)
    _push_success(
        task_queue,
        apply_request,
        run_apply_find_replace_preview_task(apply_request, PreviewTaskContext()),
        EngineName.PYTHON,
    )

    def _head_operation() -> str:
        head = _head_record(library_root)
        return head.operation if head is not None else ""

    qtbot.waitUntil(
        lambda: _head_operation() == "find-replace",
        timeout=2000,
    )


def test_editor_bar_sort_and_entry(qtbot: QtBot, tmp_path: Path) -> None:
    _library_root, task_queue, window = _ready_window(
        qtbot,
        tmp_path,
        [["名称", "数量"], ["b", 2], ["a", 1]],
    )

    # 无进行中任务时功能区条入口可用
    qtbot.mouseClick(
        _child(window, QPushButton, "bar-find-replace-button"), Qt.MouseButton.LeftButton
    )  # type: ignore[no-untyped-call]
    dialog = window._find_dialog
    assert dialog is not None
    assert dialog.isVisible()
    dialog.close()

    table = _child(window, QTableView, "preview-table")
    table.selectColumn(0)

    qtbot.mouseClick(_child(window, QPushButton, "bar-sort-asc-button"), Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    sort_request = task_queue.submitted[-1]
    assert sort_request.operation == SORT_PREVIEW_OPERATION
    sort_keys = sort_request.payload["sort_keys"]
    assert isinstance(sort_keys, list) and isinstance(sort_keys[0], dict)
    assert sort_keys[0]["column_index"] == 0
    assert sort_keys[0]["direction"] == "asc"
    # 需求第 19.1 节：单列排序明确提示按该列排序完整数据行
    assert "按 A 列排序完整数据行" in _child(window, QLabel, "banner-message").text()
    # 处理预览进行中当前预览被清空，功能区条入口应禁用，避免误触发错误提示
    assert not _child(window, QPushButton, "bar-find-replace-button").isEnabled()


def test_deduplicate_params_bar_options_flow_to_preview(qtbot: QtBot, tmp_path: Path) -> None:
    """需求第 19.2 节：去重保留规则与比较选项在参数行可调并传入任务。"""
    _library_root, task_queue, window = _ready_window(
        qtbot,
        tmp_path,
        [["名称", "数量"], ["苹果", 1], ["苹果", 2], ["香蕉", 3]],
    )
    window._one_step_processing("deduplicate", [0])
    qtbot.waitUntil(
        lambda: _child(window, QFrame, "processing-params-bar").isVisibleTo(window), timeout=500
    )
    assert "关键列 A" in _child(window, QLabel, "params-columns-label").text()

    keep = _child(window, QComboBox, "params-keep-combo")
    keep.setCurrentIndex(1)
    _child(window, QCheckBox, "params-ignore-case").setChecked(True)
    qtbot.mouseClick(
        _child(window, QPushButton, "params-confirm-button"), Qt.MouseButton.LeftButton
    )  # type: ignore[no-untyped-call]

    deduplicate_request = _last_request(task_queue, DEDUPLICATE_PREVIEW_OPERATION)
    assert deduplicate_request.payload["key_columns"] == [0]
    assert deduplicate_request.payload["keep"] == "last"
    assert deduplicate_request.payload["ignore_case"] is True
    assert deduplicate_request.payload["trim_whitespace"] is False
    result = run_deduplicate_preview_task(deduplicate_request, PreviewTaskContext())
    assert result.duplicate_groups


def test_filter_dialog_two_conditions_same_column_or() -> None:
    """DEC-021：同一列两个条件支持或者，跨列强制并且。"""
    from hyacinth.ui import FilterDialog

    dialog = FilterDialog("数据", ("A · 名称", "B · 数量"))
    payloads: list[dict[str, object]] = []
    dialog.params_submitted.connect(payloads.append)

    dialog._second.setChecked(True)
    dialog._connector.setCurrentIndex(1)
    dialog._submit()
    assert len(payloads[-1]["conditions"]) == 2  # type: ignore[arg-type]
    assert payloads[-1]["connector"] == "or"

    second_column = dialog._condition_rows[1]["column"]
    assert isinstance(second_column, QComboBox)
    second_column.setCurrentIndex(1)
    dialog._submit()
    assert payloads[-1].get("connector") == "and"


def test_find_replace_dialog_trim_whitespace_option() -> None:
    """需求第 19.4 节：查找替换支持忽略首尾空格比较。"""
    from hyacinth.ui import FindReplaceDialog

    dialog = FindReplaceDialog("数据")
    payloads: list[dict[str, object]] = []
    dialog.params_submitted.connect(lambda _sheet, params: payloads.append(params))

    dialog._find_text.setText("苹果")
    dialog._ignore_trim.setChecked(True)
    dialog._submit(replace_all=False)
    assert payloads[-1]["trim_whitespace"] is True


def test_grid_shows_data_margin_and_select_all_is_safe(qtbot: QtBot, tmp_path: Path) -> None:
    _library_root, _task_queue, window = _ready_window(
        qtbot,
        tmp_path,
        [["名称", "数量"], ["苹果", 3], ["香蕉", 2]],
    )
    table = _child(window, QTableView, "preview-table")
    model = table.model()
    assert model is not None
    # 数据 3 行 2 列 + 编辑余量 32 行 4 列，不再生成百万行逻辑网格
    assert model.rowCount() == 3 + 32
    assert model.columnCount() == 2 + 4

    table.selectAll()
    selected = table.selectionModel().selectedIndexes()
    # 全选只选数据区域，不会选中余量，更不会触碰百万行
    assert len(selected) == 3 * 2


def test_sort_dialog_two_keys_and_same_column_guard() -> None:
    """需求第 19.1 节：多列排序对话框支持两个关键字并阻止同列重复。"""
    from hyacinth.ui import SortDialog

    dialog = SortDialog("数据")
    dialog.set_columns(("A · 名称", "B · 数量", "C · 备注"))
    payloads: list[dict[str, object]] = []
    dialog.params_submitted.connect(payloads.append)

    dialog._submit()
    assert len(payloads[-1]["sort_keys"]) == 1  # type: ignore[arg-type]

    dialog._second.setChecked(True)
    dialog._key2_direction.setCurrentIndex(1)
    dialog._submit()
    keys = payloads[-1]["sort_keys"]
    assert isinstance(keys, list) and len(keys) == 2
    assert keys[1]["column_index"] == 1 and keys[1]["direction"] == "desc"

    # 次要关键字与主要关键字相同：不提交并给出提示
    dialog._key1_column.setCurrentIndex(1)
    dialog._submit()
    assert len(payloads) == 2
    assert "不能与主要关键字相同" in dialog._status.text()


def test_multi_sort_from_ribbon_submits_two_key_preview(qtbot: QtBot, tmp_path: Path) -> None:
    _library_root, task_queue, window = _ready_window(
        qtbot,
        tmp_path,
        [["名称", "数量", "备注"], ["b", 2, "x"], ["a", 1, "y"]],
    )

    qtbot.mouseClick(
        _child(window, QPushButton, "bar-sort-multi-button"), Qt.MouseButton.LeftButton
    )  # type: ignore[no-untyped-call]
    dialog = window._sort_dialog
    assert dialog is not None
    assert dialog.isVisible()
    assert dialog._key1_column.count() == 3

    dialog._second.setChecked(True)
    dialog._key1_direction.setCurrentIndex(1)
    qtbot.mouseClick(_child(dialog, QPushButton, "sort-submit-button"), Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    sort_request = _last_request(task_queue, SORT_PREVIEW_OPERATION)
    keys = sort_request.payload["sort_keys"]
    assert isinstance(keys, list) and len(keys) == 2
    assert keys[0] == {"column_index": 0, "direction": "desc"}
    assert keys[1] == {"column_index": 1, "direction": "asc"}
    assert sort_request.payload["sheet_name"] == "数据"
    assert "按 A、B 列排序完整数据行" in _child(window, QLabel, "banner-message").text()


def test_find_replace_single_item_edits_session(qtbot: QtBot, tmp_path: Path) -> None:
    """需求第 19.4 节：查找后逐项替换写入当前编辑会话，不直接生成版本。"""
    library_root, task_queue, window = _ready_window(
        qtbot,
        tmp_path,
        [["名称", "数量"], ["Apple 销售", 3], ["apple pie", 5]],
    )
    window._open_find_replace_dialog()
    window._submit_find_replace_preview(
        "数据",
        {
            "all_sheets": False,
            "mode": "values",
            "find_text": "apple",
            "replace_text": "橙子",
            "match_case": False,
            "whole_cell": False,
            "trim_whitespace": False,
            "replace_all": False,
        },
    )
    find_request = _last_request(task_queue, FIND_REPLACE_PREVIEW_OPERATION)
    find_result = run_find_replace_preview_task(find_request, PreviewTaskContext())
    _push_success(task_queue, find_request, find_result)

    matches_table = _child(window, QTableView, "find-matches-table")
    qtbot.waitUntil(lambda: matches_table.model().rowCount() == 2, timeout=2000)
    assert "找到 2 处匹配" in _child(window, QLabel, "find-dialog-status").text()

    replace_one = _child(window, QPushButton, "find-replace-one-button")
    qtbot.waitUntil(replace_one.isEnabled, timeout=2000)
    qtbot.mouseClick(replace_one, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    edits = window._workbook_preview.pending_edits()
    assert len(edits) == 1
    assert (edits[0].sheet_name, edits[0].row, edits[0].column) == ("数据", 1, 0)
    assert edits[0].value == "橙子 销售"
    # 逐项替换只进编辑会话，不生成版本节点
    head = MetadataStore(library_root).get_workbook("file-1").head_version
    assert head is not None and head.version_id == "version-1"
    # 替换后状态提示进度并跳到下一处
    assert "已替换 1/2" in _child(window, QLabel, "find-dialog-status").text()
    selection = matches_table.selectionModel()
    assert selection is not None and selection.currentIndex().row() == 1


def test_find_matches_numeric_cells_in_values_mode(tmp_path: Path) -> None:
    """值与文本模式按显示文本匹配数字单元格（对齐 Excel 查找习惯）。"""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "数据"
    sheet.append(["名称", "数量"])
    sheet.append(["苹果", 3])
    sheet.append(["香蕉", 30])
    source = tmp_path / "source.xlsx"
    workbook.save(source)
    workbook.close()

    request = TaskRequest(
        task_id="find-1",
        name="只查找",
        file_id="file-1",
        engine=None,
        operation=FIND_REPLACE_PREVIEW_OPERATION,
        payload={
            "source_path": str(source),
            "preview_path": str(tmp_path / "preview.xlsx"),
            "parent_version_id": "version-1",
            "sheet_name": "数据",
            "all_sheets": False,
            "mode": "values",
            "find_text": "3",
            "replace_text": "三",
            "match_case": False,
            "whole_cell": True,
            "trim_whitespace": False,
            "replace_all": False,
        },
    )
    result = run_find_replace_preview_task(request, PreviewTaskContext())
    # 整格匹配时 30 不算 3 的匹配；数字 3 命中且替换预览为文本"三"
    assert [(change.row, change.column, change.after) for change in result.changes] == [
        (2, 2, "三")
    ]


def test_find_replace_targets_currently_displayed_sheet(qtbot: QtBot, tmp_path: Path) -> None:
    """查找必须作用于用户正在查看的工作表，而不是第一个工作表。"""
    library_root = tmp_path / "library"
    directory = library_root / "files/file-1"
    original = directory / "original/多表.xlsx"
    working = directory / "working/current.xlsx"
    snapshot = directory / "versions/version-1/snapshot.xlsx"
    for path in (original, working, snapshot):
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        summary = workbook.active
        assert summary is not None
        summary.title = "汇总"
        summary.append(["区域", "合计"])
        summary.append(["华东", 100])
        detail = workbook.create_sheet("明细")
        detail.append(["名称", "数量"])
        detail.append(["Apple 销售", 3])
        workbook.save(path)
        workbook.close()
    version = VersionRecord(
        "version-1",
        "file-1",
        None,
        "导入原始文件",
        datetime(2026, 8, 16, 8, 0, tzinfo=UTC),
        "import",
        None,
        snapshot,
        sha256(snapshot.read_bytes()).hexdigest(),
    )
    record = ImportedWorkbook(
        "file-1",
        "多表.xlsx",
        original,
        working,
        version,
        datetime(2026, 8, 16, 8, 0, tzinfo=UTC),
    )
    MetadataStore(library_root).record_import(record)

    from hyacinth.app import create_main_window

    task_queue = FakeApplicationTaskQueue([])
    window = create_main_window(task_queue=task_queue, library_root=library_root)
    qtbot.addWidget(window)
    window.show()
    preview_request = _preview_request_of(task_queue)
    preview = run_preview_index_task(preview_request, PreviewTaskContext())
    _push_success(task_queue, preview_request, preview)
    table = _child(window, QTableView, "preview-table")
    qtbot.waitUntil(lambda: table.model() is not None, timeout=2000)

    tabs = window._workbook_preview._tabs
    assert tabs.count() == 2
    tabs.setCurrentIndex(1)  # 切到"明细"
    assert window._workbook_preview.current_sheet_name == "明细"

    window._open_find_replace_dialog()
    assert window._find_dialog is not None
    assert window._find_dialog._sheet_name == "明细"

    # 提交时同样以正在显示的工作表为准，即使对话框打开于其他工作表
    window._submit_find_replace_preview(
        "汇总",
        {
            "all_sheets": False,
            "mode": "values",
            "find_text": "Apple",
            "replace_text": "橙子",
            "match_case": False,
            "whole_cell": False,
            "trim_whitespace": False,
            "replace_all": False,
        },
    )
    find_request = _last_request(task_queue, FIND_REPLACE_PREVIEW_OPERATION)
    assert find_request.payload["sheet_name"] == "明细"
    find_result = run_find_replace_preview_task(find_request, PreviewTaskContext())
    assert len(find_result.changes) == 1
    assert find_result.changes[0].after == "橙子 销售"


def test_find_only_keeps_current_grid_preview(qtbot: QtBot, tmp_path: Path) -> None:
    """只查找不产生临时文件，不能把正在查看的数据预览清成导入空状态。"""
    _library_root, task_queue, window = _ready_window(
        qtbot,
        tmp_path,
        [["名称", "数量"], ["苹果", 3], ["香蕉", 2]],
    )
    table = _child(window, QTableView, "preview-table")
    model_before = table.model()
    assert model_before is not None

    window._submit_find_replace_preview(
        "数据",
        {
            "all_sheets": False,
            "mode": "values",
            "find_text": "苹果",
            "replace_text": "橙子",
            "match_case": False,
            "whole_cell": False,
            "trim_whitespace": False,
            "replace_all": False,
        },
    )
    find_request = _last_request(task_queue, FIND_REPLACE_PREVIEW_OPERATION)
    find_result = run_find_replace_preview_task(find_request, PreviewTaskContext())
    _push_success(task_queue, find_request, find_result)

    # 网格保持原模型，未被清成空状态
    assert table.model() is model_before
    assert window._workbook_preview._stack.currentIndex() == 1
    assert window._workbook_preview.current_preview() is not None


def test_chained_operations_create_single_version(qtbot: QtBot, tmp_path: Path) -> None:
    """链式多步处理：清空格→排序连续叠加，应用时只生成一个多步处理节点。"""
    from openpyxl import load_workbook

    from hyacinth.processing import (
        APPLY_CHAINED_PREVIEW_OPERATION,
        run_apply_chained_preview_task,
        run_sort_preview_task,
    )

    library_root, task_queue, window = _ready_window(
        qtbot,
        tmp_path,
        [["名称", "数量"], [" b ", 2], [" a ", 1]],
    )

    # 第一步：清除首尾空格
    window._one_step_processing("trim", [0])
    qtbot.waitUntil(
        lambda: _child(window, QFrame, "processing-params-bar").isVisibleTo(window), timeout=500
    )
    qtbot.mouseClick(
        _child(window, QPushButton, "params-confirm-button"), Qt.MouseButton.LeftButton
    )  # type: ignore[no-untyped-call]
    trim_request = _last_request(task_queue, TRIM_PREVIEW_OPERATION)
    _push_success(
        task_queue, trim_request, run_trim_preview_task(trim_request, PreviewTaskContext())
    )
    _wait_preview_request_count(qtbot, task_queue, 2)
    temporary_request = _last_request(task_queue, BUILD_PREVIEW_INDEX_OPERATION)
    _push_success(
        task_queue,
        temporary_request,
        run_preview_index_task(temporary_request, PreviewTaskContext()),
    )
    # Fake 队列靠轮询派发事件：等临时预览真正上屏后再发起下一步链式操作。
    table = _child(window, QTableView, "preview-table")
    qtbot.waitUntil(lambda: table.model() is not None, timeout=2000)

    # 第二步：直接在临时结果上按 A 列排序（不再要求先应用上一步）
    window._quick_sort_from_table(0, "asc")
    sort_request = _last_request(task_queue, SORT_PREVIEW_OPERATION)
    assert sort_request.payload["source_path"] == str(trim_request.payload["preview_path"])
    assert sort_request.payload["parent_version_id"] == "version-1"
    assert sort_request.payload["edits"] == []
    _push_success(
        task_queue, sort_request, run_sort_preview_task(sort_request, PreviewTaskContext())
    )
    _wait_preview_request_count(qtbot, task_queue, 3)
    chained_request = _last_request(task_queue, BUILD_PREVIEW_INDEX_OPERATION)
    _push_success(
        task_queue,
        chained_request,
        run_preview_index_task(chained_request, PreviewTaskContext()),
    )
    qtbot.waitUntil(
        lambda: "已连续 2 步处理" in _child(window, QLabel, "banner-message").text(),
        timeout=2000,
    )

    # 应用：一个节点承载两步操作
    qtbot.mouseClick(_child(window, QPushButton, "banner-apply-button"), Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]
    apply_request = _last_request(task_queue, APPLY_CHAINED_PREVIEW_OPERATION)
    steps = apply_request.payload["steps"]
    assert isinstance(steps, list) and len(steps) == 2
    assert apply_request.payload["parent_version_id"] == "version-1"
    _push_success(
        task_queue,
        apply_request,
        run_apply_chained_preview_task(apply_request, PreviewTaskContext()),
        EngineName.PYTHON,
    )

    def _new_head_id() -> str | None:
        head = _head_record(library_root)
        return head.version_id if head is not None else None

    qtbot.waitUntil(lambda: _new_head_id() not in (None, "version-1"), timeout=2000)
    head = _head_record(library_root)
    assert head is not None
    assert head.operation == "multi-step"
    assert head.parent_version_id == "version-1"
    assert "多步处理（2 项操作）" in head.name
    import json

    assert len(json.loads(head.parameters_json)["steps"]) == 2

    workbook = load_workbook(head.snapshot_path)
    try:
        sheet = workbook["数据"]
        # 先清空格再排序：a 在 b 前，且首尾空格已去除
        assert [sheet["A2"].value, sheet["A3"].value] == ["a", "b"]
        assert [sheet["B2"].value, sheet["B3"].value] == [1, 2]
    finally:
        workbook.close()


def test_preview_reload_preserves_current_sheet(qtbot: QtBot, tmp_path: Path) -> None:
    """临时预览重载后保持用户正在查看的工作表，不跳回第一张表。"""
    library_root = tmp_path / "library"
    directory = library_root / "files/file-1"
    original = directory / "original/多表.xlsx"
    working = directory / "working/current.xlsx"
    snapshot = directory / "versions/version-1/snapshot.xlsx"
    for path in (original, working, snapshot):
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        summary = workbook.active
        assert summary is not None
        summary.title = "汇总"
        summary.append(["区域", "合计"])
        summary.append(["华东", 100])
        detail = workbook.create_sheet("明细")
        detail.append(["名称", "数量"])
        detail.append([" 橙子 ", 3])
        workbook.save(path)
        workbook.close()
    version = VersionRecord(
        "version-1",
        "file-1",
        None,
        "导入原始文件",
        datetime(2026, 8, 16, 8, 0, tzinfo=UTC),
        "import",
        None,
        snapshot,
        sha256(snapshot.read_bytes()).hexdigest(),
    )
    record = ImportedWorkbook(
        "file-1",
        "多表.xlsx",
        original,
        working,
        version,
        datetime(2026, 8, 16, 8, 0, tzinfo=UTC),
    )
    MetadataStore(library_root).record_import(record)

    from hyacinth.app import create_main_window

    task_queue = FakeApplicationTaskQueue([])
    window = create_main_window(task_queue=task_queue, library_root=library_root)
    qtbot.addWidget(window)
    window.show()
    preview_request = _preview_request_of(task_queue)
    preview = run_preview_index_task(preview_request, PreviewTaskContext())
    _push_success(task_queue, preview_request, preview)
    table = _child(window, QTableView, "preview-table")
    qtbot.waitUntil(lambda: table.model() is not None, timeout=2000)

    tabs = window._workbook_preview._tabs
    assert tabs.count() == 2
    tabs.setCurrentIndex(1)  # 正在查看"明细"

    # 在"明细"上执行清除空格 → 临时预览重载后仍应停留在"明细"
    window._one_step_processing("trim", [0])
    qtbot.waitUntil(
        lambda: _child(window, QFrame, "processing-params-bar").isVisibleTo(window), timeout=500
    )
    qtbot.mouseClick(
        _child(window, QPushButton, "params-confirm-button"), Qt.MouseButton.LeftButton
    )  # type: ignore[no-untyped-call]
    trim_request = _last_request(task_queue, TRIM_PREVIEW_OPERATION)
    assert trim_request.payload["sheet_name"] == "明细"
    _push_success(
        task_queue, trim_request, run_trim_preview_task(trim_request, PreviewTaskContext())
    )
    _wait_second_preview_request(qtbot, task_queue)
    temporary_request = _last_request(task_queue, BUILD_PREVIEW_INDEX_OPERATION)
    _push_success(
        task_queue,
        temporary_request,
        run_preview_index_task(temporary_request, PreviewTaskContext()),
    )
    qtbot.waitUntil(lambda: table.model() is not None, timeout=2000)
    assert tabs.currentIndex() == 1
    assert tabs.tabText(tabs.currentIndex()) == "明细"


def test_first_operation_bakes_pending_edits(qtbot: QtBot, tmp_path: Path) -> None:
    """HEAD 预览上的未保存编辑随首个处理操作烘焙进临时结果，不再被无声清空。"""
    _library_root, task_queue, window = _ready_window(
        qtbot,
        tmp_path,
        [["名称", "数量"], [" 苹果 ", 3], [" 香蕉 ", 2]],
    )
    dirty_label = _child(window, QLabel, "dirty-edits-label")
    assert not dirty_label.isVisibleTo(window)

    window._workbook_preview.apply_cell_edit("数据", 1, 0, base_value="苹果", new_value="梨")
    qtbot.waitUntil(lambda: dirty_label.isVisibleTo(window), timeout=2000)
    assert dirty_label.text() == "1 处未保存编辑"

    window._one_step_processing("trim", [0])
    qtbot.waitUntil(
        lambda: _child(window, QFrame, "processing-params-bar").isVisibleTo(window), timeout=500
    )
    qtbot.mouseClick(
        _child(window, QPushButton, "params-confirm-button"), Qt.MouseButton.LeftButton
    )  # type: ignore[no-untyped-call]
    trim_request = _last_request(task_queue, TRIM_PREVIEW_OPERATION)
    # 编辑被烘焙进首个处理操作，而不是被丢弃
    assert trim_request.payload["edits"] == [
        {"sheet_name": "数据", "row": 1, "column": 0, "value": "梨"}
    ]
    _push_success(
        task_queue, trim_request, run_trim_preview_task(trim_request, PreviewTaskContext())
    )
    _wait_second_preview_request(qtbot, task_queue)
    temporary_request = _last_request(task_queue, BUILD_PREVIEW_INDEX_OPERATION)
    _push_success(
        task_queue,
        temporary_request,
        run_preview_index_task(temporary_request, PreviewTaskContext()),
    )
    table = _child(window, QTableView, "preview-table")
    qtbot.waitUntil(lambda: table.model() is not None, timeout=2000)
    # 烘焙完成后编辑会话清空，脏标记消失
    qtbot.waitUntil(lambda: not dirty_label.isVisibleTo(window), timeout=2000)
    assert window._workbook_preview.pending_edits() == ()


def test_cancel_temporary_result_prompts_when_edits_pending(qtbot: QtBot, tmp_path: Path) -> None:
    """临时结果上有未保存编辑时，取消前必须弹窗让用户选择。"""
    from hyacinth.app import create_main_window

    library_root = tmp_path / "library"
    _seed_file(library_root, [["名称", "数量"], [" 苹果 ", 3]])
    actions: list[str] = []

    def presenter(_parent: object, action: str, _allow_in_place: bool = False) -> str:
        actions.append(action)
        return "cancel"

    task_queue = FakeApplicationTaskQueue([])
    window = create_main_window(
        task_queue=task_queue,
        library_root=library_root,
        unsaved_changes_presenter=presenter,
    )
    qtbot.addWidget(window)
    window.show()
    preview_request = _preview_request_of(task_queue)
    _push_success(
        task_queue, preview_request, run_preview_index_task(preview_request, PreviewTaskContext())
    )
    table = _child(window, QTableView, "preview-table")
    qtbot.waitUntil(lambda: table.model() is not None, timeout=2000)

    window._one_step_processing("trim", [0])
    qtbot.waitUntil(
        lambda: _child(window, QFrame, "processing-params-bar").isVisibleTo(window), timeout=500
    )
    qtbot.mouseClick(
        _child(window, QPushButton, "params-confirm-button"), Qt.MouseButton.LeftButton
    )  # type: ignore[no-untyped-call]
    trim_request = _last_request(task_queue, TRIM_PREVIEW_OPERATION)
    _push_success(
        task_queue, trim_request, run_trim_preview_task(trim_request, PreviewTaskContext())
    )
    _wait_second_preview_request(qtbot, task_queue)
    temporary_request = _last_request(task_queue, BUILD_PREVIEW_INDEX_OPERATION)
    _push_success(
        task_queue,
        temporary_request,
        run_preview_index_task(temporary_request, PreviewTaskContext()),
    )
    qtbot.waitUntil(lambda: table.model() is not None, timeout=2000)
    assert window._processing_result is not None

    # 临时结果上再敲一处编辑，然后取消：应弹窗，用户选“取消”则链保持原状
    window._workbook_preview.apply_cell_edit("数据", 1, 0, base_value="苹果", new_value="梨")
    window._cancel_processing_workflow()
    assert actions == ["取消临时结果"]
    assert window._processing_result is not None


def test_save_in_place_updates_leaf_without_new_version(qtbot: QtBot, tmp_path: Path) -> None:
    """“就地更新此节点”：同一 version_id、哈希更新、工作副本与快照同步。"""
    from openpyxl import load_workbook

    from hyacinth.processing import (
        UPDATE_VERSION_IN_PLACE_OPERATION,
        run_update_version_in_place_task,
    )

    library_root, task_queue, window = _ready_window(
        qtbot,
        tmp_path,
        [["名称", "数量"], ["苹果", 3], ["香蕉", 2]],
    )
    head_before = _head_record(library_root)
    assert head_before is not None

    window._workbook_preview.apply_cell_edit("数据", 1, 0, base_value="苹果", new_value="梨")
    window._submit_update_in_place()
    request = _last_request(task_queue, UPDATE_VERSION_IN_PLACE_OPERATION)
    assert request.payload["version_id"] == "version-1"
    assert request.payload["expected_hash"] == head_before.content_hash
    assert request.payload["edits"] == [
        {"sheet_name": "数据", "row": 1, "column": 0, "value": "梨"}
    ]
    _push_success(
        task_queue, request, run_update_version_in_place_task(request, PreviewTaskContext())
    )

    def _head_hash() -> str:
        head = _head_record(library_root)
        assert head is not None
        return head.content_hash

    qtbot.waitUntil(lambda: _head_hash() != head_before.content_hash, timeout=2000)
    head_after = _head_record(library_root)
    assert head_after is not None
    assert head_after.version_id == "version-1"  # 不生成新节点
    workbook = load_workbook(head_after.snapshot_path)
    try:
        assert workbook["数据"]["A2"].value == "梨"
    finally:
        workbook.close()
    qtbot.waitUntil(lambda: window._workbook_preview.pending_edits() == (), timeout=2000)

    # 就地更新后重载预览，脏标记消失
    _wait_preview_request_count(qtbot, task_queue, 2)
    reload_request = _last_request(task_queue, BUILD_PREVIEW_INDEX_OPERATION)
    _push_success(
        task_queue,
        reload_request,
        run_preview_index_task(reload_request, PreviewTaskContext()),
    )
    dirty_label = _child(window, QLabel, "dirty-edits-label")
    qtbot.waitUntil(lambda: not dirty_label.isVisibleTo(window), timeout=2000)
