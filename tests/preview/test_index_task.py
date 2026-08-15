from collections.abc import Iterator
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from hyacinth.tasks import TaskRequest
from hyacinth.tasks.worker import TaskCancelled


class PreviewTaskContext:
    def __init__(self) -> None:
        self.messages: list[str] = []
        self.committed = False

    def report_progress(self, progress: float | None, message: str = "") -> None:
        self.messages.append(message)

    def check_cancelled(self) -> None:
        return

    def commit(self) -> None:
        self.committed = True

    @contextmanager
    def critical_section(self, message: str = "") -> Iterator[None]:
        self.messages.append(message)
        yield


class CancellingPreviewTaskContext(PreviewTaskContext):
    def __init__(self) -> None:
        super().__init__()
        self.check_count = 0

    def check_cancelled(self) -> None:
        self.check_count += 1
        if self.check_count >= 2:
            raise TaskCancelled


def _workbook(path: Path) -> None:
    workbook = Workbook()
    data = workbook.active
    assert data is not None
    data.title = "数据"
    data.append(["名称", "数量", "日期", "合计"])
    data.append(["苹果", 12.5, datetime(2026, 8, 15, 9, 30), "=B2*2"])
    notes = workbook.create_sheet("备注")
    notes["A1"] = "第二张工作表"
    hidden = workbook.create_sheet("内部")
    hidden.sheet_state = "hidden"
    workbook.save(path)
    workbook.close()


def test_preview_task_builds_atomic_sparse_index(tmp_path: Path) -> None:
    from hyacinth.preview import SqliteGridDataSource, run_preview_index_task

    working = tmp_path / "working" / "current.xlsx"
    working.parent.mkdir()
    _workbook(working)
    index_path = tmp_path / "cache" / "preview.sqlite"
    context = PreviewTaskContext()
    request = TaskRequest(
        task_id="preview-1",
        name="加载预览",
        file_id="file-1",
        engine=None,
        operation="build-preview-index",
        payload={"working_path": str(working), "index_path": str(index_path)},
    )

    preview = run_preview_index_task(request, context)

    assert context.committed is True
    assert index_path.is_file()
    assert [sheet.title for sheet in preview.sheets] == ["数据", "备注"]
    assert [(sheet.row_count, sheet.column_count) for sheet in preview.sheets] == [
        (2, 4),
        (1, 1),
    ]
    source = SqliteGridDataSource(index_path, preview.sheets[0])
    try:
        assert source.value_at(0, 0) == "名称"
        assert source.value_at(1, 0) == "苹果"
        assert source.value_at(1, 1) == "12.5"
        assert source.value_at(1, 2) == "2026-08-15 09:30:00"
        assert source.value_at(1, 3) == "=B2*2"
        assert source.value_at(20, 20) == ""
    finally:
        source.close()
    assert not tuple(index_path.parent.glob("*.tmp"))


def test_preview_index_maps_hidden_rows_out_of_visible_grid(tmp_path: Path) -> None:
    from hyacinth.preview import SqliteGridDataSource, run_preview_index_task

    working = tmp_path / "filtered.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["名称"])
    sheet.append(["显示 A"])
    sheet.append(["隐藏"])
    sheet.append(["显示 B"])
    sheet.row_dimensions[3].hidden = True
    workbook.save(working)
    workbook.close()
    request = TaskRequest(
        task_id="preview-hidden",
        name="加载筛选预览",
        file_id="file-1",
        engine=None,
        operation="build-preview-index",
        payload={
            "working_path": str(working),
            "index_path": str(tmp_path / "preview.sqlite"),
        },
    )

    preview = run_preview_index_task(request, PreviewTaskContext())
    assert preview.sheets[0].visible_row_count == 3
    source = SqliteGridDataSource(preview.index_path, preview.sheets[0])
    try:
        assert [source.value_at(row, 0) for row in range(4)] == [
            "名称",
            "显示 A",
            "显示 B",
            "",
        ]
    finally:
        source.close()


def test_preview_task_reuses_current_index(tmp_path: Path) -> None:
    from hyacinth.preview import run_preview_index_task

    working = tmp_path / "current.xlsx"
    _workbook(working)
    index_path = tmp_path / "preview.sqlite"
    request = TaskRequest(
        task_id="preview-1",
        name="加载预览",
        file_id="file-1",
        engine=None,
        operation="build-preview-index",
        payload={"working_path": str(working), "index_path": str(index_path)},
    )

    first = run_preview_index_task(request, PreviewTaskContext())
    first_mtime = index_path.stat().st_mtime_ns
    second_context = PreviewTaskContext()
    second = run_preview_index_task(request, second_context)

    assert second == first
    assert index_path.stat().st_mtime_ns == first_mtime
    assert second_context.committed is False


def test_cancelled_preview_build_preserves_existing_index(tmp_path: Path) -> None:
    from hyacinth.preview import run_preview_index_task

    working = tmp_path / "current.xlsx"
    _workbook(working)
    index_path = tmp_path / "preview.sqlite"
    index_path.write_bytes(b"existing-index")
    request = TaskRequest(
        task_id="preview-cancel",
        name="加载预览",
        file_id="file-1",
        engine=None,
        operation="build-preview-index",
        payload={"working_path": str(working), "index_path": str(index_path)},
    )

    with pytest.raises(TaskCancelled):
        run_preview_index_task(request, CancellingPreviewTaskContext())

    assert index_path.read_bytes() == b"existing-index"
    assert not tuple(tmp_path.glob("*.tmp"))


def test_preview_build_closes_partially_opened_workbooks(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    from unittest.mock import Mock

    from hyacinth.preview import index_task

    working = tmp_path / "current.xlsx"
    _workbook(working)
    first_workbook = Mock()
    monkeypatch.setattr(
        index_task,
        "load_workbook",
        Mock(side_effect=[first_workbook, ValueError("第二次打开失败")]),
    )
    request = TaskRequest(
        task_id="preview-fail",
        name="加载预览",
        file_id="file-1",
        engine=None,
        operation="build-preview-index",
        payload={
            "working_path": str(working),
            "index_path": str(tmp_path / "preview.sqlite"),
        },
    )

    with pytest.raises(ValueError, match="第二次打开失败"):
        index_task.run_preview_index_task(request, PreviewTaskContext())

    first_workbook.close.assert_called_once_with()
    assert not tuple(tmp_path.glob("*.tmp"))
