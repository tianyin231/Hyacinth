from pathlib import Path

import pytest
from openpyxl import load_workbook

from hyacinth.excel.com_engine import is_excel_com_available
from hyacinth.excel.contracts import EngineName

FIXTURES = Path(__file__).parents[1] / "fixtures"


@pytest.mark.integration
def test_com_engine_preserves_formula_and_basic_formatting(tmp_path: Path) -> None:
    try:
        from hyacinth.excel.com_engine import ComExcelEngine
    except ImportError:
        pytest.fail("hyacinth.excel.com_engine.ComExcelEngine is not implemented")

    if not is_excel_com_available():
        pytest.skip("Microsoft Excel COM is unavailable")

    destination = tmp_path / "com-working-copy.xlsx"
    result = ComExcelEngine().convert_xls_to_xlsx(
        FIXTURES / "legacy-fidelity.xls",
        destination,
    )

    workbook = load_workbook(destination, data_only=False)
    data = workbook["Data"]

    assert result.engine is EngineName.COM
    assert result.output_path == destination
    assert result.warnings == ()
    assert workbook.sheetnames == ["Notes", "Data"]
    assert data["D2"].value == "=B2*2"
    assert data["B2"].number_format == "0.00"
    assert data["A1"].font.bold is True
    assert "A4:B4" in data.merged_cells
