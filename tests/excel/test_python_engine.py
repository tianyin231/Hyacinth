from pathlib import Path

import pytest
import xlrd  # type: ignore[import-untyped]
from openpyxl import load_workbook

from hyacinth.excel.contracts import ConversionWarning, EngineName

FIXTURES = Path(__file__).parents[1] / "fixtures"


def test_python_engine_converts_xls_data_and_reports_limitations(tmp_path: Path) -> None:
    try:
        from hyacinth.excel.python_engine import PythonExcelEngine
    except ModuleNotFoundError:
        pytest.fail("hyacinth.excel.python_engine.PythonExcelEngine is not implemented")

    source = FIXTURES / "legacy-fidelity.xls"
    source_workbook = xlrd.open_workbook(source)
    source_date = xlrd.xldate_as_datetime(
        source_workbook.sheet_by_name("Data").cell(1, 2).value,
        source_workbook.datemode,
    )
    destination = tmp_path / "working-copy.xlsx"
    result = PythonExcelEngine().convert_xls_to_xlsx(
        source,
        destination,
    )

    workbook = load_workbook(destination, data_only=False)
    data = workbook["Data"]

    assert result.engine is EngineName.PYTHON
    assert result.output_path == destination
    assert set(result.warnings) == {
        ConversionWarning.COMPLEX_FORMATTING_MAY_BE_LOST,
        ConversionWarning.XLS_FORMULAS_MAY_BECOME_CACHED_VALUES,
    }
    assert workbook.sheetnames == ["Notes", "Data"]
    assert data["A2"].value == "Alpha"
    assert data["B2"].value == 12.5
    assert data["C2"].value == source_date
    assert data["D2"].value == 25
    assert "A4:B4" in data.merged_cells
    assert workbook["Notes"]["A1"].value == "Second sheet"
