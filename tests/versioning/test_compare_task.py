"""版本差异比较任务测试（需求第 13 节）。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from hyacinth.tasks import TaskRequest
from hyacinth.versioning.compare_task import (
    COMPARE_VERSIONS_OPERATION,
    run_compare_versions_task,
)


class _Context:
    def report_progress(self, progress: float | None, message: str = "") -> None:
        return

    def check_cancelled(self) -> None:
        return


def _write_workbook(
    path: Path,
    sheets: dict[str, list[list[object]]],
) -> Path:
    book = Workbook()
    default = book.active
    assert default is not None
    first = True
    for name, rows in sheets.items():
        sheet = default if first else book.create_sheet()
        first = False
        sheet.title = name
        for row in rows:
            sheet.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
    book.close()
    return path


def _request(tmp_path: Path, base: Path, target: Path) -> TaskRequest:
    return TaskRequest(
        task_id="compare-1",
        name="版本对比",
        file_id="file-1",
        engine=None,
        operation=COMPARE_VERSIONS_OPERATION,
        payload={
            "file_id": "file-1",
            "base_version_id": "version-1",
            "target_version_id": "version-2",
            "base_path": str(base),
            "target_path": str(target),
        },
    )


def test_compare_reports_cell_added_removed_changed(tmp_path: Path) -> None:
    base = _write_workbook(
        tmp_path / "base.xlsx",
        {"一月": [["名称", "数量"], ["apple", 2], ["banana", 1]]},
    )
    target = _write_workbook(
        tmp_path / "target.xlsx",
        {"一月": [["名称", "数量", "备注"], ["apple", 5], ["banana", 1]]},
    )

    result = run_compare_versions_task(_request(tmp_path, base, target), _Context())

    assert result.total_diffs == 2
    sheet = result.sheets[0]
    assert sheet.status == "changed"
    assert sheet.target_name == "一月"
    by_ref = {diff.ref: diff for diff in sheet.cells}
    assert by_ref["B2"].kind == "changed"
    assert by_ref["B2"].base_value == "2"
    assert by_ref["B2"].target_value == "5"
    assert by_ref["C1"].kind == "added"
    assert by_ref["C1"].base_value is None
    assert by_ref["C1"].target_value == "备注"


def test_compare_reports_removed_cell_and_formula_text(tmp_path: Path) -> None:
    base = _write_workbook(
        tmp_path / "base.xlsx",
        {"一月": [["名称", "数量"], ["合计", 10], ["备注", "旧值"]]},
    )
    target_book = Workbook()
    sheet = target_book.active
    assert sheet is not None
    sheet.title = "一月"
    sheet.append(["名称", "数量"])
    sheet.append(["合计", "=SUM(B2:B3)"])
    target = tmp_path / "target.xlsx"
    target_book.save(target)
    target_book.close()

    result = run_compare_versions_task(_request(tmp_path, base, target), _Context())

    by_ref = {diff.ref: diff for diff in result.sheets[0].cells}
    # 公式按 openpyxl 原始表达式文本参与比较
    assert by_ref["B2"].kind == "changed"
    assert by_ref["B2"].target_value == "=SUM(B2:B3)"
    assert by_ref["A3"].kind == "removed"
    assert by_ref["A3"].base_value == "备注"
    assert by_ref["B3"].kind == "removed"


def test_compare_detects_sheet_added_removed_changed(tmp_path: Path) -> None:
    base = _write_workbook(
        tmp_path / "base.xlsx",
        {"一月": [["名称"], ["apple"]], "二月": [["x"], [1]], "三月": [["y"], [2]]},
    )
    target = _write_workbook(
        tmp_path / "target.xlsx",
        {"一月": [["名称"], ["apple"]], "三月": [["y"], [3]], "四月": [["x"], [9]]},
    )

    result = run_compare_versions_task(_request(tmp_path, base, target), _Context())

    labels = {sheet.label: sheet for sheet in result.sheets}
    assert labels["一月"].status == "unchanged"
    assert labels["一月"].diff_count == 0
    assert labels["三月"].status == "changed"
    assert labels["三月"].diff_count == 1
    by_status: dict[str, list[str]] = {}
    for sheet in result.sheets:
        by_status.setdefault(sheet.status, []).append(sheet.label)
    assert by_status["removed"] == ["二月"]
    assert by_status["added"] == ["四月"]
    assert set(by_status) == {"removed", "added", "changed", "unchanged"}


def test_compare_identifies_renamed_sheet_with_identical_content(tmp_path: Path) -> None:
    rows: list[list[object]] = [["名称"], ["apple"]]
    base = _write_workbook(tmp_path / "base.xlsx", {"旧表": rows})
    target = _write_workbook(tmp_path / "target.xlsx", {"新表": rows})

    result = run_compare_versions_task(_request(tmp_path, base, target), _Context())

    assert result.total_diffs == 0
    sheet = result.sheets[0]
    assert sheet.status == "renamed"
    assert sheet.base_name == "旧表"
    assert sheet.target_name == "新表"


def test_compare_identical_versions_have_no_diffs(tmp_path: Path) -> None:
    rows: list[list[object]] = [["名称", "数量"], ["apple", 2]]
    base = _write_workbook(tmp_path / "base.xlsx", {"一月": rows})
    target = _write_workbook(tmp_path / "target.xlsx", {"一月": rows})

    result = run_compare_versions_task(_request(tmp_path, base, target), _Context())

    assert result.total_diffs == 0
    assert result.sheets[0].status == "unchanged"
    assert len(result.sheets) == 1
