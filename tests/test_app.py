import os
import subprocess
import sys
from collections.abc import Iterator
from contextlib import contextmanager
from datetime import UTC, datetime
from hashlib import sha256
from pathlib import Path

import pytest
from openpyxl import Workbook
from PySide6.QtCore import QObject, QPoint, QPointF, QSize, Qt
from PySide6.QtTest import QTest
from PySide6.QtWidgets import (
    QCheckBox,
    QComboBox,
    QFrame,
    QGraphicsProxyWidget,
    QGraphicsView,
    QLabel,
    QLineEdit,
    QListWidget,
    QMainWindow,
    QPushButton,
    QSplitter,
    QStackedWidget,
    QTabBar,
    QTableView,
    QWidget,
)
from pytestqt.qtbot import QtBot

from hyacinth.excel.contracts import EngineName
from hyacinth.library import IMPORT_WORKBOOK_OPERATION, ImportedWorkbook
from hyacinth.preview import BUILD_PREVIEW_INDEX_OPERATION, run_preview_index_task
from hyacinth.processing import (
    APPLY_DEDUPLICATE_PREVIEW_OPERATION,
    APPLY_DELETE_BLANK_ROWS_PREVIEW_OPERATION,
    APPLY_FILTER_PREVIEW_OPERATION,
    APPLY_SORT_PREVIEW_OPERATION,
    DEDUPLICATE_PREVIEW_OPERATION,
    DELETE_BLANK_ROWS_PREVIEW_OPERATION,
    FILTER_PREVIEW_OPERATION,
    SAVE_MANUAL_EDITS_OPERATION,
    SORT_PREVIEW_OPERATION,
    run_apply_deduplicate_preview_task,
    run_apply_delete_blank_rows_preview_task,
    run_apply_filter_preview_task,
    run_apply_sort_preview_task,
    run_deduplicate_preview_task,
    run_delete_blank_rows_preview_task,
    run_filter_preview_task,
    run_save_manual_edits_task,
    run_sort_preview_task,
)
from hyacinth.tasks import TaskEvent, TaskRequest, TaskState, TaskStatusWidget
from hyacinth.ui import VersionTreePanel, format_byte_size
from hyacinth.versioning import (
    CHECKOUT_VERSION_OPERATION,
    DELETE_VERSION_OPERATION,
    EXPORT_VERSION_OPERATION,
    VERSION_STORAGE_STATS_OPERATION,
    MetadataStore,
    VersionRecord,
    run_checkout_version_task,
    run_delete_version_task,
    run_export_version_task,
    run_version_storage_stats_task,
)


def _child[WidgetT: QObject](parent: QObject, child_type: type[WidgetT], name: str) -> WidgetT:
    child = parent.findChild(child_type, name)
    assert child is not None
    return child


class FakeApplicationTaskQueue:
    def __init__(self, events: list[TaskEvent]) -> None:
        self._events = events
        self.cancelled: list[str] = []
        self.submitted: list[TaskRequest] = []
        self.shutdown_called = False

    def submit(self, request: TaskRequest) -> None:
        self.submitted.append(request)

    def push_event(self, event: TaskEvent) -> None:
        self._events.append(event)

    def poll_events(self) -> tuple[TaskEvent, ...]:
        events = tuple(self._events)
        self._events.clear()
        return events

    def cancel(self, task_id: str) -> bool:
        self.cancelled.append(task_id)
        return True

    def shutdown(self, timeout: float = 1.0) -> bool:
        self.shutdown_called = True
        return True


def _preview_request_of(queue: FakeApplicationTaskQueue) -> TaskRequest:
    return next(
        request for request in queue.submitted if request.operation == BUILD_PREVIEW_INDEX_OPERATION
    )


def _submitted_without_storage_stats(queue: FakeApplicationTaskQueue) -> list[TaskRequest]:
    return [
        request
        for request in queue.submitted
        if request.operation != VERSION_STORAGE_STATS_OPERATION
    ]


class PreviewTaskContext:
    def report_progress(self, progress: float | None, message: str = "") -> None:
        return

    def check_cancelled(self) -> None:
        return

    def set_engine(self, engine: EngineName) -> None:
        return

    def commit(self) -> None:
        return

    @contextmanager
    def critical_section(self, message: str = "") -> Iterator[None]:
        yield


def _seed_versioned_workbook(
    library_root: Path,
    rows: list[list[object]] | None = None,
) -> ImportedWorkbook:
    directory = library_root / "files/file-1"
    original = directory / "original/销售.xlsx"
    working = directory / "working/current.xlsx"
    snapshot = directory / "versions/version-1/snapshot.xlsx"
    for path in (original, working, snapshot):
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "销售"
        for row in rows or [["名称", "数量"], ["apple", 2], ["banana", 1]]:
            sheet.append(row)
        workbook.save(path)
        workbook.close()
    version = VersionRecord(
        "version-1",
        "file-1",
        None,
        "导入原始文件",
        datetime(2026, 8, 15, 8, 0, tzinfo=UTC),
        "import",
        None,
        snapshot,
        sha256(snapshot.read_bytes()).hexdigest(),
    )
    record = ImportedWorkbook("file-1", "销售.xlsx", original, working, version)
    MetadataStore(library_root).record_import(record)
    return record


def test_delete_head_from_tree_switches_working_version_and_can_be_undone(
    qtbot: QtBot,
    tmp_path: Path,
) -> None:
    library_root = tmp_path / "library"
    record = _seed_versioned_workbook(library_root)
    root = record.head_version
    assert root is not None
    child_snapshot = library_root / "files/file-1/versions/version-2/snapshot.xlsx"
    child_snapshot.parent.mkdir(parents=True)
    child_snapshot.write_bytes(record.working_path.read_bytes())
    child = VersionRecord(
        "version-2",
        record.file_id,
        root.version_id,
        "多列排序",
        datetime(2026, 8, 15, 9, 0, tzinfo=UTC),
        "sort",
        EngineName.PYTHON,
        child_snapshot,
        sha256(child_snapshot.read_bytes()).hexdigest(),
    )
    store = MetadataStore(library_root)
    store.record_child_version(child, root.version_id)
    confirmations: list[tuple[str, str]] = []
    task_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window

    def confirm_delete(_parent: QWidget, title: str, message: str) -> bool:
        confirmations.append((title, message))
        return True

    window = create_main_window(
        task_queue=task_queue,
        library_root=library_root,
        confirmation_presenter=confirm_delete,
    )
    qtbot.addWidget(window)
    window.show()
    tree_panel = _child(window, VersionTreePanel, "version-tree-panel")

    tree_panel.version_delete_requested.emit(record.file_id, child.version_id)

    delete_request = task_queue.submitted[-1]
    assert delete_request.operation == DELETE_VERSION_OPERATION
    assert delete_request.payload["replacement_version_id"] == root.version_id
    assert confirmations and "HEAD 将切换到“导入原始文件”" in confirmations[0][1]
    deleted_workbook = run_delete_version_task(delete_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            delete_request.task_id,
            TaskState.SUCCEEDED,
            delete_request.name,
            delete_request.file_id,
            EngineName.PYTHON,
            result=deleted_workbook,
        )
    )

    undo = _child(window, QPushButton, "version-undo-delete-button")
    qtbot.waitUntil(lambda: undo.isVisibleTo(window), timeout=500)
    assert store.get_workbook(record.file_id).head_version == root
    assert store.get_version(record.file_id, child.version_id).deleted_at is not None
    cards = {
        str(proxy.widget().property("version-id")): proxy.widget()
        for proxy in _child(window, QGraphicsView, "version-tree-view").scene().items()
        if isinstance(proxy, QGraphicsProxyWidget) and proxy.widget() is not None
    }
    assert cards[child.version_id].property("deleted") is True

    qtbot.mouseClick(undo, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    assert store.get_version(record.file_id, child.version_id).deleted_at is None
    assert store.get_workbook(record.file_id).head_version == root


def test_create_main_window_uses_product_identity(qtbot: QtBot, tmp_path: Path) -> None:
    try:
        from hyacinth.app import create_main_window
    except ModuleNotFoundError:
        pytest.fail("hyacinth.app.create_main_window is not implemented")

    window = create_main_window(library_root=tmp_path / "library")
    qtbot.addWidget(window)

    assert isinstance(window, QMainWindow)
    assert window.windowTitle() == "风信子"
    assert window.objectName() == "main-window"
    assert not window.windowIcon().isNull()


def test_initial_window_size_prefers_1440x900_and_adapts_to_available_screen() -> None:
    from hyacinth.app import initial_window_size

    assert initial_window_size(QSize(1920, 1040)) == QSize(1440, 900)
    assert initial_window_size(QSize(1366, 728)) == QSize(1229, 655)
    assert initial_window_size(QSize(900, 600)) == QSize(1024, 640)


def test_main_window_matches_approved_workspace_shell(qtbot: QtBot, tmp_path: Path) -> None:
    from hyacinth.app import create_main_window

    window = create_main_window(library_root=tmp_path / "library")
    qtbot.addWidget(window)

    main_splitter = _child(window, QSplitter, "main-workspace-splitter")
    import_button = _child(window, QPushButton, "toolbar-import-button")
    empty_import_button = _child(window, QPushButton, "preview-import-button")
    import_button_parent = import_button.parent()
    function_stack = _child(window, QStackedWidget, "function-body-stack")
    function_panel = _child(window, QFrame, "function-panel")
    editor_frame = _child(window, QFrame, "editor-frame")
    file_library = _child(window, QFrame, "file-library")

    assert _child(window, QLabel, "app-brand").text() == "风信子"
    assert _child(window, QLabel, "document-title").text() == "未选择文件"
    assert _child(window, QFrame, "function-panel").isEnabled()
    assert _child(window, QFrame, "file-library").isEnabled()
    assert _child(window, QFrame, "version-tree-panel").isEnabled()
    assert _child(window, QFrame, "formula-bar").isEnabled()
    assert _child(window, QFrame, "format-bar").isEnabled()
    assert main_splitter.count() == 3
    # 功能面板随参数表单一起位于右侧编辑区内，左侧只保留文件列表
    assert function_panel.parent() is editor_frame
    assert file_library.parent() is main_splitter
    assert import_button_parent is not None
    assert import_button_parent.objectName() == "top-toolbar"
    assert import_button.minimumHeight() >= 30
    assert not import_button.icon().isNull()
    assert not empty_import_button.isHidden()
    assert function_stack.currentIndex() == 0
    assert _child(window, QFrame, "function-footer").isHidden()
    assert not _child(window, QPushButton, "toolbar-save-version-button").isEnabled()


def test_version_tree_focus_mode_hides_other_regions_and_restores_layout(
    qtbot: QtBot,
    tmp_path: Path,
) -> None:
    from hyacinth.app import create_main_window

    window = create_main_window(library_root=tmp_path / "library")
    qtbot.addWidget(window)
    window.show()
    main_splitter = _child(window, QSplitter, "main-workspace-splitter")
    file_library = _child(window, QFrame, "file-library")

    version_tree = _child(window, QFrame, "version-tree-panel")
    editor = _child(window, QFrame, "editor-frame")
    header = _child(window, QFrame, "application-header")
    command_bar = _child(window, QFrame, "top-toolbar")
    focus_button = _child(window, QPushButton, "version-focus-button")
    main_sizes = main_splitter.sizes()

    qtbot.mouseClick(focus_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    assert version_tree.isVisibleTo(window)
    assert header.isHidden()
    assert command_bar.isHidden()
    assert file_library.isHidden()
    assert editor.isHidden()
    assert window.statusBar().isHidden()
    assert focus_button.text() == "退出专注"
    assert focus_button.accessibleName() == "退出版本图谱专注模式"

    qtbot.mouseClick(focus_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    assert not header.isHidden()
    assert not command_bar.isHidden()
    assert not file_library.isHidden()
    assert not editor.isHidden()
    assert not window.statusBar().isHidden()
    assert main_splitter.sizes() == main_sizes
    assert focus_button.text() == "专注"


def test_focus_mode_keeps_canvas_position(qtbot: QtBot, tmp_path: Path) -> None:
    library_root = tmp_path / "library"
    _seed_versioned_workbook(library_root)
    task_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window

    window = create_main_window(task_queue=task_queue, library_root=library_root)
    qtbot.addWidget(window)
    window.show()
    view = _child(window, QGraphicsView, "version-tree-view")
    focus_button = _child(window, QPushButton, "version-focus-button")
    target = QPointF(-1200.0, -900.0)
    view.centerOn(target)
    before = view.mapToScene(view.viewport().rect().center())
    assert abs(before.x() - target.x()) < 3.0
    assert abs(before.y() - target.y()) < 3.0

    qtbot.mouseClick(focus_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    after_enter = view.mapToScene(view.viewport().rect().center())
    assert abs(after_enter.x() - before.x()) < 5.0
    assert abs(after_enter.y() - before.y()) < 5.0

    qtbot.mouseClick(focus_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    after_exit = view.mapToScene(view.viewport().rect().center())
    assert abs(after_exit.x() - before.x()) < 5.0
    assert abs(after_exit.y() - before.y()) < 5.0


def test_empty_preview_import_button_uses_normal_import_flow(
    qtbot: QtBot,
    tmp_path: Path,
) -> None:
    source = tmp_path / "空状态导入.xlsx"
    source.write_bytes(b"source")
    task_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window

    window = create_main_window(
        task_queue=task_queue,
        library_root=tmp_path / "library",
        file_picker=lambda _parent: source,
    )
    qtbot.addWidget(window)
    window.show()
    qtbot.mouseClick(  # type: ignore[no-untyped-call]
        _child(window, QPushButton, "preview-import-button"),
        Qt.MouseButton.LeftButton,
    )

    assert len(_submitted_without_storage_stats(task_queue)) == 1
    assert task_queue.submitted[0].operation == IMPORT_WORKBOOK_OPERATION


def test_manual_cell_edit_saves_new_child_version(qtbot: QtBot, tmp_path: Path) -> None:
    library_root = tmp_path / "library"
    record = _seed_versioned_workbook(library_root)
    task_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window

    window = create_main_window(task_queue=task_queue, library_root=library_root)
    qtbot.addWidget(window)
    window.show()
    preview_request = _preview_request_of(task_queue)
    preview = run_preview_index_task(preview_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            task_id=preview_request.task_id,
            state=TaskState.SUCCEEDED,
            name=preview_request.name,
            file_id=record.file_id,
            engine=None,
            result=preview,
        )
    )
    table = _child(window, QTableView, "preview-table")
    qtbot.waitUntil(lambda: table.model() is not None, timeout=500)
    model = table.model()
    assert model is not None
    edited_cell = model.index(1, 0)

    assert model.setData(edited_cell, "pear", Qt.ItemDataRole.EditRole)
    save_button = _child(window, QPushButton, "toolbar-save-version-button")
    undo_button = _child(window, QPushButton, "toolbar-undo-button")
    assert save_button.isEnabled()
    assert undo_button.isEnabled()
    qtbot.mouseClick(save_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]
    save_request = task_queue.submitted[-1]
    assert save_request.operation == SAVE_MANUAL_EDITS_OPERATION
    assert save_request.payload["edits"] == [
        {"sheet_name": "销售", "row": 1, "column": 0, "value": "pear"}
    ]

    saved = run_save_manual_edits_task(save_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            task_id=save_request.task_id,
            state=TaskState.SUCCEEDED,
            name=save_request.name,
            file_id=record.file_id,
            engine=EngineName.PYTHON,
            result=saved,
        )
    )
    qtbot.waitUntil(lambda: len(MetadataStore(library_root).list_versions(record.file_id)) == 2)

    versions = MetadataStore(library_root).list_versions(record.file_id)
    assert versions[-1].operation == "manual-edit"
    assert versions[-1].parent_version_id == versions[0].version_id
    assert not save_button.isEnabled()
    assert not undo_button.isEnabled()


def test_unsaved_cell_edits_block_close_until_discarded(qtbot: QtBot, tmp_path: Path) -> None:
    library_root = tmp_path / "library"
    record = _seed_versioned_workbook(library_root)
    task_queue = FakeApplicationTaskQueue([])
    choices = ["cancel", "discard"]
    from hyacinth.app import create_main_window

    window = create_main_window(
        task_queue=task_queue,
        library_root=library_root,
        unsaved_changes_presenter=lambda _parent, _action: choices.pop(0),
    )
    qtbot.addWidget(window)
    window.show()
    preview_request = _preview_request_of(task_queue)
    preview = run_preview_index_task(preview_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            task_id=preview_request.task_id,
            state=TaskState.SUCCEEDED,
            name=preview_request.name,
            file_id=record.file_id,
            engine=None,
            result=preview,
        )
    )
    table = _child(window, QTableView, "preview-table")
    qtbot.waitUntil(lambda: table.model() is not None, timeout=500)
    model = table.model()
    assert model is not None
    assert model.setData(model.index(1, 0), "pear", Qt.ItemDataRole.EditRole)

    assert not window.close()
    assert not task_queue.shutdown_called
    assert _child(window, QPushButton, "toolbar-save-version-button").isEnabled()
    assert window.close()
    assert task_queue.shutdown_called


def test_version_node_save_as_exports_and_reports_destination(
    qtbot: QtBot,
    tmp_path: Path,
) -> None:
    library_root = tmp_path / "library"
    record = _seed_versioned_workbook(library_root)
    version = record.head_version
    assert version is not None
    destination = tmp_path / "exports" / "自定义名称.xlsx"
    exported_paths: list[Path] = []
    task_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window

    window = create_main_window(
        task_queue=task_queue,
        library_root=library_root,
        save_as_picker=lambda _parent, _suggested: destination,
        export_presenter=lambda _parent, path: exported_paths.append(path),
    )
    qtbot.addWidget(window)
    window.show()
    export_button = _child(window, QPushButton, "toolbar-export-button")
    assert export_button.isEnabled()
    version_tree = _child(window, VersionTreePanel, "version-tree-panel")
    version_tree.version_export_requested.emit(record.file_id, version.version_id, True)
    request = task_queue.submitted[-1]
    assert request.operation == EXPORT_VERSION_OPERATION
    assert request.payload["destination_path"] == str(destination)

    result = run_export_version_task(request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            task_id=request.task_id,
            state=TaskState.SUCCEEDED,
            name=request.name,
            file_id=record.file_id,
            engine=None,
            result=result,
        )
    )
    qtbot.waitUntil(lambda: bool(exported_paths), timeout=500)

    assert exported_paths == [destination]
    assert destination.read_bytes() == record.original_path.read_bytes()


def test_main_runs_qt_event_loop() -> None:
    environment = os.environ.copy()
    environment["QT_QPA_PLATFORM"] = "offscreen"
    completed = subprocess.run(
        [
            sys.executable,
            "-c",
            (
                "from PySide6.QtCore import QTimer; "
                "from PySide6.QtWidgets import QApplication; "
                "from hyacinth.__main__ import main; "
                "app = QApplication.instance() or QApplication([]); "
                "QTimer.singleShot(0, app.quit); "
                "raise SystemExit(main([]))"
            ),
        ],
        capture_output=True,
        text=True,
        timeout=10,
        check=False,
        env=environment,
    )

    assert completed.returncode == 0, completed.stderr


def test_main_window_connects_task_queue_to_status_bar(qtbot: QtBot) -> None:
    event = TaskEvent(
        task_id="convert-1",
        state=TaskState.RUNNING,
        name="转换旧版工作簿",
        file_id="销售报表.xls",
        engine=EngineName.COM,
        progress=None,
        elapsed_seconds=1.2,
    )
    task_queue = FakeApplicationTaskQueue([event])

    from hyacinth.app import create_main_window

    window = create_main_window(task_queue=task_queue)
    qtbot.addWidget(window)
    window.show()
    status = _child(window, TaskStatusWidget, "task-status")

    qtbot.waitUntil(
        lambda: _child(status, QLabel, "task-status-state").text() == "处理中",
        timeout=500,
    )
    qtbot.mouseClick(  # type: ignore[no-untyped-call]
        _child(status, QPushButton, "task-status-cancel"),
        Qt.MouseButton.LeftButton,
    )
    window.close()

    assert task_queue.cancelled == ["convert-1"]
    assert task_queue.shutdown_called is True


def test_import_button_submits_task_and_lists_successful_result(
    qtbot: QtBot,
    tmp_path: Path,
) -> None:
    source = tmp_path / "销售报表.xlsx"
    source.write_bytes(b"source")
    library_root = tmp_path / "library"
    task_queue = FakeApplicationTaskQueue([])

    from hyacinth.app import create_main_window

    window = create_main_window(
        task_queue=task_queue,
        library_root=library_root,
        file_picker=lambda _parent: source,
    )
    qtbot.addWidget(window)
    window.show()
    qtbot.mouseClick(  # type: ignore[no-untyped-call]
        _child(window, QPushButton, "toolbar-import-button"),
        Qt.MouseButton.LeftButton,
    )

    assert len(_submitted_without_storage_stats(task_queue)) == 1
    request = task_queue.submitted[0]
    assert request.operation == IMPORT_WORKBOOK_OPERATION
    assert request.payload == {
        "source_path": str(source),
        "library_root": str(library_root),
    }

    directory = library_root / "files" / request.file_id
    result = ImportedWorkbook(
        file_id=request.file_id,
        display_name=source.name,
        original_path=directory / "original" / source.name,
        working_path=directory / "working" / "current.xlsx",
    )
    task_queue.push_event(
        TaskEvent(
            task_id=request.task_id,
            state=TaskState.SUCCEEDED,
            name=request.name,
            file_id=request.file_id,
            engine=None,
            result=result,
        )
    )
    file_list = _child(window, QListWidget, "library-file-list")
    qtbot.waitUntil(lambda: file_list.count() == 1, timeout=500)

    assert file_list.item(0).text() == source.name
    assert file_list.currentRow() == 0


def test_failed_import_shows_reason_and_keeps_import_available(
    qtbot: QtBot,
    tmp_path: Path,
) -> None:
    source = tmp_path / "损坏.xlsx"
    source.write_bytes(b"invalid")
    task_queue = FakeApplicationTaskQueue([])
    errors: list[str] = []

    from hyacinth.app import create_main_window

    window = create_main_window(
        task_queue=task_queue,
        library_root=tmp_path / "library",
        file_picker=lambda _parent: source,
        error_presenter=lambda _parent, message: errors.append(message),
    )
    qtbot.addWidget(window)
    window.show()
    button = _child(window, QPushButton, "toolbar-import-button")
    qtbot.mouseClick(button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]
    request = task_queue.submitted[0]
    task_queue.push_event(
        TaskEvent(
            task_id=request.task_id,
            state=TaskState.FAILED,
            name=request.name,
            file_id=request.file_id,
            engine=None,
            message="工作簿无法打开",
        )
    )

    qtbot.waitUntil(lambda: errors == ["工作簿无法打开"], timeout=500)
    qtbot.mouseClick(button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    assert len(_submitted_without_storage_stats(task_queue)) == 2


def test_existing_file_loads_working_copy_and_renders_selected_sheet(
    qtbot: QtBot,
    tmp_path: Path,
) -> None:
    library_root = tmp_path / "library"
    directory = library_root / "files" / "file-1"
    original = directory / "original" / "销售.xlsx"
    working = directory / "working" / "current.xlsx"
    original.parent.mkdir(parents=True)
    working.parent.mkdir(parents=True)
    original.write_bytes(b"original")
    workbook = Workbook()
    sales = workbook.active
    assert sales is not None
    sales.title = "销售"
    sales["A1"] = "一月"
    workbook.create_sheet("库存")["B2"] = 42
    workbook.save(working)
    workbook.close()
    task_queue = FakeApplicationTaskQueue([])

    from hyacinth.app import create_main_window

    window = create_main_window(task_queue=task_queue, library_root=library_root)
    qtbot.addWidget(window)
    window.show()

    assert len(_submitted_without_storage_stats(task_queue)) == 1
    request = _preview_request_of(task_queue)
    assert request.payload["working_path"] == str(working)
    assert _child(window, QLabel, "document-title").text() == "销售.xlsx"
    assert window.windowTitle() == "风信子 — 销售.xlsx"
    preview = run_preview_index_task(request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            task_id=request.task_id,
            state=TaskState.SUCCEEDED,
            name=request.name,
            file_id=request.file_id,
            engine=None,
            result=preview,
        )
    )
    table = _child(window, QTableView, "preview-table")
    tabs = _child(window, QTabBar, "preview-sheet-tabs")
    qtbot.waitUntil(lambda: table.model() is not None, timeout=500)

    assert [tabs.tabText(index) for index in range(tabs.count())] == ["销售", "库存"]
    assert table.model().data(table.model().index(0, 0)) == "一月"
    window.close()
    preview.index_path.unlink()

    assert not preview.index_path.exists()


def test_switching_files_cancels_old_preview_and_ignores_its_result(
    qtbot: QtBot,
    tmp_path: Path,
) -> None:
    library_root = tmp_path / "library"
    records: list[ImportedWorkbook] = []
    for file_id, name in (("file-1", "一.xlsx"), ("file-2", "二.xlsx")):
        directory = library_root / "files" / file_id
        original = directory / "original" / name
        working = directory / "working" / "current.xlsx"
        original.parent.mkdir(parents=True)
        working.parent.mkdir(parents=True)
        original.write_bytes(b"original")
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet["A1"] = name
        workbook.save(working)
        workbook.close()
        records.append(ImportedWorkbook(file_id, name, original, working))
    task_queue = FakeApplicationTaskQueue([])

    from hyacinth.app import create_main_window

    window = create_main_window(task_queue=task_queue, library_root=library_root)
    qtbot.addWidget(window)
    window.show()
    first_request = _preview_request_of(task_queue)
    stale_preview = run_preview_index_task(first_request, PreviewTaskContext())
    file_list = _child(window, QListWidget, "library-file-list")
    file_list.setCurrentRow(1)

    assert len(_submitted_without_storage_stats(task_queue)) == 2
    assert task_queue.cancelled == [first_request.task_id]
    task_queue.push_event(
        TaskEvent(
            task_id=first_request.task_id,
            state=TaskState.SUCCEEDED,
            name=first_request.name,
            file_id=first_request.file_id,
            engine=None,
            result=stale_preview,
        )
    )
    state = _child(window, QLabel, "preview-state")
    qtbot.wait(100)

    assert "正在加载" in state.text()


def test_sort_preview_apply_creates_child_and_refreshes_tree(
    qtbot: QtBot,
    tmp_path: Path,
) -> None:
    library_root = tmp_path / "library"
    record = _seed_versioned_workbook(library_root)
    task_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window

    window = create_main_window(task_queue=task_queue, library_root=library_root)
    qtbot.addWidget(window)
    window.show()
    initial_request = _preview_request_of(task_queue)
    base_preview = run_preview_index_task(initial_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            initial_request.task_id,
            TaskState.SUCCEEDED,
            initial_request.name,
            initial_request.file_id,
            None,
            result=base_preview,
        )
    )
    preview_button = _child(window, QPushButton, "function-preview-button")
    qtbot.waitUntil(preview_button.isEnabled, timeout=500)
    primary = _child(window, QComboBox, "sort-primary-column")
    primary.setCurrentIndex(1)
    qtbot.mouseClick(preview_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    sort_request = task_queue.submitted[-1]
    assert sort_request.operation == SORT_PREVIEW_OPERATION
    assert not _child(window, QFrame, "top-toolbar").isEnabled()
    assert not _child(window, QListWidget, "library-file-list").isEnabled()
    parent = record.head_version
    assert parent is not None
    assert sort_request.payload["source_path"] == str(parent.snapshot_path)
    sort_result = run_sort_preview_task(sort_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            sort_request.task_id,
            TaskState.SUCCEEDED,
            sort_request.name,
            sort_request.file_id,
            EngineName.PYTHON,
            result=sort_result,
        )
    )
    qtbot.waitUntil(lambda: len(_submitted_without_storage_stats(task_queue)) == 3, timeout=500)
    temporary_index_request = task_queue.submitted[-1]
    temporary_preview = run_preview_index_task(temporary_index_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            temporary_index_request.task_id,
            TaskState.SUCCEEDED,
            temporary_index_request.name,
            temporary_index_request.file_id,
            None,
            result=temporary_preview,
        )
    )
    apply_button = _child(window, QPushButton, "function-apply-button")
    banner = _child(window, QLabel, "temporary-result-banner")
    qtbot.waitUntil(apply_button.isEnabled, timeout=500)
    assert banner.isVisible()
    assert _child(window, QFrame, "top-toolbar").isEnabled()
    assert _child(window, QListWidget, "library-file-list").isEnabled()
    qtbot.mouseClick(apply_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    apply_request = task_queue.submitted[-1]
    assert apply_request.operation == APPLY_SORT_PREVIEW_OPERATION
    assert not _child(window, QFrame, "top-toolbar").isEnabled()
    assert not _child(window, QListWidget, "library-file-list").isEnabled()
    assert apply_request.payload["preview_path"] == str(sort_result.preview_path)
    assert apply_request.payload["preview_hash"] == sort_result.content_hash
    applied = run_apply_sort_preview_task(apply_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            apply_request.task_id,
            TaskState.SUCCEEDED,
            apply_request.name,
            apply_request.file_id,
            EngineName.PYTHON,
            result=applied,
        )
    )
    qtbot.waitUntil(lambda: len(_submitted_without_storage_stats(task_queue)) == 5, timeout=500)
    tree = _child(window, QGraphicsView, "version-tree-view")
    proxies = [item for item in tree.scene().items() if isinstance(item, QGraphicsProxyWidget)]
    head = MetadataStore(library_root).get_workbook("file-1").head_version

    assert len(proxies) == 2
    assert head is not None and head.parent_version_id == "version-1"
    assert not sort_result.preview_path.parent.exists()
    assert _child(window, QFrame, "top-toolbar").isEnabled()
    assert _child(window, QListWidget, "library-file-list").isEnabled()


def test_deduplicate_preview_apply_creates_child_and_shows_statistics(
    qtbot: QtBot,
    tmp_path: Path,
) -> None:
    library_root = tmp_path / "library"
    record = _seed_versioned_workbook(
        library_root,
        [
            ["名称", "类别"],
            [" Apple ", "水果"],
            ["apple", "水果"],
            ["banana", "水果"],
        ],
    )
    task_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window

    window = create_main_window(task_queue=task_queue, library_root=library_root)
    qtbot.addWidget(window)
    window.show()
    initial_request = _preview_request_of(task_queue)
    base_preview = run_preview_index_task(initial_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            initial_request.task_id,
            TaskState.SUCCEEDED,
            initial_request.name,
            initial_request.file_id,
            None,
            result=base_preview,
        )
    )
    preview_button = _child(window, QPushButton, "function-preview-button")
    qtbot.waitUntil(preview_button.isEnabled, timeout=500)
    operation = _child(window, QComboBox, "processing-operation")
    operation.setCurrentIndex(operation.findData("deduplicate"))
    columns = _child(window, QListWidget, "deduplicate-key-columns")
    columns.item(0).setSelected(True)
    _child(window, QCheckBox, "deduplicate-ignore-case").setChecked(True)
    _child(window, QCheckBox, "deduplicate-trim-whitespace").setChecked(True)
    qtbot.mouseClick(preview_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    deduplicate_request = task_queue.submitted[-1]
    assert deduplicate_request.operation == DEDUPLICATE_PREVIEW_OPERATION
    assert deduplicate_request.payload["key_columns"] == [0]
    assert deduplicate_request.payload["ignore_case"] is True
    assert deduplicate_request.payload["trim_whitespace"] is True
    parent = record.head_version
    assert parent is not None
    assert deduplicate_request.payload["source_path"] == str(parent.snapshot_path)
    result = run_deduplicate_preview_task(deduplicate_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            deduplicate_request.task_id,
            TaskState.SUCCEEDED,
            deduplicate_request.name,
            deduplicate_request.file_id,
            EngineName.PYTHON,
            result=result,
        )
    )
    qtbot.waitUntil(lambda: len(_submitted_without_storage_stats(task_queue)) == 3, timeout=500)
    temporary_index_request = task_queue.submitted[-1]
    temporary_preview = run_preview_index_task(temporary_index_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            temporary_index_request.task_id,
            TaskState.SUCCEEDED,
            temporary_index_request.name,
            temporary_index_request.file_id,
            None,
            result=temporary_preview,
        )
    )
    apply_button = _child(window, QPushButton, "function-apply-button")
    state = _child(window, QLabel, "sort-state")
    details = _child(window, QPushButton, "deduplicate-details-button")
    qtbot.waitUntil(apply_button.isEnabled, timeout=500)
    assert "1 个重复组" in state.text()
    assert "删除 1 行" in state.text()
    assert details.isEnabled()
    qtbot.mouseClick(apply_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    apply_request = task_queue.submitted[-1]
    assert apply_request.operation == APPLY_DEDUPLICATE_PREVIEW_OPERATION
    assert apply_request.payload["deleted_rows"] == 1
    applied = run_apply_deduplicate_preview_task(apply_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            apply_request.task_id,
            TaskState.SUCCEEDED,
            apply_request.name,
            apply_request.file_id,
            EngineName.PYTHON,
            result=applied,
        )
    )
    qtbot.waitUntil(lambda: len(_submitted_without_storage_stats(task_queue)) == 5, timeout=500)
    head = MetadataStore(library_root).get_workbook("file-1").head_version
    tree = _child(window, QGraphicsView, "version-tree-view")
    proxies = [item for item in tree.scene().items() if isinstance(item, QGraphicsProxyWidget)]

    assert head is not None and head.operation == "delete-duplicates"
    assert head.parent_version_id == "version-1"
    assert len(proxies) == 2
    assert not result.preview_path.parent.exists()


def test_delete_blank_rows_preview_apply_creates_child_and_shows_original_rows(
    qtbot: QtBot,
    tmp_path: Path,
) -> None:
    library_root = tmp_path / "library"
    record = _seed_versioned_workbook(
        library_root,
        [
            ["名称", "类别"],
            ["apple", "水果"],
            [None, None],
            ["banana", "水果"],
        ],
    )
    task_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window

    window = create_main_window(task_queue=task_queue, library_root=library_root)
    qtbot.addWidget(window)
    window.show()
    initial_request = _preview_request_of(task_queue)
    base_preview = run_preview_index_task(initial_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            initial_request.task_id,
            TaskState.SUCCEEDED,
            initial_request.name,
            initial_request.file_id,
            None,
            result=base_preview,
        )
    )
    preview_button = _child(window, QPushButton, "function-preview-button")
    qtbot.waitUntil(preview_button.isEnabled, timeout=500)
    operation = _child(window, QComboBox, "processing-operation")
    operation.setCurrentIndex(operation.findData("delete_blank_rows"))
    columns = _child(window, QListWidget, "blank-rows-key-columns")
    columns.item(0).setSelected(True)
    qtbot.mouseClick(preview_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    delete_request = task_queue.submitted[-1]
    assert delete_request.operation == DELETE_BLANK_ROWS_PREVIEW_OPERATION
    assert delete_request.payload["key_columns"] == [0]
    assert delete_request.payload["allow_unsafe"] is False
    parent = record.head_version
    assert parent is not None
    assert delete_request.payload["source_path"] == str(parent.snapshot_path)
    result = run_delete_blank_rows_preview_task(delete_request, PreviewTaskContext())
    assert result.deleted_row_numbers == (3,)
    task_queue.push_event(
        TaskEvent(
            delete_request.task_id,
            TaskState.SUCCEEDED,
            delete_request.name,
            delete_request.file_id,
            EngineName.PYTHON,
            result=result,
        )
    )
    qtbot.waitUntil(lambda: len(_submitted_without_storage_stats(task_queue)) == 3, timeout=500)
    temporary_index_request = task_queue.submitted[-1]
    temporary_preview = run_preview_index_task(temporary_index_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            temporary_index_request.task_id,
            TaskState.SUCCEEDED,
            temporary_index_request.name,
            temporary_index_request.file_id,
            None,
            result=temporary_preview,
        )
    )
    apply_button = _child(window, QPushButton, "function-apply-button")
    state = _child(window, QLabel, "sort-state")
    details = _child(window, QPushButton, "blank-rows-details-button")
    qtbot.waitUntil(apply_button.isEnabled, timeout=500)
    assert "删除 1 行" in state.text()
    assert details.isEnabled()
    qtbot.mouseClick(apply_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    apply_request = task_queue.submitted[-1]
    assert apply_request.operation == APPLY_DELETE_BLANK_ROWS_PREVIEW_OPERATION
    assert apply_request.payload["deleted_row_numbers"] == [3]
    applied = run_apply_delete_blank_rows_preview_task(apply_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            apply_request.task_id,
            TaskState.SUCCEEDED,
            apply_request.name,
            apply_request.file_id,
            EngineName.PYTHON,
            result=applied,
        )
    )
    qtbot.waitUntil(lambda: len(_submitted_without_storage_stats(task_queue)) == 5, timeout=500)
    head = MetadataStore(library_root).get_workbook("file-1").head_version
    tree = _child(window, QGraphicsView, "version-tree-view")
    proxies = [item for item in tree.scene().items() if isinstance(item, QGraphicsProxyWidget)]

    assert head is not None and head.operation == "delete-blank-rows"
    assert head.parent_version_id == "version-1"
    assert len(proxies) == 2
    assert not result.preview_path.parent.exists()


def test_filter_preview_apply_creates_child_and_only_shows_matching_rows(
    qtbot: QtBot,
    tmp_path: Path,
) -> None:
    library_root = tmp_path / "library"
    record = _seed_versioned_workbook(
        library_root,
        [
            ["名称", "数量"],
            ["apple", 2],
            ["apple", 5],
            ["banana", 8],
        ],
    )
    task_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window

    window = create_main_window(task_queue=task_queue, library_root=library_root)
    qtbot.addWidget(window)
    window.show()
    initial_request = _preview_request_of(task_queue)
    base_preview = run_preview_index_task(initial_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            initial_request.task_id,
            TaskState.SUCCEEDED,
            initial_request.name,
            initial_request.file_id,
            None,
            result=base_preview,
        )
    )
    preview_button = _child(window, QPushButton, "function-preview-button")
    qtbot.waitUntil(preview_button.isEnabled, timeout=500)
    operation = _child(window, QComboBox, "processing-operation")
    operation.setCurrentIndex(operation.findData("filter"))
    first_operator = _child(window, QComboBox, "filter-first-operator")
    first_operator.setCurrentIndex(first_operator.findData("contains"))
    _child(window, QLineEdit, "filter-first-value").setText("apple")
    _child(window, QCheckBox, "filter-enable-second").setChecked(True)
    second_column = _child(window, QComboBox, "filter-second-column")
    second_column.setCurrentIndex(1)
    second_type = _child(window, QComboBox, "filter-second-type")
    second_type.setCurrentIndex(second_type.findData("number"))
    second_operator = _child(window, QComboBox, "filter-second-operator")
    second_operator.setCurrentIndex(second_operator.findData("greater_than"))
    _child(window, QLineEdit, "filter-second-value").setText("3")
    qtbot.mouseClick(preview_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    filter_request = task_queue.submitted[-1]
    assert filter_request.operation == FILTER_PREVIEW_OPERATION
    assert filter_request.payload["connector"] == "and"
    assert filter_request.payload["conditions"] == [
        {
            "column_index": 0,
            "operator": "contains",
            "value_type": "text",
            "value": "apple",
            "second_value": None,
        },
        {
            "column_index": 1,
            "operator": "greater_than",
            "value_type": "number",
            "value": "3",
            "second_value": None,
        },
    ]
    parent = record.head_version
    assert parent is not None
    assert filter_request.payload["source_path"] == str(parent.snapshot_path)
    result = run_filter_preview_task(filter_request, PreviewTaskContext())
    assert result.matched_rows == 1
    assert result.total_rows == 3
    assert result.hidden_row_numbers == (2, 4)
    task_queue.push_event(
        TaskEvent(
            filter_request.task_id,
            TaskState.SUCCEEDED,
            filter_request.name,
            filter_request.file_id,
            EngineName.PYTHON,
            result=result,
        )
    )
    qtbot.waitUntil(lambda: len(_submitted_without_storage_stats(task_queue)) == 3, timeout=500)
    temporary_index_request = task_queue.submitted[-1]
    temporary_preview = run_preview_index_task(temporary_index_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            temporary_index_request.task_id,
            TaskState.SUCCEEDED,
            temporary_index_request.name,
            temporary_index_request.file_id,
            None,
            result=temporary_preview,
        )
    )
    apply_button = _child(window, QPushButton, "function-apply-button")
    state = _child(window, QLabel, "sort-state")
    table = _child(window, QTableView, "preview-table")
    qtbot.waitUntil(apply_button.isEnabled, timeout=500)
    assert "匹配 1 / 3 行" in state.text()
    assert "33.3%" in state.text()
    assert table.model().data(table.model().index(0, 0)) == "名称"
    assert table.model().data(table.model().index(1, 0)) == "apple"
    assert table.model().data(table.model().index(1, 1)) == "5"
    qtbot.mouseClick(apply_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    apply_request = task_queue.submitted[-1]
    assert apply_request.operation == APPLY_FILTER_PREVIEW_OPERATION
    assert apply_request.payload["matched_rows"] == 1
    assert apply_request.payload["total_rows"] == 3
    applied = run_apply_filter_preview_task(apply_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            apply_request.task_id,
            TaskState.SUCCEEDED,
            apply_request.name,
            apply_request.file_id,
            EngineName.PYTHON,
            result=applied,
        )
    )
    qtbot.waitUntil(lambda: len(_submitted_without_storage_stats(task_queue)) == 5, timeout=500)
    head = MetadataStore(library_root).get_workbook("file-1").head_version
    tree = _child(window, QGraphicsView, "version-tree-view")
    proxies = [item for item in tree.scene().items() if isinstance(item, QGraphicsProxyWidget)]

    assert head is not None and head.operation == "filter"
    assert head.parent_version_id == "version-1"
    assert len(proxies) == 2
    assert not result.preview_path.parent.exists()


def test_historical_preview_checkout_and_processing_create_branch(
    qtbot: QtBot,
    tmp_path: Path,
) -> None:
    library_root = tmp_path / "library"
    record = _seed_versioned_workbook(
        library_root,
        [["名称", "数量"], ["banana", 1], ["apple", 2]],
    )
    root = record.head_version
    assert root is not None
    child_snapshot = library_root / "files/file-1/versions/version-2/snapshot.xlsx"
    child_snapshot.parent.mkdir(parents=True)
    child_workbook = Workbook()
    child_sheet = child_workbook.active
    assert child_sheet is not None
    child_sheet.title = "销售"
    child_rows: list[list[str | int]] = [
        ["名称", "数量"],
        ["apple", 2],
        ["banana", 1],
    ]
    for row in child_rows:
        child_sheet.append(row)
    child_workbook.save(child_snapshot)
    child_workbook.close()
    record.working_path.write_bytes(child_snapshot.read_bytes())
    child = VersionRecord(
        "version-2",
        record.file_id,
        root.version_id,
        "已有排序分支",
        datetime(2026, 8, 15, 9, 0, tzinfo=UTC),
        "sort",
        EngineName.PYTHON,
        child_snapshot,
        sha256(child_snapshot.read_bytes()).hexdigest(),
    )
    store = MetadataStore(library_root)
    store.record_child_version(child, root.version_id)
    task_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window

    window = create_main_window(task_queue=task_queue, library_root=library_root)
    qtbot.addWidget(window)
    window.show()
    initial_request = _preview_request_of(task_queue)
    initial_preview = run_preview_index_task(initial_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            initial_request.task_id,
            TaskState.SUCCEEDED,
            initial_request.name,
            initial_request.file_id,
            None,
            result=initial_preview,
        )
    )
    preview_button = _child(window, QPushButton, "function-preview-button")
    qtbot.waitUntil(preview_button.isEnabled, timeout=500)
    tree = _child(window, QGraphicsView, "version-tree-view")
    cards = {
        str(proxy.widget().property("version-id")): proxy.widget()
        for proxy in tree.scene().items()
        if isinstance(proxy, QGraphicsProxyWidget) and proxy.widget() is not None
    }
    qtbot.mouseClick(cards[root.version_id], Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    qtbot.waitUntil(lambda: len(_submitted_without_storage_stats(task_queue)) == 2, timeout=500)
    historical_request = task_queue.submitted[-1]
    assert historical_request.operation == BUILD_PREVIEW_INDEX_OPERATION
    assert historical_request.payload["working_path"] == str(root.snapshot_path)
    assert store.get_workbook(record.file_id).head_version == child
    historical_preview = run_preview_index_task(historical_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            historical_request.task_id,
            TaskState.SUCCEEDED,
            historical_request.name,
            historical_request.file_id,
            None,
            result=historical_preview,
        )
    )
    table = _child(window, QTableView, "preview-table")
    qtbot.waitUntil(lambda: table.model() is not None, timeout=500)
    assert table.model().data(table.model().index(1, 0)) == "banana"
    assert not _child(window, QFrame, "function-panel").isEnabled()

    continue_button = _child(window, QPushButton, "version-continue-button")
    qtbot.mouseClick(continue_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]
    checkout_request = task_queue.submitted[-1]
    assert checkout_request.operation == CHECKOUT_VERSION_OPERATION
    assert checkout_request.payload["version_id"] == root.version_id
    assert checkout_request.payload["expected_head_version_id"] == child.version_id
    checked_out = run_checkout_version_task(checkout_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            checkout_request.task_id,
            TaskState.SUCCEEDED,
            checkout_request.name,
            checkout_request.file_id,
            EngineName.PYTHON,
            result=checked_out,
        )
    )
    qtbot.waitUntil(lambda: len(_submitted_without_storage_stats(task_queue)) == 4, timeout=500)
    checked_out_preview_request = task_queue.submitted[-1]
    checked_out_preview = run_preview_index_task(
        checked_out_preview_request,
        PreviewTaskContext(),
    )
    task_queue.push_event(
        TaskEvent(
            checked_out_preview_request.task_id,
            TaskState.SUCCEEDED,
            checked_out_preview_request.name,
            checked_out_preview_request.file_id,
            None,
            result=checked_out_preview,
        )
    )
    qtbot.waitUntil(preview_button.isEnabled, timeout=500)
    assert store.get_workbook(record.file_id).head_version == root

    primary = _child(window, QComboBox, "sort-primary-column")
    primary.setCurrentIndex(0)
    qtbot.mouseClick(preview_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]
    sort_request = task_queue.submitted[-1]
    assert sort_request.payload["parent_version_id"] == root.version_id
    sort_result = run_sort_preview_task(sort_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            sort_request.task_id,
            TaskState.SUCCEEDED,
            sort_request.name,
            sort_request.file_id,
            EngineName.PYTHON,
            result=sort_result,
        )
    )
    qtbot.waitUntil(lambda: len(_submitted_without_storage_stats(task_queue)) == 6, timeout=500)
    temporary_index_request = task_queue.submitted[-1]
    temporary_preview = run_preview_index_task(temporary_index_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            temporary_index_request.task_id,
            TaskState.SUCCEEDED,
            temporary_index_request.name,
            temporary_index_request.file_id,
            None,
            result=temporary_preview,
        )
    )
    apply_button = _child(window, QPushButton, "function-apply-button")
    qtbot.waitUntil(apply_button.isEnabled, timeout=500)
    qtbot.mouseClick(apply_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]
    apply_request = task_queue.submitted[-1]
    assert apply_request.payload["parent_version_id"] == root.version_id
    applied = run_apply_sort_preview_task(apply_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            apply_request.task_id,
            TaskState.SUCCEEDED,
            apply_request.name,
            apply_request.file_id,
            EngineName.PYTHON,
            result=applied,
        )
    )
    qtbot.waitUntil(lambda: len(_submitted_without_storage_stats(task_queue)) == 8, timeout=500)
    versions = store.list_versions(record.file_id)
    head = store.get_workbook(record.file_id).head_version

    assert head is not None and head.parent_version_id == root.version_id
    assert child in versions
    assert (
        len([version for version in versions if version.parent_version_id == root.version_id]) == 2
    )


def test_dragged_version_position_persists_after_reopening_tree(
    qtbot: QtBot,
    tmp_path: Path,
) -> None:
    library_root = tmp_path / "library"
    record = _seed_versioned_workbook(library_root)
    root = record.head_version
    assert root is not None
    first_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window

    first_window = create_main_window(task_queue=first_queue, library_root=library_root)
    qtbot.addWidget(first_window)
    first_window.show()
    request = _preview_request_of(first_queue)
    preview = run_preview_index_task(request, PreviewTaskContext())
    first_queue.push_event(
        TaskEvent(
            request.task_id,
            TaskState.SUCCEEDED,
            request.name,
            request.file_id,
            None,
            result=preview,
        )
    )
    tree = _child(first_window, QGraphicsView, "version-tree-view")
    proxy = next(
        item
        for item in tree.scene().items()
        if isinstance(item, QGraphicsProxyWidget)
        and item.widget() is not None
        and item.widget().property("version-id") == root.version_id
    )
    card = proxy.widget()
    assert card is not None
    initial_position = proxy.pos()
    center = tree.mapFromScene(proxy.sceneBoundingRect().center())
    qtbot.mousePress(  # type: ignore[no-untyped-call]
        tree.viewport(), Qt.MouseButton.LeftButton, pos=center
    )
    qtbot.mouseMove(tree.viewport(), pos=center + QPoint(60, 40))  # type: ignore[no-untyped-call]
    qtbot.mouseRelease(
        tree.viewport(),
        Qt.MouseButton.LeftButton,
        pos=center + QPoint(60, 40),
    )  # type: ignore[no-untyped-call]
    layouts = MetadataStore(library_root).list_version_layouts(record.file_id)
    assert layouts[root.version_id].x > initial_position.x()
    first_window.close()

    second_queue = FakeApplicationTaskQueue([])
    second_window = create_main_window(task_queue=second_queue, library_root=library_root)
    qtbot.addWidget(second_window)
    second_window.show()
    reopened_tree = _child(second_window, QGraphicsView, "version-tree-view")
    reopened_proxy = next(
        item
        for item in reopened_tree.scene().items()
        if isinstance(item, QGraphicsProxyWidget)
        and item.widget() is not None
        and item.widget().property("version-id") == root.version_id
    )

    assert reopened_proxy.pos().x() == layouts[root.version_id].x
    # 存储的 y 为泳道内容区相对坐标，渲染时加泳道内容区顶（42 + 40）
    assert reopened_proxy.pos().y() == layouts[root.version_id].y


def test_storage_status_shows_format_and_sizes_for_selected_and_previewed_version(
    qtbot: QtBot,
    tmp_path: Path,
) -> None:
    library_root = tmp_path / "library"
    record = _seed_versioned_workbook(library_root)
    root = record.head_version
    assert root is not None
    child_snapshot = library_root / "files/file-1/versions/version-2/snapshot.xlsx"
    child_snapshot.parent.mkdir(parents=True)
    child_snapshot.write_bytes(record.working_path.read_bytes() + b"extra-manual-edit-content")
    child = VersionRecord(
        "version-2",
        record.file_id,
        root.version_id,
        "手动编辑",
        datetime(2026, 8, 16, 9, 0, tzinfo=UTC),
        "manual-edit",
        None,
        child_snapshot,
        sha256(child_snapshot.read_bytes()).hexdigest(),
    )
    MetadataStore(library_root).record_child_version(child, root.version_id)
    task_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window

    window = create_main_window(task_queue=task_queue, library_root=library_root)
    qtbot.addWidget(window)
    window.show()
    format_label = _child(window, QLabel, "storage-format-pill")
    sizes_label = _child(window, QLabel, "storage-size-text")

    stats_request = task_queue.submitted[0]
    assert stats_request.operation == VERSION_STORAGE_STATS_OPERATION
    assert stats_request.payload["preview_version_id"] == child.version_id
    stats = run_version_storage_stats_task(stats_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            stats_request.task_id,
            TaskState.SUCCEEDED,
            stats_request.name,
            stats_request.file_id,
            None,
            result=stats,
        )
    )

    qtbot.waitUntil(lambda: "版本总占用" in sizes_label.text(), timeout=2000)
    expected_total = root.snapshot_path.stat().st_size + child_snapshot.stat().st_size
    assert format_label.text() == "XLSX"
    assert format_byte_size(expected_total) in sizes_label.text()
    assert format_byte_size(child_snapshot.stat().st_size) in sizes_label.text()

    tree_panel = _child(window, VersionTreePanel, "version-tree-panel")
    tree_panel.version_preview_requested.emit(record.file_id, root.version_id)

    switch_stats_request = task_queue.submitted[-2]
    assert switch_stats_request.operation == VERSION_STORAGE_STATS_OPERATION
    assert switch_stats_request.payload["preview_version_id"] == root.version_id


def test_storage_status_empty_state_without_selected_file(qtbot: QtBot, tmp_path: Path) -> None:
    task_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window

    window = create_main_window(task_queue=task_queue, library_root=tmp_path / "library")
    qtbot.addWidget(window)

    format_label = _child(window, QLabel, "storage-format-pill")
    sizes_label = _child(window, QLabel, "storage-size-text")

    assert format_label.text() == "未选择文件"
    assert sizes_label.text() == ""
    assert not any(
        request.operation == VERSION_STORAGE_STATS_OPERATION for request in task_queue.submitted
    )


def test_file_delete_from_library_moves_to_trash_and_clears_current_preview(
    qtbot: QtBot,
    tmp_path: Path,
) -> None:
    library_root = tmp_path / "library"
    record = _seed_versioned_workbook(library_root)
    task_queue = FakeApplicationTaskQueue([])
    confirmations: list[tuple[str, str]] = []
    from hyacinth.app import create_main_window

    def confirm_delete(_parent: QWidget, title: str, message: str) -> bool:
        confirmations.append((title, message))
        return True

    window = create_main_window(
        task_queue=task_queue,
        library_root=library_root,
        confirmation_presenter=confirm_delete,
    )
    qtbot.addWidget(window)
    window.show()
    file_list = _child(window, QListWidget, "library-file-list")

    assert file_list.count() == 1
    file_list.item(0).setSelected(True)
    from hyacinth.library import FileLibraryWidget

    library_panel = window.findChild(FileLibraryWidget, "file-library")
    assert library_panel is not None
    library_panel.workbook_delete_requested.emit(record)

    store = MetadataStore(library_root)
    assert confirmations and confirmations[0][0] == "删除文件"
    assert file_list.count() == 0
    assert store.list_workbooks() == ()
    assert [item.file_id for item in store.list_deleted_files()] == [record.file_id]
    assert _child(window, QLabel, "document-title").text() == "未选择文件"
    assert _child(window, QLabel, "storage-format-pill").text() == "未选择文件"
    assert record.head_version is not None
    assert record.head_version.snapshot_path.is_file()

    reopen_queue = FakeApplicationTaskQueue([])
    reopen_window = create_main_window(task_queue=reopen_queue, library_root=library_root)
    qtbot.addWidget(reopen_window)
    reopened_list = _child(reopen_window, QListWidget, "library-file-list")

    assert reopened_list.count() == 0


def test_recycle_bin_restores_file_to_library(qtbot: QtBot, tmp_path: Path) -> None:
    library_root = tmp_path / "library"
    record = _seed_versioned_workbook(library_root)
    store = MetadataStore(library_root)
    assert record.head_version is not None
    store.soft_delete_file(record.file_id, record.head_version.version_id)
    task_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window
    from hyacinth.ui import RecycleBinDialog

    window = create_main_window(task_queue=task_queue, library_root=library_root)
    qtbot.addWidget(window)
    window.show()
    recycle_button = _child(window, QPushButton, "toolbar-recycle-button")
    assert recycle_button.isEnabled()

    qtbot.mouseClick(recycle_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    dialog = window.findChild(RecycleBinDialog, "recycle-bin-dialog")
    assert dialog is not None
    bin_list = _child(dialog, QListWidget, "recycle-file-list")
    assert bin_list.count() == 1
    assert record.display_name in bin_list.item(0).text()
    bin_list.setCurrentRow(0)

    qtbot.mouseClick(
        _child(dialog, QPushButton, "recycle-restore-button"), Qt.MouseButton.LeftButton
    )  # type: ignore[no-untyped-call]

    file_list = _child(window, QListWidget, "library-file-list")
    assert file_list.count() == 1
    assert [item.file_id for item in store.list_workbooks()] == [record.file_id]
    assert store.list_deleted_files() == ()
    assert bin_list.count() == 0


def test_recycle_bin_purges_file_after_confirmation(qtbot: QtBot, tmp_path: Path) -> None:
    library_root = tmp_path / "library"
    record = _seed_versioned_workbook(library_root)
    store = MetadataStore(library_root)
    assert record.head_version is not None
    store.soft_delete_file(record.file_id, record.head_version.version_id)
    confirmations: list[tuple[str, str]] = []
    task_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window
    from hyacinth.ui import RecycleBinDialog

    def confirm_purge(_parent: QWidget, title: str, message: str) -> bool:
        confirmations.append((title, message))
        return True

    window = create_main_window(
        task_queue=task_queue,
        library_root=library_root,
        confirmation_presenter=confirm_purge,
    )
    qtbot.addWidget(window)
    window.show()

    recycle_button = _child(window, QPushButton, "toolbar-recycle-button")
    qtbot.mouseClick(recycle_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    from hyacinth.versioning import PURGE_FILE_OPERATION, run_purge_file_task

    dialog = window.findChild(RecycleBinDialog, "recycle-bin-dialog")
    assert dialog is not None
    bin_list = _child(dialog, QListWidget, "recycle-file-list")
    bin_list.setCurrentRow(0)

    qtbot.mouseClick(_child(dialog, QPushButton, "recycle-purge-button"), Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    assert confirmations and "无法恢复" in confirmations[0][1]
    purge_request = task_queue.submitted[-1]
    assert purge_request.operation == PURGE_FILE_OPERATION
    purged = run_purge_file_task(purge_request, PreviewTaskContext())
    task_queue.push_event(
        TaskEvent(
            purge_request.task_id,
            TaskState.SUCCEEDED,
            purge_request.name,
            purge_request.file_id,
            None,
            result=purged,
        )
    )

    qtbot.waitUntil(lambda: bin_list.count() == 0, timeout=2000)
    assert not (library_root / "files" / record.file_id).exists()
    assert store.list_deleted_files() == ()


def test_recycle_bin_lists_and_restores_deleted_version_node(
    qtbot: QtBot,
    tmp_path: Path,
) -> None:
    library_root = tmp_path / "library"
    record = _seed_versioned_workbook(library_root)
    root = record.head_version
    assert root is not None
    child_snapshot = library_root / "files/file-1/versions/version-2/snapshot.xlsx"
    child_snapshot.parent.mkdir(parents=True)
    child_snapshot.write_bytes(record.working_path.read_bytes() + b"child-extra")
    child = VersionRecord(
        "version-2",
        record.file_id,
        root.version_id,
        "多列排序",
        datetime(2026, 8, 16, 9, 0, tzinfo=UTC),
        "sort",
        None,
        child_snapshot,
        sha256(child_snapshot.read_bytes()).hexdigest(),
    )
    store = MetadataStore(library_root)
    store.record_child_version(child, root.version_id)
    store.soft_delete_version(record.file_id, child.version_id, child.version_id)
    task_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window
    from hyacinth.ui import RecycleBinDialog

    window = create_main_window(task_queue=task_queue, library_root=library_root)
    qtbot.addWidget(window)
    window.show()
    recycle_button = _child(window, QPushButton, "toolbar-recycle-button")

    qtbot.mouseClick(recycle_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    dialog = window.findChild(RecycleBinDialog, "recycle-bin-dialog")
    assert dialog is not None
    bin_list = _child(dialog, QListWidget, "recycle-file-list")
    assert bin_list.count() == 1
    entry_text = bin_list.item(0).text()
    assert entry_text.startswith("版本")
    assert record.display_name in entry_text
    assert "多列排序" in entry_text
    bin_list.setCurrentRow(0)

    qtbot.mouseClick(
        _child(dialog, QPushButton, "recycle-restore-button"), Qt.MouseButton.LeftButton
    )  # type: ignore[no-untyped-call]

    assert store.get_version(record.file_id, child.version_id).deleted_at is None
    assert bin_list.count() == 0
    view = _child(window, QGraphicsView, "version-tree-view")
    cards = {
        str(proxy.widget().property("version-id")): proxy.widget()
        for proxy in view.scene().items()
        if isinstance(proxy, QGraphicsProxyWidget) and proxy.widget() is not None
    }
    assert cards[child.version_id].property("deleted") is False


def test_global_canvas_renders_all_files_and_switches_on_node_click(
    qtbot: QtBot,
    tmp_path: Path,
) -> None:
    library_root = tmp_path / "library"
    record = _seed_versioned_workbook(library_root)
    second_directory = library_root / "files/file-2"
    second_original = second_directory / "original/库存.xlsx"
    second_working = second_directory / "working/current.xlsx"
    second_snapshot = second_directory / "versions/version-b/snapshot.xlsx"
    for path in (second_original, second_working, second_snapshot):
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "库存"
        sheet.append(["名称", "数量"])
        sheet.append(["螺丝", 8])
        workbook.save(path)
        workbook.close()
    second_root = VersionRecord(
        "version-b",
        "file-2",
        None,
        "导入原始文件",
        datetime(2026, 8, 16, 10, 0, tzinfo=UTC),
        "import",
        None,
        second_snapshot,
        sha256(second_snapshot.read_bytes()).hexdigest(),
    )
    second_record = ImportedWorkbook(
        "file-2",
        "库存.xlsx",
        second_original,
        second_working,
        second_root,
        datetime(2026, 8, 16, 10, 0, tzinfo=UTC),
    )
    MetadataStore(library_root).record_import(second_record)
    task_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window

    window = create_main_window(task_queue=task_queue, library_root=library_root)
    qtbot.addWidget(window)
    window.show()
    view = _child(window, QGraphicsView, "version-tree-view")

    cards = {
        str(proxy.widget().property("file-id")): proxy.widget()
        for proxy in view.scene().items()
        if isinstance(proxy, QGraphicsProxyWidget) and proxy.widget() is not None
    }
    assert set(cards) == {"file-1", "file-2"}
    assert window.windowTitle() == "风信子 — 库存.xlsx"

    QTest.mouseClick(cards["file-1"], Qt.MouseButton.LeftButton)

    assert window.windowTitle() == "风信子 — 销售.xlsx"
    assert _child(window, QLabel, "document-title").text() == "销售.xlsx"
    assert MetadataStore(library_root).get_workbook("file-2").display_name == "库存.xlsx"
    proxies = {
        str(proxy.widget().property("file-id")): proxy
        for proxy in view.scene().items()
        if isinstance(proxy, QGraphicsProxyWidget) and proxy.widget() is not None
    }
    assert record.head_version is not None
    head_center = proxies["file-1"].sceneBoundingRect().center()
    view_center = view.mapToScene(view.viewport().rect().center())
    assert abs(view_center.x() - head_center.x()) < 260.0
    assert abs(view_center.y() - head_center.y()) < 260.0

    file_list = _child(window, QListWidget, "library-file-list")
    for row in range(file_list.count()):
        if file_list.item(row).data(Qt.ItemDataRole.UserRole) == "file-2":
            file_list.setCurrentRow(row)
            break

    assert window.windowTitle() == "风信子 — 库存.xlsx"
    second_center = (
        next(
            proxy
            for proxy in view.scene().items()
            if isinstance(proxy, QGraphicsProxyWidget)
            and proxy.widget() is not None
            and str(proxy.widget().property("file-id")) == "file-2"
        )
        .sceneBoundingRect()
        .center()
    )
    view_after = view.mapToScene(view.viewport().rect().center())
    assert abs(view_after.x() - second_center.x()) < 260.0
    assert abs(view_after.y() - second_center.y()) < 260.0


def test_reset_layouts_restores_default_positions(qtbot: QtBot, tmp_path: Path) -> None:
    library_root = tmp_path / "library"
    record = _seed_versioned_workbook(library_root)
    root = record.head_version
    assert root is not None
    store = MetadataStore(library_root)
    store.save_version_layout(record.file_id, root.version_id, 900.0, -800.0, fixed=True)
    confirmations: list[tuple[str, str]] = []
    task_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window

    def confirm_reset(_parent: QWidget, title: str, message: str) -> bool:
        confirmations.append((title, message))
        return True

    window = create_main_window(
        task_queue=task_queue,
        library_root=library_root,
        confirmation_presenter=confirm_reset,
    )
    qtbot.addWidget(window)
    window.show()
    view = _child(window, QGraphicsView, "version-tree-view")
    proxies = {
        str(proxy.widget().property("version-id")): proxy
        for proxy in view.scene().items()
        if isinstance(proxy, QGraphicsProxyWidget) and proxy.widget() is not None
    }
    assert proxies[root.version_id].pos().x() == 900.0
    assert proxies[root.version_id].pos().y() == -800.0

    reset_button = _child(window, QPushButton, "version-reset-layout-button")
    qtbot.mouseClick(reset_button, Qt.MouseButton.LeftButton)  # type: ignore[no-untyped-call]

    assert confirmations and record.display_name in confirmations[0][1]
    assert store.list_version_layouts(record.file_id) == {}
    after = next(
        proxy
        for proxy in view.scene().items()
        if isinstance(proxy, QGraphicsProxyWidget)
        and proxy.widget() is not None
        and str(proxy.widget().property("version-id")) == root.version_id
    )
    assert after.pos().x() < 900.0
    assert after.pos().y() > -800.0


def test_deleted_version_can_be_purged_from_canvas(qtbot: QtBot, tmp_path: Path) -> None:
    library_root = tmp_path / "library"
    record = _seed_versioned_workbook(library_root)
    root = record.head_version
    assert root is not None
    child_snapshot = library_root / "files/file-1/versions/version-2/snapshot.xlsx"
    child_snapshot.parent.mkdir(parents=True)
    child_snapshot.write_bytes(record.working_path.read_bytes() + b"purge-me")
    child = VersionRecord(
        "version-2",
        record.file_id,
        root.version_id,
        "多列排序",
        datetime(2026, 8, 16, 9, 0, tzinfo=UTC),
        "sort",
        None,
        child_snapshot,
        sha256(child_snapshot.read_bytes()).hexdigest(),
    )
    store = MetadataStore(library_root)
    store.record_child_version(child, root.version_id)
    store.soft_delete_version(record.file_id, child.version_id, child.version_id)
    confirmations: list[tuple[str, str]] = []
    task_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window
    from hyacinth.versioning import PURGE_VERSION_OPERATION, run_purge_version_task

    def confirm_purge(_parent: QWidget, title: str, message: str) -> bool:
        confirmations.append((title, message))
        return True

    window = create_main_window(
        task_queue=task_queue,
        library_root=library_root,
        confirmation_presenter=confirm_purge,
    )
    qtbot.addWidget(window)
    window.show()
    tree_panel = _child(window, VersionTreePanel, "version-tree-panel")

    tree_panel.version_purge_requested.emit(record.file_id, child.version_id)

    assert confirmations and "无法恢复" in confirmations[0][1]
    purge_request = task_queue.submitted[-1]
    assert purge_request.operation == PURGE_VERSION_OPERATION
    purged = run_purge_version_task(purge_request, PreviewTaskContext())
    assert purged.version_id == child.version_id
    task_queue.push_event(
        TaskEvent(
            purge_request.task_id,
            TaskState.SUCCEEDED,
            purge_request.name,
            purge_request.file_id,
            None,
            result=purged,
        )
    )

    view = _child(window, QGraphicsView, "version-tree-view")
    qtbot.waitUntil(
        lambda: (
            not any(
                isinstance(proxy, QGraphicsProxyWidget)
                and proxy.widget() is not None
                and str(proxy.widget().property("version-id")) == child.version_id
                for proxy in view.scene().items()
            )
        ),
        timeout=2000,
    )
    assert not child_snapshot.parent.exists()
    with pytest.raises(ValueError):
        store.get_version(record.file_id, child.version_id)


def test_other_file_nodes_never_move_when_lane_grows(qtbot: QtBot, tmp_path: Path) -> None:
    library_root = tmp_path / "library"
    record = _seed_versioned_workbook(library_root)
    second_directory = library_root / "files/file-2"
    second_original = second_directory / "original/库存.xlsx"
    second_working = second_directory / "working/current.xlsx"
    second_snapshot = second_directory / "versions/version-b/snapshot.xlsx"
    for path in (second_original, second_working, second_snapshot):
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(b"second")
    second_root = VersionRecord(
        "version-b",
        "file-2",
        None,
        "导入原始文件",
        datetime(2026, 8, 16, 10, 0, tzinfo=UTC),
        "import",
        None,
        second_snapshot,
        sha256(second_snapshot.read_bytes()).hexdigest(),
    )
    second_record = ImportedWorkbook(
        "file-2",
        "库存.xlsx",
        second_original,
        second_working,
        second_root,
        datetime(2026, 8, 16, 10, 0, tzinfo=UTC),
    )
    store = MetadataStore(library_root)
    store.record_import(second_record)
    store.save_version_layout("file-2", "version-b", 500.0, 700.0, fixed=True)
    task_queue = FakeApplicationTaskQueue([])
    from hyacinth.app import create_main_window

    window = create_main_window(task_queue=task_queue, library_root=library_root)
    qtbot.addWidget(window)
    window.show()
    view = _child(window, QGraphicsView, "version-tree-view")

    def position_of(version_id: str) -> tuple[float, float]:
        proxy = next(
            p
            for p in view.scene().items()
            if isinstance(p, QGraphicsProxyWidget)
            and p.widget() is not None
            and str(p.widget().property("version-id")) == version_id
        )
        return (proxy.pos().x(), proxy.pos().y())

    before = position_of("version-b")

    child_snapshot = library_root / "files/file-1/versions/version-9/snapshot.xlsx"
    child_snapshot.parent.mkdir(parents=True)
    child_snapshot.write_bytes(record.working_path.read_bytes())
    root_version = record.head_version
    assert root_version is not None
    store.record_child_version(
        VersionRecord(
            "version-9",
            record.file_id,
            root_version.version_id,
            "手动编辑",
            datetime(2026, 8, 16, 11, 0, tzinfo=UTC),
            "manual-edit",
            None,
            child_snapshot,
            sha256(child_snapshot.read_bytes()).hexdigest(),
        ),
        root_version.version_id,
    )
    window._refresh_version_canvas()

    assert position_of("version-b") == before
