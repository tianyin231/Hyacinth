import os
import shutil
from collections.abc import Callable
from dataclasses import replace
from datetime import UTC, datetime
from hashlib import sha256
from pathlib import Path
from typing import Protocol
from uuid import uuid4

from openpyxl import load_workbook

from hyacinth.excel.selection import create_default_engine
from hyacinth.excel.task_handler import (
    ConversionTaskContext,
    EngineSelector,
    run_conversion_task,
)
from hyacinth.tasks import TaskRequest
from hyacinth.tasks.worker import TaskContext, TaskHandler
from hyacinth.versioning import (
    ImportedWorkbook,
    MetadataStore,
    VersionRecord,
    write_recovery_manifest,
)

IMPORT_WORKBOOK_OPERATION = "import-workbook"
COPY_CHUNK_SIZE = 1024 * 1024


class ImportTaskContext(ConversionTaskContext, Protocol):
    def commit(self) -> None: ...


class ImportMetadataStore(Protocol):
    def record_import(self, record: ImportedWorkbook) -> None: ...


MetadataStoreFactory = Callable[[Path], ImportMetadataStore]


def run_import_task(
    request: TaskRequest,
    context: ImportTaskContext,
    *,
    select_engine: EngineSelector = create_default_engine,
    metadata_store_factory: MetadataStoreFactory = MetadataStore,
) -> ImportedWorkbook:
    source = _payload_path(request, "source_path")
    library_root = _payload_path(request, "library_root")
    if source.suffix.lower() not in {".xls", ".xlsx"}:
        raise ValueError("第一版只支持 .xls 和 .xlsx 文件")

    staging_root = library_root / ".staging"
    files_root = library_root / "files"
    staging_directory = staging_root / f"{request.file_id}-{request.task_id}"
    final_directory = files_root / request.file_id

    original = staging_directory / "original" / source.name
    working = staging_directory / "working" / "current.xlsx"
    final_original = final_directory / "original" / source.name
    final_working = final_directory / "working" / "current.xlsx"
    version_id = uuid4().hex
    snapshot = staging_directory / "versions" / version_id / "snapshot.xlsx"
    final_snapshot = final_directory / "versions" / version_id / "snapshot.xlsx"
    staging_root.mkdir(parents=True, exist_ok=True)
    files_root.mkdir(parents=True, exist_ok=True)

    try:
        original.parent.mkdir(parents=True)
        working.parent.mkdir(parents=True)
        context.report_progress(None, "正在复制原始文件")
        _copy_file(source, original, context)
        context.check_cancelled()
        engine = None
        if source.suffix.lower() == ".xls":
            conversion = run_conversion_task(
                replace(
                    request,
                    payload={
                        "source_path": str(original),
                        "destination_path": str(working),
                    },
                ),
                context,
                select_engine=select_engine,
            )
            engine = conversion.engine
        else:
            _copy_file(original, working, context)
        context.check_cancelled()
        context.report_progress(None, "正在校验工作副本")
        workbook = load_workbook(working, read_only=True)
        workbook.close()
        context.check_cancelled()
        context.report_progress(None, "正在创建根版本")
        snapshot.parent.mkdir(parents=True)
        _copy_file(working, snapshot, context)
        content_hash = _file_hash(snapshot, context)
        root_version = VersionRecord(
            version_id=version_id,
            file_id=request.file_id,
            parent_version_id=None,
            name="导入原始文件",
            created_at=datetime.now(UTC),
            operation="import",
            engine=engine,
            snapshot_path=final_snapshot,
            content_hash=content_hash,
        )
        result = ImportedWorkbook(
            file_id=request.file_id,
            display_name=source.name,
            original_path=final_original,
            working_path=final_working,
            root_version=root_version,
            imported_at=root_version.created_at,
        )
        write_recovery_manifest(snapshot.parent / "manifest.json", library_root, result)
        with context.critical_section("正在安全完成导入"):
            context.commit()
            os.replace(staging_directory, final_directory)
            metadata_store_factory(library_root).record_import(result)
    finally:
        shutil.rmtree(staging_directory, ignore_errors=True)

    context.report_progress(1.0, "导入完成")
    return result


def import_workbook_task(request: TaskRequest, context: TaskContext) -> object:
    return run_import_task(request, context)


def import_task_handlers() -> dict[str, TaskHandler]:
    return {IMPORT_WORKBOOK_OPERATION: import_workbook_task}


def _copy_file(source: Path, destination: Path, context: ImportTaskContext) -> None:
    with source.open("rb") as source_file, destination.open("wb") as destination_file:
        while chunk := source_file.read(COPY_CHUNK_SIZE):
            context.check_cancelled()
            destination_file.write(chunk)
    context.check_cancelled()
    shutil.copystat(source, destination)


def _payload_path(request: TaskRequest, key: str) -> Path:
    value = request.payload.get(key)
    if not isinstance(value, str) or not value:
        raise ValueError(f"任务参数缺少路径：{key}")
    return Path(value)


def _file_hash(path: Path, context: ImportTaskContext) -> str:
    digest = sha256()
    with path.open("rb") as source:
        while chunk := source.read(COPY_CHUNK_SIZE):
            context.check_cancelled()
            digest.update(chunk)
    context.check_cancelled()
    return digest.hexdigest()
