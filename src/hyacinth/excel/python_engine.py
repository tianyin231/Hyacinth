from pathlib import Path

import xlrd  # type: ignore[import-untyped]
from openpyxl import Workbook

from hyacinth.excel.contracts import (
    ConversionResult,
    EngineName,
    capabilities_for,
)


class PythonExcelEngine:
    name = EngineName.PYTHON
    capabilities = capabilities_for(name)

    def convert_xls_to_xlsx(self, source: Path, destination: Path) -> ConversionResult:
        source_workbook = xlrd.open_workbook(str(source), formatting_info=True)
        target_workbook = Workbook()
        default_sheet = target_workbook.active
        assert default_sheet is not None
        target_workbook.remove(default_sheet)

        for source_sheet in source_workbook.sheets():
            target_sheet = target_workbook.create_sheet(source_sheet.name)
            for row_index in range(source_sheet.nrows):
                for column_index in range(source_sheet.ncols):
                    source_cell = source_sheet.cell(row_index, column_index)
                    value = _cell_value(source_cell, source_workbook.datemode)
                    target_sheet.cell(row_index + 1, column_index + 1, value)

            for row_low, row_high, column_low, column_high in source_sheet.merged_cells:
                target_sheet.merge_cells(
                    start_row=row_low + 1,
                    end_row=row_high,
                    start_column=column_low + 1,
                    end_column=column_high,
                )

        destination.parent.mkdir(parents=True, exist_ok=True)
        target_workbook.save(destination)
        return ConversionResult(
            engine=self.name,
            output_path=destination,
            warnings=self.capabilities.limitations,
        )


def _cell_value(cell: xlrd.sheet.Cell, datemode: int) -> object:
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate_as_datetime(cell.value, datemode)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    return cell.value
