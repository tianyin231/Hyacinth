import os
import sqlite3
import zipfile
from contextlib import AbstractContextManager
from dataclasses import dataclass
from datetime import date, datetime, time
from itertools import zip_longest
from pathlib import Path
from typing import Protocol, cast
from xml.etree.ElementTree import iterparse

from openpyxl import load_workbook

from hyacinth.tasks import TaskRequest
from hyacinth.tasks.worker import TaskContext, TaskHandler

BUILD_PREVIEW_INDEX_OPERATION = "build-preview-index"
INDEX_SCHEMA_VERSION = "2"
PROGRESS_ROW_INTERVAL = 256


@dataclass(frozen=True, slots=True)
class SheetPreview:
    index: int
    title: str
    row_count: int
    column_count: int
    visible_row_count: int | None = None


@dataclass(frozen=True, slots=True)
class WorkbookPreview:
    working_path: Path
    index_path: Path
    sheets: tuple[SheetPreview, ...]


class PreviewTaskContext(Protocol):
    def report_progress(self, progress: float | None, message: str = "") -> None: ...

    def check_cancelled(self) -> None: ...

    def commit(self) -> None: ...

    def critical_section(self, message: str = "") -> AbstractContextManager[None]: ...


def preview_index_path(working_path: Path) -> Path:
    return working_path.parent.parent / "cache" / "preview.sqlite"


def run_preview_index_task(
    request: TaskRequest,
    context: PreviewTaskContext,
) -> WorkbookPreview:
    working_path = _payload_path(request, "working_path")
    index_path = _payload_path(request, "index_path")
    context.check_cancelled()
    current = _current_preview(working_path, index_path)
    if current is not None:
        context.report_progress(1.0, "预览已就绪")
        return current

    index_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = index_path.with_name(f".{index_path.name}.tmp")
    temporary_path.unlink(missing_ok=True)
    try:
        context.report_progress(None, "正在读取工作簿结构")
        _build_index(working_path, temporary_path, context)
        context.check_cancelled()
        with context.critical_section("正在完成预览索引"):
            context.commit()
            os.replace(temporary_path, index_path)
    finally:
        temporary_path.unlink(missing_ok=True)

    context.report_progress(1.0, "预览已就绪")
    preview = _current_preview(working_path, index_path)
    if preview is None:
        raise RuntimeError("预览索引生成后校验失败")
    return preview


def preview_index_task(request: TaskRequest, context: TaskContext) -> object:
    return run_preview_index_task(request, context)


def preview_task_handlers() -> dict[str, TaskHandler]:
    return {BUILD_PREVIEW_INDEX_OPERATION: preview_index_task}


def _build_index(
    working_path: Path,
    destination: Path,
    context: PreviewTaskContext,
) -> None:
    source_stat = working_path.stat()
    formula_workbook = None
    value_workbook = None
    connection = None
    try:
        formula_workbook = load_workbook(working_path, read_only=True, data_only=False)
        value_workbook = load_workbook(working_path, read_only=True, data_only=True)
        connection = sqlite3.connect(destination)
        connection.execute("PRAGMA journal_mode=OFF")
        connection.execute("PRAGMA synchronous=OFF")
        _create_schema(connection)
        visible_sheets = [
            sheet for sheet in formula_workbook.worksheets if sheet.sheet_state == "visible"
        ]
        if not visible_sheets:
            raise ValueError("工作簿没有可见工作表")
        estimated_rows = sum(max(sheet.max_row or 1, 1) for sheet in visible_sheets)
        processed_rows = 0
        workbook_archive = zipfile.ZipFile(working_path)
        for sheet_index, formula_sheet in enumerate(visible_sheets):
            context.check_cancelled()
            value_sheet = value_workbook[formula_sheet.title]
            hidden_rows = _hidden_rows(
                workbook_archive,
                cast(str, getattr(formula_sheet, "_worksheet_path")),
            )
            row_count = max(formula_sheet.max_row or 1, value_sheet.max_row or 1, 1)
            column_count = max(
                formula_sheet.max_column or 1,
                value_sheet.max_column or 1,
                1,
            )
            batch: list[tuple[int, int, int, str, str | None]] = []
            rows = zip_longest(
                formula_sheet.iter_rows(values_only=True),
                value_sheet.iter_rows(values_only=True),
                fillvalue=(),
            )
            for row_index, (formula_row, value_row) in enumerate(rows):
                row_count = max(row_count, row_index + 1)
                column_count = max(column_count, len(formula_row), len(value_row), 1)
                for column_index in range(max(len(formula_row), len(value_row))):
                    formula_value = _at(formula_row, column_index)
                    cached_value = _at(value_row, column_index)
                    if formula_value is None and cached_value is None:
                        continue
                    formula = (
                        formula_value
                        if isinstance(formula_value, str) and formula_value.startswith("=")
                        else None
                    )
                    display_value = cached_value if cached_value is not None else formula_value
                    batch.append(
                        (
                            sheet_index,
                            row_index,
                            column_index,
                            _display_text(display_value),
                            formula,
                        )
                    )
                if len(batch) >= 2_000:
                    _insert_cells(connection, batch)
                    batch.clear()
                processed_rows += 1
                if processed_rows % PROGRESS_ROW_INTERVAL == 0:
                    context.check_cancelled()
                    progress = min(processed_rows / estimated_rows, 0.99)
                    context.report_progress(progress, f"正在索引 {formula_sheet.title}")
            if batch:
                _insert_cells(connection, batch)
            visible_row_count: int | None = None
            if hidden_rows:
                visible_rows = [row for row in range(row_count) if row not in hidden_rows]
                connection.executemany(
                    "INSERT INTO visible_rows VALUES (?, ?, ?)",
                    (
                        (sheet_index, visible_index, source_index)
                        for visible_index, source_index in enumerate(visible_rows)
                    ),
                )
                visible_row_count = len(visible_rows)
            connection.execute(
                "INSERT INTO sheets VALUES (?, ?, ?, ?, ?)",
                (
                    sheet_index,
                    formula_sheet.title,
                    row_count,
                    column_count,
                    visible_row_count,
                ),
            )
        latest_stat = working_path.stat()
        if (
            latest_stat.st_size != source_stat.st_size
            or latest_stat.st_mtime_ns != source_stat.st_mtime_ns
        ):
            raise RuntimeError("工作副本在读取期间发生变化，请重试")
        connection.executemany(
            "INSERT INTO metadata VALUES (?, ?)",
            (
                ("schema_version", INDEX_SCHEMA_VERSION),
                ("source_size", str(source_stat.st_size)),
                ("source_mtime_ns", str(source_stat.st_mtime_ns)),
            ),
        )
        connection.commit()
    finally:
        if connection is not None:
            connection.close()
        if formula_workbook is not None:
            formula_workbook.close()
        if value_workbook is not None:
            value_workbook.close()
        if "workbook_archive" in locals():
            workbook_archive.close()


def _create_schema(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        CREATE TABLE metadata (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        CREATE TABLE sheets (
            sheet_index INTEGER PRIMARY KEY,
            title TEXT NOT NULL,
            row_count INTEGER NOT NULL,
            column_count INTEGER NOT NULL,
            visible_row_count INTEGER
        );
        CREATE TABLE cells (
            sheet_index INTEGER NOT NULL,
            row_index INTEGER NOT NULL,
            column_index INTEGER NOT NULL,
            display_value TEXT NOT NULL,
            formula TEXT,
            PRIMARY KEY (sheet_index, row_index, column_index)
        );
        CREATE TABLE visible_rows (
            sheet_index INTEGER NOT NULL,
            visible_row_index INTEGER NOT NULL,
            source_row_index INTEGER NOT NULL,
            PRIMARY KEY (sheet_index, visible_row_index)
        );
        """
    )


def _insert_cells(
    connection: sqlite3.Connection,
    cells: list[tuple[int, int, int, str, str | None]],
) -> None:
    connection.executemany("INSERT INTO cells VALUES (?, ?, ?, ?, ?)", cells)


def _current_preview(working_path: Path, index_path: Path) -> WorkbookPreview | None:
    if not working_path.is_file() or not index_path.is_file():
        return None
    source_stat = working_path.stat()
    try:
        connection = sqlite3.connect(f"{index_path.resolve().as_uri()}?mode=ro", uri=True)
        try:
            metadata = dict(connection.execute("SELECT key, value FROM metadata"))
            if metadata != {
                "schema_version": INDEX_SCHEMA_VERSION,
                "source_size": str(source_stat.st_size),
                "source_mtime_ns": str(source_stat.st_mtime_ns),
            }:
                return None
            sheets = tuple(
                SheetPreview(*row)
                for row in connection.execute(
                    "SELECT sheet_index, title, row_count, column_count, visible_row_count "
                    "FROM sheets ORDER BY sheet_index"
                )
            )
        finally:
            connection.close()
    except (OSError, sqlite3.Error, ValueError):
        return None
    return WorkbookPreview(working_path=working_path, index_path=index_path, sheets=sheets)


def _display_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    return str(value)


def _hidden_rows(archive: zipfile.ZipFile, worksheet_path: str) -> set[int]:
    hidden: set[int] = set()
    with archive.open(worksheet_path) as source:
        for _event, element in iterparse(source, events=("end",)):
            if element.tag.endswith("}row") and element.attrib.get("hidden") in {
                "1",
                "true",
                "True",
            }:
                row_number = element.attrib.get("r")
                if row_number is not None:
                    hidden.add(int(row_number) - 1)
            element.clear()
    return hidden


def _at(row: tuple[object, ...], index: int) -> object | None:
    return row[index] if index < len(row) else None


def _payload_path(request: TaskRequest, key: str) -> Path:
    value = request.payload.get(key)
    if not isinstance(value, str) or not value:
        raise ValueError(f"任务参数缺少路径：{key}")
    return Path(value)
