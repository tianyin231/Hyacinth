"""版本差异比较后台任务（需求第 13 节）。

按工作表名称与单元格坐标比较两个版本快照的内容与公式（公式按
openpyxl 原始表达式文本参与比较，格式差异不纳入）。识别工作表的
新增、删除与重命名（一删一增且内容逐格相同视为重命名）。
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Protocol

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from hyacinth.tasks import TaskRequest
from hyacinth.tasks.worker import TaskContext, TaskHandler

COMPARE_VERSIONS_OPERATION = "compare-versions"

# 每张工作表最多上报的差异单元格数，防止极端场景把事件塞爆。
_MAX_CELLS_PER_SHEET = 20000


class CompareVersionsTaskContext(Protocol):
    def report_progress(self, progress: float | None, message: str = "") -> None: ...

    def check_cancelled(self) -> None: ...


@dataclass(frozen=True)
class CellDiff:
    ref: str
    base_value: str | None
    target_value: str | None
    kind: str  # added / removed / changed


@dataclass(frozen=True)
class SheetDiff:
    status: str  # added / removed / renamed / changed / unchanged
    base_name: str | None
    target_name: str | None
    cells: tuple[CellDiff, ...] = ()
    base_cells: dict[tuple[int, int], str | None] = field(default_factory=dict)
    target_cells: dict[tuple[int, int], str | None] = field(default_factory=dict)

    @property
    def label(self) -> str:
        if self.status == "renamed":
            return f"{self.base_name} → {self.target_name}"
        return self.target_name if self.target_name is not None else str(self.base_name)

    @property
    def diff_count(self) -> int:
        return len(self.cells)


@dataclass(frozen=True)
class CompareResult:
    file_id: str
    base_version_id: str
    target_version_id: str
    sheets: tuple[SheetDiff, ...]

    @property
    def total_diffs(self) -> int:
        return sum(sheet.diff_count for sheet in self.sheets)


def _sheet_cells(sheet: object) -> dict[tuple[int, int], str | None]:
    cells: dict[tuple[int, int], str | None] = {}
    for row in sheet.iter_rows():  # type: ignore[attr-defined]
        for cell in row:
            if cell.value is not None:
                cells[(cell.row, cell.column)] = str(cell.value)
    return cells


def _cell_ref(position: tuple[int, int]) -> str:
    row, column = position
    return f"{get_column_letter(column)}{row}"


def _diff_cells(
    base: dict[tuple[int, int], str | None],
    target: dict[tuple[int, int], str | None],
) -> tuple[CellDiff, ...]:
    diffs: list[CellDiff] = []
    for position in sorted(set(base) | set(target)):
        base_value = base.get(position)
        target_value = target.get(position)
        if base_value == target_value:
            continue
        if len(diffs) >= _MAX_CELLS_PER_SHEET:
            break
        if base_value is None:
            kind = "added"
        elif target_value is None:
            kind = "removed"
        else:
            kind = "changed"
        diffs.append(CellDiff(_cell_ref(position), base_value, target_value, kind))
    return tuple(diffs)


def run_compare_versions_task(
    request: TaskRequest,
    context: CompareVersionsTaskContext,
) -> CompareResult:
    payload = request.payload
    file_id = str(payload["file_id"])
    base_version_id = str(payload["base_version_id"])
    target_version_id = str(payload["target_version_id"])

    context.report_progress(0.1, "读取基准版本")
    base_book = load_workbook(str(payload["base_path"]), read_only=True, data_only=False)
    context.check_cancelled()
    context.report_progress(0.3, "读取目标版本")
    target_book = load_workbook(str(payload["target_path"]), read_only=True, data_only=False)
    context.check_cancelled()

    total = len(base_book.sheetnames) + len(target_book.sheetnames) + 1
    done = 0
    base_sheets: dict[str, dict[tuple[int, int], str | None]] = {}
    for name in base_book.sheetnames:
        base_sheets[name] = _sheet_cells(base_book[name])
        done += 1
        context.report_progress(0.3 + 0.6 * done / total, f"比较工作表 {name}")
        context.check_cancelled()
    target_sheets: dict[str, dict[tuple[int, int], str | None]] = {}
    for name in target_book.sheetnames:
        target_sheets[name] = _sheet_cells(target_book[name])
        done += 1
        context.report_progress(0.3 + 0.6 * done / total, f"比较工作表 {name}")
        context.check_cancelled()
    base_names = list(base_book.sheetnames)
    target_names = list(target_book.sheetnames)
    base_book.close()
    target_book.close()

    removed_names = [name for name in base_sheets if name not in target_sheets]
    added_names = [name for name in target_sheets if name not in base_sheets]

    # 重命名识别：一删一增且内容逐格相同。
    renamed_pairs: list[tuple[str, str]] = []
    for removed in list(removed_names):
        for added in list(added_names):
            if base_sheets[removed] == target_sheets[added]:
                renamed_pairs.append((removed, added))
                removed_names.remove(removed)
                added_names.remove(added)
                break

    sheets: list[SheetDiff] = []
    for removed, added in renamed_pairs:
        sheets.append(
            SheetDiff("renamed", removed, added, (), base_sheets[removed], target_sheets[added])
        )
    for name in base_names:
        if name in removed_names:
            sheets.append(SheetDiff("removed", name, None, (), base_sheets[name], {}))
    for name in target_names:
        if name in added_names:
            sheets.append(SheetDiff("added", None, name, (), {}, target_sheets[name]))
    for name in target_names:
        if name not in base_sheets:
            continue
        diffs = _diff_cells(base_sheets[name], target_sheets[name])
        status = "changed" if diffs else "unchanged"
        sheets.append(SheetDiff(status, name, name, diffs, base_sheets[name], target_sheets[name]))

    # 展示顺序：目标工作簿的表顺序优先，随后是纯删除的表。
    order = {name: index for index, name in enumerate(target_names)}
    sheets.sort(
        key=lambda sheet: (
            order.get(sheet.target_name if sheet.target_name is not None else "", len(order)),
            sheet.base_name or "",
        )
    )
    context.report_progress(1.0, "比较完成")
    return CompareResult(file_id, base_version_id, target_version_id, tuple(sheets))


def compare_versions_task(request: TaskRequest, context: TaskContext) -> object:
    return run_compare_versions_task(request, context)


def compare_version_handlers() -> dict[str, TaskHandler]:
    return {COMPARE_VERSIONS_OPERATION: compare_versions_task}
