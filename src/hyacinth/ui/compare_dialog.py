"""版本差异对比抽屉（需求第 13 节）。

以右侧覆盖式抽屉呈现（与设置抽屉一致，从右往左滑入，设置/对比互斥）。
三种视图随时切换：单表高亮 + 修改详情（默认）、左右并排同步网格、只看差异清单；
左右并排与单表高亮一致，对差异单元格用绿/红/黄高亮，点击差异格在底部显示修改前后。
最后一次查看方式由调用方持久化。
"""

from __future__ import annotations

from collections.abc import Sequence

from openpyxl.utils import coordinate_to_tuple, get_column_letter
from PySide6.QtCore import (
    QAbstractTableModel,
    QModelIndex,
    QPersistentModelIndex,
    Qt,
    Signal,
)
from PySide6.QtGui import QBrush, QColor
from PySide6.QtWidgets import (
    QAbstractItemView,
    QButtonGroup,
    QCheckBox,
    QComboBox,
    QFormLayout,
    QFrame,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QStackedWidget,
    QTableView,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from hyacinth.ui.drawer_base import SlideDrawer
from hyacinth.versioning.compare_task import CellDiff, CompareResult, SheetDiff
from hyacinth.versioning.models import VersionRecord

COMPARE_DRAWER_WIDTH = 640


def _ref_to_position(ref: str) -> tuple[int, int]:
    row, column = coordinate_to_tuple(ref)
    return row, column


VIEW_MODE_SINGLE = "single"
VIEW_MODE_SIDE = "side-by-side"

_COLOR_ADDED = QColor("#dcfce7")
_COLOR_REMOVED = QColor("#fee2e2")
_COLOR_CHANGED = QColor("#fef9c3")

_KIND_TEXT = {"added": "新增", "removed": "删除", "changed": "修改"}


def _diff_color(kind: str) -> QColor:
    return {
        "added": _COLOR_ADDED,
        "removed": _COLOR_REMOVED,
        "changed": _COLOR_CHANGED,
    }[kind]


class _SnapshotGridModel(QAbstractTableModel):
    """并排视图的快照网格：内存单元格字典 + Excel 式行列号表头。

    可选 ``diffs`` 提供差异映射，使基准/目标侧各自用绿（增）/红（删）/
    黄（改）高亮差异单元格，与单表高亮一致。
    """

    def __init__(
        self,
        cells: dict[tuple[int, int], str | None],
        diffs: dict[tuple[int, int], CellDiff] | None = None,
    ) -> None:
        super().__init__()
        self._cells = cells
        self._diffs = diffs or {}
        self._rows = max((row for row, _column in cells), default=1)
        self._columns = max((column for _row, column in cells), default=1)

    def diff_at(self, index: QModelIndex | QPersistentModelIndex) -> CellDiff | None:
        return self._diffs.get((index.row() + 1, index.column() + 1))

    def rowCount(
        self,
        parent: QModelIndex | QPersistentModelIndex = QModelIndex(),
    ) -> int:
        return 0 if parent.isValid() else self._rows + 2

    def columnCount(
        self,
        parent: QModelIndex | QPersistentModelIndex = QModelIndex(),
    ) -> int:
        return 0 if parent.isValid() else self._columns + 2

    def data(
        self,
        index: QModelIndex | QPersistentModelIndex,
        role: int = Qt.ItemDataRole.DisplayRole,
    ) -> object:
        position = (index.row() + 1, index.column() + 1)
        diff = self._diffs.get(position)
        if role == Qt.ItemDataRole.BackgroundRole:
            if diff is None:
                return None
            return QBrush(_diff_color(diff.kind))
        if role == Qt.ItemDataRole.ToolTipRole and diff is not None:
            base_text = diff.base_value if diff.base_value is not None else "（空）"
            target_text = diff.target_value if diff.target_value is not None else "（空）"
            return f"{_KIND_TEXT[diff.kind]}：{base_text} → {target_text}"
        if role != Qt.ItemDataRole.DisplayRole:
            return None
        return self._cells.get(position)

    def headerData(  # noqa: N802
        self,
        section: int,
        orientation: Qt.Orientation,
        role: int = Qt.ItemDataRole.DisplayRole,
    ) -> object:
        if role != Qt.ItemDataRole.DisplayRole:
            return None
        if orientation == Qt.Orientation.Horizontal:
            return get_column_letter(section + 1)
        return str(section + 1)


class _DiffGridModel(QAbstractTableModel):
    """单表高亮视图：目标侧网格；新增绿/删除红/修改黄。"""

    def __init__(self, sheet: SheetDiff) -> None:
        super().__init__()
        self._sheet = sheet
        self._diffs_by_position: dict[tuple[int, int], CellDiff] = {}
        for diff in sheet.cells:
            row, column = _ref_to_position(diff.ref)
            self._diffs_by_position[(row, column)] = diff
        rows = [row for row, _column in set(sheet.base_cells) | set(sheet.target_cells)]
        columns = [column for _row, column in set(sheet.base_cells) | set(sheet.target_cells)]
        self._rows = max(rows, default=1)
        self._columns = max(columns, default=1)

    def rowCount(
        self,
        parent: QModelIndex | QPersistentModelIndex = QModelIndex(),
    ) -> int:
        return 0 if parent.isValid() else self._rows + 2

    def columnCount(
        self,
        parent: QModelIndex | QPersistentModelIndex = QModelIndex(),
    ) -> int:
        return 0 if parent.isValid() else self._columns + 2

    def diff_at(self, index: QModelIndex | QPersistentModelIndex) -> CellDiff | None:
        return self._diffs_by_position.get((index.row() + 1, index.column() + 1))

    def data(
        self,
        index: QModelIndex | QPersistentModelIndex,
        role: int = Qt.ItemDataRole.DisplayRole,
    ) -> object:
        position = (index.row() + 1, index.column() + 1)
        diff = self._diffs_by_position.get(position)
        if role == Qt.ItemDataRole.BackgroundRole:
            if diff is None:
                return None
            return QBrush(_diff_color(diff.kind))
        if role == Qt.ItemDataRole.ToolTipRole and diff is not None:
            base_text = diff.base_value if diff.base_value is not None else "（空）"
            target_text = diff.target_value if diff.target_value is not None else "（空）"
            return f"{_KIND_TEXT[diff.kind]}：{base_text} → {target_text}"
        if role != Qt.ItemDataRole.DisplayRole:
            return None
        value = self._sheet.target_cells.get(position)
        if value is None:
            value = self._sheet.base_cells.get(position)
        return value

    def headerData(  # noqa: N802
        self,
        section: int,
        orientation: Qt.Orientation,
        role: int = Qt.ItemDataRole.DisplayRole,
    ) -> object:
        if role != Qt.ItemDataRole.DisplayRole:
            return None
        if orientation == Qt.Orientation.Horizontal:
            return get_column_letter(section + 1)
        return str(section + 1)


class VersionCompareDialog(SlideDrawer):
    """同一文件两个版本的差异对比抽屉（右侧覆盖，设置/对比互斥）。"""

    compare_requested = Signal(str, str)
    view_mode_changed = Signal(str)

    def __init__(
        self,
        parent: QWidget | None,
        file_name: str,
        versions: Sequence[VersionRecord],
        initial_base_id: str | None,
        initial_target_id: str | None,
        initial_view_mode: str = VIEW_MODE_SINGLE,
    ) -> None:
        super().__init__(parent, width=COMPARE_DRAWER_WIDTH, object_name="compare-drawer")
        self._versions = list(versions)
        self._result: CompareResult | None = None
        self._selected_sheet: SheetDiff | None = None
        self._syncing = False

        header = QFrame(self)
        header.setObjectName("compare-drawer-header")
        header.setFixedHeight(44)
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(16, 0, 10, 0)
        title = QLabel(f"版本对比 · {file_name}", header)
        title.setObjectName("compare-drawer-title")
        close_button = QPushButton("✕", header)
        close_button.setObjectName("compare-close-button")
        close_button.setFixedSize(30, 26)
        close_button.setToolTip("关闭对比 (Esc)")
        close_button.clicked.connect(self.close_drawer)
        header_layout.addWidget(title)
        header_layout.addStretch()
        header_layout.addWidget(close_button)

        controls = QFrame(self)
        controls.setObjectName("compare-controls")
        controls_layout = QHBoxLayout(controls)
        controls_layout.setContentsMargins(12, 8, 12, 8)
        controls_layout.setSpacing(6)
        controls_layout.addWidget(self._field_label("基准版本", controls))
        self._base_combo = QComboBox(controls)
        self._base_combo.setProperty("class", "field-control")
        self._base_combo.setSizeAdjustPolicy(
            QComboBox.SizeAdjustPolicy.AdjustToMinimumContentsLengthWithIcon
        )
        controls_layout.addWidget(self._base_combo, 1)
        controls_layout.addWidget(self._field_label("目标版本", controls))
        self._target_combo = QComboBox(controls)
        self._target_combo.setProperty("class", "field-control")
        controls_layout.addWidget(self._target_combo, 1)
        run_button = QPushButton("开始对比", controls)
        run_button.setObjectName("compare-run-button")
        run_button.setDefault(True)
        run_button.clicked.connect(self._emit_compare)
        controls_layout.addWidget(run_button)

        self._status_label = QLabel("选择两个版本后点击“开始对比”", self)
        self._status_label.setObjectName("compare-status-label")
        self._status_label.setWordWrap(True)
        status_row = QHBoxLayout()
        status_row.setContentsMargins(12, 4, 12, 2)
        status_row.addWidget(self._status_label)

        for combo in (self._base_combo, self._target_combo):
            for version in sorted(self._versions, key=lambda item: item.created_at):
                combo.addItem(_combo_text(version), version.version_id)
        if initial_base_id is not None:
            self._select_combo(self._base_combo, initial_base_id)
        if initial_target_id is not None:
            self._select_combo(self._target_combo, initial_target_id)

        mode_bar = QFrame(self)
        mode_layout = QHBoxLayout(mode_bar)
        mode_layout.setContentsMargins(12, 4, 12, 4)
        mode_layout.setSpacing(8)
        self._single_button = QPushButton("单表高亮", mode_bar)
        self._single_button.setProperty("class", "tool-button")
        self._single_button.setCheckable(True)
        self._side_button = QPushButton("左右并排", mode_bar)
        self._side_button.setProperty("class", "tool-button")
        self._side_button.setCheckable(True)
        mode_group = QButtonGroup(mode_bar)
        mode_group.addButton(self._single_button)
        mode_group.addButton(self._side_button)
        # 两个按钮的 toggled 都接同一处理器：切换时一个变 False 一个变 True，
        # 以最终状态为准（程序化 setChecked 同样生效）。
        self._single_button.toggled.connect(lambda _checked: self._handle_mode_changed())
        self._side_button.toggled.connect(lambda _checked: self._handle_mode_changed())
        self._diff_only = QCheckBox("只看差异单元格", mode_bar)
        self._diff_only.toggled.connect(self._update_stack)
        self._sync_scroll = QCheckBox("同步滚动", mode_bar)
        self._sync_scroll.setChecked(True)
        mode_layout.addWidget(self._single_button)
        mode_layout.addWidget(self._side_button)
        mode_layout.addSpacing(8)
        mode_layout.addWidget(self._diff_only)
        mode_layout.addWidget(self._sync_scroll)
        mode_layout.addStretch()

        self._sheet_list = QTableWidget(self)
        self._sheet_list.setColumnCount(1)
        self._sheet_list.setHorizontalHeaderLabels(["工作表"])
        self._sheet_list.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._sheet_list.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._sheet_list.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._sheet_list.setFixedWidth(200)
        self._sheet_list.currentCellChanged.connect(lambda *_args: self._show_sheet())
        self._sheet_list.horizontalHeader().setStretchLastSection(True)

        self._stack = QStackedWidget(self)
        # 页 0：单表高亮 + 修改详情
        single_page = QWidget(self)
        single_layout = QVBoxLayout(single_page)
        single_layout.setContentsMargins(0, 0, 0, 0)
        self._diff_table = QTableView(self)
        self._diff_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._diff_table.clicked.connect(self._handle_diff_clicked)
        single_layout.addWidget(self._diff_table, 1)
        # 页 1：左右并排（各自差异高亮）
        side_page = QWidget(self)
        side_layout = QVBoxLayout(side_page)
        side_layout.setContentsMargins(0, 0, 0, 0)
        side_split = QHBoxLayout()
        side_split.setContentsMargins(0, 0, 0, 0)
        side_split.setSpacing(6)
        side_base_box = QVBoxLayout()
        side_base_box.setContentsMargins(0, 0, 0, 0)
        side_base_label = QLabel("基准", side_page)
        side_base_label.setObjectName("compare-side-label")
        side_base_box.addWidget(side_base_label)
        self._base_view = QTableView(side_page)
        self._base_view.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._base_view.clicked.connect(self._handle_side_clicked)
        side_base_box.addWidget(self._base_view, 1)
        side_target_box = QVBoxLayout()
        side_target_box.setContentsMargins(0, 0, 0, 0)
        side_target_label = QLabel("目标", side_page)
        side_target_label.setObjectName("compare-side-label")
        side_target_box.addWidget(side_target_label)
        self._target_view = QTableView(side_page)
        self._target_view.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._target_view.clicked.connect(self._handle_side_clicked)
        side_target_box.addWidget(self._target_view, 1)
        side_split.addLayout(side_base_box, 1)
        side_split.addLayout(side_target_box, 1)
        side_layout.addLayout(side_split, 1)
        self._base_view.verticalScrollBar().valueChanged.connect(
            lambda value: self._sync_scrollbars(self._base_view, self._target_view, value)
        )
        self._target_view.verticalScrollBar().valueChanged.connect(
            lambda value: self._sync_scrollbars(self._target_view, self._base_view, value)
        )
        self._base_view.horizontalScrollBar().valueChanged.connect(
            lambda value: self._sync_scrollbars(self._base_view, self._target_view, value, True)
        )
        self._target_view.horizontalScrollBar().valueChanged.connect(
            lambda value: self._sync_scrollbars(self._target_view, self._base_view, value, True)
        )
        # 页 2：只看差异（差异单元格清单）
        self._diff_list = QTableWidget(self)
        self._diff_list.setColumnCount(4)
        self._diff_list.setHorizontalHeaderLabels(["位置", "类型", "修改前", "修改后"])
        self._diff_list.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._diff_list.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._diff_list.horizontalHeader().setStretchLastSection(True)
        self._stack.addWidget(single_page)
        self._stack.addWidget(side_page)
        self._stack.addWidget(self._diff_list)

        body = QHBoxLayout()
        body.setContentsMargins(12, 4, 12, 8)
        body.addWidget(self._sheet_list)
        body.addWidget(self._stack, 1)

        self._detail_bar = self._build_detail()

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        layout.addWidget(header)
        layout.addWidget(controls)
        layout.addLayout(status_row)
        layout.addWidget(mode_bar)
        layout.addLayout(body, 1)
        layout.addWidget(self._detail_bar)

        self._set_view_mode(initial_view_mode)

    # ── 对外接口 ─────────────────────────────────────────────────────

    def base_version_id(self) -> str | None:
        return self._current_combo_id(self._base_combo)

    def target_version_id(self) -> str | None:
        return self._current_combo_id(self._target_combo)

    def view_mode(self) -> str:
        return VIEW_MODE_SIDE if self._side_button.isChecked() else VIEW_MODE_SINGLE

    def set_busy(self) -> None:
        self._status_label.setText("对比中…")

    def set_error(self, message: str) -> None:
        self._status_label.setText(message)

    def set_result(self, result: CompareResult) -> None:
        self._result = result
        self._sheet_list.setRowCount(0)
        for sheet in result.sheets:
            row = self._sheet_list.rowCount()
            self._sheet_list.insertRow(row)
            item = QTableWidgetItem(_sheet_badge(sheet))
            item.setData(Qt.ItemDataRole.UserRole, sheet.label)
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self._sheet_list.setItem(row, 0, item)
        if result.total_diffs == 0:
            self._status_label.setText("两个版本内容一致，没有差异")
        else:
            self._status_label.setText(f"共 {result.total_diffs} 处差异")
        if self._sheet_list.rowCount() > 0:
            self._sheet_list.setCurrentCell(0, 0)
        self._show_sheet()

    # ── 内部实现 ─────────────────────────────────────────────────────

    def _field_label(self, text: str, parent: QWidget) -> QLabel:
        label = QLabel(text, parent)
        label.setProperty("class", "compare-field-label")
        return label

    def _build_detail(self) -> QWidget:
        detail = QWidget(self)
        detail.setObjectName("compare-detail-bar")
        detail_layout = QFormLayout(detail)
        detail_layout.setContentsMargins(10, 4, 10, 4)
        detail_layout.setSpacing(4)
        self._detail_ref = QLabel("—", detail)
        self._detail_kind = QLabel("—", detail)
        self._detail_base = QLabel("—", detail)
        self._detail_target = QLabel("—", detail)
        self._detail_ref.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        self._detail_base.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        self._detail_target.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        detail_layout.addRow("位置", self._detail_ref)
        detail_layout.addRow("类型", self._detail_kind)
        detail_layout.addRow("修改前", self._detail_base)
        detail_layout.addRow("修改后", self._detail_target)
        return detail

    def _emit_compare(self) -> None:
        base_id = self.base_version_id()
        target_id = self.target_version_id()
        if base_id is not None and target_id is not None and base_id != target_id:
            self.compare_requested.emit(base_id, target_id)

    def _handle_mode_changed(self) -> None:
        self.view_mode_changed.emit(self.view_mode())
        self._update_stack()

    def _set_view_mode(self, mode: str) -> None:
        # 初始化不触发持久化：initial_view_mode 本身就是持久值。
        self._single_button.blockSignals(True)
        self._side_button.blockSignals(True)
        try:
            side = mode == VIEW_MODE_SIDE
            self._side_button.setChecked(side)
            self._single_button.setChecked(not side)
        finally:
            self._single_button.blockSignals(False)
            self._side_button.blockSignals(False)
        self._diff_only.setEnabled(not side)
        self._update_stack()

    def _update_stack(self) -> None:
        if self._side_button.isChecked():
            self._stack.setCurrentIndex(1)
        elif self._diff_only.isChecked():
            self._stack.setCurrentIndex(2)
        else:
            self._stack.setCurrentIndex(0)

    def _show_sheet(self) -> None:
        item = self._sheet_list.currentItem()
        sheet = self._sheet_of_item(item)
        self._selected_sheet = sheet
        if sheet is None:
            self._diff_table.setModel(None)
            self._base_view.setModel(None)
            self._target_view.setModel(None)
            self._diff_list.setRowCount(0)
            return
        side_diffs = {_ref_to_position(diff.ref): diff for diff in sheet.cells}
        self._diff_table.setModel(_DiffGridModel(sheet))
        self._base_view.setModel(_SnapshotGridModel(sheet.base_cells, side_diffs))
        self._target_view.setModel(_SnapshotGridModel(sheet.target_cells, side_diffs))
        self._diff_table.resizeColumnsToContents()
        self._base_view.resizeColumnsToContents()
        self._target_view.resizeColumnsToContents()
        self._fill_diff_list(sheet)
        self._clear_detail()
        self._update_stack()

    def _sheet_of_item(self, item: QTableWidgetItem | None) -> SheetDiff | None:
        if item is None or self._result is None:
            return None
        for sheet in self._result.sheets:
            if sheet.label == item.data(Qt.ItemDataRole.UserRole):
                return sheet
        return None

    def _handle_diff_clicked(self, index: QModelIndex) -> None:
        model = self._diff_table.model()
        diff = model.diff_at(index) if isinstance(model, _DiffGridModel) else None
        self._show_diff_detail(diff)

    def _handle_side_clicked(self, index: QModelIndex) -> None:
        sender = self.sender()
        if not isinstance(sender, QTableView):
            return
        model = sender.model()
        if isinstance(model, _SnapshotGridModel):
            diff = model.diff_at(index)
        elif isinstance(model, _DiffGridModel):
            diff = model.diff_at(index)
        else:
            diff = None
        self._show_diff_detail(diff)

    def _show_diff_detail(self, diff: CellDiff | None) -> None:
        if diff is None:
            self._clear_detail()
            return
        self._detail_ref.setText(diff.ref)
        self._detail_kind.setText(_KIND_TEXT.get(diff.kind, diff.kind))
        self._detail_base.setText(diff.base_value if diff.base_value is not None else "（空）")
        self._detail_target.setText(
            diff.target_value if diff.target_value is not None else "（空）"
        )

    def _clear_detail(self) -> None:
        self._detail_ref.setText("—")
        self._detail_kind.setText("—")
        self._detail_base.setText("—")
        self._detail_target.setText("—")

    def _fill_diff_list(self, sheet: SheetDiff) -> None:
        self._diff_list.setRowCount(0)
        for diff in sheet.cells:
            row = self._diff_list.rowCount()
            self._diff_list.insertRow(row)
            values = (
                diff.ref,
                _KIND_TEXT.get(diff.kind, diff.kind),
                diff.base_value if diff.base_value is not None else "（空）",
                diff.target_value if diff.target_value is not None else "（空）",
            )
            for column, text in enumerate(values):
                item = QTableWidgetItem(text)
                if column == 1:
                    item.setForeground(
                        QBrush(
                            {
                                "added": QColor("#15803d"),
                                "removed": QColor("#b91c1c"),
                                "changed": QColor("#a16207"),
                            }.get(diff.kind, QColor("#2f3640"))
                        )
                    )
                self._diff_list.setItem(row, column, item)

    def _sync_scrollbars(
        self,
        source: QTableView,
        target: QTableView,
        value: int,
        horizontal: bool = False,
    ) -> None:
        if not self._sync_scroll.isChecked() or self._syncing:
            return
        self._syncing = True
        try:
            if horizontal:
                target.horizontalScrollBar().setValue(value)
            else:
                target.verticalScrollBar().setValue(value)
        finally:
            self._syncing = False

    def _select_combo(self, combo: QComboBox, version_id: str) -> None:
        for index in range(combo.count()):
            if combo.itemData(index) == version_id:
                combo.setCurrentIndex(index)
                return

    def _current_combo_id(self, combo: QComboBox) -> str | None:
        data = combo.currentData()
        return data if isinstance(data, str) else None


def _combo_text(version: VersionRecord) -> str:
    created = version.created_at.astimezone().strftime("%m-%d %H:%M")
    return f"{version.name}（{created}）"


def _sheet_badge(sheet: SheetDiff) -> str:
    prefix = {"added": "＋", "removed": "－", "renamed": "↔"}.get(sheet.status, "")
    label = f"{prefix} {sheet.label}".strip()
    if sheet.diff_count:
        return f"{label}  ·  {sheet.diff_count} 处"
    if sheet.status == "unchanged":
        return f"{label}  ·  一致"
    return label
