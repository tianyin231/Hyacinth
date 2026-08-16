import shutil
from collections.abc import Callable
from datetime import datetime
from pathlib import Path
from typing import Protocol
from uuid import uuid4

from openpyxl.utils import get_column_letter
from PySide6.QtCore import QAbstractTableModel, QPoint, QSize, QStandardPaths, Qt, QUrl
from PySide6.QtGui import (
    QCloseEvent,
    QDesktopServices,
    QGuiApplication,
    QKeySequence,
    QShortcut,
)
from PySide6.QtWidgets import (
    QFileDialog,
    QInputDialog,
    QMainWindow,
    QMessageBox,
    QSplitter,
    QVBoxLayout,
    QWidget,
)

from hyacinth.app_icon import application_icon
from hyacinth.excel.task_handler import CONVERT_XLS_OPERATION, conversion_task_handlers
from hyacinth.library import (
    IMPORT_WORKBOOK_OPERATION,
    FileLibraryWidget,
    ImportedWorkbook,
    discover_imported_workbooks,
    import_task_handlers,
)
from hyacinth.preview import (
    BUILD_PREVIEW_INDEX_OPERATION,
    SqliteGridDataSource,
    WorkbookPreview,
    WorkbookPreviewWidget,
    preview_index_path,
    preview_task_handlers,
)
from hyacinth.processing import (
    APPLY_CHAINED_PREVIEW_OPERATION,
    APPLY_DEDUPLICATE_PREVIEW_OPERATION,
    APPLY_DELETE_BLANK_ROWS_PREVIEW_OPERATION,
    APPLY_FILTER_PREVIEW_OPERATION,
    APPLY_FIND_REPLACE_PREVIEW_OPERATION,
    APPLY_SORT_PREVIEW_OPERATION,
    APPLY_TRIM_PREVIEW_OPERATION,
    DEDUPLICATE_PREVIEW_OPERATION,
    DELETE_BLANK_ROWS_PREVIEW_OPERATION,
    FILTER_PREVIEW_OPERATION,
    FIND_REPLACE_PREVIEW_OPERATION,
    SAVE_MANUAL_EDITS_OPERATION,
    SORT_PREVIEW_OPERATION,
    TRIM_PREVIEW_OPERATION,
    UPDATE_VERSION_IN_PLACE_OPERATION,
    DeduplicatePreviewResult,
    DeleteBlankRowsPreviewResult,
    FilterPreviewResult,
    FindReplacePreviewResult,
    SortPreviewResult,
    TrimPreviewResult,
    apply_version_handlers,
    deduplicate_preview_handlers,
    delete_blank_rows_preview_handlers,
    filter_preview_handlers,
    find_replace_preview_handlers,
    sort_preview_handlers,
    trim_preview_handlers,
)
from hyacinth.tasks import (
    TaskEvent,
    TaskQueue,
    TaskQueueBridge,
    TaskRequest,
    TaskState,
    TaskStatusWidget,
)
from hyacinth.tasks.qt_bridge import TaskQueuePort
from hyacinth.ui import (
    APP_STYLESHEET,
    ApplicationHeader,
    CommandBar,
    DeletedRowsModel,
    DuplicateMappingModel,
    FileVersionTree,
    FilterDialog,
    FindDetailsModel,
    FindReplaceDialog,
    ProcessingDetailsDialog,
    RecycleBinDialog,
    RecycleEntry,
    SortDialog,
    TrimDetailsModel,
    VersionStorageStatus,
    VersionTreePanel,
    WorkbookEditorFrame,
)
from hyacinth.versioning import (
    CHECKOUT_VERSION_OPERATION,
    DELETE_VERSION_OPERATION,
    EXPORT_VERSION_OPERATION,
    PURGE_FILE_OPERATION,
    PURGE_VERSION_OPERATION,
    VERSION_STORAGE_STATS_OPERATION,
    ExportedVersion,
    MetadataStore,
    VersionRecord,
    VersionStorageStats,
    checkout_version_handlers,
    delete_version_handlers,
    export_version_handlers,
    purge_file_handlers,
    purge_version_handlers,
    suggested_export_filename,
    version_storage_stats_handlers,
)


class ApplicationTaskQueue(TaskQueuePort, Protocol):
    def submit(self, request: TaskRequest) -> None: ...


FilePicker = Callable[[QWidget], Path | None]
ErrorPresenter = Callable[[QWidget, str], None]
ConfirmationPresenter = Callable[[QWidget, str, str], bool]
VersionChoicePresenter = Callable[[QWidget, tuple[VersionRecord, ...]], str | None]
UnsavedChangesPresenter = Callable[[QWidget, str, bool], str]
SaveAsPicker = Callable[[QWidget, str], Path | None]
ExportPresenter = Callable[[QWidget, Path], None]

DEFAULT_WINDOW_SIZE = QSize(1440, 900)
MINIMUM_WINDOW_SIZE = QSize(1024, 640)
AVAILABLE_SCREEN_RATIO = 0.9


def initial_window_size(available_size: QSize) -> QSize:
    return QSize(
        min(
            DEFAULT_WINDOW_SIZE.width(),
            max(MINIMUM_WINDOW_SIZE.width(), int(available_size.width() * AVAILABLE_SCREEN_RATIO)),
        ),
        min(
            DEFAULT_WINDOW_SIZE.height(),
            max(
                MINIMUM_WINDOW_SIZE.height(),
                int(available_size.height() * AVAILABLE_SCREEN_RATIO),
            ),
        ),
    )


class HyacinthMainWindow(QMainWindow):
    def __init__(
        self,
        task_queue: ApplicationTaskQueue,
        library_root: Path,
        file_picker: FilePicker,
        error_presenter: ErrorPresenter,
        confirmation_presenter: ConfirmationPresenter,
        version_choice_presenter: VersionChoicePresenter,
        unsaved_changes_presenter: UnsavedChangesPresenter,
        save_as_picker: SaveAsPicker,
        export_presenter: ExportPresenter,
    ) -> None:
        super().__init__()
        self.setObjectName("main-window")
        self.setWindowTitle("风信子")
        self.setWindowIcon(application_icon())
        self.setMinimumSize(MINIMUM_WINDOW_SIZE)
        self._apply_initial_window_geometry()
        self.setStyleSheet(APP_STYLESHEET)

        workspace_root = QWidget(self)
        workspace_root.setObjectName("workspace-root")
        self.setCentralWidget(workspace_root)

        self._library_root = library_root
        self._file_picker = file_picker
        self._error_presenter = error_presenter
        self._confirmation_presenter = confirmation_presenter
        self._version_choice_presenter = version_choice_presenter
        self._unsaved_changes_presenter = unsaved_changes_presenter
        self._save_as_picker = save_as_picker
        self._export_presenter = export_presenter
        self._task_queue = task_queue
        self._import_task_ids: set[str] = set()
        self._preview_task_id: str | None = None
        self._preview_is_temporary = False
        self._current_workbook: ImportedWorkbook | None = None
        self._processing_task_id: str | None = None
        self._processing_result: (
            SortPreviewResult
            | DeduplicatePreviewResult
            | DeleteBlankRowsPreviewResult
            | FilterPreviewResult
            | TrimPreviewResult
            | FindReplacePreviewResult
            | None
        ) = None
        self._temporary_preview: WorkbookPreview | None = None
        # 链式多步处理：临时结果上可以继续叠加处理操作，应用时只生成一个节点。
        self._processing_steps: list[dict[str, object]] = []
        self._processing_base_version_id: str | None = None
        self._processing_previous_path: Path | None = None
        self._processing_edits_baked = False
        self._processing_submitted_operation: str | None = None
        self._processing_submitted_task_name: str | None = None
        self._apply_task_id: str | None = None
        self._apply_version_id: str | None = None
        self._manual_save_task_id: str | None = None
        self._manual_save_version_id: str | None = None
        self._checkout_task_id: str | None = None
        self._delete_task_id: str | None = None
        self._delete_version_id: str | None = None
        self._delete_file_id: str | None = None
        self._previewed_version_id: str | None = None
        self._close_after_manual_save = False
        self._close_after_apply = False
        self._in_place_task_id: str | None = None
        self._close_after_in_place = False
        self._export_task_id: str | None = None
        self._storage_stats_task_id: str | None = None
        self._purge_task_id: str | None = None
        self._purge_version_task_id: str | None = None
        self._recycle_dialog: RecycleBinDialog | None = None
        self._focus_restore_main_sizes: list[int] | None = None
        self._focus_restore_left_sizes: list[int] | None = None
        self._focus_restore_hidden: tuple[bool, ...] | None = None

        self._application_header = ApplicationHeader(workspace_root)
        self._command_bar = CommandBar(workspace_root)
        self._command_bar.import_requested.connect(self._choose_import_file)
        self._command_bar.save_version_requested.connect(self._submit_manual_save)
        self._command_bar.undo_requested.connect(self._workbook_preview_undo)
        self._command_bar.redo_requested.connect(self._workbook_preview_redo)
        self._command_bar.export_requested.connect(self._export_current_version)
        self._command_bar.recycle_requested.connect(self._open_recycle_bin)
        self._filter_dialog: FilterDialog | None = None
        self._find_dialog: FindReplaceDialog | None = None
        self._sort_dialog: SortDialog | None = None
        self._file_library = FileLibraryWidget(
            discover_imported_workbooks(library_root),
            workspace_root,
        )
        self._file_library.workbook_selected.connect(self._select_workbook)
        self._file_library.workbook_delete_requested.connect(self._request_delete_file)
        self._version_tree = VersionTreePanel(workspace_root)
        self._version_tree.version_preview_requested.connect(self._preview_version)
        self._version_tree.version_continue_requested.connect(self._continue_from_version)
        self._version_tree.version_position_changed.connect(self._save_version_position)
        self._version_tree.version_delete_requested.connect(self._request_delete_version)
        self._version_tree.version_restore_requested.connect(self._restore_deleted_version)
        self._version_tree.version_export_requested.connect(self._request_export_version)
        self._version_tree.version_purge_requested.connect(self._request_purge_version)
        self._version_tree.layout_reset_requested.connect(self._request_reset_layouts)
        self._workbook_preview = WorkbookPreviewWidget(workspace_root)
        self._workbook_preview.import_requested.connect(self._choose_import_file)
        self._workbook_preview.edit_state_changed.connect(self._apply_edit_state)
        self._workbook_preview.pending_edits_changed.connect(self._apply_pending_edit_count)
        self._workbook_preview.header_sort_requested.connect(self._quick_sort_from_table)
        self._workbook_preview.header_multi_sort_requested.connect(self._open_sort_dialog)
        self._workbook_preview.header_filter_requested.connect(
            lambda _column: self._open_filter_dialog()
        )
        self._workbook_preview.processing_menu_requested.connect(self._one_step_processing)
        self._editor = WorkbookEditorFrame(self._workbook_preview, workspace_root)
        self._editor.sort_requested.connect(self._quick_sort_from_table)
        self._editor.multi_sort_requested.connect(lambda: self._open_sort_dialog())
        self._editor.one_step_requested.connect(self._one_step_processing)
        self._editor.filter_requested.connect(self._open_filter_dialog)
        self._editor.find_replace_requested.connect(self._open_find_replace_dialog)
        self._editor.apply_requested.connect(self._submit_apply_processing_preview)
        self._editor.preview_cancel_requested.connect(self._cancel_processing_workflow)
        self._editor.details_requested.connect(self._show_processing_details)
        self._editor.deduplicate_params_confirmed.connect(self._confirm_deduplicate_params)
        self._editor.trim_params_confirmed.connect(self._confirm_trim_params)
        self._editor.params_dismissed.connect(self._editor.hide_params_bar)

        self._main_splitter = QSplitter(Qt.Orientation.Horizontal, workspace_root)
        self._main_splitter.setObjectName("main-workspace-splitter")
        self._main_splitter.addWidget(self._file_library)
        self._main_splitter.addWidget(self._version_tree)
        self._main_splitter.addWidget(self._editor)
        self._main_splitter.setStretchFactor(0, 0)
        self._main_splitter.setStretchFactor(1, 0)
        self._main_splitter.setStretchFactor(2, 1)
        self._main_splitter.setSizes([260, 340, 680])
        self._version_tree.focus_mode_requested.connect(self._set_version_focus_mode)

        workspace_layout = QVBoxLayout(workspace_root)
        workspace_layout.setContentsMargins(0, 0, 0, 0)
        workspace_layout.setSpacing(0)
        workspace_layout.addWidget(self._application_header)
        workspace_layout.addWidget(self._command_bar)
        workspace_layout.addWidget(self._main_splitter, 1)

        self._task_bridge = TaskQueueBridge(task_queue, parent=self)
        self._task_status = TaskStatusWidget(self)
        self._task_bridge.event_received.connect(self._task_status.apply_event)
        self._task_bridge.event_received.connect(self._apply_import_event)
        self._task_bridge.event_received.connect(self._apply_preview_event)
        self._task_bridge.event_received.connect(self._apply_processing_event)
        self._task_bridge.event_received.connect(self._apply_apply_event)
        self._task_bridge.event_received.connect(self._apply_manual_save_event)
        self._task_bridge.event_received.connect(self._apply_in_place_event)
        self._task_bridge.event_received.connect(self._apply_checkout_event)
        self._task_bridge.event_received.connect(self._apply_delete_event)
        self._task_bridge.event_received.connect(self._apply_export_event)
        self._task_bridge.event_received.connect(self._apply_storage_stats_event)
        self._task_bridge.event_received.connect(self._apply_purge_event)
        self._task_bridge.event_received.connect(self._apply_purge_version_event)
        self._task_status.cancel_requested.connect(self._task_bridge.cancel)

        status_bar = self.statusBar()
        status_bar.setObjectName("main-status-bar")
        status_bar.setSizeGripEnabled(False)
        status_bar.setContentsMargins(0, 0, 0, 0)
        status_bar.addWidget(self._task_status, 1)
        self._storage_status = VersionStorageStatus()
        status_bar.addPermanentWidget(self._storage_status)
        self._task_bridge.start()
        find_shortcut = QShortcut(QKeySequence("Ctrl+F"), self)
        find_shortcut.activated.connect(self._open_find_replace_dialog)
        replace_shortcut = QShortcut(QKeySequence("Ctrl+H"), self)
        replace_shortcut.activated.connect(self._open_find_replace_dialog)
        current_workbook = self._file_library.current_workbook()
        if current_workbook is not None:
            self._select_workbook(current_workbook)

    def _set_version_focus_mode(self, enabled: bool) -> None:
        hidden_widgets: tuple[QWidget, ...] = (
            self._application_header,
            self._command_bar,
            self._file_library,
            self._editor,
            self.statusBar(),
        )
        # 视口尺寸变化时 QGraphicsView 会按锚点重排滚动位置；
        # 进入/退出专注前后恢复场景中心，保证画布内容不发生跳变。
        anchor = self._version_tree.focus_anchor()
        if enabled:
            self._focus_restore_main_sizes = self._main_splitter.sizes()
            self._focus_restore_hidden = tuple(widget.isHidden() for widget in hidden_widgets)
            for widget in hidden_widgets:
                widget.hide()
            self._main_splitter.setSizes([0, self._main_splitter.width(), 0])
            self._version_tree.restore_focus_anchor(anchor)
            return

        if self._focus_restore_hidden is None:
            return
        for widget, was_hidden in zip(hidden_widgets, self._focus_restore_hidden, strict=True):
            widget.setVisible(not was_hidden)
        if self._focus_restore_main_sizes is not None:
            self._main_splitter.setSizes(self._focus_restore_main_sizes)
        self._version_tree.restore_focus_anchor(anchor)
        self._focus_restore_main_sizes = None
        self._focus_restore_left_sizes = None
        self._focus_restore_hidden = None

    def _apply_initial_window_geometry(self) -> None:
        screen = QGuiApplication.primaryScreen()
        if screen is None:
            self.resize(DEFAULT_WINDOW_SIZE)
            return
        available = screen.availableGeometry()
        target = initial_window_size(available.size())
        self.resize(target)
        top_left = available.center() - QPoint(target.width() // 2, target.height() // 2)
        self.move(max(available.left(), top_left.x()), max(available.top(), top_left.y()))

    def _choose_import_file(self) -> None:
        source = self._file_picker(self)
        if source is not None:
            self.submit_import(source)

    def submit_import(self, source: Path) -> str:
        task_id = uuid4().hex
        file_id = uuid4().hex
        self._import_task_ids.add(task_id)
        self._task_queue.submit(
            TaskRequest(
                task_id=task_id,
                name=f"导入 {source.name}",
                file_id=file_id,
                engine=None,
                operation=IMPORT_WORKBOOK_OPERATION,
                payload={
                    "source_path": str(source),
                    "library_root": str(self._library_root),
                },
            )
        )
        return task_id

    def _apply_import_event(self, event: TaskEvent) -> None:
        if event.task_id not in self._import_task_ids:
            return
        if event.state is TaskState.SUCCEEDED and isinstance(event.result, ImportedWorkbook):
            self._file_library.add_workbook(event.result)
            self._import_task_ids.discard(event.task_id)
        elif event.state is TaskState.FAILED:
            self._error_presenter(self, event.message or "导入失败，请重试")
            self._import_task_ids.discard(event.task_id)
        elif event.state is TaskState.CANCELLED:
            self._import_task_ids.discard(event.task_id)

    def _select_workbook(self, workbook: ImportedWorkbook) -> None:
        if (
            self._current_workbook is not None
            and self._current_workbook.file_id != workbook.file_id
            and not self._resolve_unsaved_changes("切换文件")
        ):
            self._file_library.select_workbook(self._current_workbook.file_id)
            return
        if (
            self._current_workbook is not None
            and self._current_workbook.file_id != workbook.file_id
        ):
            self._cancel_processing_workflow(reload_base=False)
        if self._preview_task_id is not None:
            self._task_queue.cancel(self._preview_task_id)
        self._current_workbook = workbook
        self._previewed_version_id = (
            workbook.head_version.version_id if workbook.head_version is not None else None
        )
        self.setWindowTitle(f"风信子 — {workbook.display_name}")
        self._application_header.set_document_name(workbook.display_name)
        self._refresh_version_canvas(
            focus_file_id=workbook.file_id, focus_version_id=self._previewed_version_id
        )
        self._editor.clear_banner()
        self._load_preview(workbook.working_path, workbook.display_name, temporary=False)

    def _load_preview(self, source: Path, display_name: str, *, temporary: bool) -> None:
        task_id = uuid4().hex
        self._preview_task_id = task_id
        self._preview_is_temporary = temporary
        self._workbook_preview.set_loading(display_name)
        index_path = source.parent / "preview.sqlite" if temporary else preview_index_path(source)
        self._task_queue.submit(
            TaskRequest(
                task_id=task_id,
                name=f"加载 {display_name}",
                file_id=self._current_workbook.file_id if self._current_workbook else source.name,
                engine=None,
                operation=BUILD_PREVIEW_INDEX_OPERATION,
                payload={
                    "working_path": str(source),
                    "index_path": str(index_path),
                },
            )
        )

    def _apply_preview_event(self, event: TaskEvent) -> None:
        if event.task_id != self._preview_task_id:
            return
        if event.state is TaskState.SUCCEEDED and isinstance(event.result, WorkbookPreview):
            head = self._current_workbook.head_version if self._current_workbook else None
            if self._preview_is_temporary:
                # 需求第 17 节：临时结果上可继续轻量化单元格编辑，随应用统一生成版本。
                editable = True
            else:
                editable = head is not None and self._previewed_version_id == head.version_id
            self._workbook_preview.show_preview(event.result, editable=editable)
            if self._preview_is_temporary:
                self._temporary_preview = event.result
                self._editor.clear_banner()
                # 链式提交时未保存编辑已烘焙进新临时文件，加载后清空会话避免重复叠加。
                if self._processing_edits_baked:
                    self._workbook_preview.clear_edits()
                    self._processing_edits_baked = False
                self._show_processing_preview_ready()
                self._set_processing_navigation_enabled(True)
            else:
                self._editor.clear_banner()
            self._preview_task_id = None
        elif event.state is TaskState.FAILED:
            self._workbook_preview.set_error(event.message or "工作簿无法打开")
            if self._preview_is_temporary:
                self._editor.set_error(event.message or "临时结果无法打开")
                self._set_processing_navigation_enabled(True)
            self._preview_task_id = None
        elif event.state is TaskState.CANCELLED:
            was_temporary = self._preview_is_temporary
            self._preview_task_id = None
            if was_temporary:
                self._set_processing_navigation_enabled(True)
                workbook = self._current_workbook
                self._discard_processing_result()
                if workbook is not None:
                    self._load_preview(
                        workbook.working_path,
                        workbook.display_name,
                        temporary=False,
                    )
            else:
                self._workbook_preview.set_error("加载已取消，可重新选择文件")

    def _one_step_processing(self, action: str, columns: list[int]) -> None:
        if self._current_sheet_name() is None:
            self._error_presenter(self, "请先选择文件再使用处理功能")
            return
        # 去重与清空格有可调参数，先在功能区条下方原位展开确认（需求第 16 节）。
        if action == "deduplicate":
            self._editor.show_deduplicate_params(columns)
            return
        if action == "trim":
            self._editor.show_trim_params(columns)
            return
        if action == "delete_blank_rows":
            self._submit_processing_preview(
                operation=DELETE_BLANK_ROWS_PREVIEW_OPERATION,
                task_name="生成删除空白行预览",
                busy_message="正在检查空白行…",
                sheet_name=self._current_sheet_name() or "",
                parameters={
                    "key_columns": list(columns),
                    "allow_unsafe": False,
                },
            )

    def _confirm_deduplicate_params(self, parameters: object) -> None:
        if not isinstance(parameters, dict):
            return
        self._submit_processing_preview(
            operation=DEDUPLICATE_PREVIEW_OPERATION,
            task_name="生成删除重复行预览",
            busy_message="正在检查重复行…",
            sheet_name=self._current_sheet_name() or "",
            parameters=parameters,
        )

    def _confirm_trim_params(self, parameters: object) -> None:
        if not isinstance(parameters, dict):
            return
        self._submit_processing_preview(
            operation=TRIM_PREVIEW_OPERATION,
            task_name="生成清除空格预览",
            busy_message="正在清理文本空格…",
            sheet_name=self._current_sheet_name() or "",
            parameters=parameters,
        )

    def _current_sheet_name(self) -> str | None:
        return self._workbook_preview.current_sheet_name

    def _current_sheet_columns(self) -> tuple[tuple[str, ...], str | None]:
        headers, sheet = self._headers_for_current_sheet()
        return headers, sheet

    def _headers_for_current_sheet(self) -> tuple[tuple[str, ...], str | None]:
        preview = self._workbook_preview.current_preview()
        if preview is None or not preview.sheets:
            return (), None
        # 查找、筛选、排序都应作用于用户正在查看的工作表，而不是第一个工作表。
        current_name = self._workbook_preview.current_sheet_name
        sheet = next(
            (item for item in preview.sheets if item.title == current_name),
            preview.sheets[0],
        )
        source = SqliteGridDataSource(preview.index_path, sheet)
        try:
            headers = tuple(
                f"{get_column_letter(column + 1)} · {source.value_at(0, column) or '未命名列'}"
                for column in range(sheet.column_count)
            )
        finally:
            source.close()
        return headers, sheet.title

    def _open_filter_dialog(self) -> None:
        headers, sheet = self._headers_for_current_sheet()
        if sheet is None:
            self._error_presenter(self, "请先选择文件再使用筛选")
            return
        # 每次按当前工作表重建：切表后列头与目标表必须同步。
        if self._filter_dialog is not None:
            self._filter_dialog.deleteLater()
        self._filter_dialog = FilterDialog(sheet, headers, self)
        self._filter_dialog.params_submitted.connect(self._submit_filter_params)
        self._filter_dialog.show()
        self._filter_dialog.raise_()
        self._filter_dialog.activateWindow()

    def _submit_filter_params(self, payload: object) -> None:
        if not isinstance(payload, dict):
            return
        sheet_name = payload.get("sheet_name")
        if not isinstance(sheet_name, str):
            return
        conditions = payload.get("conditions")
        connector = payload.get("connector", "and")
        self._submit_processing_preview(
            operation=FILTER_PREVIEW_OPERATION,
            task_name="生成条件筛选预览",
            busy_message="正在计算匹配行…",
            sheet_name=sheet_name,
            parameters={"conditions": conditions, "connector": connector},
        )

    def _open_find_replace_dialog(self) -> None:
        _, sheet = self._headers_for_current_sheet()
        if sheet is None:
            self._error_presenter(self, "请先选择文件再使用查找替换")
            return
        if self._find_dialog is None:
            self._find_dialog = FindReplaceDialog(sheet, self)
            self._find_dialog.params_submitted.connect(self._submit_find_replace_preview)
            self._find_dialog.replace_selected_requested.connect(self._apply_single_find_replace)
        else:
            self._find_dialog.set_sheet(sheet)
        self._find_dialog.show()
        self._find_dialog.raise_()
        self._find_dialog.activateWindow()

    def _quick_sort_from_table(self, column: int, direction: str) -> None:
        sheet_name = self._current_sheet_name()
        if sheet_name is None:
            self._error_presenter(self, "请先选择文件再使用排序")
            return
        # 需求第 19.1 节：单列排序时明确提示按该列排序完整数据行。
        column_letter = get_column_letter(column + 1)
        self._submit_processing_preview(
            operation=SORT_PREVIEW_OPERATION,
            task_name="生成排序预览",
            busy_message=f"正在按 {column_letter} 列排序完整数据行…",
            sheet_name=sheet_name,
            parameters={"sort_keys": [{"column_index": column, "direction": direction}]},
        )

    def _submit_find_replace_preview(self, sheet_name: str, parameters: object) -> None:
        if not isinstance(parameters, dict):
            self._editor.set_error("查找替换参数无效，请重新输入")
            return
        replace_all = bool(parameters.get("replace_all"))
        # 范围为"当前工作表"时，以提交瞬间正在显示的工作表为准。
        target_sheet = sheet_name
        if not parameters.get("all_sheets"):
            target_sheet = self._current_sheet_name() or sheet_name
        self._submit_processing_preview(
            operation=FIND_REPLACE_PREVIEW_OPERATION,
            task_name="全部替换" if replace_all else "只查找",
            busy_message="正在替换匹配内容…" if replace_all else "正在查找匹配内容…",
            sheet_name=target_sheet,
            parameters=parameters,
        )

    def _open_sort_dialog(self, initial_column: int = 0) -> None:
        headers, sheet = self._headers_for_current_sheet()
        if sheet is None:
            self._error_presenter(self, "请先选择文件再使用多列排序")
            return
        if self._sort_dialog is None:
            self._sort_dialog = SortDialog(sheet, self)
            self._sort_dialog.params_submitted.connect(self._submit_multi_sort)
        self._sort_dialog.set_sheet(sheet)
        self._sort_dialog.set_columns(headers, initial_column)
        self._sort_dialog.show()
        self._sort_dialog.raise_()
        self._sort_dialog.activateWindow()

    def _submit_multi_sort(self, parameters: object) -> None:
        if not isinstance(parameters, dict):
            return
        sheet_name = parameters.get("sheet_name")
        if not isinstance(sheet_name, str):
            return
        keys = parameters.get("sort_keys")
        if not isinstance(keys, list) or not keys:
            return
        letters = "、".join(
            get_column_letter(int(key["column_index"]) + 1)
            for key in keys
            if isinstance(key, dict) and isinstance(key.get("column_index"), int)
        )
        self._submit_processing_preview(
            operation=SORT_PREVIEW_OPERATION,
            task_name="生成多列排序预览",
            busy_message=f"正在按 {letters} 列排序完整数据行…",
            sheet_name=sheet_name,
            parameters={"sort_keys": keys},
        )

    def _apply_single_find_replace(self, row: int) -> None:
        """需求第 19.4 节：逐项替换把选中匹配写入当前编辑会话。"""
        dialog = self._find_dialog
        if dialog is None:
            return
        change = dialog.match_at(row)
        if change is None:
            return
        if self._processing_result is not None:
            dialog.set_status("临时结果上暂不支持逐项替换，请先应用或取消临时预览", True)
            return
        if not self._workbook_preview.is_editable:
            dialog.set_status("当前为只读预览，请先从此版本继续后再逐项替换", True)
            return
        sheet_name, row_number, column_number, before, after = change
        self._workbook_preview.apply_cell_edit(
            sheet_name,
            row_number - 1,
            column_number - 1,
            base_value=before,
            new_value=after,
        )
        dialog.mark_replaced(row)

    def _submit_processing_preview(
        self,
        *,
        operation: str,
        task_name: str,
        busy_message: str,
        sheet_name: str,
        parameters: dict[str, object],
    ) -> None:
        workbook = self._current_workbook
        parent = workbook.head_version if workbook is not None else None
        if workbook is None or parent is None:
            self._editor.set_error("当前文件尚未建立可处理的根版本")
            return
        # 已有临时结果时进入链式模式：新操作以临时结果为源，可继续叠加；
        # 首个操作以 HEAD 快照为源。两种情况下未保存的单元格编辑都随任务
        # 烘焙进新的临时文件（DEC-20260816-039），处理操作永不丢编辑。
        chained = self._processing_result is not None
        pending_edits = [
            {
                "sheet_name": edit.sheet_name,
                "row": edit.row,
                "column": edit.column,
                "value": edit.value,
            }
            for edit in self._workbook_preview.pending_edits()
        ]
        if not chained:
            # 编辑已捕获进 payload，丢弃旧临时结果时保留编辑会话；
            # 任务失败或用户取消时编辑仍然在网格里，不会无声丢失。
            self._discard_processing_result(clear_edits=False)
        source_path = (
            self._processing_result.preview_path
            if chained and self._processing_result is not None
            else parent.snapshot_path
        )
        base_version_id = (
            self._processing_base_version_id if chained else parent.version_id
        ) or parent.version_id
        task_id = uuid4().hex
        preview_id = uuid4().hex
        preview_path = (
            workbook.working_path.parent.parent / ".previews" / preview_id / "result.xlsx"
        )
        self._processing_previous_path = (
            self._processing_result.preview_path
            if chained and self._processing_result is not None
            else None
        )
        self._processing_edits_baked = bool(pending_edits)
        self._processing_submitted_operation = operation
        self._processing_submitted_task_name = task_name
        self._processing_task_id = task_id
        self._editor.hide_params_bar()
        self._editor.set_busy(busy_message)
        self._set_processing_navigation_enabled(False)
        self._task_queue.submit(
            TaskRequest(
                task_id=task_id,
                name=task_name,
                file_id=workbook.file_id,
                engine=None,
                operation=operation,
                payload={
                    "source_path": str(source_path),
                    "preview_path": str(preview_path),
                    "parent_version_id": base_version_id,
                    "sheet_name": sheet_name,
                    "edits": pending_edits,
                    **parameters,
                },
            )
        )

    def _apply_processing_event(self, event: TaskEvent) -> None:
        if event.task_id != self._processing_task_id:
            return
        if (
            event.state is TaskState.SUCCEEDED
            and isinstance(
                event.result,
                FindReplacePreviewResult,
            )
            and not event.result.replace_all
        ):
            self._processing_task_id = None
            self._set_processing_navigation_enabled(True)
            # 只查找不产生新临时文件：链式状态保持不变，也不清理旧临时目录。
            self._processing_previous_path = None
            self._processing_edits_baked = False
            if self._find_dialog is not None:
                self._find_dialog.set_matches(
                    tuple(
                        (change.sheet_name, change.row, change.column, change.before, change.after)
                        for change in event.result.changes
                    )
                )
            self._editor.clear_banner()
        elif event.state is TaskState.SUCCEEDED and isinstance(
            event.result,
            (
                SortPreviewResult,
                DeduplicatePreviewResult,
                DeleteBlankRowsPreviewResult,
                FilterPreviewResult,
                TrimPreviewResult,
                FindReplacePreviewResult,
            ),
        ):
            self._processing_result = event.result
            self._processing_task_id = None
            display_name = (
                self._current_workbook.display_name if self._current_workbook else "临时结果"
            )
            if event.result.preview_path is not None:
                if self._processing_base_version_id is None:
                    self._processing_base_version_id = event.result.parent_version_id
                self._processing_steps.append(
                    {
                        "operation": self._processing_submitted_operation or "",
                        "label": self._processing_submitted_task_name or "",
                        "sheet": event.result.sheet_name,
                    }
                )
                self._load_preview(event.result.preview_path, display_name, temporary=True)
                # set_loading 已同步释放旧临时文件的 SQLite 源，可安全清理上一环。
                if self._processing_previous_path is not None:
                    shutil.rmtree(self._processing_previous_path.parent, ignore_errors=True)
                    self._processing_previous_path = None
        elif event.state is TaskState.FAILED:
            self._processing_task_id = None
            self._processing_previous_path = None
            self._processing_edits_baked = False
            self._set_processing_navigation_enabled(True)
            self._editor.set_error(event.message or "处理预览生成失败")
        elif event.state is TaskState.CANCELLED:
            self._processing_task_id = None
            self._processing_previous_path = None
            self._processing_edits_baked = False
            self._set_processing_navigation_enabled(True)
            workbook = self._current_workbook
            self._discard_processing_result()
            if workbook is not None:
                self._load_preview(workbook.working_path, workbook.display_name, temporary=False)

    def _submit_apply_processing_preview(self) -> None:
        workbook = self._current_workbook
        result = self._processing_result
        if workbook is None or result is None:
            self._editor.set_error("没有可应用的临时结果")
            return
        task_id = uuid4().hex
        version_id = uuid4().hex
        self._apply_task_id = task_id
        self._apply_version_id = version_id
        self._workbook_preview.clear_preview("正在应用临时结果…")
        self._editor.set_busy("正在创建不可变子版本…")
        self._set_processing_navigation_enabled(False)
        if len(self._processing_steps) > 1:
            # 链式多步处理：一个节点承载全部步骤，参数记录每一步。
            operation = APPLY_CHAINED_PREVIEW_OPERATION
            task_name = "应用多步处理结果"
            parameters: dict[str, object] = {"steps": list(self._processing_steps)}
        elif isinstance(result, SortPreviewResult):
            operation = APPLY_SORT_PREVIEW_OPERATION
            task_name = "应用排序结果"
            parameters = {
                "sort_keys": [
                    {"column_index": key.column_index, "direction": key.direction.value}
                    for key in result.sort_keys
                ]
            }
        elif isinstance(result, DeduplicatePreviewResult):
            operation = APPLY_DEDUPLICATE_PREVIEW_OPERATION
            task_name = "应用删除重复行结果"
            parameters = {
                "key_columns": list(result.key_columns),
                "keep": result.keep.value,
                "ignore_case": result.ignore_case,
                "trim_whitespace": result.trim_whitespace,
                "duplicate_groups": len(result.duplicate_groups),
                "deleted_rows": result.deleted_rows,
            }
        elif isinstance(result, DeleteBlankRowsPreviewResult):
            operation = APPLY_DELETE_BLANK_ROWS_PREVIEW_OPERATION
            task_name = "应用删除空白行结果"
            parameters = {
                "key_columns": list(result.key_columns),
                "allow_unsafe": result.allow_unsafe,
                "compatibility_warning": result.compatibility_warning,
                "deleted_row_numbers": list(result.deleted_row_numbers),
            }
        elif isinstance(result, TrimPreviewResult):
            operation = APPLY_TRIM_PREVIEW_OPERATION
            task_name = "应用清除空格结果"
            parameters = {
                "key_columns": list(result.key_columns),
                "collapse_spaces": result.collapse_spaces,
                "trimmed_cells": len(result.trimmed_cells),
            }
        elif isinstance(result, FindReplacePreviewResult):
            operation = APPLY_FIND_REPLACE_PREVIEW_OPERATION
            task_name = "应用查找替换结果"
            parameters = {
                "mode": result.mode.value,
                "find_text": result.find_text,
                "replace_text": result.replace_text,
                "match_case": result.match_case,
                "whole_cell": result.whole_cell,
                "trim_whitespace": result.trim_whitespace,
                "sheets": list(result.sheets),
                "replaced": len(result.changes),
            }
        else:
            operation = APPLY_FILTER_PREVIEW_OPERATION
            task_name = "应用条件筛选结果"
            parameters = {
                "conditions": [
                    {
                        "column_index": condition.column_index,
                        "operator": condition.operator.value,
                        "value_type": condition.value_type.value,
                        "value": condition.value,
                        "second_value": condition.second_value,
                    }
                    for condition in result.conditions
                ],
                "connector": result.connector.value,
                "matched_rows": result.matched_rows,
                "total_rows": result.total_rows,
            }
        manual_edits = [
            {
                "sheet_name": edit.sheet_name,
                "row": edit.row,
                "column": edit.column,
                "value": edit.value,
            }
            for edit in self._workbook_preview.pending_edits()
        ]
        self._task_queue.submit(
            TaskRequest(
                task_id=task_id,
                name=task_name,
                file_id=workbook.file_id,
                engine=None,
                operation=operation,
                payload={
                    "library_root": str(self._library_root),
                    "preview_path": str(result.preview_path),
                    "preview_hash": result.content_hash,
                    "parent_version_id": result.parent_version_id,
                    "version_id": version_id,
                    "sheet_name": result.sheet_name,
                    "edits": manual_edits,
                    **parameters,
                },
            )
        )

    def _apply_apply_event(self, event: TaskEvent) -> None:
        if event.task_id != self._apply_task_id:
            return
        if event.state is TaskState.SUCCEEDED and isinstance(event.result, ImportedWorkbook):
            self._finish_applied_version(event.result)
        elif event.state is TaskState.FAILED:
            recovered = self._recover_applied_version()
            if not recovered:
                self._restore_temporary_preview(
                    f"应用失败：{event.message or '请重试或取消临时结果'}"
                )
        elif event.state is TaskState.CANCELLED:
            self._restore_temporary_preview("应用已取消，可重试或取消临时结果")

    def _finish_applied_version(self, workbook: ImportedWorkbook) -> None:
        self._discard_processing_result()
        self._apply_task_id = None
        self._apply_version_id = None
        self._set_processing_navigation_enabled(True)
        self._current_workbook = workbook
        self._previewed_version_id = (
            workbook.head_version.version_id if workbook.head_version is not None else None
        )
        self._file_library.replace_workbook(workbook)
        self._refresh_version_canvas()
        self._load_preview(workbook.working_path, workbook.display_name, temporary=False)
        if self._close_after_apply:
            self._close_after_apply = False
            self.close()

    def _workbook_preview_undo(self) -> None:
        self._workbook_preview.undo()

    def _workbook_preview_redo(self) -> None:
        self._workbook_preview.redo()

    def _apply_edit_state(self, dirty: bool, can_undo: bool, can_redo: bool) -> None:
        # 临时结果上的编辑通过“应用生成版本”提交，此时“保存为新版本”不适用。
        self._command_bar.set_edit_state(
            dirty and not self._preview_is_temporary,
            can_undo,
            can_redo,
        )

    def _apply_pending_edit_count(self, count: int) -> None:
        self._editor.set_pending_edit_count(count)

    def _submit_manual_save(self) -> None:
        workbook = self._current_workbook
        head = workbook.head_version if workbook is not None else None
        edits = self._workbook_preview.pending_edits()
        if (
            workbook is None
            or head is None
            or not edits
            or self._manual_save_task_id is not None
            or self._previewed_version_id != head.version_id
        ):
            return
        task_id = uuid4().hex
        version_id = uuid4().hex
        self._manual_save_task_id = task_id
        self._manual_save_version_id = version_id
        self._set_processing_navigation_enabled(False)
        self._task_queue.submit(
            TaskRequest(
                task_id=task_id,
                name="保存手动编辑",
                file_id=workbook.file_id,
                engine=None,
                operation=SAVE_MANUAL_EDITS_OPERATION,
                payload={
                    "library_root": str(self._library_root),
                    "parent_version_id": head.version_id,
                    "version_id": version_id,
                    "edits": [
                        {
                            "sheet_name": edit.sheet_name,
                            "row": edit.row,
                            "column": edit.column,
                            "value": edit.value,
                        }
                        for edit in edits
                    ],
                },
            )
        )

    def _apply_manual_save_event(self, event: TaskEvent) -> None:
        if event.task_id != self._manual_save_task_id:
            return
        if event.state is TaskState.SUCCEEDED and isinstance(event.result, ImportedWorkbook):
            self._manual_save_task_id = None
            self._manual_save_version_id = None
            self._workbook_preview.clear_edits()
            self._set_processing_navigation_enabled(True)
            self._current_workbook = event.result
            head = event.result.head_version
            self._previewed_version_id = head.version_id if head is not None else None
            self._file_library.replace_workbook(event.result)
            self._refresh_version_canvas()
            self._load_preview(
                event.result.working_path,
                event.result.display_name,
                temporary=False,
            )
            if self._close_after_manual_save:
                self._close_after_manual_save = False
                self.close()
        elif event.state in {TaskState.FAILED, TaskState.CANCELLED}:
            self._manual_save_task_id = None
            self._manual_save_version_id = None
            self._close_after_manual_save = False
            self._set_processing_navigation_enabled(True)
            self._error_presenter(
                self,
                event.message
                or ("保存已取消" if event.state is TaskState.CANCELLED else "保存失败"),
            )

    def _preview_version(self, file_id: str, version_id: str) -> None:
        workbook = self._current_workbook
        if self._checkout_task_id is not None:
            return
        if workbook is None or workbook.file_id != file_id:
            try:
                target = MetadataStore(self._library_root).get_workbook(file_id)
            except ValueError as error:
                self._error_presenter(self, str(error))
                return
            if not self._resolve_unsaved_changes("切换文件"):
                return
            if self._processing_result is not None:
                self._cancel_processing_workflow(reload_base=False)
            if self._preview_task_id is not None:
                self._task_queue.cancel(self._preview_task_id)
                self._preview_task_id = None
            self._current_workbook = target
            self._previewed_version_id = None
            self._file_library.select_workbook(file_id)
            self.setWindowTitle(f"风信子 — {target.display_name}")
            self._application_header.set_document_name(target.display_name)
            self._editor.clear_banner()
            self._refresh_version_canvas(focus_file_id=file_id, focus_version_id=version_id)
            workbook = target
        elif version_id != self._previewed_version_id and not self._resolve_unsaved_changes(
            "切换版本"
        ):
            return
        try:
            version = MetadataStore(self._library_root).get_version(workbook.file_id, version_id)
        except ValueError as error:
            self._error_presenter(self, str(error))
            return
        if version.deleted_at is not None:
            self._error_presenter(self, "已删除版本只能恢复，不能预览")
            return
        if self._processing_result is not None:
            self._cancel_processing_workflow(reload_base=False)
        if self._preview_task_id is not None:
            self._task_queue.cancel(self._preview_task_id)
        self._previewed_version_id = version_id
        self._refresh_storage_stats()
        self._load_preview(version.snapshot_path, workbook.display_name, temporary=False)

    def _continue_from_version(self, file_id: str, version_id: str) -> None:
        if self._current_workbook is None or self._current_workbook.file_id != file_id:
            try:
                target = MetadataStore(self._library_root).get_workbook(file_id)
            except ValueError as error:
                self._error_presenter(self, str(error))
                return
            if not self._resolve_unsaved_changes("切换文件"):
                return
            if self._preview_task_id is not None:
                self._task_queue.cancel(self._preview_task_id)
                self._preview_task_id = None
            self._current_workbook = target
            self._previewed_version_id = None
            self._file_library.select_workbook(file_id)
            self.setWindowTitle(f"风信子 — {target.display_name}")
            self._application_header.set_document_name(target.display_name)
            self._editor.clear_banner()
        workbook = self._current_workbook
        head = workbook.head_version if workbook is not None else None
        if workbook is None or head is None or version_id == head.version_id:
            return
        if not self._resolve_unsaved_changes("切换当前工作版本"):
            return
        if self._preview_task_id is not None:
            self._task_queue.cancel(self._preview_task_id)
            self._preview_task_id = None
        task_id = uuid4().hex
        self._checkout_task_id = task_id
        self._set_processing_navigation_enabled(False)
        self._task_queue.submit(
            TaskRequest(
                task_id=task_id,
                name="从历史版本继续",
                file_id=workbook.file_id,
                engine=None,
                operation=CHECKOUT_VERSION_OPERATION,
                payload={
                    "library_root": str(self._library_root),
                    "version_id": version_id,
                    "expected_head_version_id": head.version_id,
                },
            )
        )

    def _apply_checkout_event(self, event: TaskEvent) -> None:
        if event.task_id != self._checkout_task_id:
            return
        if event.state is TaskState.SUCCEEDED and isinstance(event.result, ImportedWorkbook):
            self._checkout_task_id = None
            self._current_workbook = event.result
            head = event.result.head_version
            self._previewed_version_id = head.version_id if head is not None else None
            self._file_library.replace_workbook(event.result)
            self._refresh_version_canvas()
            self._set_processing_navigation_enabled(True)
            self._load_preview(
                event.result.working_path,
                event.result.display_name,
                temporary=False,
            )
        elif event.state is TaskState.FAILED:
            self._checkout_task_id = None
            self._set_processing_navigation_enabled(True)
            self._error_presenter(self, event.message or "当前工作版本切换失败")
            self._restore_current_head_preview()
        elif event.state is TaskState.CANCELLED:
            self._checkout_task_id = None
            self._set_processing_navigation_enabled(True)
            self._restore_current_head_preview()

    def _request_delete_version(self, file_id: str, version_id: str) -> None:
        workbook = self._current_workbook
        if workbook is not None and workbook.file_id != file_id:
            try:
                workbook = MetadataStore(self._library_root).get_workbook(file_id)
            except ValueError as error:
                self._error_presenter(self, str(error))
                return
        head = workbook.head_version if workbook is not None else None
        if workbook is None or head is None or self._delete_task_id is not None:
            return
        if (
            self._processing_result is not None
            or self._processing_task_id is not None
            or self._apply_task_id is not None
        ):
            self._error_presenter(self, "请先取消或应用当前临时预览，再删除版本")
            return
        store = MetadataStore(self._library_root)
        try:
            plan = store.plan_version_deletion(file_id, version_id)
        except ValueError as error:
            self._error_presenter(self, str(error))
            return
        replacement: VersionRecord | None = None
        if plan.requires_head_switch:
            if len(plan.replacement_candidates) == 1:
                replacement = plan.replacement_candidates[0]
            else:
                selected_id = self._version_choice_presenter(
                    self,
                    plan.replacement_candidates,
                )
                if selected_id is None:
                    return
                replacement = next(
                    (
                        candidate
                        for candidate in plan.replacement_candidates
                        if candidate.version_id == selected_id
                    ),
                    None,
                )
                if replacement is None:
                    self._error_presenter(self, "选择的新 HEAD 已不可用，请刷新版本树")
                    return
        if replacement is None:
            detail = "当前工作版本 HEAD 保持不变。"
        else:
            detail = f"删除后 HEAD 将切换到“{replacement.name}”。"
        if not self._confirmation_presenter(
            self,
            "删除版本",
            f"确定将“{plan.target.name}”移入回收状态吗？\n{detail}\n可立即撤销或稍后恢复。",
        ):
            return
        if self._preview_task_id is not None:
            self._task_queue.cancel(self._preview_task_id)
            self._preview_task_id = None
        task_id = uuid4().hex
        self._delete_task_id = task_id
        self._delete_version_id = version_id
        self._delete_file_id = file_id
        self._set_processing_navigation_enabled(False)
        self._task_queue.submit(
            TaskRequest(
                task_id=task_id,
                name=f"删除版本 {plan.target.name}",
                file_id=file_id,
                engine=None,
                operation=DELETE_VERSION_OPERATION,
                payload={
                    "library_root": str(self._library_root),
                    "version_id": version_id,
                    "expected_head_version_id": head.version_id,
                    "replacement_version_id": (
                        replacement.version_id if replacement is not None else None
                    ),
                },
            )
        )

    def _apply_delete_event(self, event: TaskEvent) -> None:
        if event.task_id != self._delete_task_id:
            return
        deleted_file_id = self._delete_file_id
        deleted_version_id = self._delete_version_id
        if event.state is TaskState.SUCCEEDED and isinstance(event.result, ImportedWorkbook):
            self._delete_task_id = None
            self._delete_version_id = None
            self._file_library.replace_workbook(event.result)
            is_current = (
                self._current_workbook is not None
                and self._current_workbook.file_id == event.result.file_id
            )
            if is_current:
                self._current_workbook = event.result
                head = event.result.head_version
                self._previewed_version_id = head.version_id if head is not None else None
            if is_current:
                self._set_processing_navigation_enabled(True)
                self._load_preview(
                    event.result.working_path,
                    event.result.display_name,
                    temporary=False,
                )
            self._refresh_version_canvas(
                focus_file_id=deleted_file_id,
                focus_version_id=event.result.head_version.version_id
                if event.result.head_version is not None
                else None,
            )
            if deleted_file_id is not None and deleted_version_id is not None:
                self._version_tree.show_delete_undo(deleted_file_id, deleted_version_id)
        elif event.state is TaskState.FAILED:
            self._delete_task_id = None
            self._delete_version_id = None
            self._delete_file_id = None
            self._set_processing_navigation_enabled(True)
            self._error_presenter(self, event.message or "版本删除失败")
            self._restore_current_head_preview()
        elif event.state is TaskState.CANCELLED:
            self._delete_task_id = None
            self._delete_version_id = None
            self._delete_file_id = None
            self._set_processing_navigation_enabled(True)
            self._restore_current_head_preview()

    def _restore_deleted_version(self, file_id: str, version_id: str) -> None:
        if self._delete_task_id is not None:
            return
        try:
            MetadataStore(self._library_root).restore_version(file_id, version_id)
            refreshed = MetadataStore(self._library_root).get_workbook(file_id)
        except ValueError as error:
            self._error_presenter(self, str(error))
            return
        self._file_library.replace_workbook(refreshed)
        if self._current_workbook is not None and self._current_workbook.file_id == file_id:
            self._current_workbook = refreshed
        self._refresh_version_canvas()
        self._version_tree.clear_delete_undo()

    def _restore_current_head_preview(self) -> None:
        workbook = self._current_workbook
        head = workbook.head_version if workbook is not None else None
        if workbook is None or head is None:
            return
        self._previewed_version_id = head.version_id
        self._refresh_version_canvas(
            focus_file_id=workbook.file_id, focus_version_id=head.version_id
        )
        self._load_preview(workbook.working_path, workbook.display_name, temporary=False)

    def _recover_applied_version(self) -> bool:
        workbook = self._current_workbook
        version_id = self._apply_version_id
        if workbook is None or version_id is None:
            return False
        store = MetadataStore(self._library_root)
        store.reconcile_manifests()
        try:
            recovered = store.get_workbook(workbook.file_id)
        except ValueError:
            return False
        head = recovered.head_version
        if head is None or head.version_id != version_id:
            return False
        self._finish_applied_version(recovered)
        return True

    def _restore_temporary_preview(self, message: str) -> None:
        self._apply_task_id = None
        self._apply_version_id = None
        self._set_processing_navigation_enabled(True)
        if self._temporary_preview is not None:
            self._workbook_preview.show_preview(self._temporary_preview, editable=True)
            self._editor.clear_banner()
            self._show_processing_preview_ready(message)
        else:
            self._editor.set_error(message)

    def _show_processing_preview_ready(self, message: str | None = None) -> None:
        result = self._processing_result
        if isinstance(result, DeduplicatePreviewResult):
            summary = f"{len(result.duplicate_groups)} 个重复组 · 将删除 {result.deleted_rows} 行"
            can_details = bool(result.duplicate_groups)
        elif isinstance(result, DeleteBlankRowsPreviewResult):
            summary = f"将删除 {len(result.deleted_row_numbers)} 行空白行"
            can_details = bool(result.deleted_row_numbers)
        elif isinstance(result, FilterPreviewResult):
            ratio = result.matched_rows / result.total_rows if result.total_rows else 0.0
            summary = f"匹配 {result.matched_rows} / {result.total_rows} 行 · {ratio:.1%}"
            can_details = False
        elif isinstance(result, TrimPreviewResult):
            summary = f"将清理 {len(result.trimmed_cells)} 个单元格"
            can_details = bool(result.trimmed_cells)
        elif isinstance(result, FindReplacePreviewResult):
            summary = f"共替换 {len(result.changes)} 处"
            can_details = True
        elif isinstance(result, SortPreviewResult):
            columns_text = "、".join(
                get_column_letter(key.column_index + 1) for key in result.sort_keys
            )
            summary = f"已按 {columns_text} 列排序完整数据行（表头不参与）"
            can_details = False
        else:
            summary = ""
            can_details = False
        chain_text = (
            f"已连续 {len(self._processing_steps)} 步处理"
            if len(self._processing_steps) > 1
            else ""
        )
        if message:
            parts = [message, chain_text, summary]
        else:
            parts = ["临时结果", chain_text, summary, "尚未生成版本"]
        text = " · ".join(part for part in parts if part)
        self._editor.set_preview_ready(text, can_details=can_details)

    def _show_processing_details(self) -> None:
        result = self._processing_result
        if result is None:
            return
        if isinstance(result, DeduplicatePreviewResult):
            title = "删除重复行明细 · 保留行与删除行"
            model: QAbstractTableModel = DuplicateMappingModel(
                tuple((group.kept_row, group.deleted_rows) for group in result.duplicate_groups)
            )
        elif isinstance(result, DeleteBlankRowsPreviewResult):
            title = "删除空白行明细 · 即将删除的行"
            model = DeletedRowsModel(result.deleted_row_numbers)
        elif isinstance(result, TrimPreviewResult):
            title = "清除空格明细 · 修改前后内容"
            model = TrimDetailsModel(
                tuple(
                    (
                        f"第 {cell.row} 行",
                        f"第 {get_column_letter(cell.column)} 列",
                        cell.before,
                        cell.after,
                    )
                    for cell in result.trimmed_cells
                )
            )
        elif isinstance(result, FindReplacePreviewResult):
            title = "查找替换明细 · 修改前后内容"
            model = FindDetailsModel(
                tuple(
                    (change.sheet_name, change.row, change.column, change.before, change.after)
                    for change in result.changes
                )
            )
        else:
            return
        dialog = ProcessingDetailsDialog(title, model, self)
        dialog.show()
        dialog.raise_()
        dialog.activateWindow()

    def _cancel_processing_workflow(self, *, reload_base: bool = True) -> None:
        if self._apply_task_id is not None:
            self._task_queue.cancel(self._apply_task_id)
            self._editor.set_busy("正在请求取消应用…")
            return
        if self._processing_task_id is not None:
            self._task_queue.cancel(self._processing_task_id)
            self._editor.set_busy("正在请求取消处理预览…")
            return
        if self._preview_task_id is not None and self._preview_is_temporary:
            self._task_queue.cancel(self._preview_task_id)
            self._editor.set_busy("正在请求取消临时结果加载…")
            return
        workbook = self._current_workbook
        # 取消会丢弃临时结果上尚未保存的单元格编辑，先让用户选择（DEC-20260816-039）。
        if self._workbook_preview.pending_edits() and not self._resolve_unsaved_changes(
            "取消临时结果"
        ):
            return
        self._discard_processing_result()
        if reload_base and workbook is not None:
            self._load_preview(workbook.working_path, workbook.display_name, temporary=False)

    def _discard_processing_result(self, *, clear_edits: bool = True) -> None:
        # 临时结果连同其上的未提交编辑一起废弃，避免残留编辑污染下一次预览。
        # 没有临时结果时（如只查找）不动网格，避免把数据预览误清成导入空状态。
        # clear_edits=False 用于编辑已随新任务提交的场合（如首个处理操作），
        # 失败/取消时编辑仍留在网格中。
        had_temporary = self._processing_result is not None
        if clear_edits:
            self._workbook_preview.clear_edits()
        if had_temporary:
            self._workbook_preview.clear_preview("正在处理…")
        self._editor.clear_banner()
        preview_path = getattr(self._processing_result, "preview_path", None)
        if had_temporary and preview_path is not None:
            shutil.rmtree(preview_path.parent, ignore_errors=True)
        self._processing_result = None
        self._temporary_preview = None
        self._processing_steps = []
        self._processing_base_version_id = None
        self._processing_previous_path = None
        self._processing_edits_baked = False

    def _refresh_version_canvas(
        self,
        *,
        focus_file_id: str | None = None,
        focus_version_id: str | None = None,
    ) -> None:
        self._refresh_storage_stats()
        store = MetadataStore(self._library_root)
        workbooks = store.list_workbooks()
        current_id = self._current_workbook.file_id if self._current_workbook is not None else None
        # 泳道顺序固定为导入顺序，不随当前文件重排，保证节点坐标稳定。
        trees: list[FileVersionTree] = []
        for workbook in workbooks:
            head = workbook.head_version
            trees.append(
                FileVersionTree(
                    file_id=workbook.file_id,
                    display_name=workbook.display_name,
                    versions=store.list_versions(workbook.file_id),
                    head_version_id=head.version_id if head is not None else None,
                    layouts=store.list_version_layouts(workbook.file_id),
                )
            )
        self._command_bar.set_version_available(self._previewed_version_id is not None)
        self._version_tree.set_workbooks(
            tuple(trees),
            current_file_id=current_id,
            focus_file_id=focus_file_id or current_id,
            focus_version_id=focus_version_id,
        )

    def _export_current_version(self) -> None:
        workbook = self._current_workbook
        head = workbook.head_version if workbook is not None else None
        version_id = self._previewed_version_id or (head.version_id if head is not None else None)
        if version_id is not None and self._current_workbook is not None:
            self._request_export_version(self._current_workbook.file_id, version_id, False)

    def _request_export_version(self, file_id: str, version_id: str, save_as: bool) -> None:
        workbook = self._current_workbook
        if workbook is None or workbook.file_id != file_id:
            try:
                workbook = MetadataStore(self._library_root).get_workbook(file_id)
            except ValueError as error:
                self._error_presenter(self, str(error))
                return
        if self._export_task_id is not None:
            return
        try:
            version = MetadataStore(self._library_root).get_version(file_id, version_id)
        except ValueError as error:
            self._error_presenter(self, str(error))
            return
        if version.deleted_at is not None:
            self._error_presenter(self, "已删除版本不能导出，请先恢复")
            return
        extension = (
            workbook.original_path.suffix
            if version.parent_version_id is None
            else version.snapshot_path.suffix
        )
        filename = suggested_export_filename(
            workbook.display_name,
            version.name,
            version.created_at.astimezone().strftime("%Y%m%d-%H%M%S"),
            extension,
        )
        payload: dict[str, object] = {
            "library_root": str(self._library_root),
            "version_id": version_id,
        }
        if save_as:
            destination = self._save_as_picker(self, filename)
            if destination is None:
                return
            payload["destination_path"] = str(destination)
        else:
            payload["destination_directory"] = str(default_export_directory())
        task_id = uuid4().hex
        self._export_task_id = task_id
        self._task_queue.submit(
            TaskRequest(
                task_id=task_id,
                name=f"导出 {version.name}",
                file_id=workbook.file_id,
                engine=None,
                operation=EXPORT_VERSION_OPERATION,
                payload=payload,
            )
        )

    def _apply_export_event(self, event: TaskEvent) -> None:
        if event.task_id != self._export_task_id:
            return
        if event.state is TaskState.SUCCEEDED and isinstance(event.result, ExportedVersion):
            self._export_task_id = None
            self._export_presenter(self, event.result.path)
        elif event.state in {TaskState.FAILED, TaskState.CANCELLED}:
            self._export_task_id = None
            self._error_presenter(
                self,
                event.message
                or ("导出已取消" if event.state is TaskState.CANCELLED else "导出失败"),
            )

    def _request_delete_file(self, workbook: ImportedWorkbook) -> None:
        if self._purge_task_id is not None:
            return
        head = workbook.head_version
        if head is None:
            self._error_presenter(self, "旧记录尚未建立版本，暂不能在软件内删除")
            return
        is_current = (
            self._current_workbook is not None
            and self._current_workbook.file_id == workbook.file_id
        )
        if is_current and not self._resolve_unsaved_changes("删除文件"):
            return
        if not self._confirmation_presenter(
            self,
            "删除文件",
            f"确定将“{workbook.display_name}”连同全部版本移入回收站吗？\n"
            "可随时在回收站中恢复；永久删除前不会清除磁盘文件。",
        ):
            return
        try:
            MetadataStore(self._library_root).soft_delete_file(workbook.file_id, head.version_id)
        except ValueError as error:
            self._error_presenter(self, str(error))
            return
        self._file_library.remove_workbook(workbook.file_id)
        if is_current:
            self._clear_current_workbook()

    def _clear_current_workbook(self) -> None:
        if self._preview_task_id is not None:
            self._task_queue.cancel(self._preview_task_id)
            self._preview_task_id = None
        self._current_workbook = None
        self._previewed_version_id = None
        self._storage_status.set_empty()
        self.setWindowTitle("风信子")
        self._application_header.set_document_name(None)
        self._command_bar.set_version_available(False)
        self._editor.clear_banner()
        self._refresh_version_canvas()
        self._workbook_preview.clear_preview()

    def _open_recycle_bin(self) -> None:
        if self._recycle_dialog is None:
            self._recycle_dialog = RecycleBinDialog(parent=self)
            self._recycle_dialog.restore_file_requested.connect(self._restore_file_from_bin)
            self._recycle_dialog.restore_version_requested.connect(self._restore_version_from_bin)
            self._recycle_dialog.purge_file_requested.connect(self._request_purge_file)
        self._refresh_recycle_bin()
        self._recycle_dialog.show()
        self._recycle_dialog.raise_()
        self._recycle_dialog.activateWindow()

    def _recycle_entries(self) -> tuple[RecycleEntry, ...]:
        store = MetadataStore(self._library_root)
        entries: list[RecycleEntry] = []
        for record in store.list_deleted_files():
            entries.append(
                RecycleEntry(
                    kind="file",
                    file_id=record.file_id,
                    file_display_name=record.display_name,
                    version_count=len(store.list_versions(record.file_id)),
                    deleted_at=record.deleted_at,
                )
            )
        for record in store.list_workbooks():
            for version in store.list_versions(record.file_id):
                if version.deleted_at is None:
                    continue
                entries.append(
                    RecycleEntry(
                        kind="version",
                        file_id=record.file_id,
                        file_display_name=record.display_name,
                        version_id=version.version_id,
                        version_name=version.name,
                        deleted_at=version.deleted_at,
                    )
                )
        entries.sort(
            key=lambda entry: entry.deleted_at or datetime.now().astimezone(), reverse=True
        )
        return tuple(entries)

    def _refresh_recycle_bin(self) -> None:
        if self._recycle_dialog is not None:
            self._recycle_dialog.refresh(self._recycle_entries())

    def _restore_file_from_bin(self, file_id: str) -> None:
        try:
            record = MetadataStore(self._library_root).restore_file(file_id)
        except ValueError as error:
            self._error_presenter(self, str(error))
            self._refresh_recycle_bin()
            return
        self._file_library.restore_workbook(record)
        self._refresh_recycle_bin()

    def _restore_version_from_bin(self, file_id: str, version_id: str) -> None:
        try:
            MetadataStore(self._library_root).restore_version(file_id, version_id)
        except ValueError as error:
            self._error_presenter(self, str(error))
            self._refresh_recycle_bin()
            return
        current = self._current_workbook
        if current is not None and current.file_id == file_id:
            refreshed = MetadataStore(self._library_root).get_workbook(file_id)
            self._current_workbook = refreshed
            self._file_library.replace_workbook(refreshed)
            self._refresh_version_canvas()
        self._refresh_recycle_bin()

    def _request_reset_layouts(self) -> None:
        store = MetadataStore(self._library_root)
        workbooks = store.list_workbooks()
        if not workbooks:
            return
        names = "、".join(workbook.display_name for workbook in workbooks[:5])
        if len(workbooks) > 5:
            names += f" 等 {len(workbooks)} 个文件"
        if not self._confirmation_presenter(
            self,
            "重整布局",
            f"将清除以下文件的手动节点位置并恢复默认排布：\n{names}\n确定继续吗？",
        ):
            return
        for workbook in workbooks:
            store.clear_version_layouts(workbook.file_id)
        self._version_tree.clear_remembered_layouts()
        self._refresh_version_canvas()

    def _request_purge_version(self, file_id: str, version_id: str) -> None:
        if self._purge_version_task_id is not None:
            return
        try:
            version = MetadataStore(self._library_root).plan_version_purge(file_id, version_id)
        except ValueError as error:
            self._error_presenter(self, str(error))
            return
        if not self._confirmation_presenter(
            self,
            "永久删除版本",
            f"确定永久删除“{version.name}”吗？\n该版本的快照文件将被清除，且无法恢复。",
        ):
            return
        task_id = uuid4().hex
        self._purge_version_task_id = task_id
        self._task_queue.submit(
            TaskRequest(
                task_id=task_id,
                name=f"永久删除版本 {version.name}",
                file_id=file_id,
                engine=None,
                operation=PURGE_VERSION_OPERATION,
                payload={
                    "library_root": str(self._library_root),
                    "version_id": version_id,
                },
            )
        )

    def _apply_purge_version_event(self, event: TaskEvent) -> None:
        if event.task_id != self._purge_version_task_id:
            return
        self._purge_version_task_id = None
        if event.state is TaskState.SUCCEEDED:
            self._refresh_version_canvas()
        elif event.state in {TaskState.FAILED, TaskState.CANCELLED}:
            self._error_presenter(
                self,
                event.message
                or (
                    "版本永久删除已取消"
                    if event.state is TaskState.CANCELLED
                    else "版本永久删除失败"
                ),
            )

    def _request_purge_file(self, file_id: str) -> None:
        if self._purge_task_id is not None or self._recycle_dialog is None:
            return
        try:
            record = MetadataStore(self._library_root).get_deleted_file(file_id)
        except ValueError as error:
            self._error_presenter(self, str(error))
            self._refresh_recycle_bin()
            return
        if not self._confirmation_presenter(
            self,
            "永久删除文件",
            f"确定永久删除“{record.display_name}”及其全部版本吗？\n"
            "磁盘文件和版本历史将被清除，且无法恢复。",
        ):
            return
        task_id = uuid4().hex
        self._purge_task_id = task_id
        self._recycle_dialog.mark_busy(True)
        self._task_queue.submit(
            TaskRequest(
                task_id=task_id,
                name=f"永久删除 {record.display_name}",
                file_id=file_id,
                engine=None,
                operation=PURGE_FILE_OPERATION,
                payload={"library_root": str(self._library_root)},
            )
        )

    def _apply_purge_event(self, event: TaskEvent) -> None:
        if event.task_id != self._purge_task_id:
            return
        self._purge_task_id = None
        if self._recycle_dialog is not None:
            self._recycle_dialog.mark_busy(False)
        if event.state is TaskState.SUCCEEDED:
            self._refresh_recycle_bin()
        elif event.state in {TaskState.FAILED, TaskState.CANCELLED}:
            self._error_presenter(
                self,
                event.message
                or ("永久删除已取消" if event.state is TaskState.CANCELLED else "永久删除失败"),
            )

    def _refresh_storage_stats(self) -> None:
        workbook = self._current_workbook
        if workbook is None:
            self._storage_status.set_empty()
            return
        task_id = uuid4().hex
        self._storage_stats_task_id = task_id
        self._task_queue.submit(
            TaskRequest(
                task_id=task_id,
                name="统计版本占用",
                file_id=workbook.file_id,
                engine=None,
                operation=VERSION_STORAGE_STATS_OPERATION,
                payload={
                    "library_root": str(self._library_root),
                    "preview_version_id": self._previewed_version_id,
                },
            )
        )

    def _apply_storage_stats_event(self, event: TaskEvent) -> None:
        if event.task_id != self._storage_stats_task_id:
            return
        if event.state is TaskState.SUCCEEDED and isinstance(event.result, VersionStorageStats):
            workbook = self._current_workbook
            file_format = (
                workbook.original_path.suffix.removeprefix(".").upper()
                if workbook is not None
                else ""
            )
            self._storage_status.set_stats(
                file_format,
                event.result.total_bytes,
                event.result.preview_bytes,
            )
            self._storage_stats_task_id = None
        elif event.state in {TaskState.FAILED, TaskState.CANCELLED}:
            self._storage_stats_task_id = None

    def _save_version_position(
        self,
        file_id: str,
        version_id: str,
        x: float,
        y: float,
    ) -> None:
        try:
            MetadataStore(self._library_root).save_version_layout(
                file_id,
                version_id,
                x,
                y,
                fixed=True,
            )
        except ValueError as error:
            self._error_presenter(self, str(error))

    def _set_processing_navigation_enabled(self, enabled: bool) -> None:
        self._command_bar.setEnabled(enabled)
        self._file_library.setEnabled(enabled)
        self._version_tree.setEnabled(enabled)
        self._editor.set_actions_enabled(enabled)

    def submit_conversion(
        self,
        source: Path,
        destination: Path,
        *,
        file_id: str | None = None,
    ) -> str:
        task_id = uuid4().hex
        self._task_queue.submit(
            TaskRequest(
                task_id=task_id,
                name="转换旧版工作簿",
                file_id=file_id or source.name,
                engine=None,
                operation=CONVERT_XLS_OPERATION,
                payload={
                    "source_path": str(source),
                    "destination_path": str(destination),
                },
            )
        )
        return task_id

    def closeEvent(self, event: QCloseEvent) -> None:
        if self._workbook_preview.pending_edits():
            choice = self._unsaved_changes_presenter(
                self, "退出软件", self._can_update_version_in_place()
            )
            if choice == "cancel":
                event.ignore()
                return
            if choice == "save":
                if self._processing_result is not None:
                    self._close_after_apply = True
                    self._submit_apply_processing_preview()
                else:
                    self._close_after_manual_save = True
                    self._submit_manual_save()
                event.ignore()
                return
            if choice == "save_in_place":
                self._close_after_in_place = True
                self._submit_update_in_place()
                event.ignore()
                return
            self._workbook_preview.clear_edits()
        if self._task_bridge.shutdown(timeout=1.0):
            self._discard_processing_result()
            self._workbook_preview.close()
            event.accept()
        else:
            event.ignore()

    def _can_update_version_in_place(self) -> bool:
        """就地修改仅限：正在预览当前 HEAD、无临时结果、且该节点没有子节点。"""
        if self._processing_result is not None or self._in_place_task_id is not None:
            return False
        workbook = self._current_workbook
        head = workbook.head_version if workbook is not None else None
        if workbook is None or head is None or self._previewed_version_id != head.version_id:
            return False
        return not MetadataStore(self._library_root).version_has_children(
            workbook.file_id, head.version_id
        )

    def _submit_update_in_place(self) -> None:
        """把未保存编辑就地写入当前叶节点（DEC-20260816-039），不生成新节点。"""
        workbook = self._current_workbook
        head = workbook.head_version if workbook is not None else None
        edits = self._workbook_preview.pending_edits()
        if workbook is None or head is None or not edits or self._in_place_task_id is not None:
            return
        task_id = uuid4().hex
        self._in_place_task_id = task_id
        self._set_processing_navigation_enabled(False)
        self._editor.set_busy("正在就地更新版本内容…")
        self._task_queue.submit(
            TaskRequest(
                task_id=task_id,
                name="就地更新版本内容",
                file_id=workbook.file_id,
                engine=None,
                operation=UPDATE_VERSION_IN_PLACE_OPERATION,
                payload={
                    "library_root": str(self._library_root),
                    "version_id": head.version_id,
                    "expected_hash": head.content_hash,
                    "edits": [
                        {
                            "sheet_name": edit.sheet_name,
                            "row": edit.row,
                            "column": edit.column,
                            "value": edit.value,
                        }
                        for edit in edits
                    ],
                },
            )
        )

    def _apply_in_place_event(self, event: TaskEvent) -> None:
        if event.task_id != self._in_place_task_id:
            return
        if event.state is TaskState.SUCCEEDED and isinstance(event.result, ImportedWorkbook):
            self._in_place_task_id = None
            self._workbook_preview.clear_edits()
            self._set_processing_navigation_enabled(True)
            self._current_workbook = event.result
            head = event.result.head_version
            self._previewed_version_id = head.version_id if head is not None else None
            self._file_library.replace_workbook(event.result)
            self._refresh_version_canvas()
            self._load_preview(
                event.result.working_path,
                event.result.display_name,
                temporary=False,
            )
            if self._close_after_in_place:
                self._close_after_in_place = False
                self.close()
        elif event.state in {TaskState.FAILED, TaskState.CANCELLED}:
            self._in_place_task_id = None
            self._close_after_in_place = False
            self._set_processing_navigation_enabled(True)
            self._editor.set_error(event.message or "就地更新未完成，编辑仍保留在表格中")

    def _resolve_unsaved_changes(self, action: str) -> bool:
        if not self._workbook_preview.pending_edits():
            return True
        allow_in_place = self._can_update_version_in_place()
        choice = self._unsaved_changes_presenter(self, action, allow_in_place)
        if choice == "discard":
            self._workbook_preview.clear_edits()
            return True
        if choice == "save_in_place":
            # 就地修改当前叶节点：不生成新节点，节点 id 与父子关系不变。
            self._submit_update_in_place()
        elif choice == "save":
            # 临时结果上的编辑只能随“应用生成版本”提交。
            if self._processing_result is not None:
                self._submit_apply_processing_preview()
            else:
                self._submit_manual_save()
        return False


def create_main_window(
    task_queue: ApplicationTaskQueue | None = None,
    *,
    library_root: Path | None = None,
    file_picker: FilePicker | None = None,
    error_presenter: ErrorPresenter | None = None,
    confirmation_presenter: ConfirmationPresenter | None = None,
    version_choice_presenter: VersionChoicePresenter | None = None,
    unsaved_changes_presenter: UnsavedChangesPresenter | None = None,
    save_as_picker: SaveAsPicker | None = None,
    export_presenter: ExportPresenter | None = None,
) -> HyacinthMainWindow:
    handlers = conversion_task_handlers()
    handlers.update(import_task_handlers())
    handlers.update(preview_task_handlers())
    handlers.update(sort_preview_handlers())
    handlers.update(deduplicate_preview_handlers())
    handlers.update(delete_blank_rows_preview_handlers())
    handlers.update(filter_preview_handlers())
    handlers.update(trim_preview_handlers())
    handlers.update(find_replace_preview_handlers())
    handlers.update(apply_version_handlers())
    handlers.update(checkout_version_handlers())
    handlers.update(delete_version_handlers())
    handlers.update(export_version_handlers())
    handlers.update(version_storage_stats_handlers())
    handlers.update(purge_file_handlers())
    handlers.update(purge_version_handlers())
    return HyacinthMainWindow(
        task_queue or TaskQueue(handlers),
        library_root or default_library_root(),
        file_picker or select_workbook_file,
        error_presenter or show_import_error,
        confirmation_presenter or confirm_action,
        version_choice_presenter or choose_replacement_version,
        unsaved_changes_presenter or ask_unsaved_changes,
        save_as_picker or select_export_path,
        export_presenter or show_export_success,
    )


def default_library_root() -> Path:
    documents = QStandardPaths.writableLocation(QStandardPaths.StandardLocation.DocumentsLocation)
    return Path(documents) / "Hyacinth"


def default_export_directory() -> Path:
    downloads = QStandardPaths.writableLocation(QStandardPaths.StandardLocation.DownloadLocation)
    return Path(downloads) if downloads else Path.home() / "Downloads"


def select_workbook_file(parent: QWidget) -> Path | None:
    file_name, _selected_filter = QFileDialog.getOpenFileName(
        parent,
        "导入 Excel 文件",
        "",
        "Excel 工作簿 (*.xlsx *.xls)",
    )
    return Path(file_name) if file_name else None


def select_export_path(parent: QWidget, filename: str) -> Path | None:
    initial = default_export_directory() / filename
    file_name, _selected_filter = QFileDialog.getSaveFileName(
        parent,
        "另存版本为",
        str(initial),
        "Excel 工作簿 (*.xlsx *.xls)",
    )
    return Path(file_name) if file_name else None


def show_import_error(parent: QWidget, message: str) -> None:
    QMessageBox.warning(parent, "导入失败", message)


def show_export_success(parent: QWidget, path: Path) -> None:
    dialog = QMessageBox(parent)
    dialog.setWindowTitle("导出完成")
    dialog.setText(f"版本已导出为 {path.name}")
    dialog.setInformativeText(str(path))
    dialog.setIcon(QMessageBox.Icon.Information)
    open_file = dialog.addButton("打开文件", QMessageBox.ButtonRole.AcceptRole)
    open_folder = dialog.addButton("打开所在位置", QMessageBox.ButtonRole.ActionRole)
    dialog.addButton("关闭", QMessageBox.ButtonRole.RejectRole)
    dialog.exec()
    if dialog.clickedButton() is open_file:
        QDesktopServices.openUrl(QUrl.fromLocalFile(str(path)))
    elif dialog.clickedButton() is open_folder:
        QDesktopServices.openUrl(QUrl.fromLocalFile(str(path.parent)))


def confirm_action(parent: QWidget, title: str, message: str) -> bool:
    return (
        QMessageBox.question(
            parent,
            title,
            message,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel,
            QMessageBox.StandardButton.Cancel,
        )
        == QMessageBox.StandardButton.Yes
    )


def choose_replacement_version(
    parent: QWidget,
    candidates: tuple[VersionRecord, ...],
) -> str | None:
    labels = [
        f"{candidate.name} · {candidate.created_at.astimezone().strftime('%Y-%m-%d %H:%M')}"
        f" · {candidate.version_id[:8]}"
        for candidate in candidates
    ]
    selected, accepted = QInputDialog.getItem(
        parent,
        "选择新的当前工作版本",
        "删除当前 HEAD 后切换到：",
        labels,
        0,
        False,
    )
    if not accepted:
        return None
    return candidates[labels.index(selected)].version_id


def ask_unsaved_changes(parent: QWidget, action: str, allow_in_place: bool = False) -> str:
    dialog = QMessageBox(parent)
    dialog.setWindowTitle("有未保存的修改")
    if allow_in_place:
        dialog.setText(f"{action}前，如何处理当前未保存的单元格修改？")
        dialog.setInformativeText(
            "“就地更新此节点”会直接修改当前版本内容，不生成新节点；"
            "“保存为新版本”会在版本树中新增一个子节点。选择“放弃”会丢弃未保存的编辑。"
        )
    else:
        dialog.setText(f"{action}前，是否将当前单元格修改保存为新版本？")
        dialog.setInformativeText("选择“放弃”会丢弃本次尚未保存的编辑。")
    dialog.setIcon(QMessageBox.Icon.Warning)
    if allow_in_place:
        in_place_button = dialog.addButton("就地更新此节点", QMessageBox.ButtonRole.AcceptRole)
        save_button = dialog.addButton("保存为新版本", QMessageBox.ButtonRole.AcceptRole)
        dialog.setDefaultButton(in_place_button)
    else:
        save_button = dialog.addButton("保存", QMessageBox.ButtonRole.AcceptRole)
        in_place_button = None
        dialog.setDefaultButton(save_button)
    discard_button = dialog.addButton("放弃", QMessageBox.ButtonRole.DestructiveRole)
    dialog.addButton("取消", QMessageBox.ButtonRole.RejectRole)
    dialog.exec()
    clicked = dialog.clickedButton()
    if clicked is in_place_button and in_place_button is not None:
        return "save_in_place"
    if clicked is save_button:
        return "save"
    if clicked is discard_button:
        return "discard"
    return "cancel"
