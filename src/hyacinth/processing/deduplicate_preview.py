"""Python 安全删除重复行临时预览任务。"""

import os
from collections import defaultdict
from contextlib import AbstractContextManager
from dataclasses import dataclass
from datetime import date, datetime, time
from enum import StrEnum
from pathlib import Path
from typing import Protocol

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from hyacinth.excel.contracts import EngineName
from hyacinth.processing.sort_preview import (
    _CellData,
    _copy_file,
    _file_hash,
    _payload_path,
    _payload_string,
    _read_used_rows,
    _validate_preview_workbook,
    _write_data_rows,
)
from hyacinth.tasks import TaskRequest
from hyacinth.tasks.worker import TaskContext, TaskHandler

DEDUPLICATE_PREVIEW_OPERATION = "deduplicate-preview"


class KeepDuplicate(StrEnum):
    FIRST = "first"
    LAST = "last"


@dataclass(frozen=True, slots=True)
class DuplicateGroup:
    kept_row: int
    deleted_rows: tuple[int, ...]


@dataclass(frozen=True, slots=True)
class DeduplicatePreviewResult:
    preview_path: Path
    source_path: Path
    parent_version_id: str
    sheet_name: str
    key_columns: tuple[int, ...]
    keep: KeepDuplicate
    ignore_case: bool
    trim_whitespace: bool
    content_hash: str
    duplicate_groups: tuple[DuplicateGroup, ...]
    deleted_rows: int
    data_rows: int
    engine: EngineName = EngineName.PYTHON


class DeduplicatePreviewTaskContext(Protocol):
    def report_progress(self, progress: float | None, message: str = "") -> None: ...

    def check_cancelled(self) -> None: ...

    def set_engine(self, engine: EngineName) -> None: ...

    def commit(self) -> None: ...

    def critical_section(self, message: str = "") -> AbstractContextManager[None]: ...


def run_deduplicate_preview_task(
    request: TaskRequest,
    context: DeduplicatePreviewTaskContext,
) -> DeduplicatePreviewResult:
    source_path = _payload_path(request, "source_path")
    preview_path = _payload_path(request, "preview_path")
    parent_version_id = _payload_string(request, "parent_version_id")
    sheet_name = _payload_string(request, "sheet_name")
    key_columns = _parse_key_columns(request)
    keep = _parse_keep(request)
    ignore_case = _payload_bool(request, "ignore_case")
    trim_whitespace = _payload_bool(request, "trim_whitespace")
    if preview_path == source_path:
        raise ValueError("预览输出路径不能与源文件相同")

    context.check_cancelled()
    context.set_engine(EngineName.PYTHON)
    preview_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = preview_path.with_name(f".{preview_path.stem}.tmp.xlsx")
    temporary_path.unlink(missing_ok=True)
    try:
        context.report_progress(None, "正在复制源工作簿")
        _copy_file(source_path, temporary_path, context)
        context.check_cancelled()
        context.report_progress(0.3, f"正在检查工作表 {sheet_name} 的重复行")
        groups, deleted_rows, data_rows = _deduplicate_copy(
            temporary_path,
            sheet_name,
            key_columns,
            keep,
            ignore_case,
            trim_whitespace,
            context,
        )
        if deleted_rows == 0:
            raise ValueError("未发现重复行，无需生成预览")
        context.check_cancelled()
        context.report_progress(0.8, "正在校验临时结果")
        _validate_preview_workbook(temporary_path)
        context.check_cancelled()
        content_hash = _file_hash(temporary_path, context)
        with context.critical_section("正在安全完成删除重复行预览"):
            context.commit()
            os.replace(temporary_path, preview_path)
    finally:
        temporary_path.unlink(missing_ok=True)

    context.report_progress(1.0, f"预览已就绪，预计删除 {deleted_rows} 行")
    return DeduplicatePreviewResult(
        preview_path=preview_path,
        source_path=source_path,
        parent_version_id=parent_version_id,
        sheet_name=sheet_name,
        key_columns=key_columns,
        keep=keep,
        ignore_case=ignore_case,
        trim_whitespace=trim_whitespace,
        content_hash=content_hash,
        duplicate_groups=groups,
        deleted_rows=deleted_rows,
        data_rows=data_rows,
    )


def deduplicate_preview_task(request: TaskRequest, context: TaskContext) -> object:
    return run_deduplicate_preview_task(request, context)


def deduplicate_preview_handlers() -> dict[str, TaskHandler]:
    return {DEDUPLICATE_PREVIEW_OPERATION: deduplicate_preview_task}


def _deduplicate_copy(
    path: Path,
    sheet_name: str,
    key_columns: tuple[int, ...],
    keep: KeepDuplicate,
    ignore_case: bool,
    trim_whitespace: bool,
    context: DeduplicatePreviewTaskContext,
) -> tuple[tuple[DuplicateGroup, ...], int, int]:
    workbook = load_workbook(path, data_only=False)
    try:
        try:
            worksheet = workbook[sheet_name]
        except KeyError as error:
            raise ValueError(f"找不到工作表：{sheet_name}") from error
        used_rows = worksheet.max_row or 0
        used_columns = worksheet.max_column or 0
        if used_rows < 2 or used_columns < 1:
            return (), 0, 0
        effective_columns = key_columns or tuple(range(used_columns))
        _reject_key_columns_out_of_range(effective_columns, used_columns)
        _reject_unsafe_structure(workbook, worksheet, used_rows, used_columns)
        rows = _read_used_rows(worksheet, used_rows, used_columns, context)
        _reject_formulas(rows)
        data_rows = rows[1:]
        groups, deleted_indexes = _find_duplicate_groups(
            data_rows,
            effective_columns,
            keep,
            ignore_case,
            trim_whitespace,
            context,
        )
        if not deleted_indexes:
            return (), 0, len(data_rows)
        retained_rows = [row for index, row in enumerate(data_rows) if index not in deleted_indexes]
        _write_data_rows(worksheet, retained_rows)
        worksheet.delete_rows(2 + len(retained_rows), len(deleted_indexes))
        workbook.save(path)
        return groups, len(deleted_indexes), len(data_rows)
    finally:
        workbook.close()


def _find_duplicate_groups(
    rows: list[list[_CellData]],
    key_columns: tuple[int, ...],
    keep: KeepDuplicate,
    ignore_case: bool,
    trim_whitespace: bool,
    context: DeduplicatePreviewTaskContext,
) -> tuple[tuple[DuplicateGroup, ...], set[int]]:
    indexes_by_key: dict[tuple[tuple[str, object], ...], list[int]] = defaultdict(list)
    for index, row in enumerate(rows):
        key = tuple(
            _normalized_value(row[column].value, ignore_case, trim_whitespace)
            for column in key_columns
        )
        indexes_by_key[key].append(index)
        if index and index % 256 == 0:
            context.check_cancelled()

    groups: list[DuplicateGroup] = []
    deleted_indexes: set[int] = set()
    for indexes in indexes_by_key.values():
        if len(indexes) < 2:
            continue
        kept_index = indexes[0] if keep is KeepDuplicate.FIRST else indexes[-1]
        removed = tuple(index for index in indexes if index != kept_index)
        deleted_indexes.update(removed)
        groups.append(
            DuplicateGroup(
                kept_row=kept_index + 2,
                deleted_rows=tuple(index + 2 for index in removed),
            )
        )
    return tuple(groups), deleted_indexes


def _normalized_value(
    value: object,
    ignore_case: bool,
    trim_whitespace: bool,
) -> tuple[str, object]:
    if value is None or value == "":
        return ("empty", "")
    if isinstance(value, str):
        normalized = value.strip() if trim_whitespace else value
        normalized = normalized.casefold() if ignore_case else normalized
        return ("text", normalized)
    if isinstance(value, bool):
        return ("bool", value)
    if isinstance(value, datetime):
        return ("datetime", value)
    if isinstance(value, date):
        return ("date", value)
    if isinstance(value, time):
        return ("time", value)
    if isinstance(value, (int, float)):
        return ("number", value)
    raise ValueError(f"不支持的去重键单元格类型：{type(value).__name__}")


def _reject_key_columns_out_of_range(
    key_columns: tuple[int, ...],
    used_columns: int,
) -> None:
    for column in key_columns:
        if column >= used_columns:
            raise ValueError(f"去重关键列 {column + 1} 超出工作表范围")


def _reject_unsafe_structure(
    workbook: object,
    worksheet: Worksheet,
    used_rows: int,
    used_columns: int,
) -> None:
    for merged in worksheet.merged_cells.ranges:
        if (
            merged.max_row >= 1
            and merged.min_row <= used_rows
            and merged.max_col >= 1
            and merged.min_col <= used_columns
        ):
            raise ValueError(f"去重区域包含合并单元格 {merged}，无法安全删除整行")
    if worksheet.tables:
        raise ValueError("工作表包含 Excel 表格，Python 安全模式暂不删除重复行")
    if getattr(worksheet, "_charts", ()):
        raise ValueError("工作表包含图表，Python 安全模式暂不删除重复行")
    defined_names = getattr(workbook, "defined_names", ())
    if defined_names:
        raise ValueError("工作簿包含命名区域，Python 安全模式暂不删除重复行")


def _reject_formulas(rows: list[list[_CellData]]) -> None:
    for row_index, row in enumerate(rows):
        for column_index, data in enumerate(row):
            if isinstance(data.value, str) and data.value.startswith("="):
                raise ValueError(
                    f"去重区域第 {row_index + 1} 行第 {column_index + 1} 列包含公式，"
                    "无法安全删除整行"
                )


def _parse_key_columns(request: TaskRequest) -> tuple[int, ...]:
    raw = request.payload.get("key_columns", [])
    if not isinstance(raw, list):
        raise ValueError("去重关键列 key_columns 必须是数组")
    columns: list[int] = []
    for value in raw:
        if not isinstance(value, int) or isinstance(value, bool) or value < 0:
            raise ValueError("去重关键列必须是大于等于 0 的整数")
        columns.append(value)
    if len(set(columns)) != len(columns):
        raise ValueError("去重关键列不能重复")
    return tuple(columns)


def _parse_keep(request: TaskRequest) -> KeepDuplicate:
    raw = request.payload.get("keep", KeepDuplicate.FIRST.value)
    if not isinstance(raw, str):
        raise ValueError("保留规则 keep 必须为 first 或 last")
    try:
        return KeepDuplicate(raw)
    except ValueError as error:
        raise ValueError("保留规则 keep 必须为 first 或 last") from error


def _payload_bool(request: TaskRequest, key: str) -> bool:
    value = request.payload.get(key, False)
    if not isinstance(value, bool):
        raise ValueError(f"任务参数 {key} 必须是布尔值")
    return value
