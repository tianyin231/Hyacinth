"""Python 安全多列排序临时预览任务。

复制源 xlsx 后对单工作表 used range 排序并保存临时预览，校验后原子发布。
"""

import os
import shutil
from contextlib import AbstractContextManager
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, time
from enum import StrEnum
from hashlib import sha256
from pathlib import Path
from typing import Any, Protocol, cast

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.comments.comments import Comment
from openpyxl.styles.cell_style import StyleArray
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet

from hyacinth.excel.contracts import EngineName
from hyacinth.tasks import TaskRequest
from hyacinth.tasks.worker import TaskContext, TaskHandler

SORT_PREVIEW_OPERATION = "sort-preview"
COPY_CHUNK_SIZE = 1024 * 1024
MAX_SORT_KEYS = 2
PROGRESS_ROW_INTERVAL = 256


class SortDirection(StrEnum):
    ASCENDING = "asc"
    DESCENDING = "desc"


@dataclass(frozen=True, slots=True)
class SortKey:
    column_index: int
    direction: SortDirection


@dataclass(frozen=True, slots=True)
class SortPreviewResult:
    preview_path: Path
    source_path: Path
    parent_version_id: str
    sheet_name: str
    sort_keys: tuple[SortKey, ...]
    content_hash: str
    engine: EngineName = EngineName.PYTHON
    data_rows: int = 0


class SortPreviewTaskContext(Protocol):
    def report_progress(self, progress: float | None, message: str = "") -> None: ...

    def check_cancelled(self) -> None: ...

    def set_engine(self, engine: EngineName) -> None: ...

    def commit(self) -> None: ...

    def critical_section(self, message: str = "") -> AbstractContextManager[None]: ...


def run_sort_preview_task(
    request: TaskRequest,
    context: SortPreviewTaskContext,
) -> SortPreviewResult:
    source_path = _payload_path(request, "source_path")
    preview_path = _payload_path(request, "preview_path")
    parent_version_id = _payload_string(request, "parent_version_id")
    sheet_name = _payload_string(request, "sheet_name")
    sort_keys = _parse_sort_keys(request)
    if preview_path == source_path:
        raise ValueError("预览输出路径不能与源文件相同")
    context.check_cancelled()
    context.set_engine(EngineName.PYTHON)
    preview_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = preview_path.with_name(f".{preview_path.stem}.tmp.xlsx")
    temporary_path.unlink(missing_ok=True)
    try:
        context.report_progress(None, "正在复制源工作簿")
        _copy_file(source_path, temporary_path, context)
        context.check_cancelled()
        context.report_progress(0.3, f"正在排序工作表 {sheet_name}")
        data_rows = _sort_copy(temporary_path, sheet_name, sort_keys, context)
        context.check_cancelled()
        context.report_progress(0.8, "正在校验临时结果")
        _validate_preview_workbook(temporary_path)
        context.check_cancelled()
        content_hash = _file_hash(temporary_path, context)
        with context.critical_section("正在安全完成排序预览"):
            context.commit()
            os.replace(temporary_path, preview_path)
    finally:
        temporary_path.unlink(missing_ok=True)

    context.report_progress(1.0, "排序预览已就绪")
    return SortPreviewResult(
        preview_path=preview_path,
        source_path=source_path,
        parent_version_id=parent_version_id,
        sheet_name=sheet_name,
        sort_keys=sort_keys,
        content_hash=content_hash,
        engine=EngineName.PYTHON,
        data_rows=data_rows,
    )


def sort_preview_task(request: TaskRequest, context: TaskContext) -> object:
    return run_sort_preview_task(request, context)


def sort_preview_handlers() -> dict[str, TaskHandler]:
    return {SORT_PREVIEW_OPERATION: sort_preview_task}


@dataclass(frozen=True, slots=True)
class _CellData:
    value: object
    style: StyleArray
    comment: Comment | None
    hyperlink: Hyperlink | None


def _sort_copy(
    path: Path,
    sheet_name: str,
    sort_keys: tuple[SortKey, ...],
    context: SortPreviewTaskContext,
) -> int:
    workbook = load_workbook(path, data_only=False)
    try:
        try:
            worksheet = workbook[sheet_name]
        except KeyError as error:
            raise ValueError(f"找不到工作表：{sheet_name}") from error
        used_rows = worksheet.max_row or 0
        used_columns = worksheet.max_column or 0
        if used_rows < 2 or used_columns < 1:
            return 0
        _reject_key_columns_out_of_range(sort_keys, used_columns)
        _reject_merged_cells(worksheet, used_rows, used_columns)
        rows = _read_used_rows(worksheet, used_rows, used_columns, context)
        _reject_formulas(rows)
        _reject_mixed_key_types(rows, sort_keys)
        data_rows = rows[1:]
        sorted_rows = _sort_data_rows(data_rows, sort_keys)
        _write_data_rows(worksheet, sorted_rows)
        workbook.save(path)
        return len(data_rows)
    finally:
        workbook.close()


def _read_used_rows(
    worksheet: Worksheet,
    used_rows: int,
    used_columns: int,
    context: SortPreviewTaskContext,
) -> list[list[_CellData]]:
    rows: list[list[_CellData]] = []
    for row in range(1, used_rows + 1):
        cells = [
            _cell_data(_worksheet_cell(worksheet, row, column))
            for column in range(1, used_columns + 1)
        ]
        rows.append(cells)
        if row % PROGRESS_ROW_INTERVAL == 0:
            context.check_cancelled()
    return rows


def _worksheet_cell(worksheet: Worksheet, row: int, column: int) -> Cell:
    return cast(Cell, worksheet.cell(row, column))


def _cell_data(cell: Cell) -> _CellData:
    return _CellData(
        value=cell.value,
        style=_style_of(cell),
        comment=cell.comment,
        hyperlink=cell.hyperlink,
    )


def _style_of(cell: Cell) -> StyleArray:
    return cast(StyleArray, getattr(cell, "_style"))


def _write_data_rows(worksheet: Worksheet, rows: list[list[_CellData]]) -> None:
    for offset, row in enumerate(rows):
        for column, data in enumerate(row):
            _write_cell_data(_worksheet_cell(worksheet, offset + 2, column + 1), data)


def _write_cell_data(cell: Cell, data: _CellData) -> None:
    cell.value = _cell_value(data.value)
    setattr(cell, "_style", copy(data.style))
    setattr(cell, "_comment", data.comment)
    if data.hyperlink is not None:
        data.hyperlink.ref = cell.coordinate
        setattr(cell, "_hyperlink", data.hyperlink)
    else:
        setattr(cell, "_hyperlink", None)


def _reject_key_columns_out_of_range(
    sort_keys: tuple[SortKey, ...],
    used_columns: int,
) -> None:
    for key in sort_keys:
        if key.column_index >= used_columns:
            raise ValueError(f"排序键列 {get_column_letter(key.column_index + 1)} 超出工作表范围")


def _reject_merged_cells(worksheet: Worksheet, used_rows: int, used_columns: int) -> None:
    for merged in worksheet.merged_cells.ranges:
        if (
            merged.max_row >= 1
            and merged.min_row <= used_rows
            and merged.max_col >= 1
            and merged.min_col <= used_columns
        ):
            raise ValueError(f"排序区域包含合并单元格 {merged}，无法安全排序")


def _reject_formulas(rows: list[list[_CellData]]) -> None:
    for row_index, row in enumerate(rows):
        for column_index, data in enumerate(row):
            if isinstance(data.value, str) and data.value.startswith("="):
                reference = _cell_reference(row_index + 1, column_index + 1)
                raise ValueError(f"排序区域 {reference} 包含公式，无法安全排序")


def _reject_mixed_key_types(
    rows: list[list[_CellData]],
    sort_keys: tuple[SortKey, ...],
) -> None:
    data_rows = rows[1:]
    for key in sort_keys:
        column = key.column_index
        categories: set[str] = set()
        for row in data_rows:
            value = row[column].value
            if _is_empty(value):
                continue
            categories.add(_type_category(value))
        if len(categories) > 1:
            raise ValueError(
                f"排序键列 {get_column_letter(column + 1)} 包含混合非空类型，无法安全排序"
            )


def _sort_data_rows(
    rows: list[list[_CellData]],
    sort_keys: tuple[SortKey, ...],
) -> list[list[_CellData]]:
    result = list(rows)
    for key in reversed(sort_keys):
        column = key.column_index
        if key.direction is SortDirection.ASCENDING:
            result = sorted(
                result,
                key=lambda row: (
                    (0, _sortable(row[column].value))
                    if not _is_empty(row[column].value)
                    else (1, _sortable(None))
                ),
            )
        else:
            result = sorted(
                result,
                key=lambda row: (
                    (1, _sortable(row[column].value))
                    if not _is_empty(row[column].value)
                    else (0, _sortable(None))
                ),
                reverse=True,
            )
    return result


def _type_category(value: object) -> str:
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, str):
        return "text"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    if isinstance(value, time):
        return "time"
    raise ValueError(f"不支持的排序单元格类型：{type(value).__name__}")


def _is_empty(value: object) -> bool:
    return value is None or value == ""


def _sortable(value: object) -> Any:
    return value


def _cell_value(value: object) -> Any:
    return value


def _cell_reference(row: int, column: int) -> str:
    return f"{get_column_letter(column)}{row}"


def _validate_preview_workbook(path: Path) -> None:
    workbook = load_workbook(path, read_only=True)
    try:
        if not workbook.sheetnames:
            raise ValueError("预览工作簿缺少工作表")
    finally:
        workbook.close()


def _parse_sort_keys(request: TaskRequest) -> tuple[SortKey, ...]:
    raw = request.payload.get("sort_keys")
    if not isinstance(raw, list) or not raw:
        raise ValueError("任务参数缺少排序键：sort_keys")
    if len(raw) > MAX_SORT_KEYS:
        raise ValueError("第一版最多支持两个排序键")
    keys: list[SortKey] = []
    for item in raw:
        if not isinstance(item, dict):
            raise ValueError("排序键必须是包含 column_index 和 direction 的对象")
        column_index = item.get("column_index")
        direction_value = item.get("direction")
        if not isinstance(column_index, int) or isinstance(column_index, bool) or column_index < 0:
            raise ValueError("排序键 column_index 必须是大于等于 0 的整数")
        if not isinstance(direction_value, str):
            raise ValueError("排序键 direction 必须为 asc 或 desc")
        try:
            direction = SortDirection(direction_value)
        except ValueError as error:
            raise ValueError("排序键 direction 必须为 asc 或 desc") from error
        keys.append(SortKey(column_index=column_index, direction=direction))
    if len({key.column_index for key in keys}) != len(keys):
        raise ValueError("排序键不能重复指定同一列")
    return tuple(keys)


def _payload_path(request: TaskRequest, key: str) -> Path:
    value = request.payload.get(key)
    if not isinstance(value, str) or not value:
        raise ValueError(f"任务参数缺少路径：{key}")
    return Path(value)


def _payload_string(request: TaskRequest, key: str) -> str:
    value = request.payload.get(key)
    if not isinstance(value, str) or not value:
        raise ValueError(f"任务参数缺少：{key}")
    return value


def _copy_file(
    source: Path,
    destination: Path,
    context: SortPreviewTaskContext,
) -> None:
    with source.open("rb") as source_file, destination.open("wb") as destination_file:
        while chunk := source_file.read(COPY_CHUNK_SIZE):
            context.check_cancelled()
            destination_file.write(chunk)
    context.check_cancelled()
    shutil.copystat(source, destination)


def _file_hash(path: Path, context: SortPreviewTaskContext) -> str:
    digest = sha256()
    with path.open("rb") as source:
        while chunk := source.read(COPY_CHUNK_SIZE):
            context.check_cancelled()
            digest.update(chunk)
    context.check_cancelled()
    return digest.hexdigest()
