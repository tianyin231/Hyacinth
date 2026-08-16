"""Python 安全查找与替换临时预览任务。"""

import os
from contextlib import AbstractContextManager
from dataclasses import dataclass
from enum import StrEnum
from pathlib import Path
from typing import Protocol

from openpyxl import Workbook, load_workbook

from hyacinth.excel.contracts import EngineName
from hyacinth.processing.deduplicate_preview import _payload_bool
from hyacinth.processing.sort_preview import (
    _copy_file,
    _file_hash,
    _payload_path,
    _payload_string,
    _validate_preview_workbook,
)
from hyacinth.tasks import TaskRequest
from hyacinth.tasks.worker import TaskContext, TaskHandler

FIND_REPLACE_PREVIEW_OPERATION = "find-replace-preview"


class FindReplaceMode(StrEnum):
    VALUES = "values"
    FORMULAS = "formulas"


@dataclass(frozen=True, slots=True)
class FindReplaceChange:
    sheet_name: str
    row: int
    column: int
    before: str
    after: str


@dataclass(frozen=True, slots=True)
class FindReplacePreviewResult:
    preview_path: Path | None
    source_path: Path
    parent_version_id: str
    sheet_name: str
    sheets: tuple[str, ...]
    mode: FindReplaceMode
    find_text: str
    replace_text: str
    match_case: bool
    whole_cell: bool
    trim_whitespace: bool
    replace_all: bool
    content_hash: str
    changes: tuple[FindReplaceChange, ...]
    engine: EngineName = EngineName.PYTHON

    @property
    def replaced(self) -> int:
        return len(self.changes) if self.replace_all else 0


class FindReplaceTaskContext(Protocol):
    def report_progress(self, progress: float | None, message: str = "") -> None: ...

    def check_cancelled(self) -> None: ...

    def set_engine(self, engine: EngineName) -> None: ...

    def commit(self) -> None: ...

    def critical_section(self, message: str = "") -> AbstractContextManager[None]: ...


def run_find_replace_preview_task(
    request: TaskRequest,
    context: FindReplaceTaskContext,
) -> FindReplacePreviewResult:
    source_path = _payload_path(request, "source_path")
    preview_path = _payload_path(request, "preview_path")
    parent_version_id = _payload_string(request, "parent_version_id")
    sheet_name = _payload_string(request, "sheet_name")
    all_sheets = _payload_bool(request, "all_sheets")
    mode_value = request.payload.get("mode", FindReplaceMode.VALUES.value)
    mode = (
        FindReplaceMode.FORMULAS
        if mode_value == FindReplaceMode.FORMULAS.value
        else FindReplaceMode.VALUES
    )
    find_text_value = request.payload.get("find_text")
    if not isinstance(find_text_value, str) or not find_text_value:
        raise ValueError("查找内容不能为空")
    find_text = find_text_value
    replace_value = request.payload.get("replace_text", "")
    replace_text = replace_value if isinstance(replace_value, str) else ""
    match_case = _payload_bool(request, "match_case")
    whole_cell = _payload_bool(request, "whole_cell")
    trim_whitespace = _payload_bool(request, "trim_whitespace")
    replace_all = _payload_bool(request, "replace_all")
    if not find_text:
        raise ValueError("查找内容不能为空")
    if mode is FindReplaceMode.FORMULAS and not replace_all:
        raise ValueError("公式模式仅支持全部替换")

    context.check_cancelled()
    context.set_engine(EngineName.PYTHON)

    workbook = load_workbook(source_path, data_only=False)
    try:
        target_sheets = _target_sheets(workbook, sheet_name, all_sheets)
    finally:
        workbook.close()

    if not replace_all:
        changes = _find_only(
            source_path,
            target_sheets,
            mode,
            find_text,
            match_case,
            whole_cell,
            trim_whitespace,
            context,
        )
        if not changes:
            raise ValueError("未找到匹配内容")
        context.report_progress(1.0, f"共找到 {len(changes)} 处匹配")
        return FindReplacePreviewResult(
            preview_path=None,
            source_path=source_path,
            parent_version_id=parent_version_id,
            sheet_name=sheet_name,
            sheets=target_sheets,
            mode=mode,
            find_text=find_text,
            replace_text=replace_text,
            match_case=match_case,
            whole_cell=whole_cell,
            trim_whitespace=trim_whitespace,
            replace_all=False,
            content_hash="",
            changes=changes,
        )

    if preview_path == source_path:
        raise ValueError("预览输出路径不能与源文件相同")
    preview_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = preview_path.with_name(f".{preview_path.stem}.tmp.xlsx")
    temporary_path.unlink(missing_ok=True)
    try:
        context.report_progress(None, "正在复制源工作簿")
        _copy_file(source_path, temporary_path, context)
        context.check_cancelled()
        context.report_progress(0.3, "正在执行全部替换")
        changes = _replace_all(
            temporary_path,
            target_sheets,
            mode,
            find_text,
            replace_text,
            match_case,
            whole_cell,
            trim_whitespace,
            context,
        )
        if not changes:
            raise ValueError("未找到匹配内容")
        context.check_cancelled()
        context.report_progress(0.8, "正在校验临时结果")
        _validate_preview_workbook(temporary_path)
        context.check_cancelled()
        content_hash = _file_hash(temporary_path, context)
        with context.critical_section("正在安全完成替换预览"):
            context.commit()
            os.replace(temporary_path, preview_path)
    finally:
        temporary_path.unlink(missing_ok=True)

    context.report_progress(1.0, f"预览已就绪，共替换 {len(changes)} 处")
    return FindReplacePreviewResult(
        preview_path=preview_path,
        source_path=source_path,
        parent_version_id=parent_version_id,
        sheet_name=sheet_name,
        sheets=target_sheets,
        mode=mode,
        find_text=find_text,
        replace_text=replace_text,
        match_case=match_case,
        whole_cell=whole_cell,
        trim_whitespace=trim_whitespace,
        replace_all=True,
        content_hash=content_hash,
        changes=changes,
    )


def find_replace_preview_task(request: TaskRequest, context: TaskContext) -> object:
    return run_find_replace_preview_task(request, context)


def find_replace_preview_handlers() -> dict[str, TaskHandler]:
    return {FIND_REPLACE_PREVIEW_OPERATION: find_replace_preview_task}


def _target_sheets(workbook: Workbook, sheet_name: str, all_sheets: bool) -> tuple[str, ...]:
    if all_sheets:
        return tuple(workbook.sheetnames)
    try:
        workbook[sheet_name]
    except KeyError as error:
        raise ValueError(f"找不到工作表：{sheet_name}") from error
    return (sheet_name,)


def _find_only(
    path: Path,
    sheets: tuple[str, ...],
    mode: FindReplaceMode,
    find_text: str,
    match_case: bool,
    whole_cell: bool,
    trim_whitespace: bool,
    context: FindReplaceTaskContext,
) -> tuple[FindReplaceChange, ...]:
    return _scan(
        path,
        sheets,
        mode,
        find_text,
        "",
        match_case,
        whole_cell,
        trim_whitespace,
        False,
        context,
    )


_TRIM_CHARS = " \t\r\n\u3000\u00a0\u2007\u202f"


def _matches(
    value: object,
    mode: FindReplaceMode,
    find_key: str,
    match_case: bool,
    whole_cell: bool,
    trim_whitespace: bool,
) -> bool:
    if not isinstance(value, str) or not value:
        return False
    if mode is FindReplaceMode.FORMULAS:
        if not value.startswith("="):
            return False
    elif value.startswith("="):
        return False
    haystack = value if match_case else value.lower()
    if trim_whitespace:
        haystack = haystack.strip(_TRIM_CHARS)
    if whole_cell:
        return haystack == find_key
    return find_key in haystack


def _apply_replace(
    value: str,
    find_text: str,
    replace_text: str,
    match_case: bool,
    whole_cell: bool,
) -> str:
    if whole_cell:
        return replace_text
    if match_case:
        return value.replace(find_text, replace_text)
    # 不区分大小写：按小写定位逐段替换，其余内容保持原文。
    lowered = value.lower()
    pieces: list[str] = []
    start = 0
    while True:
        index = lowered.find(find_text, start)
        if index < 0:
            pieces.append(value[start:])
            break
        pieces.append(value[start:index])
        pieces.append(replace_text)
        start = index + len(find_text)
    return "".join(pieces)


def _scan(
    path: Path,
    sheets: tuple[str, ...],
    mode: FindReplaceMode,
    find_text: str,
    replace_text: str,
    match_case: bool,
    whole_cell: bool,
    trim_whitespace: bool,
    replace_all: bool,
    context: FindReplaceTaskContext,
) -> tuple[FindReplaceChange, ...]:
    find_key = find_text if match_case else find_text.lower()
    workbook = load_workbook(path, data_only=False)
    try:
        changes: list[FindReplaceChange] = []
        for name in sheets:
            try:
                worksheet = workbook[name]
            except KeyError as error:
                raise ValueError(f"找不到工作表：{name}") from error
            for row in worksheet.iter_rows():
                context.check_cancelled()
                for cell in row:
                    if not _matches(
                        cell.value,
                        mode,
                        find_key,
                        match_case,
                        whole_cell,
                        trim_whitespace,
                    ):
                        continue
                    before = str(cell.value)
                    after = _apply_replace(before, find_text, replace_text, match_case, whole_cell)
                    if after == before:
                        continue
                    if replace_all:
                        cell.value = after
                    changes.append(
                        FindReplaceChange(
                            sheet_name=name,
                            row=cell.row or 0,
                            column=cell.column or 0,
                            before=before,
                            after=after,
                        )
                    )
        if replace_all and changes:
            workbook.save(path)
        return tuple(changes)
    finally:
        workbook.close()


def _replace_all(
    path: Path,
    sheets: tuple[str, ...],
    mode: FindReplaceMode,
    find_text: str,
    replace_text: str,
    match_case: bool,
    whole_cell: bool,
    trim_whitespace: bool,
    context: FindReplaceTaskContext,
) -> tuple[FindReplaceChange, ...]:
    return _scan(
        path,
        sheets,
        mode,
        find_text,
        replace_text,
        match_case,
        whole_cell,
        trim_whitespace,
        True,
        context,
    )
