"""Python 安全清除文本首尾空格临时预览任务。"""

import os
from contextlib import AbstractContextManager
from dataclasses import dataclass
from pathlib import Path
from typing import Protocol

from openpyxl import load_workbook

from hyacinth.excel.contracts import EngineName
from hyacinth.processing.deduplicate_preview import _payload_bool
from hyacinth.processing.sort_preview import (
    _copy_file,
    _file_hash,
    _payload_path,
    _payload_string,
    _validate_preview_workbook,
)
from hyacinth.tasks import TaskRequest
from hyacinth.tasks.worker import TaskContext, TaskHandler

TRIM_PREVIEW_OPERATION = "trim-preview"

_TRIM_CHARACTERS = " \t\r\n\u3000\u00a0\u2007\u202f"


@dataclass(frozen=True, slots=True)
class TrimmedCell:
    row: int
    column: int
    before: str
    after: str


@dataclass(frozen=True, slots=True)
class TrimPreviewResult:
    preview_path: Path
    source_path: Path
    parent_version_id: str
    sheet_name: str
    key_columns: tuple[int, ...]
    collapse_spaces: bool
    content_hash: str
    trimmed_cells: tuple[TrimmedCell, ...]
    engine: EngineName = EngineName.PYTHON


class TrimPreviewTaskContext(Protocol):
    def report_progress(self, progress: float | None, message: str = "") -> None: ...

    def check_cancelled(self) -> None: ...

    def set_engine(self, engine: EngineName) -> None: ...

    def commit(self) -> None: ...

    def critical_section(self, message: str = "") -> AbstractContextManager[None]: ...


def run_trim_preview_task(
    request: TaskRequest,
    context: TrimPreviewTaskContext,
) -> TrimPreviewResult:
    source_path = _payload_path(request, "source_path")
    preview_path = _payload_path(request, "preview_path")
    parent_version_id = _payload_string(request, "parent_version_id")
    sheet_name = _payload_string(request, "sheet_name")
    key_columns = _parse_key_columns(request)
    collapse_spaces = _payload_bool(request, "collapse_spaces")
    if preview_path == source_path:
        raise ValueError("预览输出路径不能与源文件相同")

    context.check_cancelled()
    context.set_engine(EngineName.PYTHON)
    preview_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = preview_path.with_name(f".{preview_path.stem}.tmp.xlsx")
    temporary_path.unlink(missing_ok=True)
    try:
        context.report_progress(None, "正在复制源工作簿")
        _copy_file(source_path, temporary_path, context)
        context.check_cancelled()
        context.report_progress(0.3, f"正在清理工作表 {sheet_name} 的文本空格")
        trimmed = _trim_worksheet(
            temporary_path,
            sheet_name,
            key_columns,
            collapse_spaces,
            context,
        )
        if not trimmed:
            raise ValueError("选区中没有需要清理的文本空格")
        context.check_cancelled()
        context.report_progress(0.8, "正在校验临时结果")
        _validate_preview_workbook(temporary_path)
        context.check_cancelled()
        content_hash = _file_hash(temporary_path, context)
        with context.critical_section("正在安全完成清除空格预览"):
            context.commit()
            os.replace(temporary_path, preview_path)
    finally:
        temporary_path.unlink(missing_ok=True)

    context.report_progress(1.0, f"预览已就绪，共清理 {len(trimmed)} 个单元格")
    return TrimPreviewResult(
        preview_path=preview_path,
        source_path=source_path,
        parent_version_id=parent_version_id,
        sheet_name=sheet_name,
        key_columns=key_columns,
        collapse_spaces=collapse_spaces,
        content_hash=content_hash,
        trimmed_cells=trimmed,
    )


def trim_preview_task(request: TaskRequest, context: TaskContext) -> object:
    return run_trim_preview_task(request, context)


def trim_preview_handlers() -> dict[str, TaskHandler]:
    return {TRIM_PREVIEW_OPERATION: trim_preview_task}


def _parse_key_columns(request: TaskRequest) -> tuple[int, ...]:
    value = request.payload.get("key_columns")
    if not isinstance(value, list) or any(
        not isinstance(item, int) or isinstance(item, bool) or item < 0 for item in value
    ):
        raise ValueError("任务参数 key_columns 必须是非负整数列表")
    return tuple(value)


def _trim_worksheet(
    path: Path,
    sheet_name: str,
    key_columns: tuple[int, ...],
    collapse_spaces: bool,
    context: TrimPreviewTaskContext,
) -> tuple[TrimmedCell, ...]:
    workbook = load_workbook(path, data_only=False)
    try:
        try:
            worksheet = workbook[sheet_name]
        except KeyError as error:
            raise ValueError(f"找不到工作表：{sheet_name}") from error
        used_rows = worksheet.max_row or 0
        used_columns = worksheet.max_column or 0
        if used_rows < 1 or used_columns < 1:
            return ()
        effective_columns = key_columns or tuple(range(used_columns))
        for column in effective_columns:
            if column >= used_columns:
                raise ValueError(f"关键列超出数据区域：第 {column + 1} 列")

        trimmed: list[TrimmedCell] = []
        total_cells = used_rows * len(effective_columns)
        processed = 0
        for row in range(1, used_rows + 1):
            for column in effective_columns:
                processed += 1
                if processed % 20000 == 0:
                    context.check_cancelled()
                    context.report_progress(
                        min(0.3 + 0.5 * processed / total_cells, 0.8),
                        "正在清理文本空格",
                    )
                cell = worksheet.cell(row=row, column=column + 1)
                value = cell.value
                if not isinstance(value, str) or value.startswith("="):
                    continue
                cleaned = _clean_text(value, collapse_spaces)
                if cleaned != value:
                    cell.value = cleaned
                    trimmed.append(
                        TrimmedCell(row=row, column=column + 1, before=value, after=cleaned)
                    )
        if trimmed:
            workbook.save(path)
        return tuple(trimmed)
    finally:
        workbook.close()


def _clean_text(value: str, collapse_spaces: bool) -> str:
    cleaned = value.strip(_TRIM_CHARACTERS)
    if collapse_spaces:
        segments = [segment for segment in cleaned.split(" ") if segment]
        cleaned = " ".join(segments)
    return cleaned
