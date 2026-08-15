"""Python 安全条件筛选临时预览任务。"""

import os
from collections.abc import Callable
from contextlib import AbstractContextManager
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from enum import StrEnum
from pathlib import Path
from typing import Any, Literal, Protocol, cast

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import to_excel
from openpyxl.worksheet.filters import CustomFilter, CustomFilters, FilterColumn
from openpyxl.worksheet.worksheet import Worksheet

from hyacinth.excel.contracts import EngineName
from hyacinth.processing.blank_rows_preview import _read_data_region
from hyacinth.processing.sort_preview import (
    _CellData,
    _copy_file,
    _file_hash,
    _payload_path,
    _payload_string,
    _validate_preview_workbook,
)
from hyacinth.tasks import TaskRequest
from hyacinth.tasks.worker import TaskContext, TaskHandler

FILTER_PREVIEW_OPERATION = "filter-preview"
MAX_FILTER_CONDITIONS = 2
type CustomFilterOperator = Literal[
    "equal",
    "notEqual",
    "greaterThan",
    "lessThan",
    "greaterThanOrEqual",
    "lessThanOrEqual",
]


class FilterConnector(StrEnum):
    AND = "and"
    OR = "or"


class FilterValueType(StrEnum):
    TEXT = "text"
    NUMBER = "number"
    DATE = "date"


class FilterOperator(StrEnum):
    EQUAL = "equal"
    NOT_EQUAL = "not_equal"
    CONTAINS = "contains"
    NOT_CONTAINS = "not_contains"
    GREATER_THAN = "greater_than"
    LESS_THAN = "less_than"
    BETWEEN = "between"
    BLANK = "blank"
    NOT_BLANK = "not_blank"


@dataclass(frozen=True, slots=True)
class FilterCondition:
    column_index: int
    operator: FilterOperator
    value_type: FilterValueType
    value: str | None = None
    second_value: str | None = None


@dataclass(frozen=True, slots=True)
class FilterPreviewResult:
    preview_path: Path
    source_path: Path
    parent_version_id: str
    sheet_name: str
    conditions: tuple[FilterCondition, ...]
    connector: FilterConnector
    matched_rows: int
    total_rows: int
    hidden_row_numbers: tuple[int, ...]
    content_hash: str
    engine: EngineName = EngineName.PYTHON

    @property
    def match_ratio(self) -> float:
        return self.matched_rows / self.total_rows if self.total_rows else 0.0


class FilterPreviewTaskContext(Protocol):
    def report_progress(self, progress: float | None, message: str = "") -> None: ...

    def check_cancelled(self) -> None: ...

    def set_engine(self, engine: EngineName) -> None: ...

    def commit(self) -> None: ...

    def critical_section(self, message: str = "") -> AbstractContextManager[None]: ...


def run_filter_preview_task(
    request: TaskRequest,
    context: FilterPreviewTaskContext,
) -> FilterPreviewResult:
    source_path = _payload_path(request, "source_path")
    preview_path = _payload_path(request, "preview_path")
    parent_version_id = _payload_string(request, "parent_version_id")
    sheet_name = _payload_string(request, "sheet_name")
    conditions = _parse_conditions(request)
    connector = _parse_connector(request, conditions)
    if preview_path == source_path:
        raise ValueError("预览输出路径不能与源文件相同")

    context.check_cancelled()
    context.set_engine(EngineName.PYTHON)
    preview_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = preview_path.with_name(f".{preview_path.stem}.tmp.xlsx")
    temporary_path.unlink(missing_ok=True)
    try:
        context.report_progress(None, "正在复制源工作簿")
        _copy_file(source_path, temporary_path, context)
        context.check_cancelled()
        context.report_progress(0.3, f"正在筛选工作表 {sheet_name}")
        matched_rows, total_rows, hidden_rows = _filter_copy(
            temporary_path,
            sheet_name,
            conditions,
            connector,
            context,
        )
        context.check_cancelled()
        context.report_progress(0.8, "正在校验临时结果")
        _validate_preview_workbook(temporary_path)
        context.check_cancelled()
        content_hash = _file_hash(temporary_path, context)
        with context.critical_section("正在安全完成条件筛选预览"):
            context.commit()
            os.replace(temporary_path, preview_path)
    finally:
        temporary_path.unlink(missing_ok=True)

    context.report_progress(1.0, f"筛选预览已就绪，匹配 {matched_rows} 行")
    return FilterPreviewResult(
        preview_path=preview_path,
        source_path=source_path,
        parent_version_id=parent_version_id,
        sheet_name=sheet_name,
        conditions=conditions,
        connector=connector,
        matched_rows=matched_rows,
        total_rows=total_rows,
        hidden_row_numbers=hidden_rows,
        content_hash=content_hash,
    )


def filter_preview_task(request: TaskRequest, context: TaskContext) -> object:
    return run_filter_preview_task(request, context)


def filter_preview_handlers() -> dict[str, TaskHandler]:
    return {FILTER_PREVIEW_OPERATION: filter_preview_task}


def _filter_copy(
    path: Path,
    sheet_name: str,
    conditions: tuple[FilterCondition, ...],
    connector: FilterConnector,
    context: FilterPreviewTaskContext,
) -> tuple[int, int, tuple[int, ...]]:
    workbook = load_workbook(path, data_only=False)
    try:
        try:
            worksheet = workbook[sheet_name]
        except KeyError as error:
            raise ValueError(f"找不到工作表：{sheet_name}") from error
        rows = _read_data_region(worksheet, context)
        if len(rows) < 2:
            raise ValueError("当前工作表没有可筛选的数据行")
        used_columns = len(rows[0])
        _validate_condition_columns(conditions, used_columns)
        _reject_unsafe_filter_region(worksheet, rows, conditions)
        data_rows = rows[1:]
        matches = tuple(_row_matches(row, conditions, connector) for row in data_rows)
        hidden_rows = tuple(
            index + 2 for index, matches_row in enumerate(matches) if not matches_row
        )
        for row_number, matches_row in enumerate(matches, start=2):
            worksheet.row_dimensions[row_number].hidden = not matches_row
        _write_auto_filter(worksheet, rows, conditions, connector)
        workbook.save(path)
        return sum(matches), len(data_rows), hidden_rows
    finally:
        workbook.close()


def _reject_unsafe_filter_region(
    worksheet: Worksheet,
    rows: list[list[_CellData]],
    conditions: tuple[FilterCondition, ...],
) -> None:
    if worksheet.auto_filter.ref:
        raise ValueError("工作表已有筛选条件，请先清除后再生成新筛选")
    if any(dimension.hidden for dimension in worksheet.row_dimensions.values()):
        raise ValueError("工作表已有隐藏行，无法区分手动隐藏与筛选隐藏")
    if worksheet.tables:
        raise ValueError("筛选区域包含 Excel 表格，首个安全切片暂不支持表格筛选")
    used_rows = len(rows)
    used_columns = len(rows[0])
    for merged in worksheet.merged_cells.ranges:
        if merged.min_row <= used_rows and merged.min_col <= used_columns:
            raise ValueError(f"筛选区域包含合并单元格 {merged}，无法安全筛选")
    for condition in conditions:
        for row_number, row in enumerate(rows[1:], start=2):
            value = row[condition.column_index].value
            if isinstance(value, str) and value.startswith("="):
                reference = f"{get_column_letter(condition.column_index + 1)}{row_number}"
                raise ValueError(f"筛选条件列 {reference} 包含公式，Python 模式无法可靠计算")


def _row_matches(
    row: list[_CellData],
    conditions: tuple[FilterCondition, ...],
    connector: FilterConnector,
) -> bool:
    results = tuple(
        _condition_matches(row[condition.column_index].value, condition) for condition in conditions
    )
    return all(results) if connector is FilterConnector.AND else any(results)


def _condition_matches(value: object, condition: FilterCondition) -> bool:
    if condition.operator is FilterOperator.BLANK:
        return _is_blank(value)
    if condition.operator is FilterOperator.NOT_BLANK:
        return not _is_blank(value)
    if condition.value_type is FilterValueType.TEXT:
        return _text_matches(value, condition)
    if condition.value_type is FilterValueType.NUMBER:
        return _number_matches(value, condition)
    return _date_matches(value, condition)


def _text_matches(value: object, condition: FilterCondition) -> bool:
    if not isinstance(value, str):
        return False
    actual = value.casefold()
    expected = _required_value(condition).casefold()
    if condition.operator is FilterOperator.EQUAL:
        return bool(actual == expected)
    if condition.operator is FilterOperator.NOT_EQUAL:
        return bool(actual != expected)
    if condition.operator is FilterOperator.CONTAINS:
        return expected in actual
    if condition.operator is FilterOperator.NOT_CONTAINS:
        return expected not in actual
    raise ValueError("文本条件仅支持等于、不等于、包含和不包含")


def _number_matches(value: object, condition: FilterCondition) -> bool:
    if isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
        return False
    actual = Decimal(str(value))
    expected = _parse_decimal(_required_value(condition))
    return _ordered_matches(actual, expected, condition, _parse_decimal)


def _date_matches(value: object, condition: FilterCondition) -> bool:
    if isinstance(value, datetime):
        actual = value.date()
    elif isinstance(value, date):
        actual = value
    else:
        return False
    expected = _parse_date(_required_value(condition))
    return _ordered_matches(actual, expected, condition, _parse_date)


def _ordered_matches(
    actual: Any,
    expected: Any,
    condition: FilterCondition,
    parser: Callable[[str], Any],
) -> bool:
    if condition.operator is FilterOperator.EQUAL:
        return bool(actual == expected)
    if condition.operator is FilterOperator.NOT_EQUAL:
        return bool(actual != expected)
    if condition.operator is FilterOperator.GREATER_THAN:
        return bool(actual > expected)
    if condition.operator is FilterOperator.LESS_THAN:
        return bool(actual < expected)
    if condition.operator is FilterOperator.BETWEEN:
        upper = parser(_required_second_value(condition))
        return bool(expected <= actual <= upper)
    raise ValueError("数字和日期条件仅支持等于、不等于、大于、小于和介于")


def _write_auto_filter(
    worksheet: Worksheet,
    rows: list[list[_CellData]],
    conditions: tuple[FilterCondition, ...],
    connector: FilterConnector,
) -> None:
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(rows[0]))}{len(rows)}"
    grouped: dict[int, list[FilterCondition]] = {}
    for condition in conditions:
        grouped.setdefault(condition.column_index, []).append(condition)
    for column, column_conditions in grouped.items():
        filters = [
            custom_filter
            for condition in column_conditions
            for custom_filter in _custom_filters(condition)
        ]
        use_and = any(
            condition.operator is FilterOperator.BETWEEN for condition in column_conditions
        )
        if len(column_conditions) > 1:
            use_and = connector is FilterConnector.AND
        worksheet.auto_filter.filterColumn.append(
            FilterColumn(
                colId=column,
                customFilters=CustomFilters(
                    _and=use_and if len(filters) > 1 else None,
                    customFilter=filters,
                ),
            )
        )


def _custom_filters(condition: FilterCondition) -> list[CustomFilter]:
    value = _native_filter_value(condition, _required_value(condition))
    operator = cast(
        CustomFilterOperator,
        {
            FilterOperator.EQUAL: "equal",
            FilterOperator.NOT_EQUAL: "notEqual",
            FilterOperator.CONTAINS: "equal",
            FilterOperator.NOT_CONTAINS: "notEqual",
            FilterOperator.GREATER_THAN: "greaterThan",
            FilterOperator.LESS_THAN: "lessThan",
            FilterOperator.BLANK: "equal",
            FilterOperator.NOT_BLANK: "notEqual",
            FilterOperator.BETWEEN: "greaterThanOrEqual",
        }[condition.operator],
    )
    if condition.operator in {FilterOperator.CONTAINS, FilterOperator.NOT_CONTAINS}:
        value = f"*{value}*"
    first = CustomFilter(operator=operator, val=value)
    if condition.operator is not FilterOperator.BETWEEN:
        return [first]
    upper = _native_filter_value(condition, _required_second_value(condition))
    return [first, CustomFilter(operator="lessThanOrEqual", val=upper)]


def _native_filter_value(condition: FilterCondition, value: str) -> str:
    if condition.operator in {FilterOperator.BLANK, FilterOperator.NOT_BLANK}:
        return ""
    if condition.value_type is FilterValueType.TEXT:
        return _escape_excel_wildcards(value)
    if condition.value_type is FilterValueType.NUMBER:
        return str(_parse_decimal(value))
    return str(to_excel(_parse_date(value)))  # type: ignore[no-untyped-call]


def _parse_conditions(request: TaskRequest) -> tuple[FilterCondition, ...]:
    raw = request.payload.get("conditions")
    if not isinstance(raw, list) or not raw:
        raise ValueError("任务参数至少需要一个筛选条件")
    if len(raw) > MAX_FILTER_CONDITIONS:
        raise ValueError("首个条件筛选切片最多支持两个条件")
    conditions = tuple(_parse_condition(item) for item in raw)
    return conditions


def _parse_condition(raw: object) -> FilterCondition:
    if not isinstance(raw, dict):
        raise ValueError("筛选条件必须是对象")
    column = raw.get("column_index")
    if not isinstance(column, int) or isinstance(column, bool) or column < 0:
        raise ValueError("筛选条件 column_index 必须是大于等于 0 的整数")
    try:
        operator = FilterOperator(str(raw.get("operator")))
        value_type = FilterValueType(str(raw.get("value_type")))
    except ValueError as error:
        raise ValueError("筛选条件的操作符或数据类型无效") from error
    value = raw.get("value")
    second_value = raw.get("second_value")
    if value is not None and not isinstance(value, str):
        raise ValueError("筛选条件 value 必须是文本")
    if second_value is not None and not isinstance(second_value, str):
        raise ValueError("筛选条件 second_value 必须是文本")
    condition = FilterCondition(column, operator, value_type, value, second_value)
    _validate_condition(condition)
    return condition


def _validate_condition(condition: FilterCondition) -> None:
    if condition.operator in {FilterOperator.BLANK, FilterOperator.NOT_BLANK}:
        return
    _required_value(condition)
    if condition.operator in {FilterOperator.CONTAINS, FilterOperator.NOT_CONTAINS}:
        if condition.value_type is not FilterValueType.TEXT:
            raise ValueError("包含和不包含仅支持文本条件")
    elif condition.value_type is FilterValueType.TEXT and condition.operator not in {
        FilterOperator.EQUAL,
        FilterOperator.NOT_EQUAL,
    }:
        raise ValueError("文本条件仅支持等于、不等于、包含和不包含")
    if condition.operator is FilterOperator.BETWEEN:
        _required_second_value(condition)
    if condition.value_type is FilterValueType.NUMBER:
        _parse_decimal(_required_value(condition))
        if condition.operator is FilterOperator.BETWEEN:
            _parse_decimal(_required_second_value(condition))
    if condition.value_type is FilterValueType.DATE:
        _parse_date(_required_value(condition))
        if condition.operator is FilterOperator.BETWEEN:
            _parse_date(_required_second_value(condition))


def _parse_connector(
    request: TaskRequest,
    conditions: tuple[FilterCondition, ...],
) -> FilterConnector:
    raw = request.payload.get("connector", "and")
    try:
        connector = FilterConnector(str(raw))
    except ValueError as error:
        raise ValueError("筛选条件连接方式必须为 and 或 or") from error
    if connector is FilterConnector.OR and len({item.column_index for item in conditions}) > 1:
        raise ValueError("原生 Excel 筛选仅允许同一列条件使用“或者”")
    if (
        len(conditions) > 1
        and len({item.column_index for item in conditions}) == 1
        and any(item.operator is FilterOperator.BETWEEN for item in conditions)
    ):
        raise ValueError("同一列的“介于”条件不能再组合第二个条件")
    return connector


def _validate_condition_columns(
    conditions: tuple[FilterCondition, ...],
    used_columns: int,
) -> None:
    for condition in conditions:
        if condition.column_index >= used_columns:
            raise ValueError(f"筛选列 {condition.column_index + 1} 超出工作表范围")


def _required_value(condition: FilterCondition) -> str:
    if condition.operator in {FilterOperator.BLANK, FilterOperator.NOT_BLANK}:
        return ""
    if condition.value is None or not condition.value:
        raise ValueError("当前筛选条件需要填写比较值")
    return condition.value


def _required_second_value(condition: FilterCondition) -> str:
    if condition.second_value is None or not condition.second_value:
        raise ValueError("“介于”条件需要填写第二个比较值")
    return condition.second_value


def _parse_decimal(value: str) -> Decimal:
    try:
        return Decimal(value)
    except InvalidOperation as error:
        raise ValueError(f"无效的数字筛选值：{value}") from error


def _parse_date(value: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as error:
        raise ValueError(f"无效的日期筛选值：{value}，请使用 YYYY-MM-DD") from error


def _is_blank(value: object) -> bool:
    return value is None or value == ""


def _escape_excel_wildcards(value: str) -> str:
    return value.replace("~", "~~").replace("*", "~*").replace("?", "~?")
