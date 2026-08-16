"""跨处理任务共享的单元格编辑解析、值转换与写入。

链式多步处理时，用户在临时结果上尚未保存的单元格编辑会以 `edits`
载荷随下一个处理任务一起烘焙进新的临时文件；应用生成版本时同样
复用该格式。坐标统一为 0 基行列。
"""

import os
import re
from datetime import date
from pathlib import Path
from typing import Protocol

from openpyxl import load_workbook

_INTEGER_PATTERN = re.compile(r"[-+]?(?:0|[1-9]\d*)")
_NUMBER_PATTERN = re.compile(r"[-+]?(?:0|[1-9]\d*)\.\d+(?:[eE][-+]?\d+)?")
_DATE_PATTERN = re.compile(r"\d{4}-\d{2}-\d{2}")


class CellEditsContext(Protocol):
    def report_progress(self, progress: float | None, message: str = "") -> None: ...

    def check_cancelled(self) -> None: ...


def excel_edit_value(value: object) -> object:
    """把编辑输入转换为 Excel 单元格值（数字、日期、布尔、公式或文本）。"""
    if not isinstance(value, str):
        return value
    if value == "":
        return None
    if value.startswith("="):
        return value
    if _INTEGER_PATTERN.fullmatch(value):
        return int(value)
    if _NUMBER_PATTERN.fullmatch(value):
        return float(value)
    if _DATE_PATTERN.fullmatch(value):
        try:
            return date.fromisoformat(value)
        except ValueError:
            return value
    if value.upper() == "TRUE":
        return True
    if value.upper() == "FALSE":
        return False
    return value


def parse_optional_edits(value: object) -> list[tuple[str, int, int, object]]:
    """解析可选的 edits 载荷；缺失或为空时返回空列表。"""
    if value is None:
        return []
    if not isinstance(value, list):
        raise ValueError("任务参数 edits 必须是数组")
    if not value:
        return []
    return parse_edits(value)


def parse_edits(value: list[object]) -> list[tuple[str, int, int, object]]:
    """解析非空 edits 载荷为 (工作表, 行, 列, 值) 元组列表。"""
    edits: list[tuple[str, int, int, object]] = []
    for item in value:
        if not isinstance(item, dict):
            raise ValueError("任务参数 edits 必须是对象数组")
        sheet_name = item.get("sheet_name")
        row = item.get("row")
        column = item.get("column")
        if not isinstance(sheet_name, str) or not sheet_name:
            raise ValueError("编辑项缺少工作表名称：sheet_name")
        if not isinstance(row, int) or isinstance(row, bool) or row < 0:
            raise ValueError("编辑项 row 必须是大于等于 0 的整数")
        if not isinstance(column, int) or isinstance(column, bool) or column < 0:
            raise ValueError("编辑项 column 必须是大于等于 0 的整数")
        if "value" not in item:
            raise ValueError("编辑项缺少 value")
        edits.append((sheet_name, row, column, item["value"]))
    return edits


def write_edits(
    workbook_path: Path,
    edits: list[tuple[str, int, int, object]],
    context: CellEditsContext,
    *,
    progress_start: float,
    progress_end: float,
    progress_message: str,
) -> None:
    """把编辑就地写入工作簿文件（先写临时文件再原子替换）。"""
    if not edits:
        return
    workbook = load_workbook(workbook_path, data_only=False)
    temporary_path = workbook_path.with_name(f".{workbook_path.stem}.edits.xlsx")
    try:
        span = progress_end - progress_start
        for index, edit in enumerate(edits):
            context.check_cancelled()
            sheet_name, row, column, value = edit
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"找不到工作表：{sheet_name}")
            cell = workbook[sheet_name].cell(row=row + 1, column=column + 1)
            cell.value = excel_edit_value(value)  # type: ignore[assignment]
            context.report_progress(
                progress_start + span * ((index + 1) / len(edits)),
                f"{progress_message} {sheet_name}!{row + 1}",
            )
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        workbook.save(temporary_path)
    finally:
        workbook.close()
    try:
        context.check_cancelled()
        os.replace(temporary_path, workbook_path)
    finally:
        temporary_path.unlink(missing_ok=True)
