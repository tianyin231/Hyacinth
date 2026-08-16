"""Python 安全删除空白行临时预览任务。"""

import os
from contextlib import AbstractContextManager
from dataclasses import dataclass
from pathlib import Path
from typing import Protocol

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from hyacinth.excel.contracts import EngineName
from hyacinth.processing.sort_preview import (
    _CellData,
    _copy_file,
    _file_hash,
    _payload_path,
    _payload_string,
    _read_used_rows,
    _validate_preview_workbook,
    _write_data_rows,
    bake_pending_edits,
)
from hyacinth.tasks import TaskRequest
from hyacinth.tasks.worker import TaskContext, TaskHandler

DELETE_BLANK_ROWS_PREVIEW_OPERATION = "delete-blank-rows-preview"


@dataclass(frozen=True, slots=True)
class DeleteBlankRowsPreviewResult:
    preview_path: Path
    source_path: Path
    parent_version_id: str
    sheet_name: str
    key_columns: tuple[int, ...]
    allow_unsafe: bool
    compatibility_warning: bool
    deleted_row_numbers: tuple[int, ...]
    content_hash: str
    data_rows: int
    engine: EngineName = EngineName.PYTHON

    @property
    def deleted_rows(self) -> int:
        return len(self.deleted_row_numbers)


class DeleteBlankRowsTaskContext(Protocol):
    def report_progress(self, progress: float | None, message: str = "") -> None: ...

    def check_cancelled(self) -> None: ...

    def set_engine(self, engine: EngineName) -> None: ...

    def commit(self) -> None: ...

    def critical_section(self, message: str = "") -> AbstractContextManager[None]: ...


def run_delete_blank_rows_preview_task(
    request: TaskRequest,
    context: DeleteBlankRowsTaskContext,
) -> DeleteBlankRowsPreviewResult:
    source_path = _payload_path(request, "source_path")
    preview_path = _payload_path(request, "preview_path")
    parent_version_id = _payload_string(request, "parent_version_id")
    sheet_name = _payload_string(request, "sheet_name")
    key_columns = _parse_key_columns(request)
    allow_unsafe = _payload_bool(request, "allow_unsafe")
    if preview_path == source_path:
        raise ValueError("预览输出路径不能与源文件相同")

    context.check_cancelled()
    context.set_engine(EngineName.PYTHON)
    preview_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = preview_path.with_name(f".{preview_path.stem}.tmp.xlsx")
    temporary_path.unlink(missing_ok=True)
    try:
        context.report_progress(None, "正在复制源工作簿")
        _copy_file(source_path, temporary_path, context)
        context.check_cancelled()
        bake_pending_edits(temporary_path, request, context)
        context.report_progress(0.3, f"正在检查工作表 {sheet_name} 的空白行")
        deleted_rows, data_rows, compatibility_warning = _delete_blank_rows_copy(
            temporary_path,
            sheet_name,
            key_columns,
            allow_unsafe,
            context,
        )
        if not deleted_rows:
            raise ValueError("未发现空白行，无需生成预览")
        context.check_cancelled()
        context.report_progress(0.8, "正在校验临时结果")
        _validate_preview_workbook(temporary_path)
        context.check_cancelled()
        content_hash = _file_hash(temporary_path, context)
        with context.critical_section("正在安全完成删除空白行预览"):
            context.commit()
            os.replace(temporary_path, preview_path)
    finally:
        temporary_path.unlink(missing_ok=True)

    context.report_progress(1.0, f"预览已就绪，预计删除 {len(deleted_rows)} 行")
    return DeleteBlankRowsPreviewResult(
        preview_path=preview_path,
        source_path=source_path,
        parent_version_id=parent_version_id,
        sheet_name=sheet_name,
        key_columns=key_columns,
        allow_unsafe=allow_unsafe,
        compatibility_warning=compatibility_warning,
        deleted_row_numbers=deleted_rows,
        content_hash=content_hash,
        data_rows=data_rows,
    )


def delete_blank_rows_preview_task(request: TaskRequest, context: TaskContext) -> object:
    return run_delete_blank_rows_preview_task(request, context)


def delete_blank_rows_preview_handlers() -> dict[str, TaskHandler]:
    return {DELETE_BLANK_ROWS_PREVIEW_OPERATION: delete_blank_rows_preview_task}


def _delete_blank_rows_copy(
    path: Path,
    sheet_name: str,
    key_columns: tuple[int, ...],
    allow_unsafe: bool,
    context: DeleteBlankRowsTaskContext,
) -> tuple[tuple[int, ...], int, bool]:
    workbook = load_workbook(path, data_only=False)
    try:
        try:
            worksheet = workbook[sheet_name]
        except KeyError as error:
            raise ValueError(f"找不到工作表：{sheet_name}") from error
        rows = _read_data_region(worksheet, context)
        if len(rows) < 2:
            return (), 0, False
        used_columns = len(rows[0])
        effective_columns = key_columns or tuple(range(used_columns))
        _reject_key_columns_out_of_range(effective_columns, used_columns)
        data_rows = rows[1:]
        deleted_indexes = tuple(
            index
            for index, row in enumerate(data_rows)
            if all(_is_blank(row[column].value) for column in effective_columns)
        )
        if not deleted_indexes:
            return (), len(data_rows), False
        deleted_row_numbers = tuple(index + 2 for index in deleted_indexes)
        _reject_merged_cells_crossing_deleted_rows(worksheet, deleted_row_numbers)
        impacts = _structural_impacts(workbook, worksheet, rows)
        if impacts and not allow_unsafe:
            raise ValueError(
                f"发现可能受影响的结构：{'、'.join(impacts)}；请检查后开启兼容模式继续生成预览"
            )
        deleted_set = set(deleted_indexes)
        retained_rows = [row for index, row in enumerate(data_rows) if index not in deleted_set]
        _write_data_rows(worksheet, retained_rows)
        worksheet.delete_rows(2 + len(retained_rows), len(deleted_indexes))
        workbook.save(path)
        return deleted_row_numbers, len(data_rows), bool(impacts)
    finally:
        workbook.close()


def _read_data_region(
    worksheet: Worksheet,
    context: DeleteBlankRowsTaskContext,
) -> list[list[_CellData]]:
    max_rows = worksheet.max_row or 0
    max_columns = worksheet.max_column or 0
    if max_rows < 1 or max_columns < 1:
        return []
    rows = _read_used_rows(worksheet, max_rows, max_columns, context)
    last_row = 0
    last_column = 0
    for row_index, row in enumerate(rows, start=1):
        for column_index, cell in enumerate(row, start=1):
            if cell.value is not None and cell.value != "":
                last_row = max(last_row, row_index)
                last_column = max(last_column, column_index)
    if last_row == 0 or last_column == 0:
        return []
    return [row[:last_column] for row in rows[:last_row]]


def _is_blank(value: object) -> bool:
    if value is None or value == "":
        return True
    return isinstance(value, str) and not value.strip()


def _structural_impacts(
    workbook: object,
    worksheet: Worksheet,
    rows: list[list[_CellData]],
) -> tuple[str, ...]:
    impacts: list[str] = []
    if any(
        isinstance(cell.value, str) and cell.value.startswith("=") for row in rows for cell in row
    ):
        impacts.append("公式")
    if worksheet.tables:
        impacts.append("Excel 表格")
    if getattr(worksheet, "_charts", ()):
        impacts.append("图表")
    if getattr(workbook, "defined_names", ()):
        impacts.append("命名区域")
    return tuple(impacts)


def _reject_merged_cells_crossing_deleted_rows(
    worksheet: Worksheet,
    deleted_row_numbers: tuple[int, ...],
) -> None:
    for merged in worksheet.merged_cells.ranges:
        if any(merged.min_row <= row <= merged.max_row for row in deleted_row_numbers):
            raise ValueError(f"合并单元格 {merged} 跨越待删除行，无法安全删除")


def _reject_key_columns_out_of_range(
    key_columns: tuple[int, ...],
    used_columns: int,
) -> None:
    for column in key_columns:
        if column >= used_columns:
            raise ValueError(f"空白判断关键列 {column + 1} 超出工作表范围")


def _parse_key_columns(request: TaskRequest) -> tuple[int, ...]:
    raw = request.payload.get("key_columns", [])
    if not isinstance(raw, list):
        raise ValueError("空白判断关键列 key_columns 必须是数组")
    columns: list[int] = []
    for value in raw:
        if not isinstance(value, int) or isinstance(value, bool) or value < 0:
            raise ValueError("空白判断关键列必须是大于等于 0 的整数")
        columns.append(value)
    if len(set(columns)) != len(columns):
        raise ValueError("空白判断关键列不能重复")
    return tuple(columns)


def _payload_bool(request: TaskRequest, key: str) -> bool:
    value = request.payload.get(key, False)
    if not isinstance(value, bool):
        raise ValueError(f"任务参数 {key} 必须是布尔值")
    return value
