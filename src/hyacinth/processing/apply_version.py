import json
import os
import re
import shutil
from collections.abc import Callable
from contextlib import AbstractContextManager
from datetime import UTC, date, datetime
from hashlib import sha256
from pathlib import Path
from typing import Protocol

from openpyxl import load_workbook

from hyacinth.excel.contracts import EngineName
from hyacinth.tasks import TaskRequest
from hyacinth.tasks.worker import TaskContext, TaskHandler
from hyacinth.versioning import (
    ImportedWorkbook,
    MetadataStore,
    VersionRecord,
    write_recovery_manifest,
)

APPLY_SORT_PREVIEW_OPERATION = "apply-sort-preview"
APPLY_DEDUPLICATE_PREVIEW_OPERATION = "apply-deduplicate-preview"
APPLY_DELETE_BLANK_ROWS_PREVIEW_OPERATION = "apply-delete-blank-rows-preview"
APPLY_FILTER_PREVIEW_OPERATION = "apply-filter-preview"
SAVE_MANUAL_EDITS_OPERATION = "save-manual-edits"
COPY_CHUNK_SIZE = 1024 * 1024


class ApplyVersionTaskContext(Protocol):
    def report_progress(self, progress: float | None, message: str = "") -> None: ...

    def check_cancelled(self) -> None: ...

    def set_engine(self, engine: EngineName) -> None: ...

    def commit(self) -> None: ...

    def critical_section(self, message: str = "") -> AbstractContextManager[None]: ...


class ApplyMetadataStore(Protocol):
    def get_workbook(self, file_id: str) -> ImportedWorkbook: ...

    def record_child_version(
        self,
        version: VersionRecord,
        expected_parent_version_id: str,
    ) -> None: ...


MetadataStoreFactory = Callable[[Path], ApplyMetadataStore]


def run_apply_sort_preview_task(
    request: TaskRequest,
    context: ApplyVersionTaskContext,
    *,
    metadata_store_factory: MetadataStoreFactory = MetadataStore,
) -> ImportedWorkbook:
    sheet_name = _payload_string(request, "sheet_name")
    sort_keys = request.payload.get("sort_keys")
    if not isinstance(sort_keys, list) or not sort_keys:
        raise ValueError("任务参数缺少排序键：sort_keys")
    return _run_apply_preview_task(
        request,
        context,
        metadata_store_factory=metadata_store_factory,
        version_name="多列排序",
        operation="sort",
        parameters={"sheet_name": sheet_name, "sort_keys": sort_keys},
    )


def run_apply_deduplicate_preview_task(
    request: TaskRequest,
    context: ApplyVersionTaskContext,
    *,
    metadata_store_factory: MetadataStoreFactory = MetadataStore,
) -> ImportedWorkbook:
    sheet_name = _payload_string(request, "sheet_name")
    key_columns = request.payload.get("key_columns")
    keep = request.payload.get("keep")
    ignore_case = request.payload.get("ignore_case")
    trim_whitespace = request.payload.get("trim_whitespace")
    duplicate_groups = request.payload.get("duplicate_groups")
    deleted_rows = request.payload.get("deleted_rows")
    if not isinstance(key_columns, list):
        raise ValueError("任务参数缺少去重关键列：key_columns")
    if keep not in {"first", "last"}:
        raise ValueError("任务参数 keep 必须为 first 或 last")
    if not isinstance(ignore_case, bool) or not isinstance(trim_whitespace, bool):
        raise ValueError("去重文本选项必须是布尔值")
    if not isinstance(duplicate_groups, int) or isinstance(duplicate_groups, bool):
        raise ValueError("任务参数 duplicate_groups 必须是整数")
    if not isinstance(deleted_rows, int) or isinstance(deleted_rows, bool) or deleted_rows < 1:
        raise ValueError("任务参数 deleted_rows 必须是大于 0 的整数")
    return _run_apply_preview_task(
        request,
        context,
        metadata_store_factory=metadata_store_factory,
        version_name="删除重复行",
        operation="delete-duplicates",
        parameters={
            "sheet_name": sheet_name,
            "key_columns": key_columns,
            "keep": keep,
            "ignore_case": ignore_case,
            "trim_whitespace": trim_whitespace,
            "duplicate_groups": duplicate_groups,
            "deleted_rows": deleted_rows,
        },
    )


def run_apply_delete_blank_rows_preview_task(
    request: TaskRequest,
    context: ApplyVersionTaskContext,
    *,
    metadata_store_factory: MetadataStoreFactory = MetadataStore,
) -> ImportedWorkbook:
    sheet_name = _payload_string(request, "sheet_name")
    key_columns = request.payload.get("key_columns")
    allow_unsafe = request.payload.get("allow_unsafe")
    compatibility_warning = request.payload.get("compatibility_warning")
    deleted_row_numbers = request.payload.get("deleted_row_numbers")
    if not isinstance(key_columns, list):
        raise ValueError("任务参数缺少空白行关键列：key_columns")
    if not isinstance(allow_unsafe, bool):
        raise ValueError("任务参数 allow_unsafe 必须是布尔值")
    if not isinstance(compatibility_warning, bool):
        raise ValueError("任务参数 compatibility_warning 必须是布尔值")
    if (
        not isinstance(deleted_row_numbers, list)
        or not deleted_row_numbers
        or any(not isinstance(row, int) or isinstance(row, bool) for row in deleted_row_numbers)
    ):
        raise ValueError("任务参数 deleted_row_numbers 必须是非空整数列表")
    return _run_apply_preview_task(
        request,
        context,
        metadata_store_factory=metadata_store_factory,
        version_name="删除空白行",
        operation="delete-blank-rows",
        parameters={
            "sheet_name": sheet_name,
            "key_columns": key_columns,
            "allow_unsafe": allow_unsafe,
            "compatibility_warning": compatibility_warning,
            "deleted_row_numbers": deleted_row_numbers,
            "deleted_rows": len(deleted_row_numbers),
        },
    )


def run_apply_filter_preview_task(
    request: TaskRequest,
    context: ApplyVersionTaskContext,
    *,
    metadata_store_factory: MetadataStoreFactory = MetadataStore,
) -> ImportedWorkbook:
    sheet_name = _payload_string(request, "sheet_name")
    conditions = request.payload.get("conditions")
    connector = request.payload.get("connector")
    matched_rows = request.payload.get("matched_rows")
    total_rows = request.payload.get("total_rows")
    if not isinstance(conditions, list) or not conditions:
        raise ValueError("任务参数 conditions 必须是非空数组")
    if connector not in {"and", "or"}:
        raise ValueError("任务参数 connector 必须为 and 或 or")
    if not isinstance(matched_rows, int) or isinstance(matched_rows, bool) or matched_rows < 0:
        raise ValueError("任务参数 matched_rows 必须是大于等于 0 的整数")
    if (
        not isinstance(total_rows, int)
        or isinstance(total_rows, bool)
        or total_rows < 1
        or matched_rows > total_rows
    ):
        raise ValueError("任务参数 total_rows 必须是不小于匹配数的正整数")
    return _run_apply_preview_task(
        request,
        context,
        metadata_store_factory=metadata_store_factory,
        version_name="条件筛选",
        operation="filter",
        parameters={
            "sheet_name": sheet_name,
            "conditions": conditions,
            "connector": connector,
            "matched_rows": matched_rows,
            "total_rows": total_rows,
        },
    )


def run_save_manual_edits_task(
    request: TaskRequest,
    context: ApplyVersionTaskContext,
    *,
    metadata_store_factory: MetadataStoreFactory = MetadataStore,
) -> ImportedWorkbook:
    library_root = _payload_path(request, "library_root")
    parent_version_id = _payload_string(request, "parent_version_id")
    version_id = _payload_string(request, "version_id")
    edits = _manual_edits(request.payload.get("edits"))
    context.set_engine(EngineName.PYTHON)
    context.check_cancelled()
    workbook_record = metadata_store_factory(library_root).get_workbook(request.file_id)
    parent = workbook_record.head_version
    if parent is None or parent.version_id != parent_version_id:
        raise ValueError("当前 HEAD 已变化，请重新编辑")

    temporary_directory = library_root / ".staging" / f"{request.file_id}-{request.task_id}-manual"
    temporary_path = temporary_directory / "edited.xlsx"
    try:
        context.report_progress(0.1, "正在应用单元格修改")
        temporary_directory.mkdir(parents=True, exist_ok=True)
        workbook = load_workbook(parent.snapshot_path, data_only=False)
        try:
            for index, edit in enumerate(edits):
                context.check_cancelled()
                sheet_name, row, column, value = edit
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"找不到工作表：{sheet_name}")
                cell = workbook[sheet_name].cell(row=row + 1, column=column + 1)
                cell.value = _excel_edit_value(value)  # type: ignore[assignment]
                context.report_progress(
                    0.1 + 0.35 * ((index + 1) / len(edits)),
                    f"正在修改 {sheet_name}!{row + 1}",
                )
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True
            workbook.calculation.calcMode = "auto"
            workbook.save(temporary_path)
        finally:
            workbook.close()
        preview_hash = _file_hash(temporary_path, context)
        derived_request = TaskRequest(
            task_id=request.task_id,
            name=request.name,
            file_id=request.file_id,
            engine=request.engine,
            operation=request.operation,
            payload={
                "library_root": str(library_root),
                "preview_path": str(temporary_path),
                "preview_hash": preview_hash,
                "parent_version_id": parent_version_id,
                "version_id": version_id,
            },
        )
        return _run_apply_preview_task(
            derived_request,
            context,
            metadata_store_factory=metadata_store_factory,
            version_name="手动编辑",
            operation="manual-edit",
            parameters={"edited_cells": len(edits)},
        )
    finally:
        shutil.rmtree(temporary_directory, ignore_errors=True)


def _run_apply_preview_task(
    request: TaskRequest,
    context: ApplyVersionTaskContext,
    *,
    metadata_store_factory: MetadataStoreFactory,
    version_name: str,
    operation: str,
    parameters: dict[str, object],
) -> ImportedWorkbook:
    library_root = _payload_path(request, "library_root")
    preview_path = _payload_path(request, "preview_path")
    preview_hash = _payload_string(request, "preview_hash")
    parent_version_id = _payload_string(request, "parent_version_id")
    version_id = _payload_string(request, "version_id")

    context.set_engine(EngineName.PYTHON)
    context.check_cancelled()
    store = metadata_store_factory(library_root)
    workbook_record = store.get_workbook(request.file_id)
    parent = workbook_record.head_version
    if parent is None or parent.version_id != parent_version_id:
        raise ValueError("当前 HEAD 已变化，请重新生成预览")

    actual_hash = _file_hash(preview_path, context)
    if actual_hash != preview_hash:
        raise ValueError("临时预览已变化，请重新生成预览")
    if actual_hash == parent.content_hash:
        raise ValueError("处理结果没有变化，无需生成新版本")
    preview_workbook = load_workbook(preview_path, read_only=True)
    preview_workbook.close()

    staging_directory = library_root / ".staging" / f"{request.file_id}-{request.task_id}"
    final_directory = library_root / "files" / request.file_id / "versions" / version_id
    if final_directory.exists():
        raise ValueError("目标版本已存在，请刷新版本记录")
    staging_snapshot = staging_directory / "snapshot.xlsx"
    final_snapshot = final_directory / "snapshot.xlsx"
    working_temporary = workbook_record.working_path.with_name(
        f".{workbook_record.working_path.name}.{request.task_id}.tmp"
    )
    child = VersionRecord(
        version_id=version_id,
        file_id=request.file_id,
        parent_version_id=parent_version_id,
        name=version_name,
        created_at=datetime.now(UTC),
        operation=operation,
        engine=EngineName.PYTHON,
        snapshot_path=final_snapshot,
        content_hash=actual_hash,
        parameters_json=json.dumps(
            parameters,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ),
    )
    result = ImportedWorkbook(
        file_id=workbook_record.file_id,
        display_name=workbook_record.display_name,
        original_path=workbook_record.original_path,
        working_path=workbook_record.working_path,
        root_version=child,
    )

    try:
        context.report_progress(0.2, "正在准备子版本快照")
        staging_snapshot.parent.mkdir(parents=True, exist_ok=True)
        _copy_file(preview_path, staging_snapshot, context)
        if _file_hash(staging_snapshot, context) != actual_hash:
            raise RuntimeError("子版本快照校验失败")
        write_recovery_manifest(staging_directory / "manifest.json", library_root, result)
        _copy_file(preview_path, working_temporary, context)
        context.check_cancelled()
        with context.critical_section("正在安全提交子版本"):
            context.commit()
            os.replace(staging_directory, final_directory)
            store.record_child_version(child, parent_version_id)
            os.replace(working_temporary, workbook_record.working_path)
    finally:
        shutil.rmtree(staging_directory, ignore_errors=True)
        working_temporary.unlink(missing_ok=True)

    context.report_progress(1.0, "新版本已创建")
    return result


def apply_sort_preview_task(request: TaskRequest, context: TaskContext) -> object:
    return run_apply_sort_preview_task(request, context)


def apply_deduplicate_preview_task(request: TaskRequest, context: TaskContext) -> object:
    return run_apply_deduplicate_preview_task(request, context)


def apply_delete_blank_rows_preview_task(request: TaskRequest, context: TaskContext) -> object:
    return run_apply_delete_blank_rows_preview_task(request, context)


def apply_filter_preview_task(request: TaskRequest, context: TaskContext) -> object:
    return run_apply_filter_preview_task(request, context)


def save_manual_edits_task(request: TaskRequest, context: TaskContext) -> object:
    return run_save_manual_edits_task(request, context)


def apply_version_handlers() -> dict[str, TaskHandler]:
    return {
        APPLY_SORT_PREVIEW_OPERATION: apply_sort_preview_task,
        APPLY_DEDUPLICATE_PREVIEW_OPERATION: apply_deduplicate_preview_task,
        APPLY_DELETE_BLANK_ROWS_PREVIEW_OPERATION: apply_delete_blank_rows_preview_task,
        APPLY_FILTER_PREVIEW_OPERATION: apply_filter_preview_task,
        SAVE_MANUAL_EDITS_OPERATION: save_manual_edits_task,
    }


def _copy_file(source: Path, destination: Path, context: ApplyVersionTaskContext) -> None:
    with source.open("rb") as source_file, destination.open("wb") as destination_file:
        while chunk := source_file.read(COPY_CHUNK_SIZE):
            context.check_cancelled()
            destination_file.write(chunk)
    context.check_cancelled()
    shutil.copystat(source, destination)


def _file_hash(path: Path, context: ApplyVersionTaskContext) -> str:
    digest = sha256()
    with path.open("rb") as source:
        while chunk := source.read(COPY_CHUNK_SIZE):
            context.check_cancelled()
            digest.update(chunk)
    context.check_cancelled()
    return digest.hexdigest()


def _payload_path(request: TaskRequest, key: str) -> Path:
    value = request.payload.get(key)
    if not isinstance(value, str) or not value:
        raise ValueError(f"任务参数缺少路径：{key}")
    return Path(value)


def _payload_string(request: TaskRequest, key: str) -> str:
    value = request.payload.get(key)
    if not isinstance(value, str) or not value:
        raise ValueError(f"任务参数缺少：{key}")
    return value


def _manual_edits(value: object) -> list[tuple[str, int, int, object]]:
    if not isinstance(value, list) or not value:
        raise ValueError("任务参数 edits 必须是非空数组")
    edits: list[tuple[str, int, int, object]] = []
    for item in value:
        if not isinstance(item, dict):
            raise ValueError("单元格修改必须是对象")
        sheet_name = item.get("sheet_name")
        row = item.get("row")
        column = item.get("column")
        cell_value = item.get("value")
        if not isinstance(sheet_name, str) or not sheet_name:
            raise ValueError("单元格修改缺少工作表名称")
        if not isinstance(row, int) or isinstance(row, bool) or row < 0:
            raise ValueError("单元格行号必须是非负整数")
        if not isinstance(column, int) or isinstance(column, bool) or column < 0:
            raise ValueError("单元格列号必须是非负整数")
        if cell_value is not None and not isinstance(cell_value, (str, int, float, bool)):
            raise ValueError("单元格值必须是文本、数字、布尔值或空值")
        edits.append((sheet_name, row, column, cell_value))
    return edits


_INTEGER_PATTERN = re.compile(r"[-+]?(?:0|[1-9]\d*)")
_NUMBER_PATTERN = re.compile(r"[-+]?(?:0|[1-9]\d*)\.\d+(?:[eE][-+]?\d+)?")
_DATE_PATTERN = re.compile(r"\d{4}-\d{2}-\d{2}")


def _excel_edit_value(value: object) -> object:
    if not isinstance(value, str):
        return value
    if value == "":
        return None
    if value.startswith("="):
        return value
    if _INTEGER_PATTERN.fullmatch(value):
        return int(value)
    if _NUMBER_PATTERN.fullmatch(value):
        return float(value)
    if _DATE_PATTERN.fullmatch(value):
        try:
            return date.fromisoformat(value)
        except ValueError:
            return value
    if value.upper() == "TRUE":
        return True
    if value.upper() == "FALSE":
        return False
    return value
