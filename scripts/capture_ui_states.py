import argparse
import tempfile
from collections.abc import Iterator
from contextlib import contextmanager
from datetime import UTC, datetime
from hashlib import sha256
from pathlib import Path

from openpyxl import Workbook
from PySide6.QtCore import QEventLoop, QPoint, Qt, QTimer
from PySide6.QtTest import QTest
from PySide6.QtWidgets import (
    QApplication,
    QCheckBox,
    QComboBox,
    QGraphicsProxyWidget,
    QGraphicsView,
    QLineEdit,
    QSplitter,
)

from hyacinth.app import ApplicationTaskQueue, create_main_window
from hyacinth.excel.contracts import EngineName
from hyacinth.preview import BUILD_PREVIEW_INDEX_OPERATION, run_preview_index_task
from hyacinth.tasks import TaskEvent, TaskRequest, TaskState
from hyacinth.ui import VersionTreePanel
from hyacinth.versioning import (
    VERSION_STORAGE_STATS_OPERATION,
    ImportedWorkbook,
    MetadataStore,
    VersionRecord,
    run_delete_version_task,
    run_version_storage_stats_task,
)


class CaptureTaskQueue(ApplicationTaskQueue):
    def __init__(self) -> None:
        self.events: list[TaskEvent] = []
        self.submitted: list[TaskRequest] = []

    def submit(self, request: TaskRequest) -> None:
        self.submitted.append(request)

    def poll_events(self) -> tuple[TaskEvent, ...]:
        events = tuple(self.events)
        self.events.clear()
        return events

    def cancel(self, task_id: str) -> bool:
        return True

    def shutdown(self, timeout: float = 1.0) -> bool:
        return True


class CaptureTaskContext:
    def report_progress(self, progress: float | None, message: str = "") -> None:
        return

    def check_cancelled(self) -> None:
        return

    def set_engine(self, engine: EngineName) -> None:
        return

    def commit(self) -> None:
        return

    @contextmanager
    def critical_section(self, message: str = "") -> Iterator[None]:
        yield


def _preview_request(task_queue: CaptureTaskQueue) -> TaskRequest:
    return next(
        request
        for request in task_queue.submitted
        if request.operation == BUILD_PREVIEW_INDEX_OPERATION
    )


def _run_storage_stats(task_queue: CaptureTaskQueue) -> None:
    for request in task_queue.submitted:
        if request.operation != VERSION_STORAGE_STATS_OPERATION:
            continue
        task_queue.events.append(
            TaskEvent(
                request.task_id,
                TaskState.SUCCEEDED,
                request.name,
                request.file_id,
                None,
                result=run_version_storage_stats_task(request, CaptureTaskContext()),
            )
        )


def _seed_workbook(library_root: Path) -> None:
    directory = library_root / "files" / "visual-file"
    original = directory / "original" / "销售报表.xlsx"
    working = directory / "working" / "current.xlsx"
    snapshot = directory / "versions" / "root-version" / "snapshot.xlsx"
    for path in (original, working, snapshot):
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "销售明细"
        sheet.append(["日期", "商品", "数量", "销售额"])
        sheet.append(["2026/08/12", "苹果", 12, 360])
        sheet.append(["2026/08/12", "蓝莓", 5, 240])
        sheet.append(["2026/08/13", "西瓜", 18, 482.4])
        workbook.save(path)
        workbook.close()
    version = VersionRecord(
        "root-version",
        "visual-file",
        None,
        "导入原始文件",
        datetime.now(UTC),
        "import",
        None,
        snapshot,
        sha256(snapshot.read_bytes()).hexdigest(),
    )
    MetadataStore(library_root).record_import(
        ImportedWorkbook(
            "visual-file",
            "销售报表.xlsx",
            original,
            working,
            version,
        )
    )


def _wait(milliseconds: int) -> None:
    loop = QEventLoop()
    QTimer.singleShot(milliseconds, loop.quit)
    loop.exec()


def _add_child_version(library_root: Path) -> None:
    store = MetadataStore(library_root)
    record = store.get_workbook("visual-file")
    root = record.head_version
    assert root is not None
    child_snapshot = (
        library_root / "files" / record.file_id / "versions" / "sorted-version" / "snapshot.xlsx"
    )
    child_snapshot.parent.mkdir(parents=True)
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "销售明细"
    sheet.append(["日期", "商品", "数量", "销售额"])
    sheet.append(["2026/08/12", "蓝莓", 5, 240])
    sheet.append(["2026/08/12", "苹果", 12, 360])
    sheet.append(["2026/08/13", "西瓜", 18, 482.4])
    workbook.save(child_snapshot)
    workbook.close()
    record.working_path.write_bytes(child_snapshot.read_bytes())
    child = VersionRecord(
        "sorted-version",
        record.file_id,
        root.version_id,
        "多列排序",
        datetime.now(UTC),
        "sort",
        EngineName.PYTHON,
        child_snapshot,
        sha256(child_snapshot.read_bytes()).hexdigest(),
    )
    store.record_child_version(child, root.version_id)


def capture_ui_states(
    output_directory: Path,
) -> tuple[Path, ...]:
    app = QApplication.instance() or QApplication([])
    output_directory.mkdir(parents=True, exist_ok=True)
    with (
        tempfile.TemporaryDirectory(prefix="hyacinth-empty-") as empty_directory,
        tempfile.TemporaryDirectory(prefix="hyacinth-populated-") as populated_directory,
    ):
        empty_window = create_main_window(
            task_queue=CaptureTaskQueue(),
            library_root=Path(empty_directory),
        )
        empty_window.show()
        app.processEvents()
        empty_path = output_directory / "default-empty-state.png"
        if not empty_window.grab().save(str(empty_path)):
            raise RuntimeError("无法保存默认空状态截图")
        empty_window.close()

        populated_root = Path(populated_directory)
        _seed_workbook(populated_root)
        task_queue = CaptureTaskQueue()
        populated_window = create_main_window(
            task_queue=task_queue,
            library_root=populated_root,
        )
        populated_window.show()
        app.processEvents()
        request = _preview_request(task_queue)
        preview = run_preview_index_task(request, CaptureTaskContext())
        task_queue.events.append(
            TaskEvent(
                request.task_id,
                TaskState.SUCCEEDED,
                request.name,
                request.file_id,
                EngineName.PYTHON,
                result=preview,
            )
        )
        _run_storage_stats(task_queue)
        _wait(150)
        app.processEvents()
        populated_path = output_directory / "fluent-shell-populated.png"
        if not populated_window.grab().save(str(populated_path)):
            raise RuntimeError("无法保存有数据状态截图")
        operation = populated_window.findChild(QComboBox, "processing-operation")
        if operation is None:
            raise RuntimeError("找不到处理功能选择器")
        operation.setCurrentIndex(operation.findData("deduplicate"))
        app.processEvents()
        deduplicate_path = output_directory / "deduplicate-configured.png"
        if not populated_window.grab().save(str(deduplicate_path)):
            raise RuntimeError("无法保存删除重复行配置截图")
        operation.setCurrentIndex(operation.findData("delete_blank_rows"))
        app.processEvents()
        blank_rows_path = output_directory / "delete-blank-rows-configured.png"
        if not populated_window.grab().save(str(blank_rows_path)):
            raise RuntimeError("无法保存删除空白行配置截图")
        operation.setCurrentIndex(operation.findData("filter"))
        first_operator = populated_window.findChild(QComboBox, "filter-first-operator")
        first_value = populated_window.findChild(QLineEdit, "filter-first-value")
        enable_second = populated_window.findChild(QCheckBox, "filter-enable-second")
        second_column = populated_window.findChild(QComboBox, "filter-second-column")
        second_type = populated_window.findChild(QComboBox, "filter-second-type")
        second_operator = populated_window.findChild(QComboBox, "filter-second-operator")
        second_value = populated_window.findChild(QLineEdit, "filter-second-value")
        if any(
            control is None
            for control in (
                first_operator,
                first_value,
                enable_second,
                second_column,
                second_type,
                second_operator,
                second_value,
            )
        ):
            raise RuntimeError("找不到条件筛选配置控件")
        assert first_operator is not None
        assert first_value is not None
        assert enable_second is not None
        assert second_column is not None
        assert second_type is not None
        assert second_operator is not None
        assert second_value is not None
        first_operator.setCurrentIndex(first_operator.findData("contains"))
        first_value.setText("苹果")
        enable_second.setChecked(True)
        second_column.setCurrentIndex(2)
        second_type.setCurrentIndex(second_type.findData("number"))
        second_operator.setCurrentIndex(second_operator.findData("greater_than"))
        second_value.setText("10")
        app.processEvents()
        filter_path = output_directory / "filter-configured.png"
        if not populated_window.grab().save(str(filter_path)):
            raise RuntimeError("无法保存条件筛选配置截图")
        populated_window.close()

        _add_child_version(populated_root)
        branch_queue = CaptureTaskQueue()
        branch_window = create_main_window(
            task_queue=branch_queue,
            library_root=populated_root,
            confirmation_presenter=lambda _parent, _title, _message: True,
        )
        branch_window.show()
        app.processEvents()
        splitter = branch_window.findChild(QSplitter, "main-workspace-splitter")
        if splitter is None:
            raise RuntimeError("找不到主界面分隔器")
        splitter.setSizes([260, 600, 580])
        initial_request = _preview_request(branch_queue)
        initial_preview = run_preview_index_task(initial_request, CaptureTaskContext())
        branch_queue.events.append(
            TaskEvent(
                initial_request.task_id,
                TaskState.SUCCEEDED,
                initial_request.name,
                initial_request.file_id,
                EngineName.PYTHON,
                result=initial_preview,
            )
        )
        _run_storage_stats(branch_queue)
        _wait(150)
        view = branch_window.findChild(QGraphicsView, "version-tree-view")
        if view is None:
            raise RuntimeError("找不到版本演化树")
        root_card = None
        for item in view.scene().items():
            if not isinstance(item, QGraphicsProxyWidget):
                continue
            widget = item.widget()
            if widget is not None and widget.property("version-id") == "root-version":
                root_card = widget
                break
        if root_card is None:
            raise RuntimeError("找不到根版本节点")
        QTest.mouseClick(root_card, Qt.MouseButton.LeftButton)
        historical_request = branch_queue.submitted[-1]
        historical_preview = run_preview_index_task(historical_request, CaptureTaskContext())
        branch_queue.events.append(
            TaskEvent(
                historical_request.task_id,
                TaskState.SUCCEEDED,
                historical_request.name,
                historical_request.file_id,
                EngineName.PYTHON,
                result=historical_preview,
            )
        )
        _run_storage_stats(branch_queue)
        _wait(150)
        branch_path = output_directory / "version-history-selected.png"
        if not branch_window.grab().save(str(branch_path)):
            raise RuntimeError("无法保存历史版本选中截图")
        center = root_card.rect().center()
        QTest.mousePress(root_card, Qt.MouseButton.LeftButton, pos=center)
        QTest.mouseMove(root_card, pos=center + QPoint(0, 160))
        QTest.mouseRelease(
            root_card,
            Qt.MouseButton.LeftButton,
            pos=center + QPoint(0, 160),
        )
        app.processEvents()
        dragged_path = output_directory / "version-node-dragged.png"
        if not branch_window.grab().save(str(dragged_path)):
            raise RuntimeError("无法保存版本节点拖动截图")
        tree_panel = branch_window.findChild(VersionTreePanel, "version-tree-panel")
        if tree_panel is None:
            raise RuntimeError("找不到版本演化树面板")
        tree_panel.version_delete_requested.emit("root-version")
        delete_request = branch_queue.submitted[-1]
        deleted_workbook = run_delete_version_task(delete_request, CaptureTaskContext())
        branch_queue.events.append(
            TaskEvent(
                delete_request.task_id,
                TaskState.SUCCEEDED,
                delete_request.name,
                delete_request.file_id,
                EngineName.PYTHON,
                result=deleted_workbook,
            )
        )
        _wait(150)
        app.processEvents()
        deleted_path = output_directory / "version-node-deleted.png"
        if not branch_window.grab().save(str(deleted_path)):
            raise RuntimeError("无法保存版本节点删除截图")
        branch_window.close()

        from hyacinth.ui import RecycleBinDialog

        recycle_store = MetadataStore(populated_root)
        recycle_record = recycle_store.get_workbook("visual-file")
        assert recycle_record.head_version is not None
        recycle_store.soft_delete_file("visual-file", recycle_record.head_version.version_id)
        recycle_window = create_main_window(
            task_queue=CaptureTaskQueue(),
            library_root=populated_root,
        )
        recycle_window.show()
        app.processEvents()
        recycle_window._open_recycle_bin()
        _wait(150)
        recycle_dialog = recycle_window.findChild(RecycleBinDialog, "recycle-bin-dialog")
        if recycle_dialog is None:
            raise RuntimeError("找不到回收站对话框")
        recycle_path = output_directory / "recycle-bin.png"
        if not recycle_dialog.grab().save(str(recycle_path)):
            raise RuntimeError("无法保存回收站截图")
        recycle_dialog.close()
        recycle_window.close()
    return (
        empty_path,
        populated_path,
        deduplicate_path,
        blank_rows_path,
        filter_path,
        recycle_path,
        branch_path,
        dragged_path,
        deleted_path,
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="抓取风信子默认、有数据与处理配置界面状态")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("output/playwright"),
        help="截图输出目录",
    )
    arguments = parser.parse_args()
    for path in capture_ui_states(arguments.output):
        print(path.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
